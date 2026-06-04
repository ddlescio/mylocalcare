import os

RUNTIME_SERVICE = os.getenv("RUNTIME_SERVICE", "web").strip().lower()
APP_RUNTIME_ROLE = "realtime" if RUNTIME_SERVICE == "chat" else "web"

from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, g, send_from_directory, abort
from whitenoise import WhiteNoise
import os
import sqlite3
import json
import uuid
import time
from flask_mail import Mail, Message
from email.utils import make_msgid
from flask_socketio import SocketIO, emit, join_room, leave_room, disconnect
from socketio import RedisManager
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv  # ✅ serve per leggere il file .env
# 🔐 Crittografia
import base64
from Crypto.Cipher import AES
from Crypto.Protocol.KDF import scrypt
from Crypto.Random import get_random_bytes
import threading
from models import calcola_media_recensioni, get_recensioni_utente
from models import (

    get_operatori, get_operatore_by_id,
    aggiungi_operatore, modifica_operatore, elimina_operatore, get_tutte_le_zone,
    get_utenti, attiva_utente, elimina_utente,
    chat_invia, chat_conversazione, chat_threads, chat_segna_letti, count_chat_non_letti,
    get_recensioni_utente, aggiungi_o_modifica_recensione, get_recensione_autore_vs_destinatario,
    calcola_media_recensioni, get_risposta_by_recensione, aggiungi_o_modifica_risposta, elimina_risposta,
    get_annunci_utente
)
from models import crea_notifica
from flask_login import login_required
from itsdangerous import URLSafeTimedSerializer, URLSafeSerializer, BadSignature, SignatureExpired
from datetime import datetime, timedelta, timezone
from services import (
    attiva_servizio,
    revoca_attivazione,
    servizio_attivo_per_annuncio,
    servizio_attivo_per_utente,
    aggiorna_servizi_scaduti,
)
import secrets
import stripe
import psycopg2
import psycopg2.extras
import psycopg2.pool as psycopg2_pool
import re
import unicodedata
import traceback
from models import fetchone_value
import os
from flask import g
from db import (insert_and_get_id)
from realtime import emit_update_notifications
from socket_registry import (
    configure_socket_registry,
    SOCKET_TTL_SECONDS,
    is_user_online,
    set_open_chat,
    get_open_chat,
    clear_open_chat,
    _bump_offline_token,
    _get_offline_token,
    _get_user_id_from_sid,
    _touch_socket_sid,
    _remove_socket_sid,
    _cleanup_user_socket_set,
    emit_to_user_sids,
    emit_to_user_room,
    _socket_user_set_key,
    _socket_sid_key,
    _socket_client_key,
)

from realtime_handlers import register_socket_lifecycle_handlers

from chat_realtime import register_chat_socket_handlers, typing_state, pagina_attiva
from realtime_auth import build_realtime_token
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from webauthn import (
    generate_registration_options,
    verify_registration_response,
    generate_authentication_options,
    verify_authentication_response,
    options_to_json,
    base64url_to_bytes,
)
from webauthn.helpers.structs import (
    AuthenticatorSelectionCriteria,
    AuthenticatorAttachment,
    UserVerificationRequirement,
    ResidentKeyRequirement,
    PublicKeyCredentialDescriptor,
)

from decimal import Decimal, InvalidOperation

# ==========================================================
# DB POOL (Postgres) + Connessione riutilizzabile per-request
# ==========================================================

_pg_pools = {}
_pg_pool_locks = {}
_pg_state_guard = threading.Lock()

def _get_pg_pid():
    return os.getpid()

def _get_pg_lock_for_pid(pid):
    with _pg_state_guard:
        lock = _pg_pool_locks.get(pid)
        if lock is None:
            lock = threading.RLock()
            _pg_pool_locks[pid] = lock
        return lock

def get_current_pg_pool():
    pid = _get_pg_pid()
    with _pg_state_guard:
        return _pg_pools.get(pid)

def _set_current_pg_pool(pool):
    pid = _get_pg_pid()
    with _pg_state_guard:
        _pg_pools[pid] = pool

def init_pg_pool():
    pid = _get_pg_pid()
    pool = get_current_pg_pool()

    if pool is not None:
        return pool

    dsn = os.getenv("DATABASE_URL")
    if not dsn:
        return None

    lock = _get_pg_lock_for_pid(pid)

    lock.acquire()

    try:
        pool = get_current_pg_pool()
        if pool is not None:
            return pool

        pool = psycopg2_pool.ThreadedConnectionPool(
            minconn=1,
            maxconn=12,
            dsn=dsn,
            connect_timeout=8,
            sslmode="require"
        )

        _set_current_pg_pool(pool)
        return pool

    except Exception:
        _set_current_pg_pool(None)
        raise

    finally:
        try:
            lock.release()
        except Exception:
            pass

def warm_pg_pool_for_current_process():
    dsn = os.getenv("DATABASE_URL")
    if not dsn:
        return

    pool = init_pg_pool()
    if pool is None:
        raise RuntimeError("PG warmup failed: pool None")

    raw = None
    try:
        raw = pool.getconn()

        raw.autocommit = True
        raw.set_session(readonly=False, autocommit=True)

        cur = raw.cursor()
        cur.execute("SELECT 1")
        cur.fetchone()
        cur.close()

    finally:
        if raw is not None:
            try:
                pool.putconn(raw)
            except Exception:
                pass

def get_cursor(conn):
    import sqlite3
    import psycopg2.extras

    # PostgreSQL (Render)
    if isinstance(conn, psycopg2.extensions.connection):
        return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # SQLite (locale)
    elif isinstance(conn, sqlite3.Connection):
        conn.row_factory = sqlite3.Row
        return conn.cursor()

    # fallback
    return conn.cursor()

def get_reset_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'])

# -- Helper per AES-GCM: salviamo ciphertext||tag in un solo campo base64 --
def gcm_pack(ciphertext: bytes, tag: bytes) -> str:
    return base64.b64encode(ciphertext + tag).decode()

def gcm_unpack(b64: str):
    raw = base64.b64decode(b64)
    if len(raw) < 16:
        raise ValueError("GCM blob troppo corto")
    return raw[:-16], raw[-16:]  # (ciphertext, tag)
# 🔐 Per cifratura end-to-end con X25519 (ECDH)
from nacl.public import PrivateKey, PublicKey

import base64
from Crypto.Cipher import AES



def encrypt_with_master(plaintext: bytes) -> tuple[str, str]:
    """
    Cifra bytes con MASTER_SECRET usando AES-GCM.
    Ritorna (ct+tag in base64 'impacchettato' con gcm_pack, nonce_base64).
    """
    cipher = AES.new(MASTER_SECRET, AES.MODE_GCM)
    ct, tag = cipher.encrypt_and_digest(plaintext)
    dek_enc_b64 = gcm_pack(ct, tag)
    nonce_b64 = base64.b64encode(cipher.nonce).decode()
    return dek_enc_b64, nonce_b64


def decrypt_with_master(enc_b64: str, nonce_b64: str) -> bytes:
    """
    Decifra quello che è stato cifrato con encrypt_with_master.
    """
    ct, tag = gcm_unpack(enc_b64)
    nonce = base64.b64decode(nonce_b64)
    cipher = AES.new(MASTER_SECRET, AES.MODE_GCM, nonce=nonce)
    return cipher.decrypt_and_verify(ct, tag)

def is_admin(user_id):
    conn = get_db_connection()

    c = get_cursor(conn)
    c.execute(
        sql("SELECT ruolo FROM utenti WHERE id=?"),
        (user_id,)
    )
    row = c.fetchone()

    return row and row["ruolo"] == "admin"

import requests
from openai import OpenAI

DAILY_BASE_URL = "https://api.daily.co/v1"


# ==========================================================
# 🤖 AI — AIUTO SCRITTURA ANNUNCI
# ==========================================================

def get_openai_client():
    """
    Crea il client OpenAI solo quando serve.
    La chiave deve stare in variabile ambiente OPENAI_API_KEY.
    """
    api_key = os.getenv("OPENAI_API_KEY", "").strip()

    if not api_key:
        raise RuntimeError("OPENAI_API_KEY non configurata")

    return OpenAI(api_key=api_key)


def crea_room_daily(nome_room: str):
    """
    Crea una room Daily tramite API.
    """
    url = f"{DAILY_BASE_URL}/rooms"

    headers = {
        "Authorization": f"Bearer {DAILY_API_KEY}",
        "Content-Type": "application/json"
    }

    payload = {
        "name": nome_room,
        "properties": {
            "enable_chat": True,
            "start_video_off": False,
            "start_audio_off": False,
            "exp": int(datetime.now().timestamp()) + 3600  # scade tra 1h
        }
    }

    response = requests.post(url, headers=headers, json=payload)

    if response.status_code != 200:
        raise RuntimeError(f"Errore creazione room Daily: {response.text}")

    return response.json()

# ==========================================================
# 🔐 SINCRONIZZAZIONE / GENERAZIONE CHIAVI X25519
# ==========================================================
def ensure_x25519_keys(user_id):
    """Verifica o genera le chiavi X25519 per l'utente (rigenera se mancano o sono incoerenti)"""
    conn = get_db_connection()

    c = get_cursor(conn)
    c.execute(sql("SELECT x25519_pub, x25519_priv_enc, x25519_priv_nonce FROM utenti WHERE id = ?"), (user_id,))
    row = c.fetchone()

    from Crypto.Cipher import AES
    import base64

    if not row or not row["x25519_pub"] or not row["x25519_priv_enc"] or not row["x25519_priv_nonce"]:
        security_log(
            "🔐 Rigenerazione chiavi X25519 utente",
            {"user_id": user_id}
        )
        from nacl.public import PrivateKey

        # Genera nuova coppia
        priv = PrivateKey.generate()
        pub = priv.public_key

        # Decifra DEK dalla sessione
        dek = base64.b64decode(session.get("dek_b64"))
        cipher = AES.new(dek, AES.MODE_GCM)
        priv_enc, tag = cipher.encrypt_and_digest(bytes(priv))
        nonce = cipher.nonce

        # Salva nel DB
        c.execute(sql("""
            UPDATE utenti
            SET x25519_pub = ?, x25519_priv_enc = ?, x25519_priv_nonce = ?
            WHERE id = ?
        """), (
            base64.b64encode(bytes(pub)).decode(),
            base64.b64encode(priv_enc + tag).decode(),
            base64.b64encode(nonce).decode(),
            user_id
        ))
        conn.commit()

        # Aggiorna sessione
        session["x25519_priv_b64"] = base64.b64encode(bytes(priv)).decode()
        session["x25519_pub_b64"] = base64.b64encode(bytes(pub)).decode()

    else:
        # Decifra chiave privata esistente e ricarica in sessione
        dek = base64.b64decode(session.get("dek_b64"))
        priv_enc_raw = base64.b64decode(row["x25519_priv_enc"])
        nonce = base64.b64decode(row["x25519_priv_nonce"])
        ct, tag = priv_enc_raw[:-16], priv_enc_raw[-16:]
        cipher = AES.new(dek, AES.MODE_GCM, nonce=nonce)
        priv_bytes = cipher.decrypt_and_verify(ct, tag)

        session["x25519_priv_b64"] = base64.b64encode(priv_bytes).decode()
        session["x25519_pub_b64"] = row["x25519_pub"]




# Slug -> chiave nel JSON + label umana
CATEGORY_MAP = {
    "operatori-benessere": ("operatori-benessere", "Operatori Benessere"),
    "operatori benessere": ("operatori-benessere", "Operatori Benessere"),

    "babysitter": ("babysitter", "Babysitter"),
    "pet-sitter": ("pet-sitter", "Pet-Sitter"),
    "petsitter": ("pet-sitter", "Pet-Sitter"),
    "caregiver": ("caregiver", "Caregiver"),
    "ripetizioni": ("ripetizioni", "Ripetizioni"),
    "aiuto-in-casa": ("aiuto-in-casa", "Aiuto in Casa"),

    "escursioni-sport": ("escursioni-sport", "Escursioni & Sport"),
    "biglietti-spettacoli": ("biglietti-spettacoli", "Biglietti Spettacoli"),
    "libri-scuola": ("libri-scuola", "Libri Scuola"),
    "caffe-parole": ("caffe-parole", "Caffè & Parole"),

    "family-kids": ("family-kids", "Family & Kids"),
    "eventi-socialita": ("eventi-socialita", "Eventi & Socialità"),
    "spazi-sale": ("spazi-sale", "Spazi & Sale"),
}

# =========================================
# MATCH: categorie (ordine = offro_1..offro_10 / cerco_1..cerco_10)
# =========================================
def to_slug(val: str) -> str:
    """Normalizza stringhe categoria in slug coerenti senza accenti."""
    if not val:
        return ""

    v = str(val).strip().lower()

    v = unicodedata.normalize("NFD", v)
    v = "".join(ch for ch in v if unicodedata.category(ch) != "Mn")

    v = v.replace("_", " ")
    v = v.replace("&", " ")
    v = re.sub(r"[^a-z0-9\s-]", " ", v)
    v = "-".join(v.split())

    return v



CATEGORIE_PREFERENZE = [
    "Operatori benessere",     # 1
    "Aiuto in casa",           # 2
    "Ripetizioni",             # 3
    "Babysitter",              # 4
    "Pet-sitter",              # 5
    "Caregiver",               # 6
    "Escursioni & Sport",      # 7
    "Biglietti spettacoli",    # 8
    "Libri scuola",            # 9
    "Caffe & parole",          # 10
    "Family & Kids",           # 11
    "Eventi & Socialità",      # 12
    "Spazi & Sale",            # 13
]

# slug -> indice 1..13
CATEGORIA_TO_INDEX = {to_slug(x): i+1 for i, x in enumerate(CATEGORIE_PREFERENZE)}

# Colonne preferenze utente usate per Offro/Cerco
PREFERENCE_COLUMNS = [
    f"{tipo}_{i}"
    for tipo in ("offro", "cerco")
    for i in range(1, len(CATEGORIE_PREFERENZE) + 1)
]

def get_utenti_profilo_incompleto():
    """
    Restituisce gli utenti attivi che non hanno selezionato
    nessuna preferenza Offro/Cerco.
    """

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        somma_colonne = " + ".join(
            [f"COALESCE({col},0)" for col in PREFERENCE_COLUMNS]
        )

        cur.execute(sql(f"""
            SELECT
                id,
                email,
                nome,
                username,
                email_notifiche
            FROM utenti
            WHERE attivo = 1
              AND sospeso = 0
              AND COALESCE(disattivato_admin,0) = 0
              AND ({somma_colonne}) = 0
        """))

        return cur.fetchall()

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def invia_reminder_profili_incompleti(dry_run=False):
    """
    Invia un reminder agli utenti attivi che non hanno ancora compilato
    nessuna preferenza Offro/Cerco.

    Canali:
    - notifica interna sempre
    - push se disponibile
    - email solo se email_notifiche = 1

    Evita doppioni se esiste già un reminder negli ultimi 7 giorni.
    """

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        utenti = get_utenti_profilo_incompleto()

        creati = 0
        saltati = 0
        push_inviate = 0
        email_inviate = 0

        for u in utenti:
            user_id = int(u["id"])

            if app.config.get("IS_POSTGRES"):
                cur.execute(sql("""
                    SELECT id
                    FROM notifiche
                    WHERE id_utente = ?
                      AND tipo = 'profilo_incompleto'
                      AND data >= CURRENT_TIMESTAMP - INTERVAL '7 days'
                    LIMIT 1
                """), (user_id,))
            else:
                cur.execute(sql("""
                    SELECT id
                    FROM notifiche
                    WHERE id_utente = ?
                      AND tipo = 'profilo_incompleto'
                      AND data >= datetime('now','-7 days')
                    LIMIT 1
                """), (user_id,))

            gia_inviata = cur.fetchone()

            if gia_inviata:
                saltati += 1
                continue

            titolo = "Completa il tuo profilo"
            messaggio = (
                "Seleziona cosa offri o cosa cerchi per ricevere annunci compatibili "
                "e suggerimenti personalizzati."
            )
            link = "/utente/dashboard"

            if not dry_run:
                cur.execute(sql("""
                    INSERT INTO notifiche (
                        id_utente,
                        titolo,
                        messaggio,
                        link,
                        tipo,
                        letta
                    )
                    VALUES (?, ?, ?, ?, ?, 0)
                """), (
                    user_id,
                    titolo,
                    messaggio,
                    link,
                    "profilo_incompleto"
                ))

                emit_update_notifications(user_id)

                try:
                    invia_push(
                        user_id,
                        titolo,
                        messaggio,
                        url=link
                    )
                    push_inviate += 1
                except Exception as e:
                    log_exception_safe(
                        "⚠️ Errore push reminder profilo incompleto",
                        e,
                        {"user_id": user_id},
                        production=True
                    )

                if int(u["email_notifiche"] or 0) == 1 and u["email"]:
                    nome = u["nome"] or u["username"] or "utente"
                    base_url = app.config.get("APP_BASE_URL", "https://www.mylocalcare.it").rstrip("/")

                    ok_email = _invia_email(
                        destinazione=u["email"],
                        oggetto="Completa il tuo profilo MyLocalCare",
                        corpo=(
                            f"Ciao {nome},\n\n"
                            "Hai creato il tuo account su MyLocalCare, ma non hai ancora indicato "
                            "cosa offri o cosa cerchi.\n\n"
                            "Completa il profilo per ricevere annunci compatibili, suggerimenti "
                            "personalizzati e notifiche più pertinenti.\n\n"
                            f"{base_url}{link}\n\n"
                            "MyLocalCare"
                        )
                    )

                    if ok_email:
                        email_inviate += 1

            creati += 1

        if not dry_run:
            conn.commit()

        return {
            "ok": True,
            "utenti_incompleti": len(utenti),
            "notifiche_create": creati,
            "saltati_per_recenti": saltati,
            "push_inviate": push_inviate,
            "email_inviate": email_inviate,
            "dry_run": dry_run
        }

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass

        log_exception_safe(
            "❌ Errore reminder profili incompleti",
            e,
            production=True
        )

        return {
            "ok": False,
            "error": str(e)
        }

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def norm_place(s: str) -> str:
    """Normalizza città/zona per confronto robusto."""
    if not s:
        return ""
    return " ".join(s.strip().lower().replace("-", " ").split())


def place_match(annuncio_zona: str, utente_citta: str) -> bool:
    """
    Match zona/città:
    True se uguali o uno contiene l'altro (case-insensitive, ripulito).
    """
    a = norm_place(annuncio_zona)
    u = norm_place(utente_citta)
    if not a or not u:
        return False
    return (a == u) or (a in u) or (u in a)


from datetime import datetime
import zoneinfo

# ==========================================================
# 1️⃣ CONFIGURAZIONE DI BASE E APP
# ==========================================================


from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")
load_dotenv(BASE_DIR / ".env.secrets")

MASTER_SECRET_HEX = os.getenv("MASTER_SECRET_KEY")
if not MASTER_SECRET_HEX:
    raise RuntimeError("MASTER_SECRET_KEY non trovata in .env / .env.secrets")

MASTER_SECRET = bytes.fromhex(MASTER_SECRET_HEX)
if len(MASTER_SECRET) != 32:
    raise RuntimeError("MASTER_SECRET_KEY deve essere lunga 32 byte (64 caratteri hex)")

DAILY_API_KEY = os.getenv("DAILY_API_KEY")
if not DAILY_API_KEY:
    raise RuntimeError("DAILY_API_KEY non trovata in .env.secrets")

app = Flask(__name__)
import json
from datetime import timedelta

app.config["APP_RUNTIME_ROLE"] = APP_RUNTIME_ROLE
app.config["IS_REALTIME_SERVER"] = (APP_RUNTIME_ROLE == "realtime")
app.config["SOCKET_BASE_URL"] = os.environ.get("SOCKET_BASE_URL", "").strip()

# ==========================================================
# 🔐 WEBAUTHN / PASSKEY CONFIG
# ==========================================================

# Dominio principale su cui vengono registrate/verificate le passkey.
# ATTENZIONE:
# - rp_id deve essere il dominio, senza https://
# - expected_origin deve essere l'origine completa con https://
WEBAUTHN_RP_ID = os.getenv("WEBAUTHN_RP_ID", "mylocalcare.it")
WEBAUTHN_RP_NAME = os.getenv("WEBAUTHN_RP_NAME", "MyLocalCare")
WEBAUTHN_EXPECTED_ORIGIN = os.getenv(
    "WEBAUTHN_EXPECTED_ORIGIN",
    "https://www.mylocalcare.it"
)

app.jinja_env.filters['from_json'] = lambda s: json.loads(s or "[]")

# ✅ 1) Chiave STABILE (niente fallback random, altrimenti dopo restart/redeploy ti slogga)
#    Su Render DEVI avere FLASK_SECRET_KEY valorizzata nelle env.
app.secret_key = os.environ["FLASK_SECRET_KEY"]

# ✅ 2) Configurazione cookie di sessione
# La configurazione completa della sessione Redis viene fatta più sotto,
# dopo l'inizializzazione di redis_client e Flask-Session.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_DOMAIN=".mylocalcare.it",
)

app.wsgi_app = WhiteNoise(app.wsgi_app, root="static/")

# ==========================================================
# 🔐 SECURITY HEADERS GLOBALI
# ==========================================================
@app.after_request
def apply_security_headers(response):
    """
    Header di sicurezza applicati a tutte le risposte HTTP.
    """

    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")

    response.headers.setdefault(
        "Permissions-Policy",
        "camera=(self), microphone=(self), geolocation=(), payment=(), usb=(), bluetooth=()"
    )

    response.headers.setdefault(
        "Strict-Transport-Security",
        "max-age=31536000; includeSubDomains"
    )

    return response

@app.before_request
def ensure_csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)


@app.context_processor
def inject_csrf_token():
    """
    Rende disponibile csrf_token() in tutti i template.
    Uso nei form:
    <input type="hidden" name="csrf_token" value="{{ csrf_token() }}">
    """
    return {
        "csrf_token": lambda: session.get("csrf_token", "")
    }


def verify_csrf():
    """
    Verifica CSRF compatibile con:
    - form HTML classici: csrf_token nel form
    - fetch/AJAX: X-CSRF-Token negli header
    - fetch JSON: X-CSRF-Token negli header
    """

    token = (
        request.headers.get("X-CSRF-Token")
        or request.headers.get("X-CSRFToken")
        or request.form.get("csrf_token")
    )

    if not token or token != session.get("csrf_token"):
        abort(403)

def privacy_debug(message, extra=None):
    """
    Log di debug sicuro:
    - in produzione non stampa nulla;
    - in locale/dev permette debug senza esporre dati personali nei log.
    """
    if os.getenv("APP_ENV", "production").lower() not in ("local", "development"):
        return

    if extra is not None:
        print(f"🧪 {message}: {extra}", flush=True)
    else:
        print(f"🧪 {message}", flush=True)

def safe_log(message, extra=None, *, level="info", production=False):
    """
    Logger centralizzato.

    Regola:
    - di default in produzione NON stampa dettagli;
    - in locale/development stampa anche extra;
    - se production=True stampa solo il messaggio essenziale, mai extra sensibili.
    """
    env = os.getenv("APP_ENV", "production").lower()

    if env not in ("local", "development"):
        if production:
            print(f"{message}", flush=True)
        return

    if extra is not None:
        print(f"{message}: {extra}", flush=True)
    else:
        print(f"{message}", flush=True)

# ==========================================================
# 🔐 SAFE LOGGING — REDAZIONE DATI SENSIBILI
# ==========================================================

SENSITIVE_LOG_KEYS = {
    "password",
    "conferma_password",
    "nuova_password",
    "password_attuale",
    "token",
    "csrf_token",
    "secret",
    "client_secret",
    "stripe_secret",
    "STRIPE_SECRET_KEY",
    "STRIPE_WEBHOOK_SECRET",
    "VAPID_PRIVATE_KEY",
    "FLASK_SECRET_KEY",
    "MASTER_SECRET_KEY",
    "DATABASE_URL",
    "MAIL_PASSWORD",
    "authorization",
    "cookie",
    "session",
    "dek_b64",
    "x25519_priv_b64",
    "x25519_priv_enc",
    "x25519_priv_nonce",
    "credential_public_key",
    "credential_id",
    "code",
    "code_hash",
}

PARTIAL_MASK_KEYS = {
    "email",
    "endpoint",
    "user_agent",
    "ip",
    "riferimento_esterno",
    "payment_intent",
    "acquisto_id",
}


def _mask_email_for_log(value):
    value = str(value or "")
    if "@" not in value:
        return "***"

    name, domain = value.split("@", 1)

    if len(name) <= 2:
        masked_name = name[:1] + "***"
    else:
        masked_name = name[:2] + "***"

    return f"{masked_name}@{domain}"


def _mask_text_for_log(value, keep_start=6, keep_end=4):
    value = str(value or "")

    if len(value) <= keep_start + keep_end:
        return "***"

    return f"{value[:keep_start]}...{value[-keep_end:]}"


def redact_for_log(value, key_name=None):
    """
    Rimuove o maschera dati sensibili prima di mandarli nei log.

    Regole:
    - chiavi altamente sensibili: sempre [REDACTED]
    - email: parzialmente mascherata
    - endpoint / User-Agent / IP / riferimenti esterni: troncati o mascherati
    - dict/list: redazione ricorsiva
    """

    key = str(key_name or "").lower()

    if key in {k.lower() for k in SENSITIVE_LOG_KEYS}:
        return "[REDACTED]"

    if isinstance(value, dict):
        return {
            k: redact_for_log(v, k)
            for k, v in value.items()
        }

    if isinstance(value, list):
        return [
            redact_for_log(v, key_name)
            for v in value
        ]

    if isinstance(value, tuple):
        return tuple(
            redact_for_log(v, key_name)
            for v in value
        )

    if value is None:
        return None

    if key == "email":
        return _mask_email_for_log(value)

    if key in {"endpoint", "user_agent", "riferimento_esterno", "payment_intent"}:
        return _mask_text_for_log(value, keep_start=10, keep_end=6)

    if key == "ip":
        return _mask_text_for_log(value, keep_start=3, keep_end=2)

    if key == "acquisto_id":
        return _mask_text_for_log(value, keep_start=2, keep_end=2)

    return value


def security_log(message, extra=None, *, production=False):
    """
    Logger sicuro per eventi tecnici e di sicurezza.

    In produzione:
    - stampa solo se production=True;
    - stampa sempre dati redatti.

    In locale/development:
    - stampa anche extra, ma comunque redatti.
    """

    env = os.getenv("APP_ENV", "production").lower()
    safe_extra = redact_for_log(extra)

    if env not in ("local", "development"):
        if production:
            if safe_extra is not None:
                print(f"{message}: {safe_extra}", flush=True)
            else:
                print(f"{message}", flush=True)
        return

    if safe_extra is not None:
        print(f"{message}: {safe_extra}", flush=True)
    else:
        print(f"{message}", flush=True)

def log_exception_safe(message, exc=None, extra=None, *, production=False):
    """
    Log sicuro per eccezioni non-Stripe.

    Regole:
    - in produzione non stampa traceback;
    - in produzione stampa solo se production=True;
    - non stampa mai repr(e) grezzo fuori dal sistema di redazione;
    - in local/development stampa anche traceback per debug.
    """

    payload = {}

    if extra:
        payload.update(extra)

    if exc is not None:
        payload.update({
            "error_type": type(exc).__name__,
            "error": repr(exc)
        })

    security_log(
        message,
        payload if payload else None,
        production=production
    )

    if os.getenv("APP_ENV", "production").lower() in ("local", "development"):
        traceback.print_exc()

import os

redis_url = os.getenv("REDIS_URL")

if not redis_url:
    raise RuntimeError("❌ REDIS_URL non configurata su Render")

SOCKET_ASYNC_MODE = "eventlet" if app.config["IS_REALTIME_SERVER"] else "threading"

APP_ENV = os.getenv("APP_ENV", "production").strip().lower()

SOCKET_CORS_ORIGINS = [
    "https://mylocalcare.it",
    "https://www.mylocalcare.it",
    "https://mylocalcare-chat.onrender.com",
]

# Origini locali solo in sviluppo, mai in produzione.
if APP_ENV in ("local", "development"):
    SOCKET_CORS_ORIGINS.extend([
        "http://127.0.0.1:5050",
        "http://localhost:5050",
    ])

socketio = SocketIO(
    app,
    async_mode=SOCKET_ASYNC_MODE,
    cors_allowed_origins=SOCKET_CORS_ORIGINS,
    message_queue=redis_url,
    channel="mylocalcare-socketio",
    allow_upgrades=True,
    ping_timeout=60,
    ping_interval=25,
    logger=app.config["IS_REALTIME_SERVER"],
    engineio_logger=app.config["IS_REALTIME_SERVER"]
)

# =====================================================
# UTENTI ONLINE (socket registry)
# =====================================================
# ==============================
# TRACKING UTENTI ONLINE
# ==============================

def is_user_online(user_id):
    return redis_client.sismember("online_users", str(user_id))

disconnect_timers = {}
recently_read_timers = {}
offline_watchdogs = {}

stripe.api_key = os.environ.get("STRIPE_SECRET_KEY")

@app.context_processor
def inject_stripe_publishable_key():
    return {
        "STRIPE_PUBLISHABLE_KEY": os.environ.get("STRIPE_PUBLISHABLE_KEY", "")
    }

# ==========================================================
# VAPID PUSH CONFIG
# ==========================================================

VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "")
VAPID_PUBLIC_KEY  = os.environ.get("VAPID_PUBLIC_KEY", "")
VAPID_CLAIM_EMAIL = os.environ.get("VAPID_CLAIM_EMAIL", "mailto:info@mylocalcare.it")
app.config["VAPID_PUBLIC_KEY"] = VAPID_PUBLIC_KEY
app.config["VAPID_CLAIM_EMAIL"] = VAPID_CLAIM_EMAIL

# Converte i "\n" in newline reali
if VAPID_PRIVATE_KEY:
    VAPID_PRIVATE_KEY = VAPID_PRIVATE_KEY.replace("\\n", "\n")

# ==========================================================
# SQL HELPER (placeholder compatibili SQLite / Postgres)
# ==========================================================

def sql(query):
    """
    Converte automaticamente i placeholder:
    SQLite  -> ?
    Postgres -> %s
    """
    if app.config.get("IS_POSTGRES"):
        return query.replace("?", "%s")
    return query

def sql_now_minus_seconds(seconds: int) -> str:
    """
    Ritorna un'espressione SQL 'now - X seconds' compatibile con:
    - Postgres: CURRENT_TIMESTAMP - INTERVAL 'X seconds'
    - SQLite:   datetime('now','-X seconds')
    """
    if app.config.get("IS_POSTGRES"):
        return f"CURRENT_TIMESTAMP - INTERVAL '{int(seconds)} seconds'"
    return f"datetime('now','-{int(seconds)} seconds')"

# ==========================================================
# 🕒 FUNZIONI TEMPO COMPATIBILI SQLite + PostgreSQL
# ==========================================================

def now_sql():
    """
    Timestamp corrente compatibile con entrambi i DB.
    """
    if app.config.get("IS_POSTGRES"):
        return "CURRENT_TIMESTAMP"
    else:
        return "datetime('now')"

def month_sql(field=None):
    """
    Restituisce YYYY-MM compatibile SQLite/Postgres.
    """
    if app.config.get("IS_POSTGRES"):
        if field:
            return f"TO_CHAR({field}, 'YYYY-MM')"
        return "TO_CHAR(NOW(), 'YYYY-MM')"
    else:
        if field:
            return f"strftime('%Y-%m', {field})"
        return "strftime('%Y-%m','now')"

def epoch_now_sql():
    """
    Timestamp unix (secondi) compatibile con entrambi i DB.
    """
    if app.config.get("IS_POSTGRES"):
        return "EXTRACT(EPOCH FROM NOW())::INT"
    else:
        return "strftime('%s','now')"


def dt_sql(field):
    """
    Normalizza conversione datetime nelle query ORDER BY.
    """
    if app.config.get("IS_POSTGRES"):
        return field
    else:
        return f"datetime({field})"

def order_datetime(field):
    return field if app.config.get("IS_POSTGRES") else f"datetime({field})"

def get_last_id(cur):
    """
    Compatibile SQLite + PostgreSQL.
    """
    if app.config.get("IS_POSTGRES"):
        return cur.fetchone()[0]
    else:
        return cur.lastrowid



from datetime import datetime
from zoneinfo import ZoneInfo

@app.teardown_request
def _release_pg_conn(exc):
    conn = getattr(g, "db_conn", None)
    if not conn:
        return

    try:
        # ⚠️ IMPORTANTE:
        # non usare putconn manuale
        # il wrapper gestisce già il rilascio al pool
        conn.close()
    except Exception:
        pass

    g.db_conn = None

@app.template_filter("dt_roma")
def dt_roma(value):
    if not value:
        return ""

    try:
        if isinstance(value, datetime):
            dt = value
        else:
            dt = datetime.fromisoformat(str(value))

        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))

        return dt.astimezone(
            ZoneInfo("Europe/Rome")
        ).strftime("%Y-%m-%d %H:%M:%S")

    except Exception:
        return str(value)

@app.template_filter("dt_roma_admin")
def dt_roma_admin(value):
    if not value:
        return ""

    try:
        if isinstance(value, datetime):
            dt = value
        else:
            dt = datetime.fromisoformat(str(value))

        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))

        return dt.astimezone(
            ZoneInfo("Europe/Rome")
        ).strftime("%d/%m/%Y %H:%M").lstrip("0").replace("/0", "/")

    except Exception:
        return str(value)

@app.template_filter("fromjson")
def fromjson_filter(value):
    if not value:
        return {}
    try:
        return json.loads(value)
    except Exception:
        return {}

from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo

def now_utc():
    return datetime.now(timezone.utc)

def parse_iso(dt_str):
    if not dt_str:
        return None
    try:
        return datetime.fromisoformat(dt_str)
    except Exception:
        return None


@app.template_filter('to_datetime')
def to_datetime_filter(value):
    if not value:
        return None
    try:
        v = str(value).replace(" ", "T")
        dt = datetime.fromisoformat(v)

        # se naive → assumo UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception as e:
        log_exception_safe(
            "❌ Errore filtro to_datetime",
            e
        )
        return None

@app.template_filter('fmt_it')
def fmt_it(value):
    try:
        dt = to_datetime_filter(value)
        if not dt:
            return value

        dt_it = dt.astimezone(ZoneInfo("Europe/Rome"))
        return dt_it.strftime("%d-%m-%Y %H:%M")
    except Exception as e:
        log_exception_safe(
            "❌ Errore filtro fmt_it",
            e
        )
        return value

@app.template_filter('fmt_it_date')
def fmt_it_date(value):
    try:
        dt = to_datetime_filter(value)
        if not dt:
            return value

        dt_it = dt.astimezone(ZoneInfo("Europe/Rome"))
        return dt_it.strftime("%d-%m-%Y")
    except Exception as e:
        log_exception_safe(
            "❌ Errore filtro fmt_it_date",
            e
        )
        return value

@app.template_filter('fmt_it_smart')
def fmt_it_smart(value):
    try:
        dt = to_datetime_filter(value)
        if not dt:
            return value

        roma = ZoneInfo("Europe/Rome")
        dt_it = dt.astimezone(roma)
        now = datetime.now(roma)

        date_it = dt_it.date()
        today = now.date()
        yesterday = (now - timedelta(days=1)).date()

        if date_it == today:
            return f"oggi {dt_it.strftime('%H:%M')}"

        if date_it == yesterday:
            return f"ieri {dt_it.strftime('%H:%M')}"

        if dt_it.isocalendar()[1] == now.isocalendar()[1] and dt_it.year == now.year:
            giorni = ["lunedì","martedì","mercoledì","giovedì","venerdì","sabato","domenica"]
            return f"{giorni[dt_it.weekday()]} {dt_it.strftime('%H:%M')}"

        return dt_it.strftime("%d-%m-%Y %H:%M")

    except Exception as e:
        log_exception_safe(
            "❌ Errore filtro fmt_it_smart",
            e
        )
        return value

@app.context_processor
def inject_session():
    """Rende disponibile la sessione Flask e la config socket in tutti i template."""
    from flask import session

    realtime_token = None
    utente_id = session.get("utente_id")
    if utente_id:
        try:
            realtime_token = build_realtime_token(utente_id)
        except Exception as e:
            log_exception_safe(
                "❌ Errore build_realtime_token",
                e,
                {"user_id": utente_id}
            )

    return dict(
        session=session,
        socket_base_url=app.config.get("SOCKET_BASE_URL", ""),
        realtime_token=realtime_token
    )

# Imposta tempo di "grazia" (in secondi) dopo la chiusura chat
app.config.setdefault('CHAT_RECENTLY_READ_TTL', 5)
# ---------------------------------------------------------
# Sessioni (Render-friendly)
# ---------------------------------------------------------
# ---------------------------------------------------------
# Sessioni Redis (Production SaaS)
# ---------------------------------------------------------

import redis
from flask_session import Session
from datetime import timedelta

redis_client = redis.from_url(os.environ["REDIS_URL"])

configure_socket_registry(redis_client, socketio)

app.config['SESSION_TYPE'] = 'redis'
app.config['SESSION_REDIS'] = redis_client

try:
    redis_client.ping()
    safe_log("✅ Redis connesso correttamente", production=True)
except Exception as e:
    log_exception_safe(
        "❌ Redis NON connesso",
        e,
        production=True
    )

app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)
app.config['SESSION_REFRESH_EACH_REQUEST'] = True

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_SAMESITE'] = "Lax"
app.config['SESSION_COOKIE_DOMAIN'] = ".mylocalcare.it"

Session(app)

@app.template_filter('safe_strip')
def safe_strip(value):
    return (value or '').strip()

@app.template_filter('datetimeformat')
def datetimeformat(value, fmt='%d %B %Y'):
    """Formatta la data per mostrare solo giorno e mese in italiano."""
    try:
        dt = datetime.fromisoformat(value)
        mesi = [
            "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
            "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"
        ]
        mese_nome = mesi[dt.month - 1]
        return f"{dt.day} {mese_nome} {dt.year}"
    except Exception:
        return value

# 🔹 Configurazione Flask-Mail letta dal file mail.env
app.config['MAIL_SERVER'] = 'smtps.aruba.it'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')

# ✅ Mittente unico e coerente per tutte le email automatiche
MAIL_FROM_ADDRESS = os.getenv('MAIL_FROM_ADDRESS', 'info@mylocalcare.it')
MAIL_FROM_NAME = os.getenv('MAIL_FROM_NAME', 'MyLocalCare')

app.config['MAIL_DEFAULT_SENDER'] = (MAIL_FROM_NAME, MAIL_FROM_ADDRESS)

app.config['MAIL_TIMEOUT'] = 20
app.config['MAIL_MAX_EMAILS'] = None
app.config['MAIL_DEBUG'] = (os.getenv("FLASK_ENV") == "development")

# 🌐 Base URL per generare link assoluti nelle email
# In produzione deve essere sempre il dominio principale con www.
app.config["APP_BASE_URL"] = os.getenv(
    "APP_BASE_URL",
    "https://www.mylocalcare.it"
).rstrip("/")

# 🔐 Salt usato per i token di reset password (mettilo anche in mail.env se vuoi)
app.config['SECURITY_PASSWORD_SALT'] = os.getenv(
    'SECURITY_PASSWORD_SALT',
    'metti-qui-una-stringa-lunga-casuale'
)

# 🔹 Inizializza Flask-Mail
mail = Mail(app)
safe_log("APP_BASE_URL configurata", production=True)
# ---------------------------------------------------------
# 📧 FUNZIONI EMAIL – UTENTE
# ---------------------------------------------------------
def build_external_url(endpoint: str, **values) -> str:
    """
    Costruisce URL assoluti usando APP_BASE_URL.

    Funziona sia dentro una normale request Flask,
    sia dentro background task come Daily Matches,
    dove non esiste una request attiva.
    """
    base = (app.config.get("APP_BASE_URL") or "https://www.mylocalcare.it").rstrip("/")

    try:
        # Caso normale: siamo dentro una request attiva
        path = url_for(endpoint, _external=False, **values)
    except RuntimeError:
        # Caso background task: non c'è request context
        with app.test_request_context(base_url=base):
            path = url_for(endpoint, _external=False, **values)

    return f"{base}{path}"

def send_async_email(app, msg):
    with app.app_context():
        mail.send(msg)

def get_reset_serializer():
    """Restituisce il serializer firmato per i token di reset password."""
    secret_key = app.config.get('SECRET_KEY') or app.secret_key
    return URLSafeTimedSerializer(secret_key)


def invia_email_sospensione(email, nome):
    """
    Invia l'email di sospensione account tramite Postmark,
    usando la funzione centrale _invia_email().
    """
    try:
        return _invia_email(
            destinazione=email,
            oggetto="Account MyLocalCare sospeso",
            corpo=(
                f"Ciao {nome},\n\n"
                "il tuo account MyLocalCare è stato sospeso.\n\n"
                "Se ritieni che si tratti di un errore, contatta l'assistenza.\n\n"
                "MyLocalCare"
            )
        )

    except Exception as e:
        log_exception_safe(
            "❌ Errore invio email sospensione",
            e,
            {"email": email},
            production=True
        )
        return False
# ==========================================================
# 2️⃣ FUNZIONE CONNESSIONE DB E MODELS
# ==========================================================

class PGCursorWrapper:
    """
    Wrapper del cursor per convertire automaticamente
    i placeholder SQLite (?) in PostgreSQL (%s)
    """

    def __init__(self, cursor):
        self.cursor = cursor

    def execute(self, query, params=None):
        query = query.replace("?", "%s")
        self.cursor.execute(query, params or ())
        return self

    def executemany(self, query, params_list):
        query = query.replace("?", "%s")
        return self.cursor.executemany(query, params_list)

    def fetchone(self):
        return self.cursor.fetchone()

    def fetchall(self):
        return self.cursor.fetchall()

    def __getattr__(self, name):
        return getattr(self.cursor, name)


class PGConnectionWrapper:
    """
    Wrapper che rende psycopg2 compatibile con stile SQLite.
    Permette di usare conn.execute() ovunque nel codice.
    """

    def __init__(self, conn):
        self.conn = conn
        self._released = False

    def execute(self, query, params=None):
        cur = self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        # converte placeholder SQLite → Postgres
        query = query.replace("?", "%s")

        cur.execute(query, params or ())
        return cur

    def cursor(self):
        return PGCursorWrapper(
            self.conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        )

    def commit(self):
        return self.conn.commit()

    def close(self):
        if self._released:
            return

        self._released = True

        try:
            pool = get_current_pg_pool()
            if pool is not None:
                pool.putconn(self.conn)
                return
        except Exception as e:
            log_exception_safe(
                "🟥 PGConnectionWrapper.close putconn error",
                e,
                production=True
            )

        try:
            self.conn.close()
        except Exception:
            pass

    def __getattr__(self, name):
        return getattr(self.conn, name)


def get_db_connection():
    from flask import has_request_context

    database_url = os.getenv("DATABASE_URL")
    app.config["IS_POSTGRES"] = bool(database_url)

    # =========================
    # POSTGRES
    # =========================
    if database_url:
        pool = init_pg_pool()

        if pool is None:
            raise RuntimeError("Pool PostgreSQL non inizializzato")

        if has_request_context():
            if hasattr(g, "db_conn") and g.db_conn is not None:
                try:
                    g.db_conn.cursor().execute("SELECT 1")
                    return g.db_conn
                except Exception:
                    try:
                        g.db_conn.close()
                    except Exception:
                        pass
                    g.db_conn = None

        raw = pool.getconn()

        if raw.closed:
            pool.putconn(raw, close=True)
            raw = pool.getconn()

        raw.autocommit = True
        raw.set_session(readonly=False, autocommit=True)

        wrapped = PGConnectionWrapper(raw)

        if has_request_context():
            g.db_conn = wrapped

        return wrapped

    # =========================
    # SQLITE
    # =========================
    else:
        if hasattr(g, "db_conn"):
            return g.db_conn

        import sqlite3

        conn = sqlite3.connect('database.db', timeout=5)
        conn.row_factory = sqlite3.Row

        conn.execute("PRAGMA foreign_keys = ON;")
        conn.execute("PRAGMA journal_mode = WAL;")
        conn.execute("PRAGMA synchronous = NORMAL;")
        conn.execute("PRAGMA busy_timeout = 5000;")

        g.db_conn = conn
        return conn

def close_db_connection(conn):
    if not conn:
        return
    try:
        conn.close()
    except:
        pass

app.config["DB_CONN_FACTORY"] = get_db_connection
app.config["IS_POSTGRES"] = bool(os.getenv("DATABASE_URL"))


# --- Middleware di protezione per login richiesto ---
def login_required(view):
    from functools import wraps
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.utente is None:
            flash("Devi accedere per vedere questa pagina.")
            return redirect(url_for('login'))
        return view(**kwargs)
    return wrapped_view

from functools import wraps
from flask import redirect, url_for, flash, g

def foto_obbligatoria(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        # Se non è loggato, si occupa già login_required
        user = g.get("utente")

        if not user:
            flash("Devi essere loggato.", "error")
            return redirect(url_for("login"))

        # ⚠️ Blocco se manca la foto profilo
        if not user["foto_profilo"]:
            flash("Per usare questa funzione devi caricare una foto profilo.", "error")
            return redirect(url_for("dashboard"))  # pagina modifica profilo

        return f(*args, **kwargs)
    return wrapper

# ==========================================================
# 🔐 ADMIN SECURITY VERSION — INVALIDAZIONE SESSIONI SENSIBILI
# ==========================================================

def ensure_admin_security_version_column():
    """
    Aggiunge alla tabella utenti una versione di sicurezza admin.

    Serve per invalidare globalmente tutte le sessioni admin sensibili
    dopo eventi critici:
    - generazione nuovi recovery code
    - uso recovery code
    - registrazione/rimozione passkey
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                ALTER TABLE utenti
                ADD COLUMN IF NOT EXISTS admin_security_version INTEGER DEFAULT 0
            """)
        else:
            cur.execute("PRAGMA table_info(utenti)")
            columns = [row["name"] for row in cur.fetchall()]

            if "admin_security_version" not in columns:
                cur.execute("""
                    ALTER TABLE utenti
                    ADD COLUMN admin_security_version INTEGER DEFAULT 0
                """)

        cur.execute(sql("""
            UPDATE utenti
            SET admin_security_version = 0
            WHERE admin_security_version IS NULL
        """))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def get_admin_security_version(user_id):
    """
    Legge la versione sicurezza admin corrente dal DB.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT COALESCE(admin_security_version, 0) AS admin_security_version
            FROM utenti
            WHERE id = ?
            LIMIT 1
        """), (int(user_id),))

        row = cur.fetchone()

        if not row:
            return 0

        return int(row["admin_security_version"] or 0)

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def bump_admin_security_version(user_id, reason="security_change"):
    """
    Incrementa la versione sicurezza admin.

    Effetto:
    - tutte le vecchie sessioni admin con versione precedente decadono;
    - la sessione corrente viene aggiornata alla nuova versione, così l'admin
      che ha eseguito l'azione non viene buttato fuori subito.

    Nota:
    questa funzione NON richiama get_admin_security_version() per evitare
    aperture/chiusure annidate della stessa connessione di request.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                UPDATE utenti
                SET admin_security_version = COALESCE(admin_security_version, 0) + 1
                WHERE id = %s
                RETURNING admin_security_version
            """, (int(user_id),))

            row = cur.fetchone()
            nuova_versione = int(row["admin_security_version"] if hasattr(row, "keys") else row[0])

        else:
            cur.execute("""
                UPDATE utenti
                SET admin_security_version = COALESCE(admin_security_version, 0) + 1
                WHERE id = ?
            """, (int(user_id),))

            cur.execute("""
                SELECT COALESCE(admin_security_version, 0) AS admin_security_version
                FROM utenti
                WHERE id = ?
                LIMIT 1
            """, (int(user_id),))

            row = cur.fetchone()
            nuova_versione = int(row["admin_security_version"] if hasattr(row, "keys") else row[0])

        conn.commit()

        session["admin_security_version"] = nuova_versione
        session.modified = True

        security_log(
            "🔐 [ADMIN SECURITY VERSION] incrementata",
            {
                "user_id": int(user_id),
                "reason": reason,
                "new_version": nuova_versione
            },
            production=True
        )

        return nuova_versione

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

# ==========================================================
# 🔐 ADMIN PASSKEYS — STORAGE CHIAVI PUBBLICHE
# ==========================================================

def ensure_admin_passkeys_table():
    """
    Crea la tabella per le passkey admin se non esiste.

    Nota importante:
    qui NON salviamo password e NON salviamo dati biometrici.
    Salviamo solo:
    - credential_id
    - chiave pubblica
    - contatore di sicurezza sign_count
    - metadati tecnici della passkey
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                CREATE TABLE IF NOT EXISTS admin_passkeys (
                    id SERIAL PRIMARY KEY,
                    utente_id INTEGER NOT NULL REFERENCES utenti(id) ON DELETE CASCADE,
                    credential_id TEXT UNIQUE NOT NULL,
                    credential_public_key TEXT NOT NULL,
                    sign_count INTEGER DEFAULT 0,
                    device_type TEXT,
                    backed_up INTEGER DEFAULT 0,
                    transports TEXT,
                    nome_dispositivo TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    last_used_at TIMESTAMP
                )
            """)

            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_admin_passkeys_utente
                ON admin_passkeys(utente_id)
            """)

        else:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS admin_passkeys (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    utente_id INTEGER NOT NULL,
                    credential_id TEXT UNIQUE NOT NULL,
                    credential_public_key TEXT NOT NULL,
                    sign_count INTEGER DEFAULT 0,
                    device_type TEXT,
                    backed_up INTEGER DEFAULT 0,
                    transports TEXT,
                    nome_dispositivo TEXT,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    last_used_at TEXT,
                    FOREIGN KEY (utente_id) REFERENCES utenti(id) ON DELETE CASCADE
                )
            """)

            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_admin_passkeys_utente
                ON admin_passkeys(utente_id)
            """)

        conn.commit()

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass

        log_exception_safe(
            "❌ Errore ensure_admin_passkeys_table",
            e,
            production=True
        )
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def get_admin_passkeys_for_user(user_id):
    """
    Recupera le passkey registrate per un admin.

    La tabella admin_passkeys deve già esistere:
    viene verificata/creata solo a bootstrap o tramite migrazione.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT
                id,
                credential_id,
                credential_public_key,
                sign_count,
                device_type,
                backed_up,
                transports,
                nome_dispositivo,
                created_at,
                last_used_at
            FROM admin_passkeys
            WHERE utente_id = ?
            ORDER BY created_at DESC
        """), (user_id,))

        return cur.fetchall()

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def get_admin_passkey_by_credential_id(credential_id_b64url):
    """
    Recupera una passkey admin tramite credential_id.
    Serve nella fase di autenticazione/sblocco.

    Nota: questa funzione viene ridefinita più sotto con una versione più sicura
    che filtra anche per user_id. Manteniamo questa solo per compatibilità interna,
    senza creare/verificare la tabella durante l'uso operativo.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT
                id,
                utente_id,
                credential_id,
                credential_public_key,
                sign_count,
                device_type,
                backed_up,
                transports,
                nome_dispositivo,
                created_at,
                last_used_at
            FROM admin_passkeys
            WHERE credential_id = ?
            LIMIT 1
        """), (credential_id_b64url,))

        return cur.fetchone()

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def update_admin_passkey_sign_count(passkey_id, sign_count):
    """
    Aggiorna il contatore WebAuthn e la data di ultimo utilizzo.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql(f"""
            UPDATE admin_passkeys
            SET sign_count = ?,
                last_used_at = {now_sql()}
            WHERE id = ?
        """), (
            int(sign_count or 0),
            int(passkey_id)
        ))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def save_admin_passkey(
    user_id,
    credential_id_b64url,
    credential_public_key_b64,
    sign_count,
    device_type=None,
    backed_up=0,
    transports=None,
    nome_dispositivo=None
):
    """
    Salva una nuova passkey admin.

    La tabella admin_passkeys deve già esistere:
    viene verificata/creata solo a bootstrap o tramite migrazione.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            INSERT INTO admin_passkeys (
                utente_id,
                credential_id,
                credential_public_key,
                sign_count,
                device_type,
                backed_up,
                transports,
                nome_dispositivo
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """), (
            user_id,
            credential_id_b64url,
            credential_public_key_b64,
            int(sign_count or 0),
            device_type,
            1 if backed_up else 0,
            json.dumps(transports or []),
            nome_dispositivo
        ))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def delete_admin_passkey_for_user(passkey_id, user_id):
    """
    Elimina una passkey admin appartenente all'utente indicato.
    Non permette di eliminare passkey di altri admin.

    La protezione contro eliminazione dell'ultima passkey viene fatta nella route.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            DELETE FROM admin_passkeys
            WHERE id = ?
              AND utente_id = ?
        """), (
            int(passkey_id),
            int(user_id)
        ))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

# ==========================================================
# 🔐 ADMIN RECOVERY CODES — CODICI DI EMERGENZA PASSKEY
# ==========================================================

def ensure_admin_recovery_codes_table():
    """
    Crea la tabella dei recovery codes admin.

    I codici NON vengono salvati in chiaro.
    Salviamo solo:
    - hash del codice
    - stato usato/non usato
    - data creazione
    - data utilizzo

    Ogni codice potrà essere usato una sola volta.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                CREATE TABLE IF NOT EXISTS admin_recovery_codes (
                    id SERIAL PRIMARY KEY,
                    utente_id INTEGER NOT NULL REFERENCES utenti(id) ON DELETE CASCADE,
                    code_hash TEXT NOT NULL,
                    used INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    used_at TIMESTAMP
                )
            """)

            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_admin_recovery_codes_utente
                ON admin_recovery_codes(utente_id)
            """)

        else:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS admin_recovery_codes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    utente_id INTEGER NOT NULL,
                    code_hash TEXT NOT NULL,
                    used INTEGER DEFAULT 0,
                    created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    used_at TEXT,
                    FOREIGN KEY (utente_id) REFERENCES utenti(id) ON DELETE CASCADE
                )
            """)

            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_admin_recovery_codes_utente
                ON admin_recovery_codes(utente_id)
            """)

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def generate_admin_recovery_codes(user_id, count=8):
    """
    Genera nuovi recovery codes per l'admin.

    Attenzione:
    - cancella/invalida i codici precedenti;
    - restituisce i codici in chiaro SOLO in questa chiamata;
    - nel DB salva solo hash.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    plain_codes = []

    try:
        cur.execute(sql("""
            DELETE FROM admin_recovery_codes
            WHERE utente_id = ?
        """), (int(user_id),))

        for _ in range(count):
            raw = secrets.token_urlsafe(18)
            code = f"LC-{raw[:6]}-{raw[6:12]}-{raw[12:18]}".upper()
            code_hash = generate_password_hash(code)

            cur.execute(sql("""
                INSERT INTO admin_recovery_codes (
                    utente_id,
                    code_hash,
                    used
                )
                VALUES (?, ?, 0)
            """), (
                int(user_id),
                code_hash
            ))

            plain_codes.append(code)

        conn.commit()
        return plain_codes

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def count_unused_admin_recovery_codes(user_id):
    """
    Conta quanti recovery codes admin sono ancora disponibili.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT COUNT(*) AS totale
            FROM admin_recovery_codes
            WHERE utente_id = ?
              AND used = 0
        """), (int(user_id),))

        row = cur.fetchone()
        return int(fetchone_value(row) or 0)

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def verify_and_consume_admin_recovery_code(user_id, plain_code):
    """
    Verifica un recovery code e lo marca come usato.

    Ritorna True se valido, False se non valido.
    """
    plain_code = (plain_code or "").strip().upper()

    if not plain_code:
        return False

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT id, code_hash
            FROM admin_recovery_codes
            WHERE utente_id = ?
              AND used = 0
            ORDER BY id ASC
        """), (int(user_id),))

        rows = cur.fetchall()

        for row in rows:
            if check_password_hash(row["code_hash"], plain_code):
                cur.execute(sql(f"""
                    UPDATE admin_recovery_codes
                    SET used = 1,
                        used_at = {now_sql()}
                    WHERE id = ?
                """), (int(row["id"]),))

                conn.commit()
                return True

        return False

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

# ==========================================================
# 🔹 ADMIN COUNTERS (Annunci e Recensioni in attesa)
# ==========================================================

from threading import Lock
db_lock = Lock()

from functools import wraps

from functools import wraps
from datetime import datetime, timezone

# ==========================================================
# 🔐 ADMIN STEP-UP AUTH
# ==========================================================

ADMIN_STEPUP_MINUTES = 15


def admin_stepup_is_valid():
    """
    Verifica se l'admin ha già fatto uno sblocco recente.
    Questo NON sostituisce admin_session_token: è un secondo livello.
    """
    raw_until = session.get("admin_stepup_until")
    fp_sessione = session.get("admin_stepup_fingerprint")
    fp_corrente = request.headers.get("User-Agent", "unknown")

    if not raw_until or not fp_sessione:
        return False

    if fp_sessione != fp_corrente:
        return False

    try:
        until = datetime.fromisoformat(raw_until)
        if until.tzinfo is None:
            until = until.replace(tzinfo=timezone.utc)
    except Exception:
        return False

    return until > datetime.now(timezone.utc)


def mark_admin_stepup_verified():
    """
    Marca lo sblocco admin come valido per pochi minuti.
    In seguito questa funzione verrà chiamata dopo WebAuthn/FaceID.
    """
    until = datetime.now(timezone.utc) + timedelta(minutes=ADMIN_STEPUP_MINUTES)

    session["admin_stepup_until"] = until.isoformat()
    session["admin_stepup_fingerprint"] = request.headers.get("User-Agent", "unknown")
    session.modified = True


def clear_admin_stepup():
    session.pop("admin_stepup_until", None)
    session.pop("admin_stepup_fingerprint", None)
    session.modified = True

def admin_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        # 1) deve essere loggato
        if not g.utente:
            flash("Devi accedere per entrare nell'area amministratore.", "error")
            return redirect(url_for("login"))

        # g.utente è uno sqlite3.Row
        row = g.utente

        # 2) deve avere ruolo = 'admin'
        ruolo = row["ruolo"] if "ruolo" in row.keys() else None
        if ruolo != "admin":
            flash("Accesso riservato agli amministratori.", "error")
            return redirect(url_for("home"))

        # 3) 🔐 Verifica token di sessione admin
        session_token = session.get("admin_session_token")

        db_token  = row["admin_session_token"]     if "admin_session_token"     in row.keys() else None
        db_expiry = row["admin_session_expiry"]    if "admin_session_expiry"    in row.keys() else None
        db_fp     = row["admin_browser_fingerprint"] if "admin_browser_fingerprint" in row.keys() else None

        # deve esistere un token in sessione
        if not session_token or not db_token:
            flash("Sessione amministratore non valida. Esegui di nuovo il login.", "error")
            session.clear()
            return redirect(url_for("login"))

        # deve coincidere con quello nel database
        if session_token != db_token:
            flash("La sessione amministratore è stata invalidata. Esegui di nuovo il login.", "error")
            session.clear()
            return redirect(url_for("login"))

        # 🔐 Verifica versione sicurezza admin
        session_admin_security_version = session.get("admin_security_version")

        db_admin_security_version = (
            row["admin_security_version"]
            if "admin_security_version" in row.keys() and row["admin_security_version"] is not None
            else 0
        )

        try:
            session_admin_security_version = int(session_admin_security_version)
            db_admin_security_version = int(db_admin_security_version)
        except Exception:
            flash("Sessione amministratore non valida. Esegui di nuovo il login.", "error")
            session.clear()
            return redirect(url_for("login"))

        if session_admin_security_version != db_admin_security_version:
            flash(
                "La sessione amministratore è stata invalidata per una modifica di sicurezza. "
                "Esegui di nuovo il login.",
                "warning"
            )
            session.clear()
            return redirect(url_for("login"))

        # 🔐 Verifica impronta browser
        session_fp = session.get("admin_browser_fingerprint")
        current_fp = request.headers.get("User-Agent", "unknown")

        if not session_fp or not db_fp:
            flash("Sessione amministratore non valida (mancano dati device).", "error")
            session.clear()
            return redirect(url_for("login"))

        # ❌ Se fingerprint diverso → blocco totale
        if session_fp != db_fp or current_fp != db_fp:
            flash("Accesso amministratore bloccato: dispositivo non riconosciuto.", "error")
            session.clear()
            return redirect(url_for("login"))

        # deve essere ancora valido (non scaduto)
        expiry_dt = None
        try:
            if not db_expiry:
                expiry_dt = None
            elif isinstance(db_expiry, datetime):
                expiry_dt = db_expiry
            elif isinstance(db_expiry, str):
                expiry_dt = datetime.fromisoformat(db_expiry)
            else:
                expiry_dt = None

            # se arriva naive, rendilo UTC
            if expiry_dt and expiry_dt.tzinfo is None:
                expiry_dt = expiry_dt.replace(tzinfo=timezone.utc)

        except Exception as e:
            print("❌ admin_required expiry parse error:", repr(e), "db_expiry=", repr(db_expiry))
            session.clear()
            flash("Sessione non valida (errore token).", "error")
            return redirect(url_for("login"))

        if not expiry_dt or expiry_dt < datetime.now(timezone.utc):
            flash("La sessione amministratore è scaduta. Accedi di nuovo.", "warning")
            session.clear()
            return redirect(url_for("login"))

        # 4) 🔐 secondo livello admin: sblocco recente richiesto
        # Alcune route tecniche devono poter funzionare anche quando lo step-up non è ancora valido.
        stepup_exempt_endpoints = {
            "admin_unlock",
            "admin_passkey_auth_options",
            "admin_passkey_auth_verify",
            "admin_recovery_code_verify",
        }

        if request.endpoint not in stepup_exempt_endpoints and not admin_stepup_is_valid():
            next_url = request.full_path if request.query_string else request.path
            return redirect(url_for("admin_unlock", next=next_url))

        # 5) tutto ok → esegui la view
        return view_func(*args, **kwargs)

    return wrapped_view

def _decimal_from_openai_amount(value):
    try:
        return Decimal(str(value or "0"))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def format_openai_euro(value):
    if value is None:
        return "—"

    value = Decimal(str(value))

    if value == 0:
        return "€ 0.000000"

    if value < Decimal("0.01"):
        return f"€ {value:.6f}"

    return f"€ {value:.4f}"


def get_openai_month_stats():
    """
    Recupera dati ufficiali OpenAI del mese corrente:
    - costo totale preciso
    - numero richieste
    - dettagli giornalieri
    """

    api_key = os.getenv("OPENAI_ADMIN_KEY")

    if not api_key:
        return {
            "ok": False,
            "error": "OPENAI_ADMIN_KEY mancante",
            "total_cost": None,
            "total_requests": None,
            "daily": []
        }

    try:
        now = datetime.now(timezone.utc)

        start_month = datetime(
            now.year,
            now.month,
            1,
            tzinfo=timezone.utc
        )

        start_time = int(start_month.timestamp())

        # OpenAI richiede un end_time successivo al bucket di partenza.
        # Usiamo l'inizio del giorno UTC successivo come limite esclusivo,
        # così il range è valido anche il primo giorno del mese.
        end_day = datetime(
            now.year,
            now.month,
            now.day,
            tzinfo=timezone.utc
        ) + timedelta(days=1)

        end_time = int(end_day.timestamp())

        headers = {
            "Authorization": f"Bearer {api_key}"
        }

        # COSTI UFFICIALI
        costs_response = requests.get(
            "https://api.openai.com/v1/organization/costs",
            headers=headers,
            params={
                "start_time": start_time,
                "end_time": end_time,
                "bucket_width": "1d",
                "limit": 31
            },
            timeout=15
        )

        if costs_response.status_code != 200:
            print("❌ OpenAI Costs API error:", costs_response.text)
            return {
                "ok": False,
                "error": costs_response.text,
                "total_cost": None,
                "total_requests": None,
                "daily": []
            }

        costs_data = costs_response.json()

        daily_map = {}
        total_cost = Decimal("0")

        for bucket in costs_data.get("data", []):
            bucket_start = bucket.get("start_time")

            giorno = datetime.fromtimestamp(
                bucket_start,
                timezone.utc
            ).strftime("%d/%m/%Y") if bucket_start else "—"

            daily_map.setdefault(giorno, {
                "data": giorno,
                "costo": Decimal("0"),
                "richieste": 0,
                "input_tokens": 0,
                "output_tokens": 0
            })

            for result in bucket.get("results", []):
                amount = result.get("amount", {})
                value = _decimal_from_openai_amount(amount.get("value"))

                daily_map[giorno]["costo"] += value
                total_cost += value

        # USAGE / RICHIESTE UFFICIALI
        usage_response = requests.get(
            "https://api.openai.com/v1/organization/usage/completions",
            headers=headers,
            params={
                "start_time": start_time,
                "end_time": end_time,
                "bucket_width": "1d",
                "limit": 31
            },
            timeout=15
        )

        total_requests = 0

        if usage_response.status_code == 200:
            usage_data = usage_response.json()

            for bucket in usage_data.get("data", []):
                bucket_start = bucket.get("start_time")

                giorno = datetime.fromtimestamp(
                    bucket_start,
                    timezone.utc
                ).strftime("%d/%m/%Y") if bucket_start else "—"

                daily_map.setdefault(giorno, {
                    "data": giorno,
                    "costo": Decimal("0"),
                    "richieste": 0,
                    "input_tokens": 0,
                    "output_tokens": 0
                })

                for result in bucket.get("results", []):
                    richieste = int(result.get("num_model_requests") or 0)
                    input_tokens = int(result.get("input_tokens") or 0)
                    output_tokens = int(result.get("output_tokens") or 0)

                    daily_map[giorno]["richieste"] += richieste
                    daily_map[giorno]["input_tokens"] += input_tokens
                    daily_map[giorno]["output_tokens"] += output_tokens

                    total_requests += richieste
        else:
            print("❌ OpenAI Usage API error:", usage_response.text)

        daily = [
            row for row in daily_map.values()
            if (
                int(row.get("richieste") or 0) > 0
                or int(row.get("input_tokens") or 0) > 0
                or int(row.get("output_tokens") or 0) > 0
                or Decimal(str(row.get("costo") or "0")) > Decimal("0")
            )
        ]

        daily = sorted(
            daily,
            key=lambda x: datetime.strptime(x["data"], "%d/%m/%Y"),
            reverse=True
        )

        for row in daily:
            row["costo_raw"] = str(row["costo"])
            row["costo"] = format_openai_euro(row["costo"])

        return {
            "ok": True,
            "error": None,
            "total_cost": total_cost,
            "total_cost_raw": str(total_cost),
            "total_cost_formatted": format_openai_euro(total_cost),
            "total_requests": total_requests,
            "daily": daily
        }

    except Exception as e:
        print("❌ Errore OpenAI stats:", e)

        return {
            "ok": False,
            "error": str(e),
            "total_cost": None,
            "total_requests": None,
            "daily": []
        }


_openai_cost_cache = {
    "value": None,
    "expires_at": 0
}


def get_openai_month_cost():
    now_ts = time.time()

    if (
        _openai_cost_cache["value"] is not None
        and _openai_cost_cache["expires_at"] > now_ts
    ):
        return _openai_cost_cache["value"]

    stats = get_openai_month_stats()

    if stats.get("ok"):
        value = stats.get("total_cost")

        _openai_cost_cache["value"] = value
        _openai_cost_cache["expires_at"] = now_ts + 300

        return value

    return None

def ensure_openai_usage_storico_table():
    conn = get_db_connection()

    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS openai_usage_giornaliero (
                data TEXT PRIMARY KEY,
                mese TEXT NOT NULL,
                costo_raw TEXT DEFAULT '0',
                richieste INTEGER DEFAULT 0,
                input_tokens INTEGER DEFAULT 0,
                output_tokens INTEGER DEFAULT 0,
                aggiornato_il TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        conn.commit()

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass

        print("❌ Impossibile creare tabella openai_usage_giornaliero:", e)

    finally:
        conn.close()

def salva_openai_usage_giornaliero(stats):
    if not stats or not stats.get("ok"):
        return

    ensure_openai_usage_storico_table()

    conn = get_db_connection()

    for row in stats.get("daily", []):
        richieste = int(row.get("richieste") or 0)
        input_tokens = int(row.get("input_tokens") or 0)
        output_tokens = int(row.get("output_tokens") or 0)
        costo_raw = str(row.get("costo_raw") or "0")

        if richieste == 0 and input_tokens == 0 and output_tokens == 0 and Decimal(costo_raw) == 0:
            continue

        data_it = row.get("data")  # es. 26/05/2026
        giorno, mese, anno = data_it.split("/")
        data_iso = f"{anno}-{mese}-{giorno}"
        mese_iso = f"{anno}-{mese}"

        conn.execute("""
            INSERT INTO openai_usage_giornaliero
                (data, mese, costo_raw, richieste, input_tokens, output_tokens, aggiornato_il)
            VALUES
                (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(data) DO UPDATE SET
                mese = excluded.mese,
                costo_raw = excluded.costo_raw,
                richieste = excluded.richieste,
                input_tokens = excluded.input_tokens,
                output_tokens = excluded.output_tokens,
                aggiornato_il = CURRENT_TIMESTAMP
        """, (
            data_iso,
            mese_iso,
            costo_raw,
            richieste,
            input_tokens,
            output_tokens
        ))

    conn.commit()
    conn.close()


def carica_openai_usage_storico():
    ensure_openai_usage_storico_table()

    conn = get_db_connection()

    rows = conn.execute("""
        SELECT
            data,
            mese,
            costo_raw,
            richieste,
            input_tokens,
            output_tokens
        FROM openai_usage_giornaliero
        ORDER BY data DESC
    """).fetchall()

    conn.close()

    daily = []

    for row in rows:
        data_iso = row["data"]
        anno, mese, giorno = data_iso.split("-")

        costo_decimal = Decimal(str(row["costo_raw"] or "0"))

        daily.append({
            "data": f"{giorno}/{mese}/{anno}",
            "mese": row["mese"],
            "costo_raw": str(costo_decimal),
            "costo": format_openai_euro(costo_decimal),
            "richieste": int(row["richieste"] or 0),
            "input_tokens": int(row["input_tokens"] or 0),
            "output_tokens": int(row["output_tokens"] or 0)
        })

    return daily

def calcola_totali_mensili_openai(daily):
    totali = {}

    for row in daily:
        mese = row.get("mese") or row["data"][3:10]

        if mese not in totali:
            totali[mese] = {
                "costo_raw": Decimal("0"),
                "richieste": 0,
                "input_tokens": 0,
                "output_tokens": 0
            }

        totali[mese]["costo_raw"] += Decimal(str(row.get("costo_raw") or "0"))
        totali[mese]["richieste"] += int(row.get("richieste") or 0)
        totali[mese]["input_tokens"] += int(row.get("input_tokens") or 0)
        totali[mese]["output_tokens"] += int(row.get("output_tokens") or 0)

    for mese, dati in totali.items():
        dati["costo"] = format_openai_euro(dati["costo_raw"])

    return totali

@app.route("/admin/sblocca", methods=["GET"])
@admin_required
def admin_unlock():
    """
    Sblocco temporaneo area admin.

    Lo sblocco admin è consentito solo tramite passkey WebAuthn.
    Questa route mostra soltanto la pagina di sblocco.

    La verifica effettiva avviene solo tramite:
    - /admin/passkey/auth/options
    - /admin/passkey/auth/verify

    Non esiste più fallback password lato backend.
    """

    next_url = request.args.get("next") or url_for("admin_dashboard")

    # 🔐 Consenti redirect solo verso percorsi interni esplicitamente autorizzati.
    # Serve per evitare open redirect, ma permette anche operazioni sensibili fuori da /admin
    # come /impostazioni/modifica-password.
    allowed_stepup_prefixes = (
        "/admin",
        "/impostazioni/modifica-password",
    )

    if (
        not next_url.startswith("/")
        or next_url.startswith("//")
        or not next_url.startswith(allowed_stepup_prefixes)
    ):
        next_url = url_for("admin_dashboard")

    return render_template("admin_unlock.html", next_url=next_url)

@app.route("/admin/passkey")
@admin_required
def admin_passkey_page():
    """
    Pagina admin per registrare e vedere le passkey abilitate
    e controllare lo stato dei codici di recupero.
    """
    user_id = int(g.utente["id"])

    passkeys = get_admin_passkeys_for_user(user_id)
    recovery_codes_disponibili = count_unused_admin_recovery_codes(user_id)

    return render_template(
        "admin_passkey.html",
        passkeys=passkeys,
        recovery_codes_disponibili=recovery_codes_disponibili,
        recovery_codes_generati=None
    )

@app.route("/admin/passkey/recovery-codes/genera", methods=["POST"])
@admin_required
def admin_generate_recovery_codes():
    """
    Genera nuovi codici di recupero admin.

    I codici vengono mostrati in chiaro solo una volta,
    subito dopo la generazione. Nel database viene salvato
    solo l'hash.
    """
    verify_csrf()

    user_id = int(g.utente["id"])

    try:
        recovery_codes = generate_admin_recovery_codes(
            user_id=user_id,
            count=8
        )

        bump_admin_security_version(
            user_id=user_id,
            reason="generate_recovery_codes"
        )

        passkeys = get_admin_passkeys_for_user(user_id)
        recovery_codes_disponibili = count_unused_admin_recovery_codes(user_id)

        flash(
            "Nuovi codici di recupero generati. Salvali ora: non saranno più mostrati.",
            "warning"
        )

        return render_template(
            "admin_passkey.html",
            passkeys=passkeys,
            recovery_codes_disponibili=recovery_codes_disponibili,
            recovery_codes_generati=recovery_codes
        )

    except Exception as e:
        log_exception_safe(
            "❌ Errore generazione recovery codes admin",
            e,
            {"user_id": user_id},
            production=True
        )

        flash("Errore durante la generazione dei codici di recupero.", "error")
        return redirect(url_for("admin_passkey_page"))

@app.route("/admin/passkey/<int:passkey_id>/elimina", methods=["POST"])
@admin_required
def admin_passkey_delete(passkey_id):
    """
    Elimina una passkey admin registrata.

    Protezione importante:
    non permette di eliminare l'ultima passkey disponibile,
    altrimenti l'admin potrebbe rimanere bloccato fuori dall'area riservata.
    """
    verify_csrf()

    user_id = int(g.utente["id"])

    try:
        passkeys = get_admin_passkeys_for_user(user_id)

        # 🔒 Sicurezza: non permettere di eliminare l'ultima passkey
        if len(passkeys) <= 1:
            flash(
                "Non puoi eliminare l’ultima passkey admin. "
                "Registra prima una nuova passkey su un altro dispositivo.",
                "warning"
            )
            return redirect(url_for("admin_passkey_page"))

        # 🔒 Sicurezza extra: verifica che la passkey richiesta appartenga davvero all'admin
        passkey_ids = {int(p["id"]) for p in passkeys}

        if int(passkey_id) not in passkey_ids:
            flash("Passkey non trovata o non autorizzata.", "error")
            return redirect(url_for("admin_passkey_page"))

        delete_admin_passkey_for_user(
            passkey_id=passkey_id,
            user_id=user_id
        )

        bump_admin_security_version(
            user_id=user_id,
            reason="delete_admin_passkey"
        )

        clear_admin_stepup()

        flash(
            "Passkey rimossa correttamente. Per sicurezza dovrai confermare di nuovo l’accesso admin.",
            "success"
        )

    except Exception as e:
        log_exception_safe(
            "❌ Errore eliminazione passkey admin",
            e,
            {
                "user_id": user_id,
                "passkey_id": passkey_id
            },
            production=True
        )
        flash("Errore durante la rimozione della passkey.", "error")

    return redirect(url_for("admin_passkey_page"))

@app.route("/admin/passkey/register/options", methods=["POST"])
@admin_required
def admin_passkey_register_options():
    """
    Genera le opzioni WebAuthn per registrare una nuova passkey admin.
    """
    verify_csrf()

    user_id = int(g.utente["id"])
    username = g.utente["username"] or f"admin-{user_id}"

    existing = get_admin_passkeys_for_user(user_id)

    exclude_credentials = []
    exclude_credentials = []
    for p in existing:
        try:
            exclude_credentials.append(
                PublicKeyCredentialDescriptor(
                    id=base64url_to_bytes(p["credential_id"])
                )
            )
        except Exception:
            pass

    options = generate_registration_options(
        rp_id=WEBAUTHN_RP_ID,
        rp_name=WEBAUTHN_RP_NAME,
        user_id=str(user_id).encode("utf-8"),
        user_name=username,
        user_display_name=username,
        exclude_credentials=exclude_credentials,
        authenticator_selection=AuthenticatorSelectionCriteria(
            authenticator_attachment=AuthenticatorAttachment.PLATFORM,
            resident_key=ResidentKeyRequirement.PREFERRED,
            user_verification=UserVerificationRequirement.REQUIRED,
        ),
    )

    session["admin_passkey_registration_challenge"] = base64.b64encode(
        options.challenge
    ).decode()

    session.modified = True

    return app.response_class(
        options_to_json(options),
        mimetype="application/json"
    )


@app.route("/admin/passkey/register/verify", methods=["POST"])
@admin_required
def admin_passkey_register_verify():
    """
    Verifica la risposta del browser e salva la passkey admin.
    """
    verify_csrf()

    challenge_b64 = session.get("admin_passkey_registration_challenge")
    if not challenge_b64:
        return jsonify({
            "ok": False,
            "error": "Sfida passkey scaduta. Riprova."
        }), 400

    expected_challenge = base64.b64decode(challenge_b64)

    data = request.get_json(silent=True) or {}

    nome_dispositivo = (
        data.pop("nome_dispositivo", None)
        or request.headers.get("User-Agent", "Dispositivo")
    )

    try:
        verification = verify_registration_response(
            credential=data,
            expected_challenge=expected_challenge,
            expected_rp_id=WEBAUTHN_RP_ID,
            expected_origin=WEBAUTHN_EXPECTED_ORIGIN,
            require_user_verification=True,
        )

        credential_id_b64url = data.get("id")

        credential_public_key_b64 = base64.b64encode(
            verification.credential_public_key
        ).decode()

        transports = []
        try:
            transports = data.get("response", {}).get("transports") or []
        except Exception:
            transports = []

        save_admin_passkey(
            user_id=int(g.utente["id"]),
            credential_id_b64url=credential_id_b64url,
            credential_public_key_b64=credential_public_key_b64,
            sign_count=verification.sign_count,
            device_type=str(verification.credential_device_type),
            backed_up=bool(verification.credential_backed_up),
            transports=transports,
            nome_dispositivo=nome_dispositivo
        )

        bump_admin_security_version(
            user_id=int(g.utente["id"]),
            reason="register_admin_passkey"
        )

        session.pop("admin_passkey_registration_challenge", None)
        clear_admin_stepup()
        session.modified = True

        return jsonify({
            "ok": True
        })

    except Exception as e:
        log_exception_safe(
            "❌ Errore registrazione passkey admin",
            e,
            {"user_id": int(g.utente["id"]) if g.utente else None},
            production=True
        )

        return jsonify({
            "ok": False,
            "error": "Registrazione passkey non riuscita."
        }), 400

def get_admin_passkey_by_credential_id(user_id, credential_id_b64url):
    """
    Recupera una passkey admin specifica tramite credential_id.

    Filtra anche per utente admin, così una passkey registrata da un admin
    non può essere usata per sbloccare l'area admin di un altro utente.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT
                id,
                utente_id,
                credential_id,
                credential_public_key,
                sign_count,
                device_type,
                backed_up,
                transports,
                nome_dispositivo,
                created_at,
                last_used_at
            FROM admin_passkeys
            WHERE utente_id = ?
              AND credential_id = ?
            LIMIT 1
        """), (
            int(user_id),
            credential_id_b64url
        ))

        return cur.fetchone()

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def update_admin_passkey_usage(passkey_id, new_sign_count):
    """
    Aggiorna contatore di sicurezza e ultimo utilizzo della passkey.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql(f"""
            UPDATE admin_passkeys
            SET sign_count = ?,
                last_used_at = {now_sql()}
            WHERE id = ?
        """), (
            int(new_sign_count or 0),
            int(passkey_id)
        ))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


@app.route("/admin/passkey/auth/options", methods=["POST"])
@admin_required
def admin_passkey_auth_options():
    """
    Genera le opzioni WebAuthn per sbloccare l'area admin con una passkey già registrata.
    """
    verify_csrf()

    user_id = int(g.utente["id"])
    passkeys = get_admin_passkeys_for_user(user_id)

    if not passkeys:
        return jsonify({
            "ok": False,
            "error": "Nessuna passkey registrata per questo admin."
        }), 400

    allow_credentials = []

    for p in passkeys:
        try:
            allow_credentials.append(
                PublicKeyCredentialDescriptor(
                    id=base64url_to_bytes(p["credential_id"])
                )
            )
        except Exception as e:
            log_exception_safe(
                "⚠️ Passkey saltata in auth/options",
                e,
                {"user_id": user_id}
            )

    if not allow_credentials:
        return jsonify({
            "ok": False,
            "error": "Nessuna passkey valida trovata."
        }), 400

    options = generate_authentication_options(
        rp_id=WEBAUTHN_RP_ID,
        allow_credentials=allow_credentials,
        user_verification=UserVerificationRequirement.REQUIRED,
    )

    session["admin_passkey_auth_challenge"] = base64.b64encode(
        options.challenge
    ).decode()

    session.modified = True

    return app.response_class(
        options_to_json(options),
        mimetype="application/json"
    )


@app.route("/admin/passkey/auth/verify", methods=["POST"])
@admin_required
def admin_passkey_auth_verify():
    """
    Verifica la passkey usata dal browser e, se valida, sblocca temporaneamente l'area admin.
    """
    verify_csrf()

    challenge_b64 = session.get("admin_passkey_auth_challenge")
    if not challenge_b64:
        return jsonify({
            "ok": False,
            "error": "Sfida passkey scaduta. Riprova."
        }), 400

    expected_challenge = base64.b64decode(challenge_b64)
    data = request.get_json(silent=True) or {}

    credential_id_b64url = data.get("id")
    if not credential_id_b64url:
        return jsonify({
            "ok": False,
            "error": "Credential ID mancante."
        }), 400

    security_log(
        "🔐 [PASSKEY AUTH] richiesta autenticazione passkey",
        {
            "user_id": int(g.utente["id"]),
            "credential_id": credential_id_b64url,
            "user_agent": request.headers.get("User-Agent", "")
        }
    )

    passkey = get_admin_passkey_by_credential_id(
        user_id=int(g.utente["id"]),
        credential_id_b64url=credential_id_b64url
    )

    if not passkey:
        return jsonify({
            "ok": False,
            "error": "Passkey non riconosciuta per questo admin."
        }), 400

    security_log(
        "✅ [PASSKEY AUTH] passkey trovata",
        {
            "passkey_id": passkey["id"],
            "nome_dispositivo": passkey["nome_dispositivo"],
            "credential_id": passkey["credential_id"],
            "last_used_at": passkey["last_used_at"]
        }
    )

    try:
        verification = verify_authentication_response(
            credential=data,
            expected_challenge=expected_challenge,
            expected_rp_id=WEBAUTHN_RP_ID,
            expected_origin=WEBAUTHN_EXPECTED_ORIGIN,
            credential_public_key=base64.b64decode(passkey["credential_public_key"]),
            credential_current_sign_count=int(passkey["sign_count"] or 0),
            require_user_verification=True,
        )

        update_admin_passkey_usage(
            passkey_id=passkey["id"],
            new_sign_count=verification.new_sign_count
        )

        session.pop("admin_passkey_auth_challenge", None)

        # ✅ Questo è il punto chiave: la passkey sostituisce la password nello step-up admin.
        mark_admin_stepup_verified()

        session.modified = True

        return jsonify({
            "ok": True
        })

    except Exception as e:
        security_log(
            "❌ Errore autenticazione passkey admin",
            {
                "user_id": int(g.utente["id"]) if g.utente else None,
                "credential_id": credential_id_b64url,
                "error_type": type(e).__name__,
                "error": repr(e),
                "user_agent": request.headers.get("User-Agent", "")
            }
        )

        if os.getenv("APP_ENV", "production").lower() in ("local", "development"):
            traceback.print_exc()

        return jsonify({
            "ok": False,
            "error": "Autenticazione passkey non riuscita."
        }), 400

# ==========================================================
# 🔐 ADMIN RECOVERY CODE — RATE LIMIT + ALERT SICUREZZA
# ==========================================================

ADMIN_RECOVERY_MAX_ATTEMPTS_USER = 5
ADMIN_RECOVERY_MAX_ATTEMPTS_IP = 12
ADMIN_RECOVERY_WINDOW_SECONDS = 15 * 60  # 15 minuti


def get_client_ip():
    """
    Recupera IP reale dietro proxy Render/Cloudflare-like.
    Non lo usiamo come unica protezione, ma come secondo fattore di rate limit.
    """
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()

    real_ip = request.headers.get("X-Real-IP", "")
    if real_ip:
        return real_ip.strip()

    return request.remote_addr or "unknown"


def admin_recovery_rate_keys(user_id):
    """
    Rate limit separato:
    - per utente admin
    - per IP
    Così blocchiamo sia brute force sul singolo account sia tentativi dallo stesso client.
    """
    ip = get_client_ip()

    return {
        "user": f"admin_recovery_rl:user:{int(user_id)}",
        "ip": f"admin_recovery_rl:ip:{ip}",
        "ip_value": ip
    }


def check_admin_recovery_rate_limit(user_id):
    """
    Ritorna:
    - True, None se può tentare
    - False, messaggio se deve essere bloccato
    """
    keys = admin_recovery_rate_keys(user_id)

    try:
        user_attempts_raw = redis_client.get(keys["user"])
        ip_attempts_raw = redis_client.get(keys["ip"])

        user_attempts = int(user_attempts_raw or 0)
        ip_attempts = int(ip_attempts_raw or 0)

        if user_attempts >= ADMIN_RECOVERY_MAX_ATTEMPTS_USER:
            ttl = redis_client.ttl(keys["user"])
            minuti = max(1, int((ttl or ADMIN_RECOVERY_WINDOW_SECONDS) / 60))
            return False, f"Troppi tentativi non riusciti. Riprova tra circa {minuti} minuti."

        if ip_attempts >= ADMIN_RECOVERY_MAX_ATTEMPTS_IP:
            ttl = redis_client.ttl(keys["ip"])
            minuti = max(1, int((ttl or ADMIN_RECOVERY_WINDOW_SECONDS) / 60))
            return False, f"Troppi tentativi da questo dispositivo/rete. Riprova tra circa {minuti} minuti."

        return True, None

    except Exception as e:
        log_exception_safe(
            "⚠️ Errore check_admin_recovery_rate_limit",
            e,
            {"user_id": user_id},
            production=True
        )
        return True, None

def register_failed_admin_recovery_attempt(user_id):
    """
    Incrementa i contatori dopo un codice errato.
    """
    keys = admin_recovery_rate_keys(user_id)

    try:
        for key in (keys["user"], keys["ip"]):
            current = redis_client.incr(key)

            # Alla prima scrittura imposto la finestra temporale.
            if current == 1:
                redis_client.expire(key, ADMIN_RECOVERY_WINDOW_SECONDS)

    except Exception as e:
        log_exception_safe(
            "⚠️ Errore register_failed_admin_recovery_attempt",
            e,
            {"user_id": user_id},
            production=True
        )


def clear_admin_recovery_rate_limit(user_id):
    """
    Cancella i contatori dopo un recovery code valido.
    """
    keys = admin_recovery_rate_keys(user_id)

    try:
        redis_client.delete(keys["user"])
        redis_client.delete(keys["ip"])
    except Exception as e:
        log_exception_safe(
            "⚠️ Errore clear_admin_recovery_rate_limit",
            e,
            {"user_id": user_id},
            production=True
        )

def notify_admin_recovery_code_used(user_id, recovery_code_id):
    """
    Avvisa l'admin che un recovery code è stato usato.

    Importante:
    - non blocca lo sblocco admin se email/notifica falliscono;
    - non include mai il codice usato;
    - salva una notifica interna e invia una email di sicurezza.
    """
    ip = get_client_ip()
    user_agent = request.headers.get("User-Agent", "unknown")
    quando = datetime.now(ZoneInfo("Europe/Rome")).strftime("%d/%m/%Y %H:%M:%S")

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = get_cursor(conn)

        cur.execute(sql("""
            SELECT email, nome, username
            FROM utenti
            WHERE id = ?
            LIMIT 1
        """), (int(user_id),))

        admin = cur.fetchone()

        if not admin:
            return

        email_admin = admin["email"]
        nome_admin = admin["nome"] or admin["username"] or "admin"

        titolo = "Codice di recupero admin utilizzato"
        messaggio = (
            "È stato utilizzato un codice di recupero per sbloccare l’area amministratore.\n\n"
            f"Data e ora: {quando}\n"
            f"IP: {ip}\n"
            f"Dispositivo/browser: {user_agent[:180]}\n\n"
            "Se sei stato tu, non devi fare nulla. "
            "Se non riconosci questa attività, accedi subito e rigenera i codici di recupero."
        )

        # Notifica interna
        try:
            _crea_notifica(
                int(user_id),
                titolo,
                messaggio,
                tipo="sicurezza",
                link=url_for("admin_passkey_page")
            )

            emit_update_notifications(int(user_id))

        except Exception as e:
            log_exception_safe(
                "⚠️ Errore notifica interna recovery code",
                e,
                {"user_id": user_id},
                production=True
            )

        # Email sicurezza
        try:
            msg = Message(
                subject="Avviso sicurezza MyLocalCare: recovery code admin usato",
                recipients=[email_admin],
                sender=app.config.get("MAIL_DEFAULT_SENDER"),
                reply_to=MAIL_FROM_ADDRESS
            )

            msg.body = (
                f"Ciao {nome_admin},\n\n"
                "ti informiamo che è stato utilizzato un codice di recupero per sbloccare "
                "temporaneamente l’area amministratore di MyLocalCare.\n\n"
                f"Data e ora: {quando}\n"
                f"IP: {ip}\n"
                f"Dispositivo/browser: {user_agent[:180]}\n\n"
                "Se sei stato tu, puoi ignorare questa email.\n\n"
                "Se non riconosci questa attività, accedi subito all’area admin, "
                "rigenera i codici di recupero e verifica le passkey registrate.\n\n"
                "MyLocalCare"
            )

            threading.Thread(
                target=send_async_email,
                args=(app, msg),
                daemon=True
            ).start()

        except Exception as e:
            log_exception_safe(
                "⚠️ Errore email recovery code used",
                e,
                {"user_id": user_id, "email": email_admin},
                production=True
            )

        security_log(
            "🔐 [ADMIN RECOVERY ALERT] alert inviato",
            {
                "user_id": int(user_id),
                "recovery_code_id": int(recovery_code_id),
                "ip": ip,
            },
            production=True
        )

    except Exception as e:
        log_exception_safe(
            "⚠️ Errore notify_admin_recovery_code_used",
            e,
            {"user_id": user_id},
            production=True
        )

    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass

        try:
            if conn:
                conn.close()
        except Exception:
            pass

def normalize_admin_recovery_code(raw_code):
    """
    Normalizza un recovery code admin accettando:
    - LC-MSYD6H-EQ0IDC-53L9F4
    - MSYD6H-EQ0IDC-53L9F4
    - LCMSYD6HEQ0IDC53L9F4
    - spazi accidentali
    """
    code = (raw_code or "").strip().upper()

    # rimuove spazi
    code = code.replace(" ", "")

    # rimuove prefisso LC- se presente
    if code.startswith("LC-"):
        code = code[3:]
    elif code.startswith("LC"):
        code = code[2:]

    # rimuove trattini
    code = code.replace("-", "")

    return code

@app.route("/admin/recovery-code/verify", methods=["POST"])
@admin_required
def admin_recovery_code_verify():
    """
    Verifica un recovery code admin monouso.

    Protezioni:
    - CSRF;
    - rate limit per admin/user_id;
    - rate limit per IP;
    - codice monouso;
    - alert email + notifica interna dopo uso valido.
    """
    verify_csrf()

    user_id = int(g.utente["id"])

    # 🔒 Rate limit PRIMA di controllare il codice
    allowed, limit_message = check_admin_recovery_rate_limit(user_id)

    if not allowed:
        security_log(
            "⛔ [ADMIN RECOVERY] rate limit attivo",
            {
                "user_id": user_id,
                "ip": get_client_ip(),
                "user_agent": request.headers.get("User-Agent", "")
            },
            production=True
        )

        return jsonify({
            "ok": False,
            "error": limit_message or "Troppi tentativi. Riprova più tardi."
        }), 429

    data = request.get_json(silent=True) or {}
    raw_code = (data.get("code") or "").strip()

    if not raw_code:
        register_failed_admin_recovery_attempt(user_id)

        return jsonify({
            "ok": False,
            "error": "Inserisci un codice di recupero."
        }), 400

    # Normalizzazione coerente:
    # accetta codice con spazi accidentali.
    # Manteniamo il comportamento attuale per non rompere i codici già generati/testati.
    normalized_code = raw_code.replace(" ", "").replace("-", "").strip().upper()

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT id, code_hash, used
            FROM admin_recovery_codes
            WHERE utente_id = ?
              AND used = 0
            ORDER BY id ASC
        """), (user_id,))

        codes = cur.fetchall()

        matched_code_id = None

        for row in codes:
            try:
                # Primo tentativo: comportamento attuale
                if check_password_hash(row["code_hash"], normalized_code):
                    matched_code_id = int(row["id"])
                    break

                # Secondo tentativo: compatibilità con codici salvati con trattini/prefisso LC-
                # Utile se alcuni codici sono stati hashati nel formato visualizzato.
                raw_upper = raw_code.strip().upper().replace(" ", "")
                if check_password_hash(row["code_hash"], raw_upper):
                    matched_code_id = int(row["id"])
                    break

            except Exception:
                continue

        if matched_code_id is None:
            register_failed_admin_recovery_attempt(user_id)

            security_log(
                "❌ [ADMIN RECOVERY] codice non valido",
                {
                    "user_id": user_id,
                    "ip": get_client_ip(),
                    "user_agent": request.headers.get("User-Agent", "")
                },
                production=True
            )

            return jsonify({
                "ok": False,
                "error": "Codice di recupero non valido o già utilizzato."
            }), 400

        cur.execute(sql(f"""
            UPDATE admin_recovery_codes
            SET used = 1,
                used_at = {now_sql()}
            WHERE id = ?
              AND utente_id = ?
              AND used = 0
        """), (
            matched_code_id,
            user_id
        ))

        # Se per concorrenza il codice fosse già stato usato tra SELECT e UPDATE,
        # non dobbiamo sbloccare.
        try:
            affected = cur.rowcount
        except Exception:
            affected = 1

        if affected != 1:
            conn.rollback()
            register_failed_admin_recovery_attempt(user_id)

            return jsonify({
                "ok": False,
                "error": "Codice di recupero già utilizzato."
            }), 400

        conn.commit()

        # ✅ Codice valido: azzera rate limit
        clear_admin_recovery_rate_limit(user_id)

        # 🔐 Un recovery code è un evento critico:
        # invalida tutte le altre sessioni admin sensibili.
        bump_admin_security_version(
            user_id=user_id,
            reason="recovery_code_used"
        )

        # ✅ Sblocca temporaneamente admin solo per la sessione corrente
        mark_admin_stepup_verified()
        session.modified = True

        security_log(
            "✅ [ADMIN RECOVERY] codice recovery usato",
            {
                "user_id": user_id,
                "recovery_code_id": matched_code_id,
                "ip": get_client_ip(),
                "user_agent": request.headers.get("User-Agent", "")
            },
            production=True
        )

        # ✅ Alert sicurezza, senza bloccare lo sblocco se fallisce
        try:
            notify_admin_recovery_code_used(
                user_id=user_id,
                recovery_code_id=matched_code_id
            )
        except Exception as e:
            log_exception_safe(
                "⚠️ Alert recovery code non inviato",
                e,
                {"user_id": user_id},
                production=True
            )

        return jsonify({
            "ok": True
        })

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass

        security_log(
            "❌ Errore verifica recovery code admin",
            {
                "user_id": user_id,
                "error_type": type(e).__name__,
                "error": repr(e),
                "ip": get_client_ip(),
                "user_agent": request.headers.get("User-Agent", "")
            },
            production=True
        )

        if os.getenv("APP_ENV", "production").lower() in ("local", "development"):
            traceback.print_exc()

        return jsonify({
            "ok": False,
            "error": "Errore durante la verifica del codice di recupero."
        }), 500

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

@app.route("/admin/counters")
@admin_required
def admin_counters():
    cache = app.config["_ADMIN_COUNTERS_CACHE"]
    ttl = app.config["_ADMIN_COUNTERS_TTL"]
    now = time.time()

    if cache["payload"] is not None and (now - cache["ts"] < ttl):
        return jsonify(cache["payload"])

    def get_count(cur, query, params=None, step=""):
        """
        Esegue una query COUNT/valore singolo e legge sempre la colonna alias 'valore'.
        Evita fetchone_value(), row[0], list(row.values())[0].
        """
        cur.execute(sql(query), params or ())
        row = cur.fetchone()

        if not row:
            return 0

        try:
            return int(row["valore"] or 0)
        except Exception as e:
            raise RuntimeError(f"Errore lettura valore admin_counters step={step}: {repr(e)}")

    step = "start"

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = get_cursor(conn)

        step = "annunci"
        pending_annunci = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM annunci
            WHERE stato = 'in_attesa'
        """, step=step)

        step = "recensioni"
        pending_recensioni = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM recensioni
            WHERE stato = 'in_attesa'
        """, step=step)

        step = "risposte_recensioni"
        pending_risposte = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM risposte_recensioni
            WHERE stato = 'in_attesa'
        """, step=step)

        pending_recensioni_totali = pending_recensioni + pending_risposte

        step = "utenti"

        try:
            totale_utenti = get_count(cur, """
                SELECT COUNT(*) AS valore
                FROM utenti
                WHERE attivo = 1
                  AND sospeso = 0
                  AND COALESCE(disattivato_admin, 0) = 0
                  AND COALESCE(email, '') NOT LIKE ?
                  AND COALESCE(username, '') NOT LIKE ?
            """, (
                "deleted_user_%@mylocalcare.local",
                "utente_eliminato_%"
            ), step=step)

        except Exception as e:
            log_exception_safe(
                "❌ Errore admin_counters solo step=utenti",
                e,
                production=True
            )

            totale_utenti = 0

        step = "messaggi"
        messaggi_non_letti = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM messaggi_chat
            WHERE letto = 0
        """, step=step)

        step = "video_minuti"

        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                SELECT COALESCE((
                    SELECT minuti_totali
                    FROM video_limiti_mensili
                    WHERE mese = TO_CHAR(NOW(), 'YYYY-MM')
                    LIMIT 1
                ), 0) AS valore
            """)
        else:
            cur.execute(f"""
                SELECT COALESCE((
                    SELECT minuti_totali
                    FROM video_limiti_mensili
                    WHERE mese = {month_sql()}
                    LIMIT 1
                ), 0) AS valore
            """)

        row_video = cur.fetchone()

        try:
            video_minuti = int(row_video["valore"] or 0) if row_video else 0
        except Exception as e:
            raise RuntimeError(f"Errore lettura valore admin_counters step=video_minuti: {repr(e)}")

        step = "openai_costo"
        openai_costo = get_openai_month_cost()

        step = "statistiche_annunci_totali"

        annunci_totali = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM annunci
            WHERE COALESCE(stato, '') <> 'eliminato'
        """, step=step)

        step = "statistiche_utenti_con_annunci"
        utenti_con_annunci = get_count(cur, """
            SELECT COUNT(DISTINCT utente_id) AS valore
            FROM annunci
            WHERE COALESCE(stato, '') <> 'eliminato'
        """, step=step)

        step = "statistiche_utenti_senza_annunci"
        utenti_senza_annunci = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM utenti u
            WHERE u.attivo = 1
              AND u.sospeso = 0
              AND COALESCE(u.disattivato_admin, 0) = 0
              AND COALESCE(u.email, '') NOT LIKE ?
              AND COALESCE(u.username, '') NOT LIKE ?
              AND NOT EXISTS (
                  SELECT 1
                  FROM annunci a
                  WHERE a.utente_id = u.id
                    AND COALESCE(a.stato, '') <> 'eliminato'
              )
        """, (
            "deleted_user_%@mylocalcare.local",
            "utente_eliminato_%"
        ), step=step)

        step = "statistiche_utenti_recensiti"
        utenti_recensiti = get_count(cur, """
            SELECT COUNT(DISTINCT id_destinatario) AS valore
            FROM recensioni
            WHERE stato = 'approvato'
        """, step=step)

        step = "statistiche_recensioni_con_risposta"
        recensioni_con_risposta = get_count(cur, """
            SELECT COUNT(DISTINCT id_recensione) AS valore
            FROM risposte_recensioni
        """, step=step)

        step = "statistiche_chat_totali"
        chat_totali = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM (
                SELECT
                    CASE
                        WHEN mittente_id < destinatario_id THEN mittente_id
                        ELSE destinatario_id
                    END AS a,
                    CASE
                        WHEN mittente_id > destinatario_id THEN mittente_id
                        ELSE destinatario_id
                    END AS b
                FROM messaggi_chat
                GROUP BY a, b
            ) AS chat_uniche
        """, step=step)

        step = "statistiche_messaggi_inviati"
        messaggi_inviati = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM messaggi_chat
        """, step=step)

        step = "statistiche_notifiche_ricevute"
        notifiche_ricevute = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM notifiche
        """, step=step)

        step = "statistiche_notifiche_da_leggere"
        notifiche_da_leggere = get_count(cur, """
            SELECT COUNT(*) AS valore
            FROM notifiche
            WHERE letta = 0
        """, step=step)

        payload = {
            "utenti": totale_utenti,
            "annunci": pending_annunci,
            "recensioni": pending_recensioni_totali,
            "risposte": pending_risposte,
            "messaggi": messaggi_non_letti,
            "totale": pending_annunci + pending_recensioni_totali,
            "video_minuti": video_minuti,

            "openai_costo": format_openai_euro(openai_costo),

            "statistiche": {
                "utenti_attivi": totale_utenti,
                "annunci_totali": annunci_totali,
                "utenti_con_annunci": utenti_con_annunci,
                "utenti_senza_annunci": utenti_senza_annunci,
                "utenti_recensiti": utenti_recensiti,
                "recensioni_con_risposta": recensioni_con_risposta,
                "chat_totali": chat_totali,
                "messaggi_inviati": messaggi_inviati,
                "messaggi_non_letti": messaggi_non_letti,
                "notifiche_ricevute": notifiche_ricevute,
                "notifiche_da_leggere": notifiche_da_leggere
            }
        }

        cache["payload"] = payload
        cache["ts"] = now

        return jsonify(payload)

    except Exception as e:
        log_exception_safe(
            f"❌ Errore admin_counters step={step}",
            e,
            production=True
        )

        payload = {
            "ok": False,
            "error": "admin_counters_failed",
            "step": step,
            "utenti": 0,
            "annunci": 0,
            "recensioni": 0,
            "risposte": 0,
            "messaggi": 0,
            "totale": 0,
            "video_minuti": 0,
            "openai_costo": "—"
        }

        return jsonify(payload), 200

    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass

        try:
            if conn:
                conn.close()
        except Exception:
            pass

@app.route("/admin/counters/page")
@admin_required
def admin_counters_page():
    return render_template("admin_counters_page.html")

@app.route("/admin/openai-usage")
@admin_required
def admin_openai_usage():
    stats = get_openai_month_stats()

    if stats.get("ok"):
        salva_openai_usage_giornaliero(stats)
        stats["daily"] = carica_openai_usage_storico()
        stats["monthly_totals"] = calcola_totali_mensili_openai(stats["daily"])

    return render_template(
        "admin_openai_usage.html",
        stats=stats
    )

@app.route("/admin/openai-usage/export")
@admin_required
def admin_openai_usage_export():
    daily = carica_openai_usage_storico()

    wb = Workbook()
    ws = wb.active
    ws.title = "OpenAI Usage"

    headers = [
        "Data",
        "Mese",
        "Costo",
        "Richieste",
        "Input token",
        "Output token"
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="E2E8F0")
    header_font = Font(bold=True, color="0F172A")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in daily:
        ws.append([
            row.get("data"),
            row.get("mese"),
            float(row.get("costo_raw") or 0),
            int(row.get("richieste") or 0),
            int(row.get("input_tokens") or 0),
            int(row.get("output_tokens") or 0)
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border

    for cell in ws["C"]:
        cell.number_format = '€ #,##0.000000'

    for cell in ws["D"]:
        cell.number_format = '#,##0'

    for cell in ws["E"]:
        cell.number_format = '#,##0'

    for cell in ws["F"]:
        cell.number_format = '#,##0'

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="openai_usage_storico.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/cron/openai-usage/monthly-save")
def cron_openai_usage_monthly_save():
    secret = request.args.get("secret")

    if not secret or secret != os.getenv("OPENAI_CRON_SECRET"):
        abort(403)

    stats = get_openai_month_stats()

    if stats.get("ok"):
        salva_openai_usage_giornaliero(stats)

        return jsonify({
            "ok": True,
            "message": "Storico OpenAI salvato correttamente",
            "total_requests": stats.get("total_requests"),
            "total_cost": stats.get("total_cost_raw")
        })

    return jsonify({
        "ok": False,
        "error": stats.get("error")
    }), 500

# ==========================================================
# NOTIFICHE: LETTURA SINGOLA
# ==========================================================
@app.route('/notifiche/leggi/<int:id>', methods=["GET", "POST"])
@login_required
def leggi_notifica(id):
    if request.method == "POST":
        verify_csrf()

    segna_notifica_letta(id, g.utente['id'])

    # 🔔 aggiorna il badge in tempo reale
    emit_update_notifications(g.utente['id'])

    if request.method == "POST":
        return jsonify({"success": True})

    flash("Notifica segnata come letta.")
    return redirect(url_for('notifiche'))

# ==========================================================
# ANNUNCI – VISTA SINGOLA + TOGGLE STATO
# ==========================================================
@app.route("/admin/annunci/<int:id>")
@admin_required
def admin_visualizza_annuncio(id):
    conn = get_db_connection()
    c = get_cursor(conn)
    c.execute(sql("""
        SELECT a.*, u.nome, u.cognome, u.email, u.username
        FROM annunci a
        JOIN utenti u ON a.utente_id = u.id
        WHERE a.id = ?
    """), (id,))
    annuncio = c.fetchone()


    if not annuncio:
        return "Annuncio non trovato", 404

    return render_template("admin_visualizza_annuncio.html", annuncio=annuncio)


@app.route("/admin/annunci/toggle/<int:id>")
@admin_required
def toggle_annuncio(id):
    conn = get_db_connection()
    c = get_cursor(conn)
    c.execute(sql("SELECT stato FROM annunci WHERE id = ?"), (id,))
    row = c.fetchone()
    if not row:

        flash("Annuncio non trovato.", "error")
        return redirect(url_for("admin_annunci"))

    nuovo_stato = "disattivato" if row["stato"] == "approvato" else "approvato"
    c.execute(sql("UPDATE annunci SET stato = ? WHERE id = ?"), (nuovo_stato, id))
    conn.commit()


    flash(f"Annuncio {nuovo_stato}.", "info")
    next_url = request.args.get("next")
    if next_url and next_url.startswith("/admin/annunci"):
        return redirect(next_url)

    return redirect(url_for("admin_annunci"))

# ==========================
#   SEZIONI ADMIN COMPLETE
# ==========================

# 🔹 ROOT ADMIN → reindirizza sempre alla dashboard unificata
@app.route("/admin")
@admin_required
def admin():
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    """Dashboard principale operatori"""
    categoria = request.args.get('categoria', '').strip()
    if categoria:
        categoria = categoria.replace("-", " ").replace("%26", "&").strip().lower()

    zona = request.args.get('zona')

    operatori = get_operatori(categoria, zona)
    utenti = get_utenti()

    return render_template(
        "admin_dashboard.html",
        operatori=operatori,
        utenti=utenti,
        categoria=categoria,
        zona=zona
    )

from flask import render_template, request, redirect, url_for, flash
import sqlite3
from datetime import datetime, timedelta
# ==========================================================
# 🎥 ADMIN — VIDEO CALLS STORICO COMPLETO
# ==========================================================
def get_daily_meetings_month():
    api_key = os.getenv("DAILY_API_KEY")

    if not api_key:
        return {
            "ok": False,
            "error": "DAILY_API_KEY mancante",
            "meetings": [],
            "total_meetings": 0,
            "total_duration_seconds": 0,
            "total_participant_minutes": 0
        }

    try:
        now = datetime.now(timezone.utc)
        start_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc)

        start_time = int(start_month.timestamp())
        end_time = int(now.timestamp())

        response = requests.get(
            "https://api.daily.co/v1/meetings",
            headers={
                "Authorization": f"Bearer {api_key}"
            },
            params={
                "timeframe_start": start_time,
                "timeframe_end": end_time,
                "limit": 100
            },
            timeout=15
        )

        if response.status_code != 200:
            return {
                "ok": False,
                "error": response.text,
                "meetings": [],
                "total_meetings": 0,
                "total_duration_seconds": 0,
                "total_participant_minutes": 0
            }

        data = response.json()
        meetings = []

        total_duration_seconds = 0
        total_participant_minutes = 0

        for item in data.get("data", []):
            start_ts = item.get("start_time")
            duration = int(item.get("duration") or 0)
            room = item.get("room") or "—"
            ongoing = bool(item.get("ongoing"))

            participants = item.get("participants") or {}
            participant_count = len(participants) if isinstance(participants, dict) else 0

            participant_minutes = 0
            if participants and isinstance(participants, dict):
                for _, p in participants.items():
                    join_time = p.get("join_time")
                    leave_time = p.get("leave_time")

                    if join_time and leave_time:
                        participant_minutes += max(0, int((leave_time - join_time) / 60))

            if participant_minutes == 0 and participant_count > 0 and duration > 0:
                participant_minutes = int((duration * participant_count) / 60)

            total_duration_seconds += duration
            total_participant_minutes += participant_minutes

            meetings.append({
                "id": item.get("id"),
                "room": room,
                "start_date": datetime.fromtimestamp(start_ts, timezone.utc).strftime("%d/%m/%Y") if start_ts else "—",
                "start_time": datetime.fromtimestamp(start_ts, timezone.utc).strftime("%H:%M") if start_ts else "—",
                "duration_minutes": int(duration / 60),
                "ongoing": ongoing,
                "participant_count": participant_count,
                "participant_minutes": participant_minutes
            })

        return {
            "ok": True,
            "error": None,
            "meetings": meetings,
            "total_meetings": len(meetings),
            "total_duration_seconds": total_duration_seconds,
            "total_participant_minutes": total_participant_minutes
        }

    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "meetings": [],
            "total_meetings": 0,
            "total_duration_seconds": 0,
            "total_participant_minutes": 0
        }


@app.route("/admin/video-calls")
@admin_required
def admin_video_calls():

    conn = get_db_connection()
    cur = get_cursor(conn)

    rows = cur.execute(sql(f"""
        SELECT
            v.*,
            u1.username AS user1_name,
            u2.username AS user2_name,
            {month_sql("v.created_at")} AS mese
        FROM video_call_log v
        LEFT JOIN utenti u1 ON u1.id = v.utente_1
        LEFT JOIN utenti u2 ON u2.id = v.utente_2
        ORDER BY v.created_at DESC
    """)).fetchall()

    limiti = cur.execute(sql("""
        SELECT mese, minuti_totali, costo_totale_cent
        FROM video_limiti_mensili
    """)).fetchall()

    limiti_dict = {l["mese"]: l for l in limiti}

    def to_local_dt(value):
        if not value:
            return None

        if isinstance(value, datetime):
            return value + timedelta(hours=1)

        if isinstance(value, str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f"):
                try:
                    return datetime.strptime(value, fmt) + timedelta(hours=1)
                except ValueError:
                    pass

        return None

    mesi = {}

    for r in rows:
        mese = r["mese"]

        if mese not in mesi:
            limite = limiti_dict.get(mese)

            used = limite["minuti_totali"] if limite and limite["minuti_totali"] is not None else 0
            costo = limite["costo_totale_cent"] if limite and limite["costo_totale_cent"] is not None else 0

            mesi[mese] = {
                "calls": [],
                "participant_used": used,
                "participant_remaining": max(0, 10000 - used),
                "costo": costo
            }

        call = dict(r)

        start_local = to_local_dt(call.get("created_at"))
        if start_local:
            call["start_date"] = start_local.strftime("%d/%m/%Y")
            call["start_time"] = start_local.strftime("%H:%M:%S")
        else:
            call["start_date"] = "-"
            call["start_time"] = "-"

        end_local = to_local_dt(call.get("ended_at"))
        if end_local:
            call["end_time"] = end_local.strftime("%H:%M:%S")
        else:
            call["end_time"] = "-"

        mesi[mese]["calls"].append(call)

    daily_stats = get_daily_meetings_month()

    return render_template(
        "admin_video_calls.html",
        mesi=mesi,
        video_calls_enabled=is_video_calls_enabled(),
        daily_stats=daily_stats
    )

# ---------------------------------------------------------
# 💰 ADMIN - SERVIZI (MONETIZZAZIONE) - SOLO CONFIG (STEP 3)
# ---------------------------------------------------------
@app.route("/admin/servizi")
@admin_required
def admin_servizi():
    conn = get_db_connection()

    c = get_cursor(conn)

    c.execute(sql("""
        SELECT id, codice, nome, descrizione, ambito, target,
               durata_default_giorni, ripetibile, attivabile_admin, attivo, created_at
        FROM servizi
        ORDER BY created_at DESC
    """))
    servizi = c.fetchall()


    return render_template(
    "admin_servizi.html",
    servizi=servizi,
    tab="servizi"
)


@app.route("/admin/servizi/nuovo", methods=["GET", "POST"])
@admin_required
def admin_servizi_nuovo():
    if request.method == "POST":
        codice = (request.form.get("codice") or "").strip()
        nome = (request.form.get("nome") or "").strip()
        descrizione = (request.form.get("descrizione") or "").strip()

        ambito = (request.form.get("ambito") or "").strip()
        target = (request.form.get("target") or "").strip()

        durata_raw = (request.form.get("durata_default_giorni") or "").strip()
        durata = int(durata_raw) if durata_raw.isdigit() else None

        ripetibile = 1 if request.form.get("ripetibile") == "1" else 0
        attivabile_admin = 1 if request.form.get("attivabile_admin") == "1" else 0
        attivo = 1 if request.form.get("attivo") == "1" else 0

        # validazioni minime (solo per evitare record rotti)
        allowed_ambiti = {"annuncio", "profilo", "chat", "homepage"}
        allowed_target = {"offre", "cerca", "entrambi"}

        if not codice or not nome:
            flash("Codice e Nome sono obbligatori.", "error")
            return redirect(url_for("admin_servizi_nuovo"))

        if ambito not in allowed_ambiti:
            flash("Ambito non valido.", "error")
            return redirect(url_for("admin_servizi_nuovo"))

        if target not in allowed_target:
            flash("Target non valido.", "error")
            return redirect(url_for("admin_servizi_nuovo"))

        conn = get_db_connection()
        c = get_cursor(conn)
        try:
            c.execute(sql("""
                INSERT INTO servizi
                (codice, nome, descrizione, ambito, target,
                 durata_default_giorni, ripetibile, attivabile_admin, attivo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """), (codice, nome, descrizione, ambito, target,
                  durata, ripetibile, attivabile_admin, attivo))
            conn.commit()
            flash("Servizio creato.", "success")
            return redirect(url_for("admin_servizi"))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Codice già esistente (deve essere univoco).", "error")
            return redirect(url_for("admin_servizi_nuovo"))
        finally:
            try:
                conn.close()
            except:
                pass


    # GET
    return render_template("admin_servizi_form.html", servizio=None)


@app.route("/admin/servizi/<int:id>/modifica", methods=["GET", "POST"])
@admin_required
def admin_servizi_modifica(id):
    conn = get_db_connection()

    c = get_cursor(conn)

    c.execute(sql("SELECT * FROM servizi WHERE id = ?"), (id,))
    servizio = c.fetchone()

    if not servizio:

        flash("Servizio non trovato.", "error")
        return redirect(url_for("admin_servizi"))

    if request.method == "POST":
        codice = (request.form.get("codice") or "").strip()
        nome = (request.form.get("nome") or "").strip()
        descrizione = (request.form.get("descrizione") or "").strip()

        ambito = (request.form.get("ambito") or "").strip()
        target = (request.form.get("target") or "").strip()

        durata_raw = (request.form.get("durata_default_giorni") or "").strip()
        durata = int(durata_raw) if durata_raw.isdigit() else None

        ripetibile = 1 if request.form.get("ripetibile") == "1" else 0
        attivabile_admin = 1 if request.form.get("attivabile_admin") == "1" else 0
        attivo = 1 if request.form.get("attivo") == "1" else 0

        allowed_ambiti = {"annuncio", "profilo", "chat", "homepage"}
        allowed_target = {"offre", "cerca", "entrambi"}

        if not codice or not nome:

            flash("Codice e Nome sono obbligatori.", "error")
            return redirect(url_for("admin_servizi_modifica", id=id))

        if ambito not in allowed_ambiti:

            flash("Ambito non valido.", "error")
            return redirect(url_for("admin_servizi_modifica", id=id))

        if target not in allowed_target:

            flash("Target non valido.", "error")
            return redirect(url_for("admin_servizi_modifica", id=id))

        try:
            c.execute(sql("""
                UPDATE servizi
                SET codice = ?, nome = ?, descrizione = ?,
                    ambito = ?, target = ?,
                    durata_default_giorni = ?,
                    ripetibile = ?, attivabile_admin = ?, attivo = ?
                WHERE id = ?
            """), (codice, nome, descrizione, ambito, target,
                  durata, ripetibile, attivabile_admin, attivo, id))
            conn.commit()
            flash("Servizio aggiornato.", "success")
            return redirect(url_for("admin_servizi"))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash("Codice già esistente (deve essere univoco).", "error")
            return redirect(url_for("admin_servizi_modifica", id=id))
        finally:
            try:
                conn.close()
            except:
                pass



    return render_template("admin_servizi_form.html", servizio=servizio)


@app.route("/admin/servizi/<int:id>/toggle")
@admin_required
def admin_servizi_toggle(id):
    conn = get_db_connection()

    c = get_cursor(conn)

    c.execute(sql("SELECT attivo FROM servizi WHERE id = ?"), (id,))
    row = c.fetchone()
    if not row:

        flash("Servizio non trovato.", "error")
        return redirect(url_for("admin_servizi"))

    nuovo = 0 if row["attivo"] == 1 else 1
    c.execute(sql("UPDATE servizi SET attivo = ? WHERE id = ?"), (nuovo, id))
    conn.commit()


    flash("Stato servizio aggiornato.", "success")
    return redirect(url_for("admin_servizi"))

@app.route("/admin/servizi/<int:servizio_id>/elimina", methods=["POST"])
@admin_required
def admin_servizi_elimina(servizio_id):
    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("""
        UPDATE servizi
        SET attivo = 0
        WHERE id = ?
    """), (servizio_id,))

    conn.commit()


    flash("Servizio disattivato correttamente.", "success")
    return redirect(url_for("admin_servizi"))

def crea_acquisto_admin(conn, utente_id, tipo, ref_id, annuncio_id=None, prezzo_id=None):
    cur = get_cursor(conn)

    acquisto_id = insert_and_get_id(
        cur,
        f"""
        INSERT INTO acquisti (
            utente_id,
            tipo,
            ref_id,
            prezzo_id,
            annuncio_id,
            importo_cent,
            metodo,
            stato,
            riferimento_esterno,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, 0, 'admin', 'paid', ?, {now_sql()})
        """,
        (
            int(utente_id),
            tipo,
            int(ref_id),
            int(prezzo_id) if prezzo_id else None,
            int(annuncio_id) if annuncio_id else None,
            f"admin-{uuid.uuid4().hex[:12]}"
        )
    )

    return acquisto_id

@app.route("/admin/toggle-servizio", methods=["POST"])
@admin_required
def admin_toggle_servizio():
    verify_csrf()

    data = request.get_json(silent=True) or {}

    codice_servizio = (data.get("codice_servizio") or "").strip()
    annuncio_id = data.get("annuncio_id")
    utente_id = data.get("utente_id")

    if not codice_servizio or not utente_id:
        return jsonify({"ok": False, "error": "Parametri mancanti"}), 400

    conn = get_db_connection()


    try:
        # 1️⃣ servizio
        servizio = conn.execute(sql("""
            SELECT id, ambito, attivo
            FROM servizi
            WHERE codice = ?
            LIMIT 1
        """), (codice_servizio,)).fetchone()

        if not servizio or servizio["attivo"] != 1:
            return jsonify({"ok": False, "error": "Servizio non valido o disattivo"}), 400

        ambito = servizio["ambito"]

        # 🔒 Alcuni servizi sono SEMPRE di profilo anche se vengono cliccati
        # dalla pagina admin_annunci. Replichiamo il comportamento della pagina admin_utenti:
        # - contatti
        # - badge_affidabilita
        # non devono mai richiedere annuncio_id.
        if codice_servizio in ("contatti", "badge_affidabilita"):
            ambito = "profilo"

        # 2️⃣ cerca attivazione attiva
        if ambito == "annuncio":

            if not annuncio_id:
                return jsonify({"ok": False, "error": "annuncio_id obbligatorio"}), 400

            attiva = conn.execute(sql(f"""
                SELECT id
                FROM attivazioni_servizi
                WHERE servizio_id = ?
                  AND annuncio_id = ?
                  AND stato = 'attivo'
                  AND data_inizio <= {now_sql()}
                  AND (data_fine IS NULL OR data_fine > {now_sql()})
                LIMIT 1
            """), (servizio["id"], annuncio_id)).fetchone()

        else:

            # ✅ Replica esatta del comportamento di /admin/utenti:
            # per contatti e badge_affidabilita controlliamo il servizio
            # sull'utente, senza pretendere annuncio_id IS NULL.
            # Questo evita falsi "non attivo" se esiste già una vecchia
            # attivazione collegata all'utente.
            if codice_servizio in ("contatti", "badge_affidabilita"):
                attiva = conn.execute(sql(f"""
                    SELECT id
                    FROM attivazioni_servizi
                    WHERE servizio_id = ?
                      AND utente_id = ?
                      AND stato = 'attivo'
                      AND data_inizio <= {now_sql()}
                      AND (data_fine IS NULL OR data_fine > {now_sql()})
                    LIMIT 1
                """), (servizio["id"], utente_id)).fetchone()

            else:
                attiva = conn.execute(sql(f"""
                    SELECT id
                    FROM attivazioni_servizi
                    WHERE servizio_id = ?
                      AND utente_id = ?
                      AND annuncio_id IS NULL
                      AND stato = 'attivo'
                      AND data_inizio <= {now_sql()}
                      AND (data_fine IS NULL OR data_fine > {now_sql()})
                    LIMIT 1
                """), (servizio["id"], utente_id)).fetchone()

        # 3️⃣ toggle
        if attiva:
            ok, msg = revoca_attivazione(attiva["id"], eseguito_da="admin")

            security_log(
                "🟦 /admin/toggle-servizio revoca",
                {
                    "ok": ok,
                    "msg": msg,
                    "codice_servizio": codice_servizio,
                    "utente_id": utente_id,
                    "annuncio_id": annuncio_id,
                    "attivazione_id": attiva["id"]
                },
                production=True
            )

            return jsonify({
                "ok": ok,
                "azione": "disattivato",
                "messaggio": msg
            })
        else:
            # =====================================================
            # FIX: "contatti" deve comportarsi come servizio profilo
            # anche se nel DB il suo ambito storico risulta diverso.
            #
            # Non modifichiamo il DB:
            # inseriamo manualmente l'attivazione profilo
            # con utente_id valorizzato e annuncio_id NULL,
            # esattamente come poi viene letta da admin_utenti/admin_annunci.
            # =====================================================
            if codice_servizio == "contatti":
                try:
                    cur = get_cursor(conn)

                    acquisto_admin_id = crea_acquisto_admin(
                        conn=conn,
                        utente_id=int(utente_id),
                        tipo="servizio",
                        ref_id=int(servizio["id"]),
                        annuncio_id=None,
                        prezzo_id=None
                    )

                    att_id = insert_and_get_id(
                        cur,
                        f"""
                        INSERT INTO attivazioni_servizi (
                            acquisto_id,
                            servizio_id,
                            utente_id,
                            annuncio_id,
                            stato,
                            data_inizio,
                            data_fine,
                            attivato_da
                        )
                        VALUES (?, ?, ?, NULL, 'attivo', {now_sql()}, NULL, ?)
                        """,
                        (
                            acquisto_admin_id,
                            int(servizio["id"]),
                            int(utente_id),
                            "admin"
                        )
                    )

                    conn.commit()

                    ok = True
                    msg = "Servizio attivato."

                except Exception as e:
                    try:
                        conn.rollback()
                    except Exception:
                        pass

                    log_exception_safe(
                        "❌ Errore attivazione manuale servizio contatti",
                        e,
                        {
                            "codice_servizio": codice_servizio,
                            "utente_id": utente_id,
                            "annuncio_id": annuncio_id,
                            "servizio_id": servizio["id"]
                        },
                        production=True
                    )

                    return jsonify({
                        "ok": False,
                        "error": "Errore durante l’attivazione del servizio contatti."
                    }), 500

            else:
                acquisto_admin_id = crea_acquisto_admin(
                    conn=conn,
                    utente_id=int(utente_id),
                    tipo="servizio",
                    ref_id=int(servizio["id"]),
                    annuncio_id=int(annuncio_id) if ambito == "annuncio" else None,
                    prezzo_id=None
                )

                ok, msg, att_id = attiva_servizio(
                    conn=conn,
                    utente_id=int(utente_id),
                    codice_servizio=codice_servizio,
                    annuncio_id=int(annuncio_id) if ambito == "annuncio" else None,
                    acquisto_id=acquisto_admin_id,
                    attivato_da="admin",
                    note="Attivazione manuale admin"
                )

                if ok:
                    conn.commit()
                else:
                    conn.rollback()

            # 🔔 NOTIFICA URGENTE — SOLO SE HA SENSO
            if ok and codice_servizio == "annuncio_urgente" and annuncio_id:
                try:
                    notifica_urgente(
                        annuncio_id=int(annuncio_id),
                        attivazione_id=att_id,
                        eseguito_da="admin"
                    )
                except Exception as e:
                    # ⚠️ Non blocca il toggle se la notifica fallisce
                    print(f"⚠️ Errore notifica urgente: {e}")

            security_log(
                "🟩 /admin/toggle-servizio attivazione",
                {
                    "ok": ok,
                    "msg": msg,
                    "codice_servizio": codice_servizio,
                    "utente_id": utente_id,
                    "annuncio_id": annuncio_id,
                    "ambito": ambito,
                    "attivazione_id": att_id
                },
                production=True
            )

            return jsonify({
                "ok": ok,
                "azione": "attivato",
                "messaggio": msg,
                "attivazione_id": att_id
            })

    except Exception as e:
        log_exception_safe(
            "❌ ERRORE /admin/toggle-servizio",
            e,
            {
                "codice_servizio": codice_servizio,
                "utente_id": utente_id,
                "annuncio_id": annuncio_id,
                "ambito": ambito if "ambito" in locals() else None,
                "servizio_id": servizio["id"] if "servizio" in locals() and servizio else None
            },
            production=True
        )

        return jsonify({
            "ok": False,
            "error": str(e),
            "debug": {
                "codice_servizio": codice_servizio,
                "utente_id": utente_id,
                "annuncio_id": annuncio_id,
                "ambito": ambito if "ambito" in locals() else None
            }
        }), 500

@app.route("/admin/servizi/<int:servizio_id>/piani")
@admin_required
def admin_servizi_piani(servizio_id):
    conn = get_db_connection()


    servizio = conn.execute(sql("""
        SELECT id, codice, nome
        FROM servizi
        WHERE id = ?
    """), (servizio_id,)).fetchone()

    if not servizio:

        flash("Servizio non trovato.", "error")
        return redirect(url_for("admin_servizi"))

    piani = conn.execute(sql("""
        SELECT *
        FROM servizi_piani
        WHERE servizio_id = ?
        ORDER BY ordine ASC, durata_giorni ASC
    """), (servizio_id,)).fetchall()



    return render_template(
        "admin_servizi_piani.html",
        servizio=servizio,
        piani=piani
    )

# ===============================
# 📦 LISTA PACCHETTI
# ===============================
@app.route("/admin/pacchetti")
@admin_required
def admin_pacchetti():
    conn = get_db_connection()


    pacchetti = conn.execute(sql("""
        SELECT *
        FROM pacchetti
        ORDER BY created_at DESC
    """)).fetchall()



    return render_template(
        "admin_pacchetti.html",
        pacchetti=pacchetti,
        tab="pacchetti"
    )


# ===============================
# ➕ NUOVO PACCHETTO
# ===============================
@app.route("/admin/pacchetti/nuovo", methods=["GET", "POST"])
@admin_required
def admin_pacchetti_nuovo():
    conn = get_db_connection()


    servizi = conn.execute(sql("""
        SELECT id, codice, nome
        FROM servizi
        WHERE attivo = 1
        ORDER BY nome
    """)).fetchall()

    if request.method == "POST":
        codice = request.form.get("codice")
        nome = request.form.get("nome")
        descrizione = request.form.get("descrizione")
        attivo = 1 if request.form.get("attivo") else 0
        servizi_selezionati = request.form.getlist("servizi")

        cur = get_cursor(conn)

        pacchetto_id = insert_and_get_id(
            cur,
            """
            INSERT INTO pacchetti (codice, nome, descrizione, attivo)
            VALUES (?, ?, ?, ?)
            """,
            (codice, nome, descrizione, attivo)
        )

        for sid in servizi_selezionati:
            cur.execute(sql("""
                INSERT INTO pacchetti_servizi (pacchetto_id, servizio_id)
                VALUES (?, ?)
            """), (pacchetto_id, sid))

        conn.commit()


        flash("Pacchetto creato.", "success")
        return redirect(url_for("admin_pacchetti"))


    return render_template(
        "admin_pacchetti_form.html",
        pacchetto=None,
        servizi=servizi,
        servizi_attivi=[],
        tab="pacchetti"
    )


# ===============================
# ✏️ MODIFICA PACCHETTO
# ===============================
@app.route("/admin/pacchetti/<int:id>/modifica", methods=["GET", "POST"])
@admin_required
def admin_pacchetti_modifica(id):
    conn = get_db_connection()

    pacchetto = conn.execute(
        sql("SELECT * FROM pacchetti WHERE id = ?"),
        (id,)
    ).fetchone()

    if not pacchetto:
        flash("Pacchetto non trovato.", "error")
        return redirect(url_for("admin_pacchetti"))

    servizi = conn.execute(
        sql("""
            SELECT id, codice, nome
            FROM servizi
            WHERE attivo = 1
            ORDER BY nome
        """)
    ).fetchall()

    servizi_attivi = [
        r["servizio_id"]
        for r in conn.execute(
            sql("""
                SELECT servizio_id
                FROM pacchetti_servizi
                WHERE pacchetto_id = ?
            """),
            (id,)
        ).fetchall()
    ]

    if request.method == "POST":
        codice = request.form.get("codice")
        nome = request.form.get("nome")
        descrizione = request.form.get("descrizione")
        attivo = 1 if request.form.get("attivo") else 0
        servizi_selezionati = request.form.getlist("servizi")

        cur = get_cursor(conn)

        # UPDATE pacchetto
        cur.execute(
            sql("""
                UPDATE pacchetti
                SET codice = ?, nome = ?, descrizione = ?, attivo = ?
                WHERE id = ?
            """),
            (codice, nome, descrizione, attivo, id)
        )

        # RESET servizi collegati
        cur.execute(
            sql("DELETE FROM pacchetti_servizi WHERE pacchetto_id = ?"),
            (id,)
        )

        # RE-INSERIMENTO servizi selezionati
        for sid in servizi_selezionati:
            cur.execute(
                sql("""
                    INSERT INTO pacchetti_servizi (pacchetto_id, servizio_id)
                    VALUES (?, ?)
                """),
                (id, sid)
            )

        conn.commit()

        flash("Pacchetto aggiornato.", "success")
        return redirect(url_for("admin_pacchetti"))

    return render_template(
        "admin_pacchetti_form.html",
        pacchetto=pacchetto,
        servizi=servizi,
        servizi_attivi=servizi_attivi,
        tab="pacchetti"
    )

# ===============================
# 🔁 TOGGLE PACCHETTO
# ===============================
@app.route("/admin/pacchetti/<int:id>/toggle")
@admin_required
def admin_toggle_pacchetto_tabella(id):
    conn = get_db_connection()

    conn.execute(sql("""
        UPDATE pacchetti
        SET attivo = CASE WHEN attivo = 1 THEN 0 ELSE 1 END
        WHERE id = ?
    """), (id,))

    conn.commit()


    flash("Stato pacchetto aggiornato.", "success")
    return redirect(url_for("admin_pacchetti"))

# ===============================
# ➕ NUOVO PIANO SERVIZIO
# ===============================
@app.route("/admin/servizi/<int:servizio_id>/piani/nuovo", methods=["GET", "POST"])
@admin_required
def admin_servizi_piani_nuovo(servizio_id):
    conn = get_db_connection()


    servizio = conn.execute(
        "SELECT id, nome FROM servizi WHERE id = ?",
        (servizio_id,)
    ).fetchone()

    if not servizio:

        flash("Servizio non trovato.", "error")
        return redirect(url_for("admin_servizi"))

    if request.method == "POST":
        # 🔢 durata: numero o permanente (NULL)
        if request.form.get("permanente"):
            durata_giorni = None
        else:
            durata_giorni = int(request.form.get("durata_giorni"))

        # 💶 conversione € → cent
        prezzo_euro_raw = request.form.get("prezzo_euro", "0").replace(",", ".")
        prezzo_cent = int(round(float(prezzo_euro_raw) * 100))

        data = (
            servizio_id,
            request.form.get("codice"),
            request.form.get("nome"),
            request.form.get("descrizione"),
            durata_giorni,
            prezzo_cent,
            int(request.form.get("ordine", 1)),
            1 if request.form.get("evidenziato") else 0,
            1 if request.form.get("consigliato") else 0,
            1 if request.form.get("attivo") else 0,
        )

        try:
            conn.execute(sql("""
                INSERT INTO servizi_piani
                (servizio_id, codice, nome, descrizione,
                 durata_giorni, prezzo_cent, ordine,
                 evidenziato, consigliato, attivo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """), data)
            conn.commit()
            flash("Piano creato.", "success")
            return redirect(
                url_for("admin_servizi_piani", servizio_id=servizio_id)
            )
        except sqlite3.IntegrityError:
            flash("Codice piano già esistente per questo servizio.", "error")


    return render_template(
        "admin_servizi_piani_form.html",
        servizio=servizio,
        piano=None
    )


# ===============================
# ✏️ MODIFICA PIANO SERVIZIO
# ===============================
@app.route("/admin/servizi/piani/<int:piano_id>/modifica", methods=["GET", "POST"])
@admin_required
def admin_servizi_piani_modifica(piano_id):
    conn = get_db_connection()


    piano = conn.execute(sql("""
        SELECT p.*, s.nome AS servizio_nome
        FROM servizi_piani p
        JOIN servizi s ON s.id = p.servizio_id
        WHERE p.id = ?
    """), (piano_id,)).fetchone()

    if not piano:

        flash("Piano non trovato.", "error")
        return redirect(url_for("admin_servizi"))

    if request.method == "POST":
        # 🔢 durata: numero o permanente (NULL)
        if request.form.get("permanente"):
            durata_giorni = None
        else:
            durata_giorni = int(request.form.get("durata_giorni"))

        # 💶 conversione € → cent
        prezzo_euro_raw = request.form.get("prezzo_euro", "0").replace(",", ".")
        prezzo_cent = int(round(float(prezzo_euro_raw) * 100))

        conn.execute(sql("""
            UPDATE servizi_piani
            SET codice = ?, nome = ?, descrizione = ?,
                durata_giorni = ?, prezzo_cent = ?, ordine = ?,
                evidenziato = ?, consigliato = ?, attivo = ?
            WHERE id = ?
        """), (
            request.form.get("codice"),
            request.form.get("nome"),
            request.form.get("descrizione"),
            durata_giorni,
            prezzo_cent,
            int(request.form.get("ordine", 1)),
            1 if request.form.get("evidenziato") else 0,
            1 if request.form.get("consigliato") else 0,
            1 if request.form.get("attivo") else 0,
            piano_id
        ))

        conn.commit()

        flash("Piano aggiornato.", "success")
        return redirect(
            url_for("admin_servizi_piani", servizio_id=piano["servizio_id"])
        )


    return render_template(
        "admin_servizi_piani_form.html",
        servizio={
            "id": piano["servizio_id"],
            "nome": piano["servizio_nome"]
        },
        piano=piano
    )

@app.route("/admin/servizi/piani/<int:piano_id>/toggle")
@admin_required
def admin_servizi_piani_toggle(piano_id):
    conn = get_db_connection()

    conn.execute(sql("""
        UPDATE servizi_piani
        SET attivo = CASE WHEN attivo = 1 THEN 0 ELSE 1 END
        WHERE id = ?
    """), (piano_id,))

    conn.commit()


    flash("Stato piano aggiornato.", "success")
    return redirect(request.referrer or url_for("admin_servizi"))

PACCHETTI = {
    "pacchetto_visibilita": [
        "boost_lista",
        "badge_evidenza",
        "contatti",
    ],
    "visibilita_premium": [
        "boost_lista",
        "vetrina_annuncio",
        "badge_evidenza",
        "contatti",
    ],
}

@app.route("/admin/toggle-pacchetto", methods=["POST"])
@admin_required
def admin_toggle_pacchetto():
    verify_csrf()

    data = request.get_json(silent=True) or {}

    codice_pacchetto = (data.get("codice_pacchetto") or "").strip()
    utente_id = data.get("utente_id")
    annuncio_id = data.get("annuncio_id")

    if not codice_pacchetto or not utente_id:
        return jsonify({"ok": False, "error": "Parametri mancanti"}), 400

    if codice_pacchetto not in PACCHETTI:
        return jsonify({"ok": False, "error": "Pacchetto non valido"}), 400

    servizi_pacchetto = PACCHETTI[codice_pacchetto]

    conn = get_db_connection()


    try:
        # =========================
        # 1️⃣ verifica se il pacchetto è già attivo
        # (basta che UNO dei servizi sia attivo)
        # =========================
        pacchetto_attivo = False
        attivazioni_attive = []

        for codice_servizio in servizi_pacchetto:
            servizio = conn.execute(sql("""
                SELECT id, ambito
                FROM servizi
                WHERE codice = ?
                  AND attivo = 1
            """), (codice_servizio,)).fetchone()

            if not servizio:
                continue

            ambito = servizio["ambito"]

            if ambito == "annuncio":
                row = conn.execute(sql(f"""
                    SELECT id
                    FROM attivazioni_servizi
                    WHERE servizio_id = ?
                      AND annuncio_id = ?
                      AND stato = 'attivo'
                      AND data_inizio <= {now_sql()}
                      AND (data_fine IS NULL OR data_fine > {now_sql()})
                    LIMIT 1
                """), (servizio["id"], annuncio_id)).fetchone()
            else:
                row = conn.execute(sql(f"""
                    SELECT id
                    FROM attivazioni_servizi
                    WHERE servizio_id = ?
                      AND utente_id = ?
                      AND annuncio_id IS NULL
                      AND stato = 'attivo'
                      AND data_inizio <= {now_sql()}
                      AND (data_fine IS NULL OR data_fine > {now_sql()})
                    LIMIT 1
                """), (servizio["id"], utente_id)).fetchone()

            if row:
                pacchetto_attivo = True
                attivazioni_attive.append(row["id"])

        # =========================
        # 2️⃣ SE ATTIVO → REVOCA
        # =========================
        if pacchetto_attivo:
            for att_id in attivazioni_attive:
                revoca_attivazione(
                    attivazione_id=att_id,
                    eseguito_da="admin",
                    note=f"Revoca pacchetto {codice_pacchetto}"
                )


            return jsonify({
                "ok": True,
                "azione": "disattivato",
                "pacchetto": codice_pacchetto
            })

        # =========================
        # 3️⃣ SE NON ATTIVO → ATTIVA TUTTI I SERVIZI
        # =========================
        pacchetto_db = conn.execute(sql("""
            SELECT id
            FROM pacchetti
            WHERE codice = ?
            LIMIT 1
        """), (codice_pacchetto,)).fetchone()

        if not pacchetto_db:
            return jsonify({"ok": False, "error": "Pacchetto non trovato nel database"}), 400

        acquisto_admin_id = crea_acquisto_admin(
            conn=conn,
            utente_id=int(utente_id),
            tipo="pacchetto",
            ref_id=int(pacchetto_db["id"]),
            annuncio_id=int(annuncio_id) if annuncio_id else None,
            prezzo_id=None
        )

        attivati = []

        for codice_servizio in servizi_pacchetto:
            ok, msg, att_id = attiva_servizio(
                conn=conn,
                utente_id=int(utente_id),
                codice_servizio=codice_servizio,
                annuncio_id=int(annuncio_id) if annuncio_id else None,
                acquisto_id=acquisto_admin_id,
                attivato_da="admin",
                note=f"Attivazione tramite pacchetto {codice_pacchetto}"
            )

            if ok:
                attivati.append(att_id)

        conn.commit()

        return jsonify({
            "ok": True,
            "azione": "attivato",
            "pacchetto": codice_pacchetto,
            "attivazioni": attivati
        })

    except Exception as e:

        return jsonify({"ok": False, "error": str(e)}), 500

# ===============================
# 📦 LISTA PIANI PACCHETTO
# ===============================
@app.route("/admin/pacchetti/<int:pacchetto_id>/piani")
@admin_required
def admin_pacchetti_piani(pacchetto_id):
    conn = get_db_connection()


    pacchetto = conn.execute(
        "SELECT id, nome FROM pacchetti WHERE id = ?",
        (pacchetto_id,)
    ).fetchone()

    if not pacchetto:

        flash("Pacchetto non trovato.", "error")
        return redirect(url_for("admin_pacchetti"))

    piani = conn.execute(sql("""
        SELECT *
        FROM pacchetti_piani
        WHERE pacchetto_id = ?
        ORDER BY ordine ASC, created_at ASC
    """), (pacchetto_id,)).fetchall()



    return render_template(
        "admin_pacchetti_piani.html",
        pacchetto=pacchetto,
        piani=piani
    )

# ===============================
# ➕ NUOVO PIANO PACCHETTO
# ===============================
@app.route("/admin/pacchetti/<int:pacchetto_id>/piani/nuovo", methods=["GET", "POST"])
@admin_required
def admin_pacchetti_piani_nuovo(pacchetto_id):
    conn = get_db_connection()


    pacchetto = conn.execute(
        "SELECT id, nome FROM pacchetti WHERE id = ?",
        (pacchetto_id,)
    ).fetchone()

    if not pacchetto:

        flash("Pacchetto non trovato.", "error")
        return redirect(url_for("admin_pacchetti"))

    if request.method == "POST":
        durata = None if request.form.get("permanente") else int(request.form["durata_giorni"])
        prezzo_cent = int(round(float(request.form["prezzo_euro"].replace(",", ".")) * 100))

        data = (
            pacchetto_id,
            request.form["codice"],
            request.form["nome"],
            request.form.get("descrizione"),
            durata,
            prezzo_cent,
            int(request.form.get("ordine", 1)),
            1 if request.form.get("evidenziato") else 0,
            1 if request.form.get("consigliato") else 0,
            1 if request.form.get("attivo") else 0,
        )

        try:
            conn.execute(sql("""
                INSERT INTO pacchetti_piani
                (pacchetto_id, codice, nome, descrizione,
                 durata_giorni, prezzo_cent, ordine,
                 evidenziato, consigliato, attivo)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """), data)
            conn.commit()
            flash("Piano pacchetto creato.", "success")
            return redirect(url_for("admin_pacchetti_piani", pacchetto_id=pacchetto_id))
        except sqlite3.IntegrityError:
            flash("Codice piano già esistente per questo pacchetto.", "error")


    return render_template(
        "admin_pacchetti_piani_form.html",
        pacchetto=pacchetto,
        piano=None
    )

# ===============================
# ✏️ MODIFICA PIANO PACCHETTO
# ===============================
@app.route("/admin/pacchetti/piani/<int:id>/modifica", methods=["GET", "POST"])
@admin_required
def admin_pacchetti_piani_modifica(id):
    conn = get_db_connection()


    piano = conn.execute(
        "SELECT * FROM pacchetti_piani WHERE id = ?",
        (id,)
    ).fetchone()

    if not piano:

        flash("Piano non trovato.", "error")
        return redirect(url_for("admin_pacchetti"))

    pacchetto = conn.execute(
        "SELECT id, nome FROM pacchetti WHERE id = ?",
        (piano["pacchetto_id"],)
    ).fetchone()

    if request.method == "POST":
        durata = None if request.form.get("permanente") else int(request.form["durata_giorni"])
        prezzo_cent = int(round(float(request.form["prezzo_euro"].replace(",", ".")) * 100))

        conn.execute(sql("""
            UPDATE pacchetti_piani
            SET codice = ?, nome = ?, descrizione = ?,
                durata_giorni = ?, prezzo_cent = ?, ordine = ?,
                evidenziato = ?, consigliato = ?, attivo = ?
            WHERE id = ?
        """), (
            request.form["codice"],
            request.form["nome"],
            request.form.get("descrizione"),
            durata,
            prezzo_cent,
            int(request.form.get("ordine", 1)),
            1 if request.form.get("evidenziato") else 0,
            1 if request.form.get("consigliato") else 0,
            1 if request.form.get("attivo") else 0,
            id
        ))
        conn.commit()
        flash("Piano aggiornato.", "success")
        return redirect(url_for("admin_pacchetti_piani", pacchetto_id=pacchetto["id"]))


    return render_template(
        "admin_pacchetti_piani_form.html",
        pacchetto=pacchetto,
        piano=piano
    )

# ===============================
# 🔁 TOGGLE ATTIVO PIANO PACCHETTO
# ===============================
@app.route("/admin/pacchetti/piani/<int:piano_id>/toggle")
@admin_required
def admin_pacchetti_piani_toggle(piano_id):
    conn = get_db_connection()


    piano = conn.execute(
        "SELECT id, pacchetto_id, attivo FROM pacchetti_piani WHERE id = ?",
        (piano_id,)
    ).fetchone()

    if not piano:

        flash("Piano non trovato.", "error")
        return redirect(url_for("admin_pacchetti"))

    nuovo_stato = 0 if piano["attivo"] else 1

    conn.execute(
        "UPDATE pacchetti_piani SET attivo = ? WHERE id = ?",
        (nuovo_stato, piano_id)
    )
    conn.commit()


    flash("Stato piano aggiornato.", "success")
    return redirect(
        url_for("admin_pacchetti_piani", pacchetto_id=piano["pacchetto_id"])
    )

def riordina_filtri_categoria(c, categoria):
    c.execute(sql("""
        SELECT id
        FROM filtri_categoria
        WHERE categoria = ?
        ORDER BY ordine ASC, id ASC
    """), (categoria,))

    rows = c.fetchall()

    for index, row in enumerate(rows, start=1):
        row = dict(row)
        c.execute(sql("""
            UPDATE filtri_categoria
            SET ordine = ?
            WHERE id = ?
        """), (index, row["id"]))

@app.route("/admin/filtri-categoria")
@admin_required
def admin_filtri_categoria():
    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("""
        SELECT id, categoria, filtro, ordine, attivo
        FROM filtri_categoria
        ORDER BY categoria ASC, ordine ASC, filtro ASC
    """))

    rows = [dict(row) for row in c.fetchall()]
    conn.close()

    filtri_per_categoria = {}
    for row in rows:
        categoria = row["categoria"]
        if categoria not in filtri_per_categoria:
            filtri_per_categoria[categoria] = []
        filtri_per_categoria[categoria].append(row)

    return render_template(
        "admin_filtri_categoria.html",
        filtri_per_categoria=filtri_per_categoria,
        categoria_aperta=request.args.get("open", "")
    )

@app.route("/admin/filtri-categoria/aggiungi", methods=["POST"])
@admin_required
def admin_aggiungi_filtro_categoria():
    categoria = request.form.get("categoria", "").strip()
    filtro = request.form.get("filtro", "").strip()
    ordine = request.form.get("ordine", "0").strip()

    if not categoria or not filtro:
        flash("Categoria e filtro sono obbligatori.", "error")
        return redirect(url_for("admin_filtri_categoria"))

    try:
        ordine = int(ordine)
    except ValueError:
        ordine = 0

    conn = get_db_connection()
    c = get_cursor(conn)

    try:
        c.execute(sql("""
            INSERT INTO filtri_categoria (categoria, filtro, ordine, attivo)
            VALUES (?, ?, ?, 1)
        """), (categoria, filtro, ordine))
        conn.commit()
        flash("Filtro aggiunto correttamente.", "success")
    except Exception as e:
        conn.rollback()
        flash("Errore: filtro già presente o non valido.", "error")
        print("Errore aggiunta filtro categoria:", e)
    finally:
        conn.close()

    return redirect(url_for("admin_filtri_categoria", open=categoria))


@app.route("/admin/filtri-categoria/<int:id>/modifica", methods=["POST"])
@admin_required
def admin_modifica_filtro_categoria(id):
    filtro = request.form.get("filtro", "").strip()
    ordine = request.form.get("ordine", "0").strip()

    if not filtro:
        flash("Il nome del filtro non può essere vuoto.", "error")
        return redirect(url_for("admin_filtri_categoria"))

    try:
        ordine = int(ordine)
    except ValueError:
        ordine = 0

    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("""
        SELECT categoria
        FROM filtri_categoria
        WHERE id = ?
    """), (id,))
    row = c.fetchone()

    if not row:
        conn.close()
        flash("Filtro non trovato.", "error")
        return redirect(url_for("admin_filtri_categoria"))

    categoria = dict(row)["categoria"]

    c.execute(sql("""
        UPDATE filtri_categoria
        SET filtro = ?, ordine = ?
        WHERE id = ?
    """), (filtro, ordine, id))

    riordina_filtri_categoria(c, categoria)

    conn.commit()
    conn.close()

    flash("Filtro aggiornato.", "success")
    return redirect(url_for("admin_filtri_categoria", open=categoria))

@app.route("/admin/filtri-categoria/<int:id>/sposta/<direzione>", methods=["POST"])
@admin_required
def admin_sposta_filtro_categoria(id, direzione):
    if direzione not in ("su", "giu"):
        return redirect(url_for("admin_filtri_categoria"))

    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("""
        SELECT id, categoria, ordine
        FROM filtri_categoria
        WHERE id = ?
    """), (id,))

    filtro = c.fetchone()

    if not filtro:
        conn.close()
        return redirect(url_for("admin_filtri_categoria"))

    filtro = dict(filtro)
    categoria = filtro["categoria"]

    riordina_filtri_categoria(c, categoria)

    c.execute(sql("""
        SELECT id, ordine
        FROM filtri_categoria
        WHERE id = ?
    """), (id,))
    filtro = dict(c.fetchone())

    ordine_corrente = filtro["ordine"]
    ordine_target = ordine_corrente - 1 if direzione == "su" else ordine_corrente + 1

    c.execute(sql("""
        SELECT id, ordine
        FROM filtri_categoria
        WHERE categoria = ?
          AND ordine = ?
    """), (categoria, ordine_target))

    vicino = c.fetchone()

    if vicino:
        vicino = dict(vicino)

        c.execute(sql("""
            UPDATE filtri_categoria
            SET ordine = ?
            WHERE id = ?
        """), (ordine_target, id))

        c.execute(sql("""
            UPDATE filtri_categoria
            SET ordine = ?
            WHERE id = ?
        """), (ordine_corrente, vicino["id"]))

    riordina_filtri_categoria(c, categoria)

    conn.commit()
    conn.close()

    return redirect(url_for("admin_filtri_categoria", open=categoria))

@app.route("/admin/filtri-categoria/<int:id>/toggle", methods=["POST"])
@admin_required
def admin_toggle_filtro_categoria(id):
    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("""
        SELECT categoria
        FROM filtri_categoria
        WHERE id = ?
    """), (id,))
    row = c.fetchone()

    if not row:
        conn.close()
        flash("Filtro non trovato.", "error")
        return redirect(url_for("admin_filtri_categoria"))

    categoria = dict(row)["categoria"]

    c.execute(sql("""
        UPDATE filtri_categoria
        SET attivo = CASE WHEN attivo = 1 THEN 0 ELSE 1 END
        WHERE id = ?
    """), (id,))

    conn.commit()
    conn.close()

    flash("Stato filtro aggiornato.", "success")
    return redirect(url_for("admin_filtri_categoria", open=categoria))


@app.route("/admin/filtri-categoria/<int:id>/elimina", methods=["POST"])
@admin_required
def admin_elimina_filtro_categoria(id):
    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("""
        SELECT categoria
        FROM filtri_categoria
        WHERE id = ?
    """), (id,))
    row = c.fetchone()

    if not row:
        conn.close()
        flash("Filtro non trovato.", "error")
        return redirect(url_for("admin_filtri_categoria"))

    categoria = dict(row)["categoria"]

    c.execute(sql("""
        DELETE FROM filtri_categoria
        WHERE id = ?
    """), (id,))

    riordina_filtri_categoria(c, categoria)

    conn.commit()
    conn.close()

    flash("Filtro eliminato.", "success")
    return redirect(url_for("admin_filtri_categoria", open=categoria))

# ==========================================================
# GESTIONE UTENTI
# ==========================================================
@app.route('/admin/toggle_utente/<int:id>')
@admin_required
def toggle_utente(id):
    conn = get_db_connection()
    cur = get_cursor(conn)
    cur.execute(sql("SELECT attivo FROM utenti WHERE id = ?"), (id,))
    user = cur.fetchone()
    if user:
        nuovo_stato = 0 if user['attivo'] else 1
        conn.execute(sql("UPDATE utenti SET attivo = ? WHERE id = ?"), (nuovo_stato, id))
        conn.commit()
        flash("Utente {} correttamente.".format("disattivato" if nuovo_stato == 0 else "attivato"))

    return redirect(url_for('admin'))


@app.route('/admin/elimina_utente/<int:id>')
@admin_required
def elimina_utente_route(id):
    elimina_utente(id)
    flash("Utente eliminato correttamente. Email e username sono stati liberati.")
    return redirect(url_for('admin'))

# ==========================================================
# GESTIONE OPERATORI (vecchio pannello)
# ==========================================================
@app.route('/admin/nuovo', methods=['GET', 'POST'])
@admin_required
def admin_nuovo():
    if request.method == 'POST':
        nome = request.form['nome']
        categoria = request.form['categoria']
        zona = request.form['zona']
        servizi = request.form['servizi']
        prezzo = request.form['prezzo']
        bio = request.form['bio']
        filtri = request.form.getlist('filtri_categoria')
        filtri_str = ", ".join(filtri)
        aggiungi_operatore(nome, categoria, zona, servizi, prezzo, bio, filtri_str)
        return redirect(url_for('admin'))
    return redirect(url_for('admin'))


@app.route('/admin/modifica/<int:id>', methods=['GET', 'POST'])
@admin_required
def admin_modifica(id):
    operatore = get_operatore_by_id(id)
    if not operatore:
        return "Operatore non trovato", 404

    if request.method == 'POST':
        nome = request.form['nome']
        categoria = request.form['categoria']
        zona = request.form['zona']
        servizi = request.form['servizi']
        prezzo = request.form['prezzo']
        bio = request.form['bio']
        filtri = request.form.getlist('filtri_categoria')
        filtri_str = ", ".join(filtri)
        modifica_operatore(id, nome, categoria, zona, servizi, prezzo, bio, filtri_str)
        return redirect(url_for("admin", successo_modifica=True))

    operatore = dict(operatore)
    operatore["filtri_categoria"] = [
        f.strip() for f in operatore.get("filtri_categoria", "").split(",") if f.strip()
    ]
    return render_template('modifica_operatore.html', operatore=operatore, successo=False)


# ==========================================================
# ADMIN – LISTA UTENTI AVANZATA
# ==========================================================
@app.route("/admin/utenti")
@admin_required
def admin_utenti():
    """Elenco utenti registrati con filtri"""

    nome = request.args.get("nome", "").strip() or ""
    email = request.args.get("email", "").strip() or ""
    citta = request.args.get("citta", "").strip() or ""
    provincia = request.args.get("provincia", "").strip() or ""
    stato = request.args.get("stato", "").strip() or ""

    has_filters = any([nome, email, citta, provincia, stato])

    conn = get_db_connection()
    c = get_cursor(conn)

    # Pattern eliminati compatibili con vecchio e nuovo formato
    deleted_email_pattern_1 = "deleted_user_%@deleted.local"
    deleted_email_pattern_2 = "deleted_user_%@mylocalcare.local"
    deleted_username_pattern = "utente_eliminato_%"

    # =====================================================
    # UTENTI TOTALI REALI
    # Esclude gli account anonimizzati/eliminati.
    # =====================================================
    c.execute(sql("""
        SELECT COUNT(*) AS totale
        FROM utenti
        WHERE COALESCE(email, '') NOT ILIKE ?
          AND COALESCE(email, '') NOT ILIKE ?
          AND COALESCE(username, '') NOT ILIKE ?
    """), (
        deleted_email_pattern_1,
        deleted_email_pattern_2,
        deleted_username_pattern
    ))

    totale_utenti = fetchone_value(c.fetchone())

    # =====================================================
    # LISTA UTENTI
    # Default: NON mostra eliminati.
    # Se stato = eliminato, mostra SOLO eliminati.
    # =====================================================
    query = f"""
        SELECT
            u.id,
            u.nome,
            u.cognome,
            u.citta,
            u.provincia,
            u.email,
            u.username,
            u.attivo,
            u.sospeso,
            u.data_creazione,

            CASE
                WHEN COALESCE(u.email, '') ILIKE ?
                  OR COALESCE(u.email, '') ILIKE ?
                  OR COALESCE(u.username, '') ILIKE ?
                THEN 1
                ELSE 0
            END AS eliminato,

            (
              SELECT 1
              FROM attivazioni_servizi a
              JOIN servizi s ON s.id = a.servizio_id
              WHERE a.utente_id = u.id
                AND s.codice = 'contatti'
                AND a.stato = 'attivo'
                AND a.data_inizio <= {now_sql()}
                AND (a.data_fine IS NULL OR a.data_fine > {now_sql()})
              LIMIT 1
            ) AS contatti_attivi,

            (
              SELECT 1
              FROM attivazioni_servizi a
              JOIN servizi s ON s.id = a.servizio_id
              WHERE a.utente_id = u.id
                AND s.codice = 'badge_affidabilita'
                AND a.stato = 'attivo'
                AND a.data_inizio <= {now_sql()}
                AND (a.data_fine IS NULL OR a.data_fine > {now_sql()})
              LIMIT 1
            ) AS affidabilita_attiva

        FROM utenti u
        WHERE 1=1
    """

    params = [
        deleted_email_pattern_1,
        deleted_email_pattern_2,
        deleted_username_pattern
    ]

    # Stato eliminato: mostra solo anonimizzati/eliminati.
    if stato == "eliminato":
        query += """
            AND (
                COALESCE(u.email, '') ILIKE ?
                OR COALESCE(u.email, '') ILIKE ?
                OR COALESCE(u.username, '') ILIKE ?
            )
        """
        params.extend([
            deleted_email_pattern_1,
            deleted_email_pattern_2,
            deleted_username_pattern
        ])

    else:
        # Default e tutti gli altri stati: esclude eliminati.
        query += """
            AND COALESCE(u.email, '') NOT ILIKE ?
            AND COALESCE(u.email, '') NOT ILIKE ?
            AND COALESCE(u.username, '') NOT ILIKE ?
        """
        params.extend([
            deleted_email_pattern_1,
            deleted_email_pattern_2,
            deleted_username_pattern
        ])

    if nome:
        query += " AND (LOWER(u.nome) LIKE ? OR LOWER(u.cognome) LIKE ?)"
        like = f"%{nome.lower()}%"
        params.extend([like, like])

    if email:
        query += " AND LOWER(u.email) LIKE ?"
        params.append(f"%{email.lower()}%")

    if citta:
        query += " AND LOWER(u.citta) LIKE ?"
        params.append(f"%{citta.lower()}%")

    if provincia:
        query += " AND LOWER(u.provincia) LIKE ?"
        params.append(f"%{provincia.lower()}%")

    if stato == "attivo":
        query += " AND u.attivo = 1 AND u.sospeso = 0"
    elif stato == "sospeso":
        query += " AND u.sospeso = 1"
    elif stato == "non_attivo":
        query += " AND u.attivo = 0"
    elif stato == "eliminato":
        pass

    query += " ORDER BY u.id DESC"

    c.execute(sql(query), params)
    utenti = c.fetchall()
    totale_filtrati = len(utenti)

    return render_template(
        "admin_utenti.html",
        utenti=utenti,
        nome=nome,
        email=email,
        citta=citta,
        provincia=provincia,
        stato=stato,
        totale_utenti=totale_utenti,
        totale_filtrati=totale_filtrati,
        has_filters=has_filters
    )

@app.route("/admin/utenti/toggle/<int:id>")
@admin_required
def toggle_utente_admin(id):
    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("SELECT attivo, sospeso FROM utenti WHERE id = ?"), (id,))
    row = c.fetchone()

    if not row:

        flash("Utente non trovato.", "error")
        return redirect(url_for('admin_utenti'))

    attivo, sospeso = row["attivo"], row["sospeso"]

    # 🔄 LOGICA DI ATTIVAZIONE:
    # • Se sospeso → NON può essere attivato
    # • Se attivo → disattiva
    # • Se non attivo → attiva
    if sospeso == 1:
        flash("Impossibile attivare un utente sospeso.", "error")
    else:
        nuovo_stato = 0 if attivo == 1 else 1
        c.execute(sql("UPDATE utenti SET attivo = ? WHERE id = ?"), (nuovo_stato, id))
        conn.commit()

        flash("Stato utente aggiornato.", "success")


    return redirect(url_for("admin_utenti"))

# ==========================================================
# ADMIN – RECENSIONI E RISPOSTE
# ==========================================================
@app.route("/admin/recensioni")
@admin_required
def admin_recensioni():
    """Gestione recensioni e relative risposte con filtri"""
    from models import get_tutte_recensioni_con_risposte

    autore = (request.args.get("autore") or "").strip().lower()
    destinatario = (request.args.get("destinatario") or "").strip().lower()
    voto = (request.args.get("voto") or "").strip()
    stato = (request.args.get("stato") or "").strip().lower()

    try:
        recensioni = get_tutte_recensioni_con_risposte()
    except Exception as e:
        print("Errore caricando recensioni:", e)
        recensioni = []

    recensioni_dict = []
    for r in recensioni:
        if isinstance(r, dict):
            recensioni_dict.append(r)
        else:
            try:
                recensioni_dict.append(dict(r))
            except Exception:
                recensioni_dict.append(r)

    def match(r):
        ok = True

        if autore:
            username_autore = (r.get("autore_username") or "").lower()
            ok = ok and (autore in username_autore)

        if destinatario:
            username_dest = (r.get("dest_username") or "").lower()
            ok = ok and (destinatario in username_dest)

        if voto:
            try:
                ok = ok and str(int(voto)) == str(r.get("voto"))
            except Exception:
                return False

        if stato:
            stato_recensione = str(r.get("stato") or "").lower()
            stato_risposta = str(r.get("risposta_stato") or "").lower()

            ok = ok and (
                stato_recensione == stato
                or stato_risposta == stato
            )

        return ok

    recensioni_filtrate = [r for r in recensioni_dict if match(r)]

    totale_recensioni = len(recensioni_dict)
    totale_filtrate = len(recensioni_filtrate)

    totale_in_attesa = sum(
        1 for r in recensioni_dict
        if str(r.get("stato") or "").lower() == "in_attesa"
    )

    totale_risposte_in_attesa = sum(
        1 for r in recensioni_dict
        if r.get("risposta_id") and str(r.get("risposta_stato") or "").lower() == "in_attesa"
    )

    return render_template(
        "admin_recensioni.html",
        recensioni=recensioni_filtrate,
        active_page="recensioni",
        totale_recensioni=totale_recensioni,
        totale_filtrate=totale_filtrate,
        totale_in_attesa=totale_in_attesa,
        totale_risposte_in_attesa=totale_risposte_in_attesa
    )

from datetime import datetime, timedelta

@app.route("/admin/debug/reminder-profili-incompleti")
@login_required
@admin_required
def admin_debug_reminder_profili_incompleti():
    risultato = invia_reminder_profili_incompleti(dry_run=True)
    return jsonify(risultato)

@app.route("/admin/debug/reminder-profilo-incompleto/<int:user_id>")
@login_required
@admin_required
def admin_debug_reminder_profilo_incompleto_singolo(user_id):
    titolo = "Completa il tuo profilo"
    messaggio = (
        "Seleziona cosa offri o cosa cerchi per ricevere annunci compatibili "
        "e suggerimenti personalizzati."
    )
    link = "/utente/dashboard"

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT id, email, nome, username, email_notifiche
            FROM utenti
            WHERE id = ?
              AND attivo = 1
              AND sospeso = 0
              AND COALESCE(disattivato_admin,0) = 0
            LIMIT 1
        """), (user_id,))

        u = cur.fetchone()

        if not u:
            return jsonify({
                "ok": False,
                "error": "utente_non_trovato_o_non_attivo"
            }), 404

        cur.execute(sql("""
            INSERT INTO notifiche (
                id_utente,
                titolo,
                messaggio,
                link,
                tipo,
                letta
            )
            VALUES (?, ?, ?, ?, ?, 0)
        """), (
            user_id,
            titolo,
            messaggio,
            link,
            "profilo_incompleto_test"
        ))

        conn.commit()

        emit_update_notifications(user_id)

        push_ok = False
        email_ok = False

        try:
            invia_push(
                user_id,
                titolo,
                messaggio,
                url=link
            )
            push_ok = True
        except Exception as e:
            log_exception_safe(
                "⚠️ Errore push test reminder profilo incompleto",
                e,
                {"user_id": user_id},
                production=True
            )

        if int(u["email_notifiche"] or 0) == 1 and u["email"]:
            nome = u["nome"] or u["username"] or "utente"
            base_url = app.config.get("APP_BASE_URL", "https://www.mylocalcare.it").rstrip("/")

            email_ok = _invia_email(
                destinazione=u["email"],
                oggetto="Completa il tuo profilo MyLocalCare",
                corpo=(
                    f"Ciao {nome},\n\n"
                    "Hai creato il tuo account su MyLocalCare, ma non hai ancora indicato "
                    "cosa offri o cosa cerchi.\n\n"
                    "Completa il profilo per ricevere annunci compatibili, suggerimenti "
                    "personalizzati e notifiche più pertinenti.\n\n"
                    f"{base_url}{link}\n\n"
                    "MyLocalCare"
                )
            )

        return jsonify({
            "ok": True,
            "user_id": user_id,
            "push_ok": push_ok,
            "email_ok": email_ok,
            "email_notifiche": int(u["email_notifiche"] or 0),
            "email": u["email"]
        })

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass

        log_exception_safe(
            "❌ Errore test singolo reminder profilo incompleto",
            e,
            {"user_id": user_id},
            production=True
        )

        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

@app.route("/admin/acquisti")
@login_required
@admin_required
def admin_acquisti():
    conn = get_db_connection()

    # filtri lato server
    q = (request.args.get("q") or "").strip().lower()
    filtro_tipo = (request.args.get("tipo") or "").strip().lower()
    filtro_stato = (request.args.get("stato") or "").strip().lower()
    filtro_metodo = (request.args.get("metodo") or "").strip().lower()

    rows = conn.execute(sql("""
        SELECT
            a.id              AS acquisto_id,
            a.tipo,
            a.importo_cent,
            a.metodo,
            a.stato,
            a.created_at,
            a.annuncio_id,
            a.ref_id,
            a.prezzo_id,

            u.id              AS utente_id,
            u.email,
            u.username        AS username,

            an.categoria       AS annuncio_categoria,
            an.provincia       AS annuncio_provincia,

            sp.servizio_id     AS servizio_id_base,
            s.nome             AS servizio_nome,
            s.codice           AS servizio_codice,
            sp.durata_giorni   AS durata_servizio_giorni,

            p.nome             AS pacchetto_nome,
            pp.durata_giorni   AS durata_pacchetto_giorni

        FROM acquisti a
        JOIN utenti u
          ON u.id = a.utente_id

        LEFT JOIN annunci an
          ON an.id = a.annuncio_id

        LEFT JOIN servizi_piani sp
          ON a.tipo = 'servizio'
         AND sp.id = a.prezzo_id

        LEFT JOIN servizi s
          ON s.id = sp.servizio_id

        LEFT JOIN pacchetti p
          ON a.tipo = 'pacchetto'
         AND p.id = a.ref_id

        LEFT JOIN pacchetti_piani pp
          ON a.tipo = 'pacchetto'
         AND pp.id = a.prezzo_id

        WHERE a.stato = 'paid'
        ORDER BY a.created_at DESC
        LIMIT 1000
    """)).fetchall()

    acquisti = []

    for r in rows:
        a = dict(r)

        a["categoria_visuale"] = a.get("annuncio_categoria") or "—"
        a["provincia_visuale"] = a.get("annuncio_provincia") or "—"

        dettagli_rows = conn.execute(sql("""
            SELECT
                ats.id          AS attivazione_id,
                ats.data_inizio,
                ats.data_fine,
                ats.stato,
                sv.id           AS servizio_id,
                sv.codice       AS servizio_codice,
                sv.nome         AS servizio_nome
            FROM attivazioni_servizi ats
            JOIN servizi sv
              ON sv.id = ats.servizio_id
            WHERE ats.acquisto_id = ?
            ORDER BY sv.nome ASC, ats.id DESC
        """), (a["acquisto_id"],)).fetchall()

        dettagli_servizi = [dict(x) for x in dettagli_rows]
        a["dettagli_servizi"] = dettagli_servizi
        a["numero_attivazioni"] = len(dettagli_servizi)

        stati = {d["stato"] for d in dettagli_servizi}

        if "attivo" in stati:
            a["stato_visuale"] = "attivo"
        elif "rinnovato" in stati:
            a["stato_visuale"] = "rinnovato"
        elif "revocato" in stati:
            a["stato_visuale"] = "revocato"
        elif a["stato"] in ("creato", "pending"):
            a["stato_visuale"] = "in_attesa"
        else:
            a["stato_visuale"] = "scaduto"

        if a["tipo"] == "pacchetto":
            a["data_fine"] = None
            a["ha_sottodettaglio"] = len(dettagli_servizi) > 0
            a["durata_iniziale_giorni"] = a.get("durata_pacchetto_giorni")
        else:
            a["ha_sottodettaglio"] = False
            a["data_fine"] = dettagli_servizi[0]["data_fine"] if dettagli_servizi else None
            a["durata_iniziale_giorni"] = a.get("durata_servizio_giorni")

        # filtro testuale
        testo_ricerca = " ".join([
            str(a.get("email") or ""),
            str(a.get("tipo") or ""),
            str(a.get("pacchetto_nome") or ""),
            str(a.get("servizio_nome") or ""),
            str(a.get("metodo") or ""),
            str(a.get("annuncio_id") or ""),
            str(a.get("annuncio_categoria") or ""),
            str(a.get("annuncio_provincia") or "")
        ]).lower()

        if q and q not in testo_ricerca:
            continue

        if filtro_tipo and a["tipo"] != filtro_tipo:
            continue

        if filtro_stato and a["stato_visuale"] != filtro_stato:
            continue

        if filtro_metodo and (a.get("metodo") or "").lower() != filtro_metodo:
            continue

        acquisti.append(a)

    return render_template(
        "admin_acquisti.html",
        acquisti=acquisti,
        tab="acquisti"
    )


@app.route("/admin/acquisti/export.xlsx")
@login_required
@admin_required
def admin_acquisti_export():
    conn = get_db_connection()

    q = (request.args.get("q") or "").strip().lower()
    filtro_tipo = (request.args.get("tipo") or "").strip().lower()
    filtro_stato = (request.args.get("stato") or "").strip().lower()
    filtro_metodo = (request.args.get("metodo") or "").strip().lower()

    def excel_safe_dt(value):
        if value is None:
            return None

        if hasattr(value, "tzinfo"):
            if value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            return value.strftime("%d/%m/%Y %H:%M:%S")

        return value

    rows = conn.execute(sql("""
        SELECT
            a.id              AS acquisto_id,
            a.tipo,
            a.importo_cent,
            a.metodo,
            a.stato,
            a.created_at,
            a.annuncio_id,
            a.ref_id,
            a.prezzo_id,

            u.id              AS utente_id,
            u.email,
            u.username        AS username,

            an.categoria       AS annuncio_categoria,
            an.provincia       AS annuncio_provincia,

            sp.servizio_id     AS servizio_id_base,
            s.nome             AS servizio_nome,
            s.codice           AS servizio_codice,
            sp.durata_giorni   AS durata_servizio_giorni,

            p.nome             AS pacchetto_nome,
            pp.durata_giorni   AS durata_pacchetto_giorni

        FROM acquisti a
        JOIN utenti u
          ON u.id = a.utente_id

        LEFT JOIN annunci an
          ON an.id = a.annuncio_id

        LEFT JOIN servizi_piani sp
          ON a.tipo = 'servizio'
         AND sp.id = a.prezzo_id

        LEFT JOIN servizi s
          ON s.id = sp.servizio_id

        LEFT JOIN pacchetti p
          ON a.tipo = 'pacchetto'
         AND p.id = a.ref_id

        LEFT JOIN pacchetti_piani pp
          ON a.tipo = 'pacchetto'
         AND pp.id = a.prezzo_id

        WHERE a.stato = 'paid'
        ORDER BY a.created_at DESC
        LIMIT 5000
    """)).fetchall()

    records = []

    for r in rows:
        a = dict(r)

        dettagli_rows = conn.execute(sql("""
            SELECT
                ats.id          AS attivazione_id,
                ats.data_inizio,
                ats.data_fine,
                ats.stato,
                sv.id           AS servizio_id,
                sv.codice       AS servizio_codice,
                sv.nome         AS servizio_nome
            FROM attivazioni_servizi ats
            JOIN servizi sv
              ON sv.id = ats.servizio_id
            WHERE ats.acquisto_id = ?
            ORDER BY sv.nome ASC, ats.id DESC
        """), (a["acquisto_id"],)).fetchall()

        dettagli_servizi = [dict(x) for x in dettagli_rows]
        stati = {d["stato"] for d in dettagli_servizi}

        if "attivo" in stati:
            stato_visuale = "attivo"
        elif "rinnovato" in stati:
            stato_visuale = "rinnovato"
        elif "revocato" in stati:
            stato_visuale = "revocato"
        elif a["stato"] in ("creato", "pending"):
            stato_visuale = "in_attesa"
        else:
            stato_visuale = "scaduto"

        durata_iniziale = (
            a.get("durata_pacchetto_giorni")
            if a["tipo"] == "pacchetto"
            else a.get("durata_servizio_giorni")
        )

        testo_ricerca = " ".join([
            str(a.get("email") or ""),
            str(a.get("tipo") or ""),
            str(a.get("pacchetto_nome") or ""),
            str(a.get("servizio_nome") or ""),
            str(a.get("metodo") or ""),
            str(a.get("annuncio_id") or ""),
            str(a.get("annuncio_categoria") or ""),
            str(a.get("annuncio_provincia") or "")
        ]).lower()

        if q and q not in testo_ricerca:
            continue
        if filtro_tipo and a["tipo"] != filtro_tipo:
            continue
        if filtro_stato and stato_visuale != filtro_stato:
            continue
        if filtro_metodo and (a.get("metodo") or "").lower() != filtro_metodo:
            continue

        if a["tipo"] == "pacchetto" and dettagli_servizi:
            for d in dettagli_servizi:
                records.append([
                    excel_safe_dt(a["created_at"]),
                    a["email"],
                    a["utente_id"],
                    a["tipo"],
                    a.get("pacchetto_nome") or a.get("servizio_nome"),
                    a.get("annuncio_id"),
                    a.get("annuncio_categoria") or "",
                    a.get("annuncio_provincia") or "",
                    (a["importo_cent"] or 0) / 100,
                    durata_iniziale,
                    d.get("servizio_nome"),
                    excel_safe_dt(d.get("data_fine")),
                    d.get("stato"),
                    a.get("metodo"),
                    a["acquisto_id"]
                ])
        else:
            records.append([
                excel_safe_dt(a["created_at"]),
                a["email"],
                a["utente_id"],
                a["tipo"],
                a.get("pacchetto_nome") or a.get("servizio_nome"),
                a.get("annuncio_id"),
                a.get("annuncio_categoria") or "",
                a.get("annuncio_provincia") or "",
                (a["importo_cent"] or 0) / 100,
                durata_iniziale,
                a.get("servizio_nome"),
                excel_safe_dt(dettagli_servizi[0]["data_fine"] if dettagli_servizi else None),
                stato_visuale,
                a.get("metodo"),
                a["acquisto_id"]
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Storico Acquisti"

    ws.append([
        "Data acquisto",
        "Email utente",
        "ID utente",
        "Tipo",
        "Oggetto acquisto",
        "Annuncio",
        "Categoria annuncio",
        "Provincia annuncio",
        "Importo €",
        "Durata iniziale (giorni)",
        "Servizio",
        "Scadenza",
        "Stato",
        "Metodo",
        "ID acquisto"
    ])

    for row in records:
        ws.append(row)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="storico_acquisti.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==========================================================
# ADMIN – STATISTICHE
# ==========================================================
@app.route("/admin/statistiche")
@admin_required
def admin_statistiche():
    conn = get_db_connection()
    c = get_cursor(conn)

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM utenti
        WHERE attivo = 1
          AND sospeso = 0
          AND COALESCE(disattivato_admin, 0) = 0
          AND COALESCE(email, '') NOT LIKE ?
          AND COALESCE(username, '') NOT LIKE ?
    """), (
        "deleted_user_%@mylocalcare.local",
        "utente_eliminato_%"
    ))
    utenti_attivi = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM annunci
        WHERE COALESCE(stato, '') <> 'eliminato'
    """))
    annunci_totali = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(DISTINCT utente_id) AS valore
        FROM annunci
        WHERE COALESCE(stato, '') <> 'eliminato'
    """))
    utenti_con_annunci = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM utenti
        WHERE attivo = 1
          AND sospeso = 0
          AND COALESCE(disattivato_admin, 0) = 0
          AND COALESCE(email, '') NOT LIKE ?
          AND COALESCE(username, '') NOT LIKE ?
          AND id NOT IN (
              SELECT DISTINCT utente_id
              FROM annunci
              WHERE COALESCE(stato, '') <> 'eliminato'
          )
    """), (
        "deleted_user_%@mylocalcare.local",
        "utente_eliminato_%"
    ))
    utenti_senza_annunci = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(DISTINCT id_destinatario) AS valore
        FROM recensioni
        WHERE stato = 'approvato'
    """))
    utenti_recensiti = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(DISTINCT id_recensione) AS valore
        FROM risposte_recensioni
    """))
    recensioni_con_risposta = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM (
            SELECT
                CASE
                    WHEN mittente_id < destinatario_id THEN mittente_id
                    ELSE destinatario_id
                END AS a,
                CASE
                    WHEN mittente_id > destinatario_id THEN mittente_id
                    ELSE destinatario_id
                END AS b
            FROM messaggi_chat
            GROUP BY a, b
        ) AS chat_uniche
    """))
    chat_totali = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM messaggi_chat
    """))
    messaggi_inviati = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM messaggi_chat
        WHERE letto = 0
    """))
    messaggi_non_letti = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM notifiche
    """))
    notifiche_ricevute = fetchone_value(c.fetchone())

    c.execute(sql("""
        SELECT COUNT(*) AS valore
        FROM notifiche
        WHERE letta = 0
    """))
    notifiche_da_leggere = fetchone_value(c.fetchone())

    return render_template(
        "admin_statistiche.html",
        utenti_attivi=utenti_attivi,
        annunci_totali=annunci_totali,
        utenti_con_annunci=utenti_con_annunci,
        utenti_senza_annunci=utenti_senza_annunci,
        utenti_recensiti=utenti_recensiti,
        recensioni_con_risposta=recensioni_con_risposta,
        chat_totali=chat_totali,
        messaggi_inviati=messaggi_inviati,
        messaggi_non_letti=messaggi_non_letti,
        notifiche_ricevute=notifiche_ricevute,
        notifiche_da_leggere=notifiche_da_leggere
    )

# ==========================================================
# ADMIN – NOTIFICHE DI SISTEMA
# ==========================================================
@app.route("/admin/notifiche", methods=["GET"])
@admin_required
def admin_notifiche():
    """Gestione invio e storico notifiche admin"""
    conn = get_db_connection()

    stats = conn.execute(sql("""
        SELECT
            COUNT(*) AS totali,
            SUM(CASE WHEN letta = 0 THEN 1 ELSE 0 END) AS non_lette,
            SUM(CASE WHEN letta = 1 THEN 1 ELSE 0 END) AS lette
        FROM notifiche
    """)).fetchone()

    # ✅ STORICO NOTIFICHE INVIATE DALL’ADMIN
    notifiche_admin = conn.execute(sql(f"""
        SELECT
            id,
            titolo,
            messaggio,
            link,
            tipo_invio,
            tab_attivo,
            filtro_json,
            destinatari_count,
            destinatari_json,
            created_at
        FROM notifiche_admin
        ORDER BY {order_datetime("created_at")} DESC
        LIMIT 50
    """)).fetchall()

    # ✅ lista utenti (per selezione multipla)
    utenti = conn.execute(sql("""
        SELECT id, nome, cognome, email, username
        FROM utenti
        WHERE sospeso = 0 AND attivo = 1
        ORDER BY nome, cognome
    """)).fetchall()



    # ✅ categorie (da JSON, non dal DB)
    json_path = os.path.join(app.root_path, "static", "data", "filtri_categoria.json")
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    categorie = sorted(list(data.keys()))

    daily_matches_settings = get_daily_matches_settings()

    return render_template(
        "admin_notifiche.html",
        stats=stats,
        utenti=utenti,
        categorie=categorie,
        notifiche_admin=notifiche_admin,
        daily_matches_settings=daily_matches_settings
    )

@app.route("/admin/notifiche/daily-matches/salva", methods=["POST"])
@admin_required
def admin_daily_matches_salva():
    verify_csrf()

    enabled = request.form.get("daily_matches_enabled") == "1"
    time_value = (request.form.get("daily_matches_time") or "08:00").strip()
    channel = (request.form.get("daily_matches_channel") or "internal").strip()

    try:
        set_daily_matches_settings(
            enabled=enabled,
            time_value=time_value,
            channel=channel
        )

        flash("Impostazioni Daily Matches aggiornate.", "success")

    except Exception as e:
        log_exception_safe(
            "❌ Errore salvataggio impostazioni Daily Matches",
            e,
            production=True
        )

        flash("Errore durante il salvataggio delle impostazioni Daily Matches.", "error")

    return redirect(url_for("admin_notifiche"))


@app.route("/admin/notifiche/daily-matches/test", methods=["POST"])
@admin_required
def admin_daily_matches_test():
    verify_csrf()

    settings = get_daily_matches_settings()

    try:
        result = processa_match_nuovi_annunci(
            channel=settings.get("channel", "internal")
        )

        if result.get("ok"):
            flash(
                "Test Daily Matches completato: "
                f"{result.get('annunci_processati', 0)} annunci processati, "
                f"{result.get('utenti_notificati', 0)} utenti notificati, "
                f"{result.get('match_creati', 0)} match creati.",
                "success"
            )
        else:
            flash("Test Daily Matches non riuscito.", "error")

    except Exception as e:
        log_exception_safe(
            "❌ Errore test Daily Matches",
            e,
            production=True
        )

        flash("Errore durante il test Daily Matches.", "error")

    return redirect(url_for("admin_notifiche"))

@app.route("/admin/notifiche/invia", methods=["POST"])
@admin_required
def admin_invia_notifica():

    tipo_invio = request.form.getlist("tipo_invio")
    titolo = request.form.get("titolo", "").strip()
    messaggio = request.form.get("messaggio", "").strip()
    link = request.form.get("link", "").strip() or None

    if not titolo or not messaggio:
        flash("Titolo e messaggio sono obbligatori.", "error")
        return redirect(url_for("admin_notifiche"))

    # 1️⃣ FILTRA DESTINATARI
    destinatari = _filtra_utenti(request.form)

    if not destinatari:
        flash("Nessun destinatario trovato.", "error")
        return redirect(url_for("admin_notifiche"))

    # 2️⃣ SALVA STORICO ADMIN (PRIMA DELL’INVIO)
    try:
        # ✅ snapshot completo del form (liste incluse)
        filtro_snapshot = request.form.to_dict(flat=False)

        # ❌ rimuovi contenuto e campi non "filtro"
        for k in ("titolo", "messaggio", "link", "tipo_invio"):
            filtro_snapshot.pop(k, None)

        tab_attivo = request.form.get("tab-attivo") or "n/a"
        tipo_invio_str = ",".join(tipo_invio) if isinstance(tipo_invio, list) else str(tipo_invio)

        # ✅ destinatari solo se u-multipli
        destinatari_snapshot = None
        if tab_attivo == "u-multipli":
            destinatari_snapshot = []
            for u in destinatari:
                destinatari_snapshot.append({
                    "id": u["id"],
                    "username": u["username"] if "username" in u.keys() else None,
                    "email": u["email"] if "email" in u.keys() else None,
                    "nome": u["nome"] if "nome" in u.keys() else None,
                    "cognome": u["cognome"] if "cognome" in u.keys() else None,
                })

        conn = get_db_connection()
        c = get_cursor(conn)
        c.execute(sql("""
            INSERT INTO notifiche_admin (
                titolo,
                messaggio,
                link,
                tipo_invio,
                tab_attivo,
                filtro_json,
                destinatari_count,
                destinatari_json
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """), (
            titolo,
            messaggio,
            link,
            tipo_invio_str,
            tab_attivo,
            json.dumps(filtro_snapshot, ensure_ascii=False),
            len(destinatari),
            json.dumps(destinatari_snapshot, ensure_ascii=False) if destinatari_snapshot else None
        ))
        conn.commit()


    except Exception as e:
        print("❌ ERRORE storico notifiche_admin:", e)

    # 3️⃣ INVIO NOTIFICHE
    invia_notifica_interna = (
        "notifica" in tipo_invio
        or "entrambi" in tipo_invio
    )

    invia_email = (
        "email" in tipo_invio
        or "entrambi" in tipo_invio
    )

    # Regola richiesta:
    # - solo notifica interna => anche push
    # - email + notifica interna => niente push
    # - solo email => niente push
    invia_push_notification = (
        invia_notifica_interna
        and not invia_email
    )

    inviati = 0

    for user in destinatari:

        if invia_notifica_interna:
            _crea_notifica(
                user["id"],
                titolo,
                messaggio,
                tipo="generica",
                link=link
            )

            emit_update_notifications(user["id"])

            if invia_push_notification:
                try:
                    invia_push(
                        user["id"],
                        titolo,
                        messaggio
                    )
                except Exception as e:
                    log_exception_safe(
                        "⚠️ Errore push notifica admin",
                        e,
                        {"user_id": user["id"]},
                        production=True
                    )

        if invia_email:
            _invia_email(
                destinazione=user["email"],
                oggetto=titolo,
                corpo=f"{messaggio}\n\n{link or ''}"
            )

        inviati += 1

    flash(f"Notifica inviata a {inviati} utenti.", "success")
    return redirect(url_for("admin_notifiche"))

import sqlite3

# Deve combaciare con l'ordine usato nel template INFO (loop.index 1..8)
CATEGORIE_INFO = [
    "Operatori benessere",             # 1
    "Aiuto in casa",                   # 2
    "Ripetizioni",                     # 3
    "Babysitter",                      # 4
    "Pet-sitter",                      # 5
    "Caregiver",                       # 6
    "escursioni-sport",                # 7
    "Biglietti spettacoli",            # 8
    "Libri scuola",                    # 9
    "Caffe & parole",                  # 10
]

def _norm(s: str) -> str:
    return (s or "").strip().lower()

def _parse_list_from_form(form, key: str):
    # accetta list dal form (checkbox multiple) oppure stringa singola
    vals = form.getlist(key)
    if vals:
        return [v.strip() for v in vals if v and v.strip()]
    v = (form.get(key) or "").strip()
    return [v] if v else []

def _parse_zone_terms(form):
    """
    Supporta:
      - input singolo "zona" (stringa)
      - oppure più input "zone" (checkbox / multiple)
    Inoltre, se l'admin incolla "Milano, Pavia" splitto per virgola.
    """
    zones = _parse_list_from_form(form, "zone")
    if not zones:
        raw = (form.get("zona") or "").strip()
        zones = [raw] if raw else []
    out = []
    for z in zones:
        for part in z.split(","):
            p = part.strip()
            if p:
                out.append(p)
    # dedup preservando ordine
    seen = set()
    dedup = []
    for z in out:
        k = z.lower()
        if k not in seen:
            seen.add(k)
            dedup.append(z)
    return dedup

def _categorie_to_info_indexes(categorie_selezionate):
    """
    Trasforma categorie (stringhe) in indici 1..8 per offro_i / cerco_i.
    """
    idx = []
    for cat in categorie_selezionate:
        c = _norm(cat)
        if c in CATEGORIE_INFO:
            idx.append(CATEGORIE_INFO.index(c) + 1)
    # dedup
    return sorted(set(idx))

def _is_checked(form, key: str) -> bool:
    """
    Gestisce checkbox Flask:
    - se spuntato arriva key in form con value spesso "on" o "1"
    - se non spuntato NON arriva proprio
    """
    v = form.get(key)
    return v is not None and str(v).strip().lower() not in ("0", "false", "off", "")


def _filtra_utenti(form):
    conn = get_db_connection()

    c = get_cursor(conn)

    # ------------------------------------------------------------
    # 0) Base filter per utenti
    # ------------------------------------------------------------
    BASE_UTENTI_WHERE = "u.sospeso = 0 AND u.attivo = 1"

    # ------------------------------------------------------------
    # 1) Tab attiva (decide quale modalità usare)
    # ------------------------------------------------------------
    tab = (form.get("tab-attivo") or "").strip()  # es: u-singolo, u-multipli, u-zona, u-categoria, u-avanzato, u-tutti

    # ------------------------------------------------------------
    # 2) Utente singolo
    # ------------------------------------------------------------
    if tab == "u-singolo":
        singolo = (form.get("utente_singolo") or "").strip()
        if singolo:
            valore = singolo.lower()
            c.execute(sql(f"""
                SELECT u.*
                FROM utenti u
                WHERE (LOWER(u.email) = ? OR LOWER(u.username) = ?)
                  AND {BASE_UTENTI_WHERE}
                LIMIT 1
            """), (valore, valore))
            row = c.fetchone()

            return [row] if row else []

        return []

    # ------------------------------------------------------------
    # 3) Selezione multipla
    # ------------------------------------------------------------
    if tab == "u-multipli":
        multipli = form.getlist("utenti_multipli")
        multipli = [m for m in multipli if str(m).strip().isdigit()]
        if not multipli:

            return []
        placeholders = ",".join(["?"] * len(multipli))
        c.execute(sql(f"""
            SELECT u.*
            FROM utenti u
            WHERE u.id IN ({placeholders})
              AND {BASE_UTENTI_WHERE}
        """), multipli)
        rows = c.fetchall()

        return rows

    # ------------------------------------------------------------
    # 4) Tutti
    # ------------------------------------------------------------
    if tab == "u-tutti":
        c.execute(sql(f"""
            SELECT u.*
            FROM utenti u
            WHERE {BASE_UTENTI_WHERE}
            ORDER BY u.nome, u.cognome
        """))
        rows = c.fetchall()

        return rows

    # ------------------------------------------------------------
    # 5) FILTRO AVANZATO (ZONA / CATEGORIA / INCROCIO)
    #    Lo usiamo sia per u-zona, u-categoria, sia per un futuro u-avanzato.
    # ------------------------------------------------------------

    # ---- input
    zone_terms = _parse_zone_terms(form)                  # lista stringhe
    # 🔹 normalizzazione zona: prefisso principale (NO JSON)
    zone_prefixes = []

    for z in zone_terms:
        z = _norm(z)
        if not z:
            continue

        # prendo solo la parte prima di separatori
        for sep in ["–", "-", ","]:
            if sep in z:
                z = z.split(sep)[0].strip()
                break

        if z and z not in zone_prefixes:
            zone_prefixes.append(z)

    categorie_sel = _parse_list_from_form(form, "categorie")  # lista stringhe (es: "babysitter", "operatori benessere")
    categorie_sel_norm = [_norm(x) for x in categorie_sel if _norm(x)]

    # ---- toggle (checkbox) – la UI li manderà così allo step 2
    # per ora, metto default "sensati" in base alla tab:
    include_zona_info    = _is_checked(form, "zona_info")
    include_zona_annunci = _is_checked(form, "zona_annunci")

    include_cat_cerco    = _is_checked(form, "cat_cerco_info")
    include_cat_offro    = _is_checked(form, "cat_offro_info")
    include_cat_annunci  = _is_checked(form, "cat_annunci")

    # Default automatici se la UI non li manda ancora:
    if tab == "u-zona":
        if ("zona_info" not in form) and ("zona_annunci" not in form):
            include_zona_info = True
            include_zona_annunci = True

    if tab == "u-categoria":
        if not (include_cat_cerco or include_cat_offro or include_cat_annunci):
            include_cat_cerco = True
            include_cat_offro = True
            include_cat_annunci = True

    # Se in futuro avrai tab "u-avanzato", lì non setto default: lo decidi dalla UI.
    # Qui però evitiamo "nessun filtro" per errore.
    has_zona = len(zone_prefixes) > 0
    has_cat  = len(categorie_sel_norm) > 0

    # Se tab è zona/categoria ma input vuoto => nessun destinatario
    if (tab in ["u-zona", "u-categoria"] or tab == "u-avanzato"):
        if not has_zona and not has_cat:

            return []

    # ------------------------------------------------------------
    # 5A) Subquery: utenti per ZONA
    # ------------------------------------------------------------
    zona_info_sql = None
    zona_info_params = []

    zona_annunci_sql = None
    zona_annunci_params = []

    if has_zona:

        # ---------------- INFO ----------------
        if include_zona_info and zone_prefixes:
            conds = " OR ".join(
                ["LOWER(u.citta) LIKE ?"] * len(zone_prefixes)
            )
            zona_info_sql = f"""
                SELECT u.id AS uid
                FROM utenti u
                WHERE {BASE_UTENTI_WHERE}
                  AND ({conds})
            """
            zona_info_params = [f"{z}%" for z in zone_prefixes]

        # ---------------- ANNUNCI ----------------
        if include_zona_annunci and zone_prefixes:
            conds = " OR ".join(
                ["LOWER(a.zona) LIKE ?"] * len(zone_prefixes)
            )
            zona_annunci_sql = f"""
                SELECT a.utente_id AS uid
                FROM annunci a
                JOIN utenti u ON u.id = a.utente_id
                WHERE {BASE_UTENTI_WHERE}
                  AND ({conds})
            """
            zona_annunci_params = [f"{z}%" for z in zone_prefixes]

    # 🔒 BLOCCO DURO: zona richiesta ma nessun match reale
    zona_sql = None
    zona_params = []

    if has_zona:

        if include_zona_info and not include_zona_annunci:
            # SOLO INFO
            if not zona_info_sql:

                return []
            zona_sql = zona_info_sql
            zona_params = zona_info_params

        elif include_zona_annunci and not include_zona_info:
            # SOLO ANNUNCI
            if not zona_annunci_sql:

                return []
            zona_sql = zona_annunci_sql
            zona_params = zona_annunci_params

        elif include_zona_info and include_zona_annunci:
            # INFO + ANNUNCI
            if not zona_info_sql and not zona_annunci_sql:

                return []

            queries = []
            if zona_info_sql:
                queries.append(zona_info_sql)
            if zona_annunci_sql:
                queries.append(zona_annunci_sql)

            zona_sql = " UNION ".join(queries)
            zona_params = zona_info_params + zona_annunci_params

        else:
            # zona richiesta ma nessun checkbox valido

            return []

        # BLOCCO DURO FINALE
        count = conn.execute(
            f"SELECT COUNT(*) FROM ({zona_sql})",
            zona_params
        ).fetchone()[0]

        if count == 0:

            return []

    # ------------------------------------------------------------
    # 5B) Subquery: utenti per CATEGORIA
    # ------------------------------------------------------------
    cat_subqueries = []
    cat_params = []

    if has_cat:
        # ---- categoria da ANNUNCI (annunci.categoria)
        if include_cat_annunci:
            placeholders = ",".join(["?"] * len(categorie_sel_norm))
            cat_subqueries.append(f"""
                SELECT a.utente_id AS uid
                FROM annunci a
                JOIN utenti u ON u.id = a.utente_id
                WHERE {BASE_UTENTI_WHERE}
                  AND LOWER(a.categoria) IN ({placeholders})
            """)
            cat_params.extend(categorie_sel_norm)

        # ---- categoria da INFO (offro_i / cerco_i)
        idx = _categorie_to_info_indexes(categorie_sel_norm)

        # costruiamo condizioni OR tipo: (u.cerco_1=1 OR u.cerco_3=1 ...)
        if idx and include_cat_cerco:
            conds = " OR ".join([f"COALESCE(u.cerco_{i},0)=1" for i in idx])
            cat_subqueries.append(f"""
                SELECT u.id AS uid
                FROM utenti u
                WHERE {BASE_UTENTI_WHERE}
                  AND ({conds})
            """)

        if idx and include_cat_offro:
            conds = " OR ".join([f"COALESCE(u.offro_{i},0)=1" for i in idx])
            cat_subqueries.append(f"""
                SELECT u.id AS uid
                FROM utenti u
                WHERE {BASE_UTENTI_WHERE}
                  AND ({conds})
            """)

    cat_sql = None
    if cat_subqueries:
        cat_sql = " UNION ".join(cat_subqueries)

    # ------------------------------------------------------------
    # 5C) Combinazione: solo zona / solo categoria / incrocio (AND)
    # ------------------------------------------------------------
    final_ids_sql = None
    final_params = []

    if has_zona and has_cat:
        if not zona_sql or not cat_sql:

            return []
        final_ids_sql = f"""
            SELECT uid FROM (
                {zona_sql}
                INTERSECT
                {cat_sql}
            )
        """
        final_params = zona_params + cat_params

    elif has_zona:
        if not zona_sql:

            return []
        final_ids_sql = f"""
            SELECT uid FROM (
                {zona_sql}
            )
        """
        final_params = zona_params

    elif has_cat:
        if not cat_sql:

            return []
        final_ids_sql = f"""
            SELECT uid FROM (
                {cat_sql}
            )
        """
        final_params = cat_params

    else:

        return []
    # ------------------------------------------------------------
    # 5D) Query finale utenti
    # ------------------------------------------------------------
    c.execute(sql(f"""
        WITH destinatari AS (
            {final_ids_sql}
        )
        SELECT u.*
        FROM utenti u
        JOIN destinatari d ON d.uid = u.id
        WHERE {BASE_UTENTI_WHERE}
        ORDER BY u.nome, u.cognome
    """), final_params)

    rows = c.fetchall()

    return rows


def _crea_notifica(id_utente, titolo, messaggio, tipo="generica", link=None):
    conn = get_db_connection()
    try:
        c = get_cursor(conn)

        c.execute(sql("""
            INSERT INTO notifiche (
                id_utente,
                titolo,
                messaggio,
                tipo,
                link,
                letta
            ) VALUES (?, ?, ?, ?, ?, 0)
        """), (id_utente, titolo, messaggio, tipo, link))

        conn.commit()

    finally:
        try:
            conn.close()
        except:
            pass

def get_daily_matches_settings():
    """
    Legge configurazione Daily Matches da app_settings.

    Chiavi usate:
    - daily_matches_enabled: 1/0
    - daily_matches_time: HH:MM
    - daily_matches_channel: internal / email / both
    """
    defaults = {
        "enabled": True,
        "time": "08:00",
        "channel": "internal"
    }

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT chiave, valore
            FROM app_settings
            WHERE chiave IN (
                'daily_matches_enabled',
                'daily_matches_time',
                'daily_matches_channel'
            )
        """))

        rows = cur.fetchall()
        values = {r["chiave"]: r["valore"] for r in rows}

        enabled_raw = str(values.get("daily_matches_enabled", "1")).strip()
        time_raw = str(values.get("daily_matches_time", defaults["time"])).strip()
        channel_raw = str(values.get("daily_matches_channel", defaults["channel"])).strip()

        if not re.match(r"^\d{2}:\d{2}$", time_raw):
            time_raw = defaults["time"]

        if channel_raw not in ("internal", "email", "both"):
            channel_raw = defaults["channel"]

        return {
            "enabled": enabled_raw == "1",
            "time": time_raw,
            "channel": channel_raw
        }

    except Exception as e:
        log_exception_safe(
            "⚠️ Errore lettura configurazione Daily Matches",
            e,
            production=True
        )
        return defaults

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def set_daily_matches_settings(enabled: bool, time_value: str, channel: str):
    """
    Salva configurazione Daily Matches in app_settings.
    """
    time_value = (time_value or "08:00").strip()
    channel = (channel or "internal").strip()

    if not re.match(r"^\d{2}:\d{2}$", time_value):
        time_value = "08:00"

    if channel not in ("internal", "email", "both"):
        channel = "internal"

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        settings = {
            "daily_matches_enabled": "1" if enabled else "0",
            "daily_matches_time": time_value,
            "daily_matches_channel": channel
        }

        for chiave, valore in settings.items():
            if app.config.get("IS_POSTGRES"):
                cur.execute("""
                    INSERT INTO app_settings (chiave, valore)
                    VALUES (%s, %s)
                    ON CONFLICT (chiave)
                    DO UPDATE SET valore = EXCLUDED.valore
                """, (chiave, valore))
            else:
                cur.execute("""
                    INSERT INTO app_settings (chiave, valore)
                    VALUES (?, ?)
                    ON CONFLICT(chiave)
                    DO UPDATE SET valore = excluded.valore
                """, (chiave, valore))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def invia_email_daily_match(user_id, categorie_count):
    """
    Invia email riepilogativa Daily Matches all'utente.
    Non blocca il processo se fallisce.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT email, nome, username, email_notifiche
            FROM utenti
            WHERE id = ?
              AND attivo = 1
              AND sospeso = 0
              AND COALESCE(disattivato_admin, 0) = 0
            LIMIT 1
        """), (int(user_id),))

        user = cur.fetchone()

        if not user:
            return False

        if int(user["email_notifiche"] or 0) != 1:
            return False

        righe = []
        for categoria, count in categorie_count.items():
            label = "nuovo annuncio" if count == 1 else "nuovi annunci"
            righe.append(f"- {categoria}: {count} {label}")

        nome = user["nome"] or user["username"] or "utente"

        home_url = f"{app.config.get('APP_BASE_URL', 'https://www.mylocalcare.it').rstrip('/')}/home"

        corpo = (
            f"Ciao {nome},\n\n"
            "abbiamo trovato nuovi annunci compatibili con le tue preferenze:\n\n"
            + "\n".join(righe)
            + "\n\nPuoi consultarli su MyLocalCare:\n"
            + home_url
            + "\n\nMyLocalCare"
        )

        return _invia_email(
            destinazione=user["email"],
            oggetto="Nuovi annunci compatibili su MyLocalCare",
            corpo=corpo
        )

    except Exception as e:
        log_exception_safe(
            "⚠️ Errore invio email Daily Matches",
            e,
            {"user_id": user_id},
            production=True
        )
        return False

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def processa_match_nuovi_annunci(channel=None):
    """
    Processa solo annunci approvati con match_da_processare = 1.

    Regola anti-duplicazione:
    - ogni annuncio viene notificato una sola volta;
    - dopo il processamento viene portato a match_da_processare = 0.

    Logica:
    - annuncio OFFRO  -> notifica utenti che CERCANO quella categoria;
    - annuncio CERCO  -> notifica utenti che OFFRONO quella categoria.

    channel:
    - internal = solo notifica interna
    - email    = solo email
    - both     = entrambe
    """
    settings = get_daily_matches_settings()
    channel = channel or settings.get("channel", "internal")

    if channel not in ("internal", "email", "both"):
        channel = "internal"

    conn = get_db_connection()
    cur = get_cursor(conn)

    annunci_processati = []
    notifiche_per_utente = {}
    match_ids_creati = []

    try:
        cur.execute(sql("""
            SELECT
                id,
                utente_id,
                categoria,
                tipo_annuncio,
                zona,
                provincia,
                titolo
            FROM annunci
            WHERE stato = 'approvato'
              AND COALESCE(match_da_processare, 0) = 1
            ORDER BY data_pubblicazione ASC
        """))

        nuovi = cur.fetchall()

        if not nuovi:
            security_log(
                "ℹ️ Daily Matches: nessun nuovo annuncio da processare",
                production=True
            )
            return {
                "ok": True,
                "annunci_processati": 0,
                "utenti_notificati": 0,
                "match_creati": 0
            }

        match_creati = 0

        for annuncio in nuovi:
            annuncio_id = int(annuncio["id"])
            autore_id = int(annuncio["utente_id"])
            categoria = (annuncio["categoria"] or "").strip()
            tipo_annuncio = (annuncio["tipo_annuncio"] or "").strip().lower()
            zona = (annuncio["zona"] or "").strip()
            provincia = (annuncio["provincia"] or "").strip()
            titolo = (annuncio["titolo"] or "").strip()

            idx = CATEGORIA_TO_INDEX.get(to_slug(categoria))

            if not idx:
                annunci_processati.append(annuncio_id)
                continue

            if tipo_annuncio == "offro":
                colonna_match = f"cerco_{idx}"
                tipo_match = "cerco"
            elif tipo_annuncio == "cerco":
                colonna_match = f"offro_{idx}"
                tipo_match = "offro"
            else:
                annunci_processati.append(annuncio_id)
                continue

            cur.execute(sql(f"""
                SELECT id, citta, provincia, email_notifiche
                FROM utenti
                WHERE COALESCE({colonna_match}, 0) = 1
                  AND attivo = 1
                  AND sospeso = 0
                  AND COALESCE(disattivato_admin, 0) = 0
                  AND id != ?
            """), (autore_id,))

            utenti = cur.fetchall()

            for utente in utenti:
                uid = int(utente["id"])

                citta_utente = utente["citta"] or ""
                provincia_utente = utente["provincia"] or ""

                match_zona = False

                if zona and citta_utente:
                    match_zona = place_match(zona, citta_utente)

                if not match_zona and provincia and provincia_utente:
                    match_zona = norm_place(provincia) == norm_place(provincia_utente)

                if not match_zona:
                    continue

                try:
                    if app.config.get("IS_POSTGRES"):
                        cur.execute(sql("""
                            INSERT INTO match_utenti (
                                utente_cerca_id,
                                utente_offre_id,
                                categoria,
                                zona,
                                annuncio_id
                            )
                            VALUES (?, ?, ?, ?, ?)
                            RETURNING id
                        """), (
                            uid if tipo_match == "cerco" else autore_id,
                            autore_id if tipo_match == "cerco" else uid,
                            categoria,
                            zona,
                            annuncio_id
                        ))

                        row_match = cur.fetchone()
                        if row_match and row_match["id"]:
                            match_ids_creati.append(int(row_match["id"]))

                    else:
                        cur.execute(sql("""
                            INSERT INTO match_utenti (
                                utente_cerca_id,
                                utente_offre_id,
                                categoria,
                                zona,
                                annuncio_id
                            )
                            VALUES (?, ?, ?, ?, ?)
                        """), (
                            uid if tipo_match == "cerco" else autore_id,
                            autore_id if tipo_match == "cerco" else uid,
                            categoria,
                            zona,
                            annuncio_id
                        ))

                        if cur.lastrowid:
                            match_ids_creati.append(int(cur.lastrowid))

                    match_creati += 1

                except Exception as e:
                    # Se il match esiste già o l'inserimento fallisce per un singolo utente,
                    # non blocchiamo l'intero ciclo.
                    log_exception_safe(
                        "⚠️ Daily Matches: match singolo non inserito",
                        e,
                        {
                            "annuncio_id": annuncio_id,
                            "user_id": uid
                        }
                    )

                notifiche_per_utente.setdefault(uid, {})
                notifiche_per_utente[uid].setdefault(categoria, 0)
                notifiche_per_utente[uid][categoria] += 1

            annunci_processati.append(annuncio_id)

        # =====================================================
        # NOTIFICHE RIEPILOGATIVE: massimo 1 per utente
        # =====================================================
        utenti_notificati = 0

        for user_id, categorie_count in notifiche_per_utente.items():
            righe = []

            for categoria, n in categorie_count.items():
                label = "nuovo annuncio" if n == 1 else "nuovi annunci"
                righe.append(f"• {categoria} ({n} {label})")

            messaggio = (
                "Abbiamo trovato nuovi annunci compatibili con le tue preferenze:\n"
                + "\n".join(righe)
            )

            if channel in ("internal", "both"):
                link_notifica = "/home"

                cur.execute(sql("""
                    INSERT INTO notifiche (
                        id_utente,
                        titolo,
                        messaggio,
                        link,
                        tipo,
                        letta
                    )
                    VALUES (?, ?, ?, ?, ?, 0)
                """), (
                    int(user_id),
                    "Nuovi annunci compatibili",
                    messaggio,
                    link_notifica,
                    "match"
                ))

                emit_update_notifications(int(user_id))

                # REGOLA DAILY MATCHES:
                # - internal = notifica interna + push
                # - both = email + notifica interna, NO push
                # - email = solo email, NO push
                if channel == "internal":
                    try:
                        invia_push(
                            int(user_id),
                            "Nuovi annunci compatibili",
                            messaggio,
                            url=link_notifica
                        )
                    except Exception as e:
                        log_exception_safe(
                            "⚠️ Errore push Daily Matches",
                            e,
                            {"user_id": int(user_id)},
                            production=True
                        )

            if channel in ("email", "both"):
                invia_email_daily_match(
                    user_id=int(user_id),
                    categorie_count=categorie_count
                )

            utenti_notificati += 1

        # =====================================================
        # SEGNA MATCH COME NOTIFICATI
        # =====================================================
        if match_ids_creati:
            placeholders_match = ",".join(["?"] * len(match_ids_creati))

            cur.execute(sql(f"""
                UPDATE match_utenti
                SET notificato = 1
                WHERE id IN ({placeholders_match})
            """), match_ids_creati)

        # =====================================================
        # SEGNA ANNUNCI COME PROCESSATI
        # =====================================================
        if annunci_processati:
            placeholders = ",".join(["?"] * len(annunci_processati))

            cur.execute(sql(f"""
                UPDATE annunci
                SET match_da_processare = 0
                WHERE id IN ({placeholders})
            """), annunci_processati)

        conn.commit()

        result = {
            "ok": True,
            "annunci_processati": len(annunci_processati),
            "utenti_notificati": utenti_notificati,
            "match_creati": match_creati,
            "match_notificati": len(match_ids_creati)
        }

        security_log(
            "✅ Daily Matches completato",
            result,
            production=True
        )

        return result

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass

        log_exception_safe(
            "❌ Errore Daily Matches",
            e,
            production=True
        )

        return {
            "ok": False,
            "error": "daily_matches_failed",
            "annunci_processati": 0,
            "utenti_notificati": 0,
            "match_creati": 0
        }

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def _invia_email(destinazione, oggetto, corpo=None, html_template=None, html=None, **kwargs):
    """
    Funzione centralizzata invio email MyLocalCare tramite Postmark.

    Usata per:
    - conferma registrazione
    - reset password
    - sospensione account
    - notifiche email inviate dall'admin

    Nota:
    - non usa Flask-Mail;
    - Flask-Mail resta solo per la funzione send_async_email(), che invia alert admin già funzionanti.
    """

    postmark_token = os.getenv("POSTMARK_SERVER_TOKEN", "").strip()

    if not destinazione or not str(destinazione).strip():
        security_log(
            "❌ Invio email annullato: destinatario mancante",
            {
                "oggetto": oggetto
            },
            production=True
        )
        return False

    if not oggetto or not str(oggetto).strip():
        security_log(
            "❌ Invio email annullato: oggetto mancante",
            {
                "destinazione": destinazione
            },
            production=True
        )
        return False

    if not postmark_token:
        security_log(
            "❌ POSTMARK_SERVER_TOKEN mancante",
            {
                "destinazione": destinazione,
                "oggetto": oggetto
            },
            production=True
        )
        return False

    try:
        html_finale = None

        # ✅ HTML da template, se presente
        if html_template:
            try:
                html_finale = render_template(html_template, **kwargs)
            except Exception as e:
                log_exception_safe(
                    "❌ Errore rendering template email",
                    e,
                    {
                        "html_template": html_template,
                        "destinazione": destinazione,
                        "oggetto": oggetto
                    },
                    production=True
                )

                # Se il template fallisce, non blocchiamo per forza l'invio:
                # inviamo almeno il testo se presente.
                html_finale = None

        # ✅ HTML diretto, se passato esplicitamente
        elif html:
            html_finale = html

        # ✅ Testo fallback
        text_finale = (corpo or "").strip()

        if not text_finale and html_finale:
            text_finale = (
                "Hai ricevuto una comunicazione da MyLocalCare.\n\n"
                "Apri questa email in formato HTML per visualizzarla correttamente."
            )

        if not text_finale and not html_finale:
            security_log(
                "❌ Invio email annullato: contenuto vuoto",
                {
                    "destinazione": destinazione,
                    "oggetto": oggetto,
                    "html_template": html_template
                },
                production=True
            )
            return False

        from_address = os.getenv("MAIL_FROM_ADDRESS", MAIL_FROM_ADDRESS).strip()
        from_name = os.getenv("MAIL_FROM_NAME", MAIL_FROM_NAME).strip() or "MyLocalCare"
        message_stream = os.getenv("POSTMARK_MESSAGE_STREAM", "outbound").strip() or "outbound"

        mittente = f"{from_name} <{from_address}>"

        payload = {
            "From": mittente,
            "To": str(destinazione).strip(),
            "Subject": str(oggetto).strip(),
            "TextBody": text_finale,
            "MessageStream": message_stream,
            "ReplyTo": from_address,
            "TrackOpens": False,
            "TrackLinks": "None"
        }

        if html_finale:
            payload["HtmlBody"] = html_finale

        response = requests.post(
            "https://api.postmarkapp.com/email",
            headers={
                "Accept": "application/json",
                "Content-Type": "application/json",
                "X-Postmark-Server-Token": postmark_token
            },
            json=payload,
            timeout=15
        )

        try:
            result = response.json()
        except Exception:
            result = {
                "raw": response.text[:500]
            }

        if response.status_code != 200:
            security_log(
                "❌ Errore invio email tramite Postmark",
                {
                    "status_code": response.status_code,
                    "destinazione": destinazione,
                    "oggetto": oggetto,
                    "result": result
                },
                production=True
            )
            return False

        security_log(
            "✅ Email inviata tramite Postmark",
            {
                "destinazione": destinazione,
                "oggetto": oggetto,
                "message_id": result.get("MessageID") if isinstance(result, dict) else None
            },
            production=True
        )

        return True

    except requests.exceptions.Timeout as e:
        log_exception_safe(
            "❌ Timeout invio email tramite Postmark",
            e,
            {
                "destinazione": destinazione,
                "oggetto": oggetto
            },
            production=True
        )
        return False

    except Exception as e:
        log_exception_safe(
            "❌ Eccezione invio email tramite Postmark",
            e,
            {
                "destinazione": destinazione,
                "oggetto": oggetto
            },
            production=True
        )
        return False

def _normalizza_lista(value):
    if not value:
        return []
    if isinstance(value, list):
        return [v for v in value if v.strip()]
    return [value] if value.strip() else []

# ==========================================================
# NOTIFICHE SERVIZIO URGENTE
# ==========================================================
import sqlite3

def invia_email_urgente_match(user_id, annuncio_id, categoria, tipo_annuncio, luogo, username, titolo):
    """
    Invia email per annuncio urgente compatibile.
    Parte solo se l'utente ha email_notifiche = 1.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("""
            SELECT email, nome, username, email_notifiche
            FROM utenti
            WHERE id = ?
              AND attivo = 1
              AND sospeso = 0
              AND COALESCE(disattivato_admin, 0) = 0
              AND email_notifiche = 1
            LIMIT 1
        """), (int(user_id),))

        user = cur.fetchone()

        if not user:
            return False

        nome = user["nome"] or user["username"] or "utente"

        url = f"{app.config.get('APP_BASE_URL', 'https://www.mylocalcare.it').rstrip('/')}/annuncio/{annuncio_id}"

        corpo = (
            f"Ciao {nome},\n\n"
            "c'è un annuncio urgente compatibile con le tue preferenze:\n\n"
            f"Categoria: {categoria}\n"
            f"Tipo: {tipo_annuncio}\n"
            f"Zona: {luogo}\n"
            f"Pubblicato da: @{username}\n"
            f"Titolo: {titolo}\n\n"
            "Puoi visualizzarlo qui:\n"
            f"{url}\n\n"
            "MyLocalCare"
        )

        return _invia_email(
            destinazione=user["email"],
            oggetto="Annuncio urgente compatibile su MyLocalCare",
            corpo=corpo
        )

    except Exception as e:
        log_exception_safe(
            "⚠️ Errore invio email annuncio urgente",
            e,
            {
                "user_id": user_id,
                "annuncio_id": annuncio_id
            },
            production=True
        )
        return False

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

def notifica_urgente(annuncio_id, attivazione_id=None, eseguito_da="admin", conn=None):
    """
    Invia notifiche per un annuncio urgente.
    Viene chiamata:
    - da admin (toggle)
    - da acquisto servizio urgente
    - da riattivazione futura
    """

    own_conn = conn is None
    if own_conn:
        conn = get_db_connection()
    c = get_cursor(conn)

    try:
        # ---------------------------------------------------------
        # 1️⃣ Recupera annuncio + verifica servizio urgente ATTIVO
        # ---------------------------------------------------------
        c.execute(sql(f"""
            SELECT
                a.id,
                a.utente_id,
                a.categoria,
                a.tipo_annuncio,
                a.provincia,
                a.zona,
                a.titolo,
                u.username
            FROM annunci a
            JOIN utenti u ON u.id = a.utente_id
            JOIN attivazioni_servizi act ON act.annuncio_id = a.id
            JOIN servizi s ON s.id = act.servizio_id
            WHERE a.id = ?
              AND a.stato = 'approvato'
              AND s.codice = 'annuncio_urgente'
              AND act.stato = 'attivo'
              AND act.data_inizio <= {now_sql()}
              AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
        """), (annuncio_id,))
        annuncio = c.fetchone()

        if not annuncio:
            print("⚠️ Annuncio non valido o non urgente.", flush=True)
            return

        # ✅ lettura corretta da row/dict-like
        annuncio_id = int(annuncio["id"])
        autore_id = int(annuncio["utente_id"])
        categoria = annuncio["categoria"]
        tipo_annuncio = annuncio["tipo_annuncio"]
        provincia = annuncio["provincia"]
        zona = annuncio["zona"]
        titolo = annuncio["titolo"]
        username = annuncio["username"]

        tipo_opposto = "cerco" if tipo_annuncio == "offro" else "offro"
        luogo = zona or provincia

        notificati = set()

        # ---------------------------------------------------------
        # 2️⃣ PRIORITÀ 1 — ANNUNCI COMPATIBILI
        # ---------------------------------------------------------
        c.execute(sql("""
            SELECT DISTINCT a.utente_id
            FROM annunci a
            JOIN utenti u ON u.id = a.utente_id
            WHERE a.stato = 'approvato'
              AND a.tipo_annuncio = ?
              AND a.categoria = ?
              AND a.provincia = ?
              AND a.utente_id != ?
              AND u.sospeso = 0
              AND u.disattivato_admin = 0
              AND u.attivo = 1
              AND u.email_notifiche = 1
        """), (tipo_opposto, categoria, provincia, autore_id))

        for row in c.fetchall():
            uid = int(row["utente_id"] if isinstance(row, dict) or hasattr(row, "keys") else row[0])
            notificati.add(uid)

        # ---------------------------------------------------------
        # 3️⃣ PRIORITÀ 2 — INFO UTENTE (match per CATEGORIA specifica)
        # ---------------------------------------------------------
        categoria_slug = to_slug(categoria)
        categoria_index = CATEGORIA_TO_INDEX.get(categoria_slug)

        if categoria_index:
            colonna = f"{tipo_opposto}_{categoria_index}"

            c.execute(sql(f"""
                SELECT id
                FROM utenti
                WHERE provincia = ?
                  AND id != ?
                  AND sospeso = 0
                  AND (disattivato_admin IS NULL OR disattivato_admin = 0)
                  AND attivo = 1
                  AND email_notifiche = 1
                  AND {colonna} = 1
            """), (provincia, autore_id))

            for row in c.fetchall():
                uid = int(row["id"] if isinstance(row, dict) or hasattr(row, "keys") else row[0])
                notificati.add(uid)

        if not notificati:
            print("ℹ️ Nessun destinatario compatibile.", flush=True)
            return

        # ---------------------------------------------------------
        # 4️⃣ Inserimento notifiche (TESTO DEFINITIVO VISIBILE)
        # ---------------------------------------------------------
        messaggio = (
            "Annuncio urgente in zona\n"
            f"{categoria}|{tipo_annuncio}|{luogo}|{username}|{titolo}"
        )

        for uid in notificati:
            c.execute(sql("""
                INSERT INTO notifiche (
                    id_utente,
                    titolo,
                    messaggio,
                    link,
                    tipo
                ) VALUES (?, ?, ?, ?, ?)
            """), (
                uid,
                "urgente",  # titolo tecnico, non visibile
                messaggio,
                f"/annuncio/{annuncio_id}",
                "urgente"
            ))

        # ---------------------------------------------------------
        # 5️⃣ Storico servizio (audit)
        # ---------------------------------------------------------
        if attivazione_id:
            c.execute(sql("""
                INSERT INTO storico_servizi (
                    attivazione_id,
                    azione,
                    eseguito_da,
                    note
                ) VALUES (?, 'notifica_inviata', ?, ?)
            """), (
                attivazione_id,
                eseguito_da,
                f"Inviate {len(notificati)} notifiche urgenti"
            ))

        conn.commit()

        print(f"✅ Notifica urgente inviata a {len(notificati)} utenti.", flush=True)

        # ---------------------------------------------------------
        # 6️⃣ EMISSIONE SOCKET REALTIME + PUSH
        # ---------------------------------------------------------
        for uid in notificati:
            emit_update_notifications(uid)

            try:
                invia_push(
                    uid,
                    "Annuncio urgente in zona",
                    f"{categoria} - {tipo_annuncio} - {luogo}",
                    url=f"/annuncio/{annuncio_id}"
                )
            except Exception as e:
                log_exception_safe(
                    "⚠️ Errore push annuncio urgente",
                    e,
                    {
                        "user_id": uid,
                        "annuncio_id": annuncio_id,
                        "attivazione_id": attivazione_id
                    },
                    production=True
                )

            try:
                invia_email_urgente_match(
                    user_id=uid,
                    annuncio_id=annuncio_id,
                    categoria=categoria,
                    tipo_annuncio=tipo_annuncio,
                    luogo=luogo,
                    username=username,
                    titolo=titolo
                )
            except Exception as e:
                log_exception_safe(
                    "⚠️ Errore email annuncio urgente",
                    e,
                    {
                        "user_id": uid,
                        "annuncio_id": annuncio_id,
                        "attivazione_id": attivazione_id
                    },
                    production=True
                )

    except Exception as e:
        conn.rollback()
        print(f"❌ Errore in notifica_urgente: {repr(e)}", flush=True)
        traceback.print_exc()

    finally:
        if own_conn:
            try:
                conn.close()
            except Exception:
                pass

def redirect_admin_recensioni_next():
    """
    Redirect sicuro verso admin_recensioni conservando eventuali filtri.
    Usa prima ?next=..., poi request.referrer, poi fallback pulito.
    """
    from urllib.parse import urlparse

    next_url = request.args.get("next") or request.referrer

    if next_url:
        parsed = urlparse(next_url)

        # Caso 1: path interno tipo /admin/recensioni?stato=in_attesa
        if not parsed.netloc and parsed.path == url_for("admin_recensioni"):
            query = f"?{parsed.query}" if parsed.query else ""
            return redirect(parsed.path + query)

        # Caso 2: referrer assoluto stesso dominio
        if parsed.netloc == request.host and parsed.path == url_for("admin_recensioni"):
            query = f"?{parsed.query}" if parsed.query else ""
            return redirect(parsed.path + query)

    return redirect(url_for("admin_recensioni"))

def redirect_admin_annunci_next():
    """
    Redirect sicuro verso admin_annunci conservando eventuali filtri.

    Priorità:
    1. ?next=/admin/annunci?...
    2. request.referrer se arriva da /admin/annunci?...
    3. fallback a /admin/annunci
    """
    from urllib.parse import urlparse

    next_url = request.args.get("next") or request.referrer

    if next_url:
        parsed = urlparse(next_url)

        # Caso 1: path interno tipo /admin/annunci?stato=in_attesa
        if not parsed.netloc and parsed.path == url_for("admin_annunci"):
            query = f"?{parsed.query}" if parsed.query else ""
            return redirect(parsed.path + query)

        # Caso 2: referrer assoluto stesso dominio
        if parsed.netloc == request.host and parsed.path == url_for("admin_annunci"):
            query = f"?{parsed.query}" if parsed.query else ""
            return redirect(parsed.path + query)

    return redirect(url_for("admin_annunci"))

# ==========================================================
# APPROVA / RIFIUTA RECENSIONI E RISPOSTE
# ==========================================================
@app.route("/admin/recensioni/approva/<int:id>")
@admin_required
def approva_recensione(id):
    from models import approva_elemento

    try:
        approva_elemento("recensioni", id)

        conn = get_db_connection()
        c = get_cursor(conn)
        c.execute(sql("""
            SELECT r.id_autore, r.id_destinatario, r.ultima_modifica, u.username
            FROM recensioni r
            JOIN utenti u ON r.id_autore = u.id
            WHERE r.id = ?
        """), (id,))
        row = c.fetchone()

        if row:
            id_autore = list(row.values())[0]
            id_destinatario = list(row.values())[1]
            ultima_modifica = list(row.values())[2]
            username_autore = list(row.values())[3] or "utente"

            if ultima_modifica:
                messaggio = f"@{username_autore} ha modificato la sua recensione su di te"
            else:
                messaggio = f"Hai ricevuto una nuova recensione da @{username_autore}"

            crea_notifica(
                id_destinatario,
                messaggio,
                link=url_for("mie_recensioni_ricevute")
            )

            emit_update_notifications(id_destinatario)

        # 🔁 Aggiorna counters admin (recensioni in attesa)
        invalidate_admin_counters()

        flash("✅ Recensione approvata e notifica inviata!", "success")

    except Exception as e:
        flash(f"Errore durante l'approvazione: {e}", "danger")

    return redirect_admin_recensioni_next()

@app.route("/admin/recensioni/rifiuta/<int:id>")
@admin_required
def rifiuta_recensione(id):
    from models import rifiuta_elemento
    try:
        rifiuta_elemento("recensioni", id)

        conn = get_db_connection()
        c = get_cursor(conn)
        c.execute(sql("SELECT id_autore FROM recensioni WHERE id = ?"), (id,))
        row = c.fetchone()


        if row:
            id_autore = list(row.values())[0]

            messaggio_notifica = "La tua recensione è stata rifiutata per contenuto poco appropriato. Modificala e inviala di nuovo. ❌"
            link_notifica = url_for("mie_recensioni")

            crea_notifica(
                id_autore,
                messaggio_notifica,
                link=link_notifica
            )

            emit_update_notifications(id_autore)

            try:
                invia_push(
                    id_autore,
                    "Recensione rifiutata ❌",
                    messaggio_notifica,
                    url=link_notifica
                )
            except Exception as e:
                log_exception_safe("⚠️ Errore push rifiuto recensione", e, {"user_id": id_autore}, production=True)

        # 🔁 Aggiorna counters admin (recensioni in attesa)
        invalidate_admin_counters()

        flash("❌ Recensione rifiutata!", "warning")

    except Exception as e:
        flash(f"Errore durante il rifiuto: {e}", "danger")

    return redirect_admin_recensioni_next()

@app.route("/admin/risposte/approva/<int:id>")
@admin_required
def approva_risposta(id):
    from models import approva_elemento
    try:
        approva_elemento("risposte_recensioni", id)

        conn = get_db_connection()
        c = get_cursor(conn)
        c.execute(sql("""
            SELECT r.id_autore, u.username
            FROM recensioni r
            JOIN risposte_recensioni rr ON rr.id_recensione = r.id
            JOIN utenti u ON rr.id_autore = u.id
            WHERE rr.id = ?
        """), (id,))
        row = c.fetchone()


        if row:
            id_autore = list(row.values())[0]
            username_risposta = list(row.values())[1] or "utente"

            crea_notifica(
                id_autore,
                f"La tua recensione a @{username_risposta} ha ricevuto una risposta 💬",
                link=url_for("mie_recensioni")
            )

            emit_update_notifications(id_autore)

        # 🔁 Aggiorna counters admin (recensioni in attesa)
        invalidate_admin_counters()

        flash("✅ Risposta approvata e notifica inviata!", "success")

    except Exception as e:
        flash(f"Errore durante l'approvazione della risposta: {e}", "danger")

    return redirect_admin_recensioni_next()

@app.route("/admin/risposte/rifiuta/<int:id>")
@admin_required
def rifiuta_risposta(id):
    from models import rifiuta_elemento
    try:
        # 1️⃣ Imposto lo stato della risposta a "rifiutata"
        rifiuta_elemento("risposte_recensioni", id)

        # 2️⃣ Recupero l'autore della risposta
        conn = get_db_connection()
        c = get_cursor(conn)
        c.execute(sql("SELECT id_autore FROM risposte_recensioni WHERE id = ?"), (id,))
        row = c.fetchone()


        if row:
            id_autore = list(row.values())[0]

            # 3️⃣ Creo la notifica per l'autore della risposta
            crea_notifica(
                id_autore,
                "La tua risposta a una recensione è stata rifiutata per contenuto poco appropriato. Modificala e inviala di nuovo. ❌",
                link=url_for("mie_recensioni_ricevute")  # 👉 o la route dove vede le sue risposte
            )

            # 4️⃣ Aggiorno in tempo reale il badge notifiche (Socket.IO)
            emit_update_notifications(id_autore)

        # 🔁 Aggiorna counters admin (recensioni in attesa)
        invalidate_admin_counters()

        flash("❌ Risposta rifiutata!", "warning")

    except Exception as e:
        flash(f"Errore durante il rifiuto della risposta: {e}", "danger")

    return redirect_admin_recensioni_next()
# ==========================================================
# ADMIN – LISTA ANNUNCI
# ==========================================================
@app.route("/admin/annunci")
@admin_required
def admin_annunci():

    # =========================
    # FILTRI
    # =========================
    utente = (request.args.get("utente") or "").strip().lower()
    categoria = (request.args.get("categoria") or "").strip()
    zona = (request.args.get("zona") or "").strip()
    provincia = (request.args.get("provincia") or "").strip()
    stato = (request.args.get("stato") or "").strip().lower()

    # 🔁 NUOVO: offro / cerco
    tipo_annuncio = (request.args.get("tipo_annuncio") or "").strip().lower()
    if tipo_annuncio not in ("offro", "cerco"):
        tipo_annuncio = ""

    # =========================
    # DB
    # =========================
    conn = get_db_connection()

    c = get_cursor(conn)

    query = f"""
        SELECT
            a.id,
            a.titolo,
            a.descrizione,
            a.media,
            a.categoria,
            a.tipo_annuncio,
            a.zona,
            a.provincia,
            a.stato,
            a.data_pubblicazione,
            a.utente_id,
            u.nome,
            u.cognome,
            u.email,
            u.username,

            /* BOOST LISTA */
            CASE WHEN EXISTS (
                SELECT 1
                FROM attivazioni_servizi act
                JOIN servizi s ON s.id = act.servizio_id
                WHERE act.annuncio_id = a.id
                  AND s.codice = 'boost_lista'
                  AND act.stato = 'attivo'
                  AND act.data_inizio <= {now_sql()}
                  AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
            ) THEN 1 ELSE 0 END AS has_boost_lista,

            /* VETRINA */
            CASE WHEN EXISTS (
                SELECT 1
                FROM attivazioni_servizi act
                JOIN servizi s ON s.id = act.servizio_id
                WHERE act.annuncio_id = a.id
                  AND s.codice = 'vetrina_annuncio'
                  AND act.stato = 'attivo'
                  AND act.data_inizio <= {now_sql()}
                  AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
            ) THEN 1 ELSE 0 END AS has_vetrina,

            /* BADGE EVIDENZA */
            CASE WHEN EXISTS (
                SELECT 1
                FROM attivazioni_servizi act
                JOIN servizi s ON s.id = act.servizio_id
                WHERE act.annuncio_id = a.id
                  AND s.codice = 'badge_evidenza'
                  AND act.stato = 'attivo'
                  AND act.data_inizio <= {now_sql()}
                  AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
            ) THEN 1 ELSE 0 END AS has_badge_evidenza,

            /* ANNUNCIO URGENTE */
            CASE WHEN EXISTS (
                SELECT 1
                FROM attivazioni_servizi act
                JOIN servizi s ON s.id = act.servizio_id
                WHERE act.annuncio_id = a.id
                  AND s.codice = 'annuncio_urgente'
                  AND act.stato = 'attivo'
                  AND act.data_inizio <= {now_sql()}
                  AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
            ) THEN 1 ELSE 0 END AS has_urgente,

            /* BADGE AFFIDABILITÀ */
            CASE WHEN EXISTS (
                SELECT 1
                FROM attivazioni_servizi act
                JOIN servizi s ON s.id = act.servizio_id
                WHERE act.utente_id = a.utente_id
                  AND act.annuncio_id IS NULL
                  AND s.codice = 'badge_affidabilita'
                  AND act.stato = 'attivo'
            ) THEN 1 ELSE 0 END AS has_affidabilita,

            /* CONTATTI PROFILO */
            CASE WHEN EXISTS (
                SELECT 1
                FROM attivazioni_servizi act
                JOIN servizi s ON s.id = act.servizio_id
                WHERE act.utente_id = a.utente_id
                  AND act.annuncio_id IS NULL
                  AND s.codice = 'contatti'
                  AND act.stato = 'attivo'
            ) THEN 1 ELSE 0 END AS has_contatti,

            /* PACCHETTO VISIBILITÀ
               Attivo solo se TUTTI i servizi collegati al pacchetto risultano attivi.
               I servizi profilo vengono controllati su utente_id + annuncio_id NULL.
               I servizi annuncio vengono controllati su annuncio_id.
            */
            CASE WHEN EXISTS (
                SELECT 1
                FROM pacchetti p
                WHERE p.codice = 'pacchetto_visibilita'
                  AND p.attivo = 1
                  AND NOT EXISTS (
                    SELECT 1
                    FROM pacchetti_servizi ps
                    JOIN servizi s ON s.id = ps.servizio_id
                    WHERE ps.pacchetto_id = p.id
                      AND NOT EXISTS (
                        SELECT 1
                        FROM attivazioni_servizi act
                        WHERE act.servizio_id = s.id
                          AND act.utente_id = a.utente_id
                          AND (
                            (
                              s.ambito = 'profilo'
                              AND act.annuncio_id IS NULL
                            )
                            OR
                            (
                              s.ambito <> 'profilo'
                              AND act.annuncio_id = a.id
                            )
                          )
                          AND act.stato = 'attivo'
                          AND act.data_inizio <= {now_sql()}
                          AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
                      )
                  )
            ) THEN 1 ELSE 0 END AS has_pacchetto_visibilita,

            /* VISIBILITÀ PREMIUM
               Nel database il codice reale è 'visibilita_premium',
               non 'pacchetto_premium'.
            */
            CASE WHEN EXISTS (
                SELECT 1
                FROM pacchetti p
                WHERE p.codice = 'visibilita_premium'
                  AND p.attivo = 1
                  AND NOT EXISTS (
                    SELECT 1
                    FROM pacchetti_servizi ps
                    JOIN servizi s ON s.id = ps.servizio_id
                    WHERE ps.pacchetto_id = p.id
                      AND NOT EXISTS (
                        SELECT 1
                        FROM attivazioni_servizi act
                        WHERE act.servizio_id = s.id
                          AND act.utente_id = a.utente_id
                          AND (
                            (
                              s.ambito = 'profilo'
                              AND act.annuncio_id IS NULL
                            )
                            OR
                            (
                              s.ambito <> 'profilo'
                              AND act.annuncio_id = a.id
                            )
                          )
                          AND act.stato = 'attivo'
                          AND act.data_inizio <= {now_sql()}
                          AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
                      )
                  )
            ) THEN 1 ELSE 0 END AS has_visibilita_premium

        FROM annunci a
        JOIN utenti u ON a.utente_id = u.id
        WHERE 1=1
    """
    params = []

    # =========================
    # APPLICAZIONE FILTRI
    # =========================
    if utente:
        like = f"%{utente}%"
        query += " AND LOWER(u.username) LIKE ?"
        params.append(like)

    if categoria:
        query += " AND a.categoria = ?"
        params.append(categoria)

    if tipo_annuncio:
        query += " AND a.tipo_annuncio = ?"
        params.append(tipo_annuncio)

    if provincia:
        query += " AND LOWER(a.provincia) = LOWER(?)"
        params.append(provincia)

    if zona:
        query += " AND LOWER(a.zona) LIKE ?"
        params.append(f"%{zona.lower()}%")

    if stato:
        query += " AND a.stato = ?"
        params.append(stato)

    query += " ORDER BY a.data_pubblicazione DESC"

    # =========================
    # EXEC
    # =========================
    c.execute(sql(query), params)
    annunci = [dict(row) for row in c.fetchall()]

    # =========================
    # CATEGORIE (select admin)
    # =========================
    with open("static/data/filtri_categoria.json", encoding="utf-8") as f:
        categorie = list(json.load(f).keys())

    # =========================
    # CONTATORI
    # =========================
    c.execute(sql("SELECT COUNT(*) AS totale FROM annunci"))
    totale_annunci = c.fetchone()["totale"]

    c.execute(sql("SELECT COUNT(*) AS totale FROM annunci WHERE stato = ?"), ("in_attesa",))
    totale_in_attesa = c.fetchone()["totale"]

    totale_filtrati = len(annunci)

    return render_template(
        "admin_annunci.html",
        annunci=annunci,
        categorie=categorie,
        totale_annunci=totale_annunci,
        totale_filtrati=totale_filtrati,
        totale_in_attesa=totale_in_attesa
    )

@app.route("/admin/annunci/approva/<int:id>")
@admin_required
def approva_annuncio(id):
    conn = get_db_connection()

    c = get_cursor(conn)

    c.execute(sql(f"""
        UPDATE annunci
        SET stato = 'approvato',
            approvato_il = {now_sql()},
            match_da_processare = 1
        WHERE id = ?
    """), (id,))

    c.execute(sql("SELECT utente_id FROM annunci WHERE id = ?"), (id,))
    row = c.fetchone()

    utente_id = None
    if row:
        utente_id = row["utente_id"]
        c.execute(sql("UPDATE utenti SET visibile_pubblicamente = 1 WHERE id = ?"), (utente_id,))
        conn.commit()



    if utente_id:
        messaggio_notifica = "Il tuo annuncio è stato approvato ed è ora visibile su MyLocalCare ✅"
        link_notifica = url_for("dashboard") + "?tab=annunci"

        crea_notifica(
            utente_id,
            messaggio_notifica,
            link=link_notifica
        )

        emit_update_notifications(utente_id)

        try:
            invia_push(
                utente_id,
                "Annuncio approvato ✅",
                messaggio_notifica,
                url=link_notifica
            )
        except Exception as e:
            log_exception_safe("⚠️ Errore push approvazione annuncio", e, {"user_id": utente_id}, production=True)

    # 🔁 Aggiorna counters admin (annunci in attesa)
    invalidate_admin_counters()

    return redirect_admin_annunci_next()

@app.route("/admin/annunci/rifiuta/<int:id>")
@admin_required
def rifiuta_annuncio(id):
    conn = get_db_connection()
    c = get_cursor(conn)

    # 1️⃣ Update stato
    c.execute(sql("""
        UPDATE annunci
        SET stato = 'rifiutato'
        WHERE id = ?
    """), (id,))

    # 2️⃣ Recupero utente DOPO update (come approva)
    c.execute(sql("SELECT utente_id FROM annunci WHERE id = ?"), (id,))
    row = c.fetchone()

    utente_id = None
    if row:
        utente_id = row["utente_id"]
        conn.commit()

    # 3️⃣ Notifica dopo commit completo
    if utente_id:
        messaggio_notifica = (
            "Il tuo annuncio è stato rifiutato perché non conforme alle linee guida di MyLocalCare. "
            "Puoi modificarlo e ripubblicarlo. ❌"
        )
        link_notifica = url_for("dashboard") + "?tab=annunci"

        crea_notifica(
            utente_id,
            messaggio_notifica,
            link=link_notifica
        )

        emit_update_notifications(utente_id)

        try:
            invia_push(
                utente_id,
                "Annuncio rifiutato ❌",
                messaggio_notifica,
                url=link_notifica
            )
        except Exception as e:
            log_exception_safe("⚠️ Errore push rifiuto annuncio", e, {"user_id": utente_id}, production=True)

    invalidate_admin_counters()

    flash("Annuncio rifiutato ❌", "warning")

    return redirect_admin_annunci_next()

# ==========================================================
# NOTIFICHE - FUNZIONI DI SUPPORTO (AGGIUNTA)
# ==========================================================
def get_notifiche_utente(user_id):
    conn = get_db_connection()
    notifiche = conn.execute(
        "SELECT * FROM notifiche WHERE id_utente = ? ORDER BY data DESC",
        (user_id,)
    ).fetchall()

    return notifiche

def conta_non_lette(user_id):
    conn = get_db_connection()
    c = get_cursor(conn)
    c.execute(sql("SELECT COUNT(*) FROM notifiche WHERE id_utente = ? AND letta = 0"), (user_id,))
    count = fetchone_value(c.fetchone())

    return count

def segna_notifica_letta(notifica_id, user_id):
    conn = get_db_connection()
    conn.execute(sql(f"""
        UPDATE notifiche
        SET letta = 1,
            data_lettura = {now_sql()}
        WHERE id = ? AND id_utente = ?
    """), (notifica_id, user_id))
    conn.commit()

@app.route("/notifiche/segna_tutte_lette", methods=["POST"])
def segna_tutte_lette_route():
    verify_csrf()

    if "utente_id" not in session:
        return jsonify({"success": False}), 403

    from models import segna_tutte_lette
    segna_tutte_lette(session["utente_id"])

    # 🔔 Aggiorna il badge in tempo reale
    emit_update_notifications(session["utente_id"])

    return jsonify({"success": True})

@app.route("/notifiche/elimina_tutte", methods=["POST"])
def elimina_tutte_notifiche_route():
    verify_csrf()

    if "utente_id" not in session:
        return jsonify({"success": False}), 403

    from models import elimina_tutte_notifiche
    elimina_tutte_notifiche(session["utente_id"])

    # 🔔 Aggiorna il badge in tempo reale
    emit_update_notifications(session["utente_id"])

    return jsonify({"success": True})

@app.route("/notifiche/elimina/<int:id>", methods=["POST"])
@login_required
def elimina_notifica_singola_route(id):
    verify_csrf()

    conn = get_db_connection()
    conn.execute(
        sql("DELETE FROM notifiche WHERE id = ? AND id_utente = ?"),
        (id, g.utente["id"])
    )
    conn.commit()

    # 🔔 aggiorna il badge realtime
    emit_update_notifications(g.utente["id"])

    return jsonify({"success": True})

def pulisci_notifiche_vecchie(giorni=None):
    """
    Elimina automaticamente le notifiche lette da più di X giorni.

    Regole:
    - NON elimina notifiche non lette;
    - NON elimina notifiche lette senza data_lettura;
    - default: 30 giorni;
    - compatibile PostgreSQL + SQLite;
    - sicura anche se chiamata da background task.
    """

    if giorni is None:
        giorni = app.config.get("NOTIFICHE_TTL_GIORNI", 30)

    try:
        giorni = int(giorni)
    except Exception:
        giorni = 30

    if giorni < 1:
        giorni = 30

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = get_cursor(conn)

        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                DELETE FROM notifiche
                WHERE letta = 1
                  AND data_lettura IS NOT NULL
                  AND data_lettura < CURRENT_TIMESTAMP - (? * INTERVAL '1 day')
            """, (giorni,))
        else:
            cur.execute("""
                DELETE FROM notifiche
                WHERE letta = 1
                  AND data_lettura IS NOT NULL
                  AND data_lettura < datetime('now', '-' || ? || ' days')
            """, (giorni,))

        eliminate = cur.rowcount if cur.rowcount is not None else 0

        conn.commit()

        security_log(
            "🧹 Pulizia notifiche vecchie completata",
            {
                "giorni": giorni,
                "eliminate": eliminate
            },
            production=True
        )

        return eliminate

    except Exception as e:
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass

        log_exception_safe(
            "⚠️ Errore pulizia notifiche vecchie",
            e,
            {
                "giorni": giorni
            },
            production=True
        )

        return 0

    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass

        try:
            if conn:
                conn.close()
        except Exception:
            pass


def pulizia_notifiche_background_loop():
    """
    Pulizia automatica periodica delle notifiche lette.

    Usa Redis come lock distribuito, così su Render:
    - anche se ci sono più worker web;
    - anche se il servizio viene scalato;
    - anche se il task parte più volte;

    la pulizia viene eseguita da un solo processo alla volta.
    """

    # Aspetta un minuto dopo l'avvio del worker,
    # così non rallenta il boot dell'app.
    socketio.sleep(60)

    while True:
        lock_key = "lock:pulizia_notifiche_vecchie"
        lock_token = secrets.token_hex(16)

        try:
            lock_acquisito = redis_client.set(
                lock_key,
                lock_token,
                nx=True,
                ex=600
            )

            if lock_acquisito:
                with app.app_context():
                    pulisci_notifiche_vecchie()

        except Exception as e:
            log_exception_safe(
                "⚠️ Errore loop pulizia notifiche vecchie",
                e,
                production=True
            )

        finally:
            try:
                valore_lock = redis_client.get(lock_key)

                if valore_lock:
                    if isinstance(valore_lock, bytes):
                        valore_lock = valore_lock.decode("utf-8")

                    if valore_lock == lock_token:
                        redis_client.delete(lock_key)

            except Exception:
                pass

        # Ripete la pulizia ogni 6 ore.
        socketio.sleep(6 * 60 * 60)

# 🧹 Avvia pulizia automatica notifiche vecchie solo sul servizio web.
# Il servizio realtime/chat non deve occuparsi della manutenzione DB ordinaria.
if APP_RUNTIME_ROLE == "web":
    socketio.start_background_task(pulizia_notifiche_background_loop)

# ==========================================================
# 🔁 DAILY MATCHES — LOOP AUTOMATICO GIORNALIERO
# ==========================================================

def daily_matches_background_loop():
    """
    Esegue il controllo Daily Matches ogni giorno all'orario configurato.

    Protezioni:
    - gira solo sul servizio web;
    - usa Redis lock per evitare doppie esecuzioni su più worker;
    - salva in Redis la data dell'ultima esecuzione per non ripetere lo stesso giorno.
    """

    socketio.sleep(30)

    while True:
        try:
            with app.app_context():
                settings = get_daily_matches_settings()

                if not settings.get("enabled", True):
                    socketio.sleep(60)
                    continue

                now_rome = datetime.now(ZoneInfo("Europe/Rome"))
                today_key = now_rome.strftime("%Y-%m-%d")
                current_hm = now_rome.strftime("%H:%M")
                target_hm = settings.get("time", "08:00")

                last_run_key = "daily_matches:last_run_date"
                lock_key = "lock:daily_matches"

                last_run = redis_client.get(last_run_key)

                if isinstance(last_run, bytes):
                    last_run = last_run.decode("utf-8")

                if current_hm >= target_hm and last_run != today_key:
                    lock_token = secrets.token_hex(16)

                    lock_acquired = redis_client.set(
                        lock_key,
                        lock_token,
                        nx=True,
                        ex=1800
                    )

                    if lock_acquired:
                        try:
                            result = processa_match_nuovi_annunci(
                                channel=settings.get("channel", "internal")
                            )

                            if result.get("ok"):
                                redis_client.set(
                                    last_run_key,
                                    today_key,
                                    ex=60 * 60 * 48
                                )

                            security_log(
                                "✅ Daily Matches automatico eseguito",
                                {
                                    "date": today_key,
                                    "time": current_hm,
                                    "settings": settings,
                                    "result": result
                                },
                                production=True
                            )

                        finally:
                            try:
                                valore_lock = redis_client.get(lock_key)

                                if isinstance(valore_lock, bytes):
                                    valore_lock = valore_lock.decode("utf-8")

                                if valore_lock == lock_token:
                                    redis_client.delete(lock_key)

                            except Exception:
                                pass

        except Exception as e:
            log_exception_safe(
                "⚠️ Errore loop Daily Matches",
                e,
                production=True
            )

        socketio.sleep(60)


# Avvio automatico solo sul servizio web.
if APP_RUNTIME_ROLE == "web":
    socketio.start_background_task(daily_matches_background_loop)

# Rende disponibile in tutti i template un helper per le notifiche non lette
@app.context_processor
def utility_functions():
    def unread_chat():
        """Conta solo i messaggi chat non letti"""
        if g.utente:
            return count_chat_non_letti(g.utente['id'])
        return 0

    def unread_notifications():
        """Conta solo le notifiche non lette"""
        if g.utente:
            return conta_non_lette(g.utente['id'])
        return 0

    # (Mantieni unread_count solo se serve un totale combinato)
    def unread_count():
        """Totale generale (chat + notifiche), se serve in futuro"""
        if g.utente:
            return unread_chat() + unread_notifications()
        return 0

    return dict(
        unread_chat=unread_chat,
        unread_notifications=unread_notifications,
        unread_count=unread_count
    )

@app.context_processor
def inject_helpers():
    from models import get_risposta_by_recensione
    return dict(get_risposta_by_recensione=get_risposta_by_recensione)

@app.context_processor
def inject_pytz():
    import pytz
    return dict(pytz=pytz)

@app.route('/notifiche/unread_count')
@login_required
def notifiche_unread_count():
    return jsonify({"count": conta_non_lette(g.utente['id'])})

# ==========================================================
# 🔹 CACHE per ADMIN COUNTERS
# ==========================================================
from time import time

# Inizializza cache e TTL (5 secondi di durata)
app.config.setdefault("_ADMIN_COUNTERS_CACHE", {"ts": 0, "payload": None})
app.config.setdefault("_ADMIN_COUNTERS_TTL", 1)  # secondi

# ==========================================================
# 🧹 Helper: resetta cache admin e aggiorna il badge live
# ==========================================================
def invalidate_admin_counters():
    """Pulisce la cache counters e forza aggiornamento admin live."""
    try:
        app.config["_ADMIN_COUNTERS_CACHE"] = {"ts": 0, "payload": None}
        socketio.emit("update_admin_counters", namespace="/")
        print("♻️ Cache admin counters invalidata e badge aggiornato.")
    except Exception as e:
        print(f"⚠️ Errore invalidate_admin_counters: {e}")


# ==========================================================
# 3️⃣ MIDDLEWARE E DASHBOARD UTENTE
# ==========================================================

# --- Middleware per proteggere pagine riservate ---
@app.before_request
def load_logged_in_user():
    # disponibile sempre nei template
    g.path = request.path
    g.utente = None

    # NON interrogare il DB per richieste statiche o infrastrutturali
    if request.endpoint == "static":
        return

    if request.path.startswith("/static/"):
        return

    if request.path.startswith("/socket.io"):
        return

    if request.path in {
        "/service-worker.js",
        "/robots.txt",
        "/favicon.ico",
        "/manifest.json",
    }:
        return

    user_id = session.get("utente_id")
    if user_id is None:
        return

    try:
        conn = get_db_connection()
        cur = get_cursor(conn)
        cur.execute(sql("SELECT * FROM utenti WHERE id = ?"), (user_id,))
        g.utente = cur.fetchone()
    except Exception as e:
        print(f"⚠️ load_logged_in_user errore: {e}")
        g.utente = None

# --- Dashboard Utente ---
@app.route("/annuncio/<int:id>/modifica", methods=["GET", "POST"])
@login_required
def modifica_annuncio(id):
    conn = get_db_connection()

    c = get_cursor(conn)

    annuncio = c.execute(
        "SELECT * FROM annunci WHERE id = ?",
        (id,)
    ).fetchone()

    # 🔒 Sicurezza: annuncio esistente e di proprietà dell’utente
    if not annuncio or annuncio["utente_id"] != g.utente["id"]:

        flash("Non puoi modificare questo annuncio.", "error")
        return redirect(url_for("dashboard"))

    # =========================================================
    # 📤 POST
    # =========================================================
    if request.method == "POST":

        # 🧩 Protezione extra contro manomissione ID
        id_form = request.form.get("id_annuncio")
        if id_form and str(id_form) != str(id):

            flash("Tentativo di modifica non autorizzato.", "error")
            return redirect(url_for("dashboard"))

        # 🔹 CAMPI BASE
        titolo = request.form.get("titolo", "").strip()
        descrizione = request.form.get("descrizione", "").strip()
        raw_categoria = request.form.get("categoria", "").strip()
        categoria = to_slug(raw_categoria)
        tipo_annuncio = request.form.get("tipo_annuncio", "").strip().lower()

        # 🔹 ZONA + PROVINCIA
        zona = request.form.get("zona", "").strip()
        provincia = request.form.get("provincia", "").strip()

        # 🔹 ALTRI CAMPI
        prezzo = request.form.get("prezzo", "").strip()
        telefono = request.form.get("telefono", "").strip()
        email = request.form.get("email", "").strip()
        bio = request.form.get("bio_utente", "").strip()
        filtri = request.form.getlist("filtri_categoria")

        # =====================================================
        # 🛡️ VALIDAZIONI
        # =====================================================

        if not titolo:

            flash("Inserisci un titolo per l’annuncio.", "warning")
            return redirect(url_for("modifica_annuncio", id=id))

        if not descrizione:

            flash("Inserisci una descrizione dettagliata.", "warning")
            return redirect(url_for("modifica_annuncio", id=id))

        if contiene_contatti_nel_testo(descrizione):

            flash(
                "Non inserire telefono, email o altri contatti nella descrizione. "
                "Usa i campi Telefono ed Email nella sezione Dettagli e contatti.",
                "warning"
            )
            return redirect(url_for("modifica_annuncio", id=id))

        if tipo_annuncio not in ("offro", "cerco"):

            flash("Devi specificare se l’annuncio è 'Offro' oppure 'Cerco'.", "warning")
            return redirect(url_for("modifica_annuncio", id=id))

        if not zona or not provincia:

            flash("Seleziona un comune valido dall’elenco.", "warning")
            return redirect(url_for("modifica_annuncio", id=id))

        # =====================================================
        # 📸 MEDIA – gestione immagini
        # =====================================================
        media_attuale = [
            img.strip()
            for img in (annuncio["media"] or "").split(",")
            if img.strip()
        ]

        immagini_da_cancellare = request.form.getlist("cancellate")

        immagini_rimanenti = [
            img for img in media_attuale
            if img not in immagini_da_cancellare
        ]

        # 🗑️ Elimina immagini rimosse
        for foto in immagini_da_cancellare:
            percorso_file = os.path.join("static", foto)
            if os.path.exists(percorso_file):
                try:
                    os.remove(percorso_file)
                except Exception as e:
                    print(f"⚠️ Impossibile eliminare {percorso_file}: {e}")

        # 📸 Upload nuove immagini
        nuove_foto = request.files.getlist("media")
        upload_dir = os.path.join("static", "uploads", "annunci")
        os.makedirs(upload_dir, exist_ok=True)

        for foto in nuove_foto:
            if foto and foto.filename:

                if not foto.mimetype.startswith("image/"):
                    flash("Puoi caricare solo immagini.", "warning")
                    return redirect(url_for("modifica_annuncio", id=id))

                if len(immagini_rimanenti) >= 4:
                    flash("Puoi avere al massimo 4 foto per annuncio.", "warning")
                    return redirect(url_for("modifica_annuncio", id=id))

                nome_file = f"{uuid.uuid4().hex}_{foto.filename}"
                percorso = os.path.join(upload_dir, nome_file)
                foto.save(percorso)

                immagini_rimanenti.append(f"uploads/annunci/{nome_file}")

        media_finale = ",".join(immagini_rimanenti)

        # Foto card opzionale: se non scelta, resta NULL/vuota e in Cerca userai foto profilo
        foto_card = request.form.get("foto_card_path", "").strip()

        if foto_card and foto_card not in immagini_rimanenti:
            foto_card = ""

        # =====================================================
        # 💾 UPDATE DB
        # =====================================================
        c.execute(sql("""
            UPDATE annunci
            SET
                titolo = ?,
                descrizione = ?,
                categoria = ?,
                tipo_annuncio = ?,
                zona = ?,
                provincia = ?,
                prezzo = ?,
                telefono = ?,
                email = ?,
                bio_utente = ?,
                media = ?,
                foto_card = ?,
                filtri_categoria = ?,
                stato = 'in_attesa'
            WHERE id = ?
        """), (
            titolo,
            descrizione,
            categoria,
            tipo_annuncio,
            zona,
            provincia,
            prezzo,
            telefono,
            email,
            bio,
            media_finale,
            foto_card,
            ",".join(filtri),
            id
        ))

        conn.commit()


        # 🔁 Aggiorna contatori admin
        invalidate_admin_counters()

        notifica_admin_evento(
            titolo="Annuncio modificato in attesa",
            messaggio=f"Annuncio modificato da revisionare: {titolo}",
            link=url_for("admin_annunci", stato="in_attesa"),
            push=True
        )

        flash("✅ Annuncio aggiornato con successo (sarà revisionato).", "success")
        return redirect(url_for("dashboard"))

    # =========================================================
    # 📥 GET
    # =========================================================

    return render_template(
        "modifica_annuncio.html",
        modalita="modifica",
        annuncio=annuncio,
        filtri_per_categoria=get_filtri_categoria_da_db()
    )

# --- Dashboard: carica anche i nuovi campi (sostituisci la tua dashboard() attuale) ---
@app.route('/utente/dashboard')
@login_required
def dashboard():
    conn = get_db_connection()
    ut = conn.execute(sql("""
        SELECT id, nome, cognome, email, username, citta, lingue, frase,
               telefono, email_pubblica, indirizzo_studio,
               sito_web, instagram, facebook, linkedin,
               orari, preferenze_contatto,
               visibile_pubblicamente, visibile_in_chat,
               media_recensioni, numero_recensioni, foto_profilo, copertina, foto_galleria,
               offro_1, offro_2, offro_3, offro_4, offro_5, offro_6, offro_7, offro_8, offro_9, offro_10, offro_11, offro_12, offro_13,
               cerco_1, cerco_2, cerco_3, cerco_4, cerco_5, cerco_6, cerco_7, cerco_8, cerco_9, cerco_10, cerco_11, cerco_12, cerco_13,
               esperienza_1, esperienza_2, esperienza_3,
               studio_1, studio_2, studio_3, certificazioni,
               descrizione
        FROM utenti
        WHERE id = ?
    """), (session["utente_id"],)).fetchone()



    # Se non trovato, gestisci come preferisci
    if not ut:
        flash("Utente non trovato.", "error")
        return redirect(url_for("home"))

    utente = dict(ut)
    # cast sicuro dei flag (possono essere None/0/1 o stringhe)
    for i in range(1, 14):
        utente[f"offro_{i}"] = int(utente.get(f"offro_{i}") or 0)
        utente[f"cerco_{i}"] = int(utente.get(f"cerco_{i}") or 0)

    # Calcoli recensioni
    media, totale = calcola_media_recensioni(utente['id'])

    # 🟡 Badge Fiducia Top automatico / manuale
    utente["affidabilita_top"] = 1 if (
        servizio_attivo_per_utente(
            utente_id=utente["id"],
            codice_servizio="badge_affidabilita"
        )
        or
        (
            float(media or 0) >= 4
            and int(totale or 0) >= 4
        )
    ) else 0

    # 🔹 Carica recensioni ricevute e scritte per la tab "Recensioni"
    recensioni_ricevute = get_recensioni_utente(utente['id'])
    recensioni_scritte = get_recensioni_scritte(utente['id'])

    # 🔹 Carica gli annunci dell'utente loggato
    conn = get_db_connection()

    c = get_cursor(conn)
    c.execute(sql("""
        SELECT id, titolo, categoria, descrizione, zona, filtri_categoria,
               data_pubblicazione, stato
        FROM annunci
        WHERE utente_id = ?
          AND COALESCE(stato, '') <> 'eliminato'
        ORDER BY data_pubblicazione DESC
    """), (session["utente_id"],))
    annunci = [dict(r) for r in c.fetchall()]


    # 🔹 Ritorna la dashboard con gli annunci caricati
    return render_template(
        'dashboard.html',
        utente=utente,
        user=g.utente,
        annunci=annunci,
        media_recensioni=media,
        totale_recensioni=totale,
        recensioni_ricevute=recensioni_ricevute,
        recensioni_scritte=recensioni_scritte,
        pubblico=False,
        page="profilo"
    )

# --- Aggiorna INFO (tab) ---
@app.route("/utente/update_info", methods=["POST"])
@login_required
def utente_update_info():
    user_id = session.get("utente_id")
    if not user_id:
        flash("Errore: sessione utente non valida.", "error")
        return redirect(url_for("dashboard"))

    conn = get_db_connection()
    c = get_cursor(conn)

    # 🔹 Campi base
    citta = request.form.get("citta", "")
    citta = citta.strip()
    provincia = get_provincia_from_comune(citta) if citta else None
    lingue = request.form.get("lingue", "")
    frase = request.form.get("frase", "")
    descrizione = request.form.get("descrizione", "")
    telefono = request.form.get("telefono", "")
    indirizzo_studio = request.form.get("indirizzo_studio", "")
    sito_web = request.form.get("sito_web", "")
    instagram = request.form.get("instagram", "")
    facebook = request.form.get("facebook", "")
    linkedin = request.form.get("linkedin", "")
    orari = request.form.get("orari", "")
    preferenze_contatto = request.form.get("preferenze_contatto", "")
    # ✅ Non aggiornare mai "email" da questo form (non viene inviato)
    # ✅ email_pubblica: se nel form non c'è, la lasciamo invariata
    c.execute(sql("SELECT email, email_pubblica FROM utenti WHERE id = ?"), (user_id,))
    row = c.fetchone()
    email_db = list(row.values())[0] if row else ""
    email_pubblica_db = list(row.values())[1] if row else ""

    email = email_db  # resta quella vera dell’account

    email_pubblica_form = request.form.get("email_pubblica", "").strip()
    email_pubblica = email_pubblica_form if email_pubblica_form != "" else (email_pubblica_db or "")

    esperienza_1 = request.form.get("esperienza_1", "")
    esperienza_2 = request.form.get("esperienza_2", "")
    esperienza_3 = request.form.get("esperienza_3", "")
    studio_1 = request.form.get("studio_1", "")
    studio_2 = request.form.get("studio_2", "")
    studio_3 = request.form.get("studio_3", "")
    certificazioni = request.form.get("certificazioni", "")

    # 🔹 Checkbox attività — se un campo non arriva nel POST, conserva il valore già salvato
    c.execute(sql("""
        SELECT
            offro_1, offro_2, offro_3, offro_4, offro_5, offro_6, offro_7,
            offro_8, offro_9, offro_10, offro_11, offro_12, offro_13,
            cerco_1, cerco_2, cerco_3, cerco_4, cerco_5, cerco_6, cerco_7,
            cerco_8, cerco_9, cerco_10, cerco_11, cerco_12, cerco_13
        FROM utenti
        WHERE id = ?
    """), (user_id,))

    valori_attuali = c.fetchone() or {}

    def _cb(name):
        vals = request.form.getlist(name)

        if vals:
            return int(vals[-1])

        return int(valori_attuali.get(name) or 0)

    offro = [_cb(f"offro_{i}") for i in range(1, 14)]
    cerco = [_cb(f"cerco_{i}") for i in range(1, 14)]

    privacy_debug("update_info attività", {
        "user_id": user_id,
        "offro_count": sum(offro),
        "cerco_count": sum(cerco)
    })

    # 🔹 Query esplicita e completa
    # 🔹 Query SOLO per TAB "Info di base"
    query_update = """
        UPDATE utenti SET
            citta = ?,
            provincia = ?,
            lingue = ?,
            frase = ?,
            offro_1 = ?, offro_2 = ?, offro_3 = ?, offro_4 = ?, offro_5 = ?,
            offro_6 = ?, offro_7 = ?, offro_8 = ?, offro_9 = ?, offro_10 = ?,
            offro_11 = ?, offro_12 = ?, offro_13 = ?,
            cerco_1 = ?, cerco_2 = ?, cerco_3 = ?, cerco_4 = ?, cerco_5 = ?,
            cerco_6 = ?, cerco_7 = ?, cerco_8 = ?, cerco_9 = ?, cerco_10 = ?,
            cerco_11 = ?, cerco_12 = ?, cerco_13 = ?
        WHERE id = ?
    """

    valori = (
        citta, provincia, lingue, frase,
        *offro, *cerco,
        user_id
    )

    try:
        c.execute(sql(query_update), valori)
        conn.commit()

        # 🔁 ALLINEA LA MACRO-AREA COME IN LANDING
        if citta:
            provincia = get_provincia_from_comune(citta)

            if provincia:
                session["macro_area"] = provincia
                session["macro_comune"] = citta

                # ⭐ SINCRONIZZA ANCHE IL DB
                conn2 = get_db_connection()
                conn2.execute(
                    "UPDATE utenti SET macro_area = ? WHERE id = ?",
                    (provincia, user_id)
                )
                conn2.commit()
                conn2.close()

    except Exception as e:
        conn.rollback()
        print("❌ ERRORE SALVATAGGIO:", e)

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({
                "ok": False,
                "error": str(e)
            }), 500

        flash(f"Errore nel salvataggio: {e}", "error")

    finally:
        try:
            conn.close()
        except:
            pass

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({
            "ok": True
        })

    flash("✅ Modifiche salvate con successo.", "success")
    return redirect(url_for("dashboard"))

@app.route("/utente/update_esperienza", methods=["POST"])
@login_required
def utente_update_esperienza():
    privacy_debug("update_esperienza chiamata", {
        "user_id": session.get("utente_id"),
        "campi_presenti": list(request.form.keys())
    })
    user_id = session.get("utente_id")
    if not user_id:
        flash("Sessione non valida.", "error")
        return redirect(url_for("login"))

    conn = get_db_connection()
    c = get_cursor(conn)

    esperienza_1 = request.form.get("esperienza_1", "")
    esperienza_2 = request.form.get("esperienza_2", "")
    esperienza_3 = request.form.get("esperienza_3", "")
    studio_1 = request.form.get("studio_1", "")
    studio_2 = request.form.get("studio_2", "")
    studio_3 = request.form.get("studio_3", "")
    certificazioni = request.form.get("certificazioni", "")

    privacy_debug("update_esperienza dati letti", {
        "user_id": user_id,
        "ha_esperienze": any([esperienza_1, esperienza_2, esperienza_3]),
        "ha_formazione": any([studio_1, studio_2, studio_3]),
        "ha_certificazioni": bool(certificazioni)
    })

    try:
        c.execute(sql("""
            UPDATE utenti SET
                esperienza_1 = ?, esperienza_2 = ?, esperienza_3 = ?,
                studio_1 = ?, studio_2 = ?, studio_3 = ?,
                certificazioni = ?
            WHERE id = ?
        """), (esperienza_1, esperienza_2, esperienza_3, studio_1, studio_2, studio_3, certificazioni, user_id))
        conn.commit()
        flash("✅ Esperienza e formazione aggiornate con successo.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"❌ Errore durante il salvataggio: {e}", "error")
    finally:
        try:
            conn.close()
        except:
            pass


    return redirect(url_for("dashboard") + "#tab-info")

# ---------------------------------------------------------
# 📞 AGGIORNA CONTATTI UTENTE
# ---------------------------------------------------------
@app.route("/utente/update_contatti", methods=["POST"])
@login_required
def utente_update_contatti():
    privacy_debug("update_contatti chiamata", {
        "user_id": session.get("utente_id"),
        "campi_presenti": list(request.form.keys())
    })

    user_id = session.get("utente_id")
    conn = get_db_connection()
    c = get_cursor(conn)

    telefono = request.form.get("telefono", "")
    email_pubblica = request.form.get("email_pubblica", "")
    indirizzo_studio = request.form.get("indirizzo_studio", "")
    sito_web = request.form.get("sito_web", "")
    instagram = request.form.get("instagram", "")
    facebook = request.form.get("facebook", "")
    linkedin = request.form.get("linkedin", "")
    orari = request.form.get("orari", "")
    preferenze_contatto = request.form.get("preferenze_contatto", "")

    try:
        c.execute(sql("""
            UPDATE utenti SET
                telefono=?, email_pubblica=?, indirizzo_studio=?, sito_web=?,
                instagram=?, facebook=?, linkedin=?, orari=?, preferenze_contatto=?
            WHERE id=?
        """), (telefono, email_pubblica, indirizzo_studio, sito_web,
              instagram, facebook, linkedin, orari, preferenze_contatto, user_id))
        conn.commit()
        flash("✅ Contatti aggiornati con successo.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"❌ Errore durante il salvataggio dei contatti: {e}", "error")
    finally:
        try:
            conn.close()
        except:
            pass


    return redirect(url_for("dashboard"))

# ---------------------------------------------------------
# ✏️ AGGIORNA DESCRIZIONE UTENTE
# ---------------------------------------------------------
@app.route("/utente/update_descrizione", methods=["POST"])
@login_required
def utente_update_descrizione():
    privacy_debug("update_descrizione chiamata", {
        "user_id": session.get("utente_id"),
        "ha_descrizione": bool(request.form.get("descrizione", "").strip())
    })

    user_id = session.get("utente_id")
    if not user_id:
        flash("Sessione non valida.", "error")
        return redirect(url_for("login"))

    descrizione = request.form.get("descrizione", "").strip()

    conn = get_db_connection()
    c = get_cursor(conn)

    try:
        c.execute(sql("UPDATE utenti SET descrizione = ? WHERE id = ?"), (descrizione, user_id))
        conn.commit()
        flash("✅ Descrizione aggiornata con successo.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"❌ Errore durante il salvataggio della descrizione: {e}", "error")
    finally:
        try:
            conn.close()
        except:
            pass


    return redirect(url_for("dashboard") + "#tab-descrizione")


# --- Aggiorna GALLERIA (tab) ---
@app.route('/utente/update_galleria', methods=['POST'])
@login_required
def utente_update_galleria():
    privacy_debug("update_galleria chiamata", {
        "user_id": session.get("utente_id"),
        "campi_presenti": list(request.form.keys()),
        "numero_file": len(request.files)
    })

    MAX_FOTO_GALLERIA = 4

    conn = get_db_connection()
    c = get_cursor(conn)

    # --- Recupera galleria esistente ---
    c.execute(sql("SELECT foto_galleria FROM utenti WHERE id = ?"), (g.utente['id'],))
    row = c.fetchone()

    correnti = []
    if row and row['foto_galleria']:
        try:
            correnti = json.loads(row['foto_galleria'])
        except Exception:
            correnti = [p for p in row['foto_galleria'].split(',') if p.strip()]

    # normalizza eventuali vuoti / duplicati
    correnti = [p.strip() for p in correnti if p and str(p).strip()]
    correnti = list(dict.fromkeys(correnti))

    # --- Rimuovi selezionate ---
    to_remove = request.form.getlist("remove")
    to_remove = [p.strip() for p in to_remove if p and str(p).strip()]
    correnti = [p for p in correnti if p not in to_remove]

    # --- Calcola slot disponibili DOPO le rimozioni ---
    slot_disponibili = max(0, MAX_FOTO_GALLERIA - len(correnti))

    # --- Aggiungi nuove immagini SOLO entro il limite ---
    uploaded_files = request.files.getlist("foto_galleria")

    # Salviamo la galleria nello stesso spazio servito da /static/uploads/...
    # perché nel template le immagini vengono lette con url_for('static', filename=...)
    upload_dir = os.path.join(app.static_folder, "uploads", "profili", "galleria")
    os.makedirs(upload_dir, exist_ok=True)

    file_validi = []
    for file in uploaded_files:
        if not file or not file.filename:
            continue

        estensione = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if estensione not in {"jpg", "jpeg", "png", "gif", "webp"}:
            continue

        file_validi.append((file, estensione))

    # se stanno provando a caricare più del consentito, tieni solo i primi disponibili
    file_da_salvare = file_validi[:slot_disponibili]
    file_scartati = len(file_validi) - len(file_da_salvare)

    for file, estensione in file_da_salvare:
        nome_file = f"u{g.utente['id']}_{uuid.uuid4().hex}.{estensione}"
        percorso = os.path.join(upload_dir, nome_file)
        file.save(percorso)
        correnti.append(f"uploads/profili/galleria/{nome_file}")

    # sicurezza finale assoluta lato server
    correnti = correnti[:MAX_FOTO_GALLERIA]

    # --- Salva nel DB ---
    c.execute(
        sql("UPDATE utenti SET foto_galleria = ? WHERE id = ?"),
        (json.dumps(correnti), g.utente['id'])
    )
    conn.commit()

    # --- Messaggi utente ---
    if file_scartati > 0:
        flash(
            f"⚠️ Hai raggiunto il limite massimo di {MAX_FOTO_GALLERIA} foto. "
            f"Alcuni file non sono stati caricati.",
            "warning"
        )
    elif to_remove and not file_da_salvare:
        flash("✅ Foto rimosse correttamente.", "success")
    elif file_da_salvare or to_remove:
        flash("✅ Galleria aggiornata correttamente 📸", "success")
    else:
        flash("ℹ️ Nessuna modifica effettuata.", "info")

    return redirect(url_for("dashboard") + "#tab-foto")

@app.route('/annuncio/<int:id>/elimina', methods=["POST"])
@login_required
def elimina_annuncio(id):
    conn = get_db_connection()
    cur = get_cursor(conn)

    cur.execute(sql("SELECT * FROM annunci WHERE id = ?"), (id,))
    annuncio = cur.fetchone()

    if not annuncio or annuncio["utente_id"] != g.utente["id"]:
        flash("Non puoi eliminare questo annuncio.", "error")
        return redirect(url_for("dashboard"))

    # Soft delete: non cancelliamo fisicamente l'annuncio perché può avere acquisti collegati.
    # Lo marchiamo come eliminato per preservare storico pagamenti e vincoli FK.
    cur.execute(
        sql("UPDATE annunci SET stato = ? WHERE id = ? AND utente_id = ?"),
        ("eliminato", id, g.utente["id"])
    )

    conn.commit()

    # 🔁 Se l'annuncio era 'in_attesa', i counters admin vanno aggiornati
    invalidate_admin_counters()

    flash("Annuncio eliminato con successo.", "success")
    return redirect(url_for("dashboard"))

# --- Foto Profilo ---

import werkzeug
from werkzeug.utils import secure_filename

# --- Configurazione cartelle upload su disco persistente Render ---
BASE_UPLOAD_DIR = os.getenv("UPLOAD_BASE_DIR", "/uploads")

UPLOAD_FOLDER = os.path.join(BASE_UPLOAD_DIR, "profili")
UPLOAD_COPERTINE_FOLDER = os.path.join(UPLOAD_FOLDER, "copertine")
UPLOAD_GALLERIA_FOLDER = os.path.join(UPLOAD_FOLDER, "galleria")

# Le cartelle upload vanno create solo sul servizio web.
# Il servizio realtime/chat non deve provare a scrivere su /uploads.
if APP_RUNTIME_ROLE != "realtime":
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(UPLOAD_COPERTINE_FOLDER, exist_ok=True)
    os.makedirs(UPLOAD_GALLERIA_FOLDER, exist_ok=True)

app.config["UPLOAD_BASE_DIR"] = BASE_UPLOAD_DIR
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["UPLOAD_COPERTINE_FOLDER"] = UPLOAD_COPERTINE_FOLDER
app.config["UPLOAD_GALLERIA_FOLDER"] = UPLOAD_GALLERIA_FOLDER

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# --- Upload foto profilo ---
@app.route('/utente/foto', methods=['GET', 'POST'])
@login_required
def upload_foto():
    if request.method == 'POST':
        if 'foto' not in request.files:
            flash("Nessun file selezionato.")
            return redirect(request.url)

        file = request.files['foto']
        if file.filename == '':
            flash("Seleziona un file valido.")
            return redirect(request.url)

        if not (file and allowed_file(file.filename)):
            flash("Formato file non valido. Usa JPG, PNG o GIF.")
            return redirect(request.url)

        user_id = g.utente['id']
        conn = None
        cur = None

        try:
            conn = get_db_connection()
            cur = get_cursor(conn)

            upload_dir = os.path.join(app.static_folder, "uploads", "profili")
            os.makedirs(upload_dir, exist_ok=True)

            # elimina eventuali vecchie foto profilo dello stesso utente, qualunque estensione abbiano
            possibili_vecchie = [
                os.path.join(upload_dir, f"utente_{user_id}.jpg"),
                os.path.join(upload_dir, f"utente_{user_id}.jpeg"),
                os.path.join(upload_dir, f"utente_{user_id}.png"),
                os.path.join(upload_dir, f"utente_{user_id}.gif"),
                os.path.join(upload_dir, f"utente_{user_id}.webp"),
            ]

            for vecchio_file in possibili_vecchie:
                if os.path.exists(vecchio_file):
                    try:
                        os.remove(vecchio_file)
                    except Exception as e:
                        print(f"⚠️ Errore eliminando vecchia foto profilo {vecchio_file}: {e}", flush=True)

            # nome fisso coerente con ciò che la dashboard sta già chiedendo nei log
            filename = f"utente_{user_id}.png"
            file_path = os.path.join(upload_dir, filename)
            file.save(file_path)

            percorso_db = f"uploads/profili/{filename}"

            cur.execute(
                sql("UPDATE utenti SET foto_profilo = ? WHERE id = ?"),
                (percorso_db, user_id)
            )
            conn.commit()

            flash("Foto profilo aggiornata con successo.")
            return redirect(url_for('dashboard'))

        except Exception as e:
            if conn:
                try:
                    conn.rollback()
                except Exception:
                    pass

            print(f"❌ Errore upload_foto user={user_id}: {e}", flush=True)
            traceback.print_exc()
            flash("Errore durante il salvataggio della foto profilo.")
            return redirect(request.url)

        finally:
            try:
                if cur:
                    cur.close()
            except Exception:
                pass

            try:
                if conn:
                    conn.close()
            except Exception:
                pass

    return render_template('upload_foto.html', utente=g.utente)

# --- Upload copertina profilo ---
@app.route('/utente/copertina', methods=['POST'])
@login_required
def upload_copertina():
    if 'file' not in request.files:
        flash("Nessun file selezionato.")
        return redirect(request.referrer or url_for('dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash("Seleziona un file valido.")
        return redirect(request.referrer or url_for('dashboard'))

    allowed_extensions = {'png', 'jpg', 'jpeg', 'webp'}
    estensione = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else None

    if estensione not in allowed_extensions:
        flash("Formato non valido. Usa JPG, PNG o WEBP.")
        return redirect(request.referrer or url_for('dashboard'))

    user_id = g.utente['id']
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = get_cursor(conn)

        upload_dir = os.path.join(app.static_folder, "uploads", "profili", "copertine")
        os.makedirs(upload_dir, exist_ok=True)

        # elimina eventuali vecchie copertine dello stesso utente, qualunque estensione abbiano
        possibili_vecchie = [
            os.path.join(upload_dir, f"copertina_{user_id}.jpg"),
            os.path.join(upload_dir, f"copertina_{user_id}.jpeg"),
            os.path.join(upload_dir, f"copertina_{user_id}.png"),
            os.path.join(upload_dir, f"copertina_{user_id}.webp"),
        ]

        for vecchio_file in possibili_vecchie:
            if os.path.exists(vecchio_file):
                try:
                    os.remove(vecchio_file)
                except Exception as e:
                    print(f"⚠️ Errore eliminando vecchia copertina {vecchio_file}: {e}", flush=True)

        # nome fisso SEMPRE .jpg per evitare mismatch tra DB e file reale
        filename = f"copertina_{user_id}.jpg"
        file_path = os.path.join(upload_dir, filename)
        file.save(file_path)

        percorso_db = f"uploads/profili/copertine/{filename}"

        cur.execute(
            sql("UPDATE utenti SET copertina = ? WHERE id = ?"),
            (percorso_db, user_id)
        )
        conn.commit()

        flash("Copertina aggiornata con successo 📸", "success")
        return redirect(request.referrer or url_for('dashboard'))

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass

        print(f"❌ Errore upload_copertina user={user_id}: {e}", flush=True)
        traceback.print_exc()
        flash("Errore durante il salvataggio della copertina.", "error")
        return redirect(request.referrer or url_for('dashboard'))

    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass

        try:
            if conn:
                conn.close()
        except Exception:
            pass

@app.route('/rimuovi_copertina', methods=['POST'])
@login_required
def rimuovi_copertina():
    user_id = g.utente['id']

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = get_cursor(conn)

        cur.execute(sql("SELECT copertina FROM utenti WHERE id = ?"), (user_id,))
        row = cur.fetchone()

        if row and row['copertina']:
            file_path = os.path.join(app.static_folder, row['copertina'])
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    print(f"⚠️ Errore eliminando la copertina: {e}")

        conn.execute(sql("UPDATE utenti SET copertina = NULL WHERE id = ?"), (user_id,))
        conn.commit()

        g.utente['copertina'] = None

        flash("Copertina rimossa. Tornerà il fondo di default 💙", "info")
        return redirect(request.referrer or url_for('dashboard'))

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        print(f"❌ Errore rimuovi_copertina: {e}")
        traceback.print_exc()
        flash("Errore durante la rimozione della copertina.", "error")
        return redirect(request.referrer or url_for('dashboard'))

    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass

        try:
            if conn:
                conn.close()
        except Exception:
            pass

@app.route('/utente/messaggi')
@login_required
def utente_messaggi():
    """Mostra solo le chat dirette tra utenti"""
    from models import chat_threads
    threads = chat_threads(g.utente['id'])
    return render_template('utente_messaggi.html', threads=threads, utente=g.utente)


# --- Pagina "I miei annunci" ---
@app.route('/utente/annunci')
@login_required
def utente_annunci():
    conn = get_db_connection()
    annunci = conn.execute(sql("""
        SELECT * FROM annunci
        WHERE utente_id = ?
        ORDER BY data_pubblicazione DESC
    """), (g.utente['id'],)).fetchall()

    return render_template('utente_annunci.html', annunci=annunci, utente=g.utente)
# ==========================================================
# RECENSIONI – AREA PERSONALE UTENTE
# ==========================================================
from models import get_recensioni_scritte, elimina_recensione

@app.route("/mie-recensioni")
@login_required
def mie_recensioni():
    """Mostra tutte le recensioni scritte dall’utente loggato"""
    user_id = g.utente["id"]
    recensioni = get_recensioni_scritte(user_id)

    # ✅ Recupera e svuota eventuali flash pendenti
    flashed = get_flashed_messages(with_categories=True)
    session.modified = True

    return render_template(
        "mie_recensioni.html",
        recensioni=recensioni,
        utente=g.utente,
        flashed=flashed
    )

@app.route("/elimina-recensione/<int:id>", methods=["POST"])
@login_required
def elimina_recensione_route(id):
    """Permette all’utente di eliminare una propria recensione"""
    elimina_recensione(id, id_autore=g.utente["id"])
    flash("Recensione eliminata con successo ✅", "success")
    return redirect(url_for("mie_recensioni"))

@app.route("/modifica_recensione", methods=["POST"])
@login_required
def modifica_recensione():
    from models import aggiungi_o_modifica_recensione

    id_recensione = request.form.get("id_recensione")
    testo = request.form.get("testo", "").strip()
    voto = int(request.form.get("voto", 0))

    # Recupera il destinatario della recensione
    conn = get_db_connection()
    c = get_cursor(conn)
    c.execute(sql("""
        SELECT id_destinatario
        FROM recensioni
        WHERE id = ? AND id_autore = ?
    """), (id_recensione, g.utente["id"]))
    row = c.fetchone()


    if not row:
        flash("Recensione non trovata o non autorizzata.", "error")
        return redirect(url_for("mie_recensioni"))

    id_destinatario = list(row.values())[0]

    # 🟦 DECISIONE AUTOMATICA DELLO STATO
    # - Se il testo è vuoto → la recensione è approvata subito
    # - Se c'è testo → deve essere moderata
    if testo == "":
        stato = "approvato"
    else:
        stato = "in_attesa"

    # 🟩 Usa la funzione di utility mantenendo stato coerente
    aggiungi_o_modifica_recensione(
        id_autore=g.utente["id"],
        id_destinatario=id_destinatario,
        voto=voto,
        testo=testo,
        stato=stato
    )

    # 🔔 Aggiorna counter admin solo se serve moderazione
    if stato == "in_attesa":
        invalidate_admin_counters()

        notifica_admin_evento(
            titolo="Recensione modificata in attesa",
            messaggio="Una recensione modificata è in attesa di approvazione.",
            link=url_for("admin_recensioni", stato="in_attesa"),
            push=True
        )

    # 🔵 Messaggio coerente con stato scelto
    if stato == "approvato":
        flash("⭐ Recensione aggiornata con successo.", "success")
    else:
        flash("✏️ Modifica inviata e in attesa di approvazione.", "success")

    return redirect(url_for("mie_recensioni"))

# =========================================================
# 🔔 PUSH NOTIFICATIONS
# =========================================================

from pywebpush import webpush, WebPushException
import requests
import json
import os
from psycopg2.extras import RealDictCursor


def invia_push(user_id, title, body, url=None):
    conn = None
    cur = None

    try:
        security_log(
            "🔔 [invia_push] START",
            {
                "user_id": user_id,
                "title_present": bool(title),
                "runtime_realtime": app.config.get("IS_REALTIME_SERVER"),
                "vapid_private_present": bool(VAPID_PRIVATE_KEY),
                "vapid_claim": VAPID_CLAIM_EMAIL
            }
        )

        if not VAPID_PRIVATE_KEY:
            security_log(
                "❌ [invia_push] VAPID_PRIVATE_KEY mancante",
                {"user_id": user_id},
                production=True
            )
            return

        database_url = os.getenv("DATABASE_URL")
        if not database_url:
            security_log(
                "❌ [invia_push] DATABASE_URL mancante",
                {"user_id": user_id},
                production=True
            )
            return

        conn = psycopg2.connect(
            database_url,
            cursor_factory=RealDictCursor,
            connect_timeout=5
        )
        conn.autocommit = True
        cur = conn.cursor()

        cur.execute("""
            SELECT endpoint, p256dh, auth
            FROM push_subscriptions
            WHERE utente_id = %s
        """, (user_id,))

        subs = cur.fetchall()

        if not subs:
            security_log(
                "⚠️ [invia_push] nessuna subscription trovata",
                {"user_id": user_id}
            )
            return

        security_log(
            "🔔 [invia_push] subscription trovate",
            {
                "user_id": user_id,
                "subscriptions_count": len(subs)
            }
        )

        for sub in subs:
            endpoint = sub["endpoint"]

            try:
                push_url = url or "/notifiche"

                webpush(
                    subscription_info={
                        "endpoint": endpoint,
                        "keys": {
                            "p256dh": sub["p256dh"],
                            "auth": sub["auth"],
                        },
                    },
                    data=json.dumps({
                        "title": title,
                        "body": body,
                        "url": push_url
                    }),
                    vapid_private_key=VAPID_PRIVATE_KEY,
                    vapid_claims={
                        "sub": VAPID_CLAIM_EMAIL
                    },
                    timeout=5
                )

                security_log(
                    "✅ [invia_push] push inviata",
                    {
                        "user_id": user_id,
                        "endpoint": endpoint
                    }
                )

            except WebPushException as e:
                status_code = None
                response_text = None

                try:
                    if e.response is not None:
                        status_code = e.response.status_code
                        response_text = e.response.text
                except Exception:
                    pass

                security_log(
                    "❌ [invia_push] WebPushException",
                    {
                        "user_id": user_id,
                        "status_code": status_code,
                        "endpoint": endpoint,
                        "error_type": type(e).__name__
                    },
                    production=True
                )

                should_delete_subscription = False

                if status_code in (404, 410):
                    should_delete_subscription = True

                if status_code in (400, 403):
                    error_text = (response_text or "").lower()

                    if (
                        "vapid" in error_text
                        or "vapidpkhashmismatch" in error_text
                        or "credentials used to create the subscriptions" in error_text
                    ):
                        should_delete_subscription = True

                if should_delete_subscription:
                    security_log(
                        "🧹 [invia_push] subscription non più valida, eliminazione",
                        {
                            "user_id": user_id,
                            "status_code": status_code,
                            "endpoint": endpoint
                        },
                        production=True
                    )

                    cur.execute("""
                        DELETE FROM push_subscriptions
                        WHERE endpoint = %s
                    """, (endpoint,))

            except requests.exceptions.Timeout:
                security_log(
                    "⚠️ [invia_push] timeout",
                    {
                        "user_id": user_id,
                        "endpoint": endpoint
                    },
                    production=True
                )

            except Exception as e:
                log_exception_safe(
                    "❌ [invia_push] errore generico",
                    e,
                    {
                        "user_id": user_id,
                        "endpoint": endpoint
                    },
                    production=True
                )

        security_log(
            "🔔 [invia_push] END",
            {"user_id": user_id}
        )

    except Exception as e:
        log_exception_safe(
            "❌ [invia_push] errore fatale",
            e,
            {"user_id": user_id},
            production=True
        )

    finally:
        try:
            if cur is not None:
                cur.close()
        except Exception as e:
            log_exception_safe(
                "⚠️ [invia_push] errore chiusura cur",
                e
            )

        try:
            if conn is not None:
                conn.close()
        except Exception as e:
            log_exception_safe(
                "⚠️ [invia_push] errore chiusura conn",
                e
            )

def notifica_admin_evento(titolo, messaggio, link=None, push=True):
    """
    Crea una notifica interna per tutti gli admin attivi
    e, se possibile, invia anche una push.

    Non usa Postmark.
    Non invia email.
    """
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = get_cursor(conn)

        cur.execute(sql("""
            SELECT id
            FROM utenti
            WHERE ruolo = 'admin'
              AND attivo = 1
              AND sospeso = 0
              AND COALESCE(disattivato_admin, 0) = 0
        """))

        admins = cur.fetchall()

        if not admins:
            return

        for admin in admins:
            admin_id = int(admin["id"])

            try:
                _crea_notifica(
                    admin_id,
                    titolo,
                    messaggio,
                    tipo="admin",
                    link=link
                )

                emit_update_notifications(admin_id)

                if push:
                    invia_push(
                        admin_id,
                        titolo,
                        messaggio,
                        url=link or url_for("admin_dashboard")
                    )

            except Exception as e:
                log_exception_safe(
                    "⚠️ Errore notifica_admin_evento per singolo admin",
                    e,
                    {"admin_id": admin_id},
                    production=True
                )

    except Exception as e:
        log_exception_safe(
            "❌ Errore notifica_admin_evento",
            e,
            production=True
        )

    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass

        try:
            if conn:
                conn.close()
        except Exception:
            pass

@app.route("/internal/push/send", methods=["POST"])
def internal_push_send():
    try:
        internal_token = request.headers.get("X-Internal-Token", "")
        expected_token = os.environ.get("INTERNAL_PUSH_TOKEN", "")

        if not expected_token:
            security_log(
                "❌ [internal_push_send] INTERNAL_PUSH_TOKEN mancante",
                production=True
            )
            return jsonify({"ok": False, "error": "internal token not configured"}), 500

        if internal_token != expected_token:
            security_log(
                "❌ [internal_push_send] token interno non valido",
                {
                    "ip": get_client_ip(),
                    "user_agent": request.headers.get("User-Agent", "")
                },
                production=True
            )

            return jsonify({"ok": False, "error": "unauthorized"}), 403

        data = request.get_json(silent=True) or {}

        user_id = data.get("user_id")
        title = data.get("title")
        body = data.get("body")
        push_url = data.get("url") or None
        push_url = data.get("url")

        if not user_id or not title or body is None:

            security_log(
                "❌ [internal_push_send] payload non valido",
                {
                    "keys": list(data.keys()) if isinstance(data, dict) else [],
                    "user_id_present": bool(data.get("user_id")) if isinstance(data, dict) else False,
                    "title_present": bool(data.get("title")) if isinstance(data, dict) else False,
                    "body_present": data.get("body") is not None if isinstance(data, dict) else False
                },
                production=True
            )
            return jsonify({"ok": False, "error": "invalid payload"}), 400

        security_log(
            "🔔 [internal_push_send] richiesta ricevuta",
            {
                "user_id": user_id,
                "title_present": bool(title)
            }
        )

        invia_push(
            int(user_id),
            str(title),
            str(body),
            url=push_url
        )

        security_log(
            "✅ [internal_push_send] invia_push completata",
            {"user_id": user_id}
        )

        return jsonify({"ok": True})

    except Exception as e:
        log_exception_safe(
            "❌ [internal_push_send] errore fatale",
            e,
            production=True
        )
        return jsonify({"ok": False, "error": "Errore interno invio push."}), 500

# =========================================================
# 🔔 PUSH SUBSCRIBE PUBBLICO
# Associa sempre la subscription all'utente attualmente loggato
# =========================================================
@app.route("/push/debug", methods=["POST"])
@login_required
def push_debug():
    try:
        data = request.get_json(silent=True) or {}

        privacy_debug("push_debug", {
            "user_id": session.get("utente_id"),
            "g_utente_id": g.utente["id"] if getattr(g, "utente", None) else None,
            "step": data.get("step"),
            "ok": data.get("ok"),
            "details": data.get("details"),
            "user_agent": request.headers.get("User-Agent", "")
        })

        return jsonify({"ok": True})

    except Exception as e:
        log_exception_safe(
            "❌ [push_debug] errore",
            e
        )
        return jsonify({
            "ok": False,
            "error": "Errore debug push."
        }), 500

@app.route("/push/subscribe", methods=["POST"])
@login_required
def push_subscribe():
    conn = None
    cur = None

    try:
        user_id = g.utente["id"]
        data = request.get_json(silent=True) or {}

        endpoint = data.get("endpoint")
        keys = data.get("keys") or {}
        p256dh = keys.get("p256dh")
        auth = keys.get("auth")

        if not endpoint or not p256dh or not auth:
            security_log(
                "❌ [push_subscribe] payload non valido",
                {
                    "user_id": user_id,
                    "has_endpoint": bool(endpoint),
                    "has_p256dh": bool(p256dh),
                    "has_auth": bool(auth)
                }
            )

            return jsonify({
                "ok": False,
                "error": "payload subscription non valido"
            }), 400

        conn = get_db_connection()
        cur = conn.cursor()

        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                INSERT INTO push_subscriptions (
                    utente_id,
                    endpoint,
                    p256dh,
                    auth,
                    user_agent,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (endpoint)
                DO UPDATE SET
                    utente_id = EXCLUDED.utente_id,
                    p256dh = EXCLUDED.p256dh,
                    auth = EXCLUDED.auth,
                    user_agent = EXCLUDED.user_agent,
                    updated_at = CURRENT_TIMESTAMP
            """, (
                user_id,
                endpoint,
                p256dh,
                auth,
                request.headers.get("User-Agent", "")
            ))

        else:
            cur.execute("""
                INSERT INTO push_subscriptions (
                    utente_id,
                    endpoint,
                    p256dh,
                    auth,
                    user_agent,
                    created_at,
                    updated_at
                )
                VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT(endpoint)
                DO UPDATE SET
                    utente_id = excluded.utente_id,
                    p256dh = excluded.p256dh,
                    auth = excluded.auth,
                    user_agent = excluded.user_agent,
                    updated_at = CURRENT_TIMESTAMP
            """, (
                user_id,
                endpoint,
                p256dh,
                auth,
                request.headers.get("User-Agent", "")
            ))

        conn.commit()

        security_log(
            "✅ [push_subscribe] subscription sincronizzata",
            {
                "user_id": user_id,
                "endpoint": endpoint
            }
        )

        return jsonify({
            "ok": True,
            "user_id": user_id
        })

    except Exception as e:
        log_exception_safe(
            "❌ [push_subscribe] errore",
            e,
            {
                "user_id": g.utente["id"] if getattr(g, "utente", None) else None
            },
            production=True
        )
        return jsonify({
            "ok": False,
            "error": "Errore durante la registrazione delle notifiche push."
        }), 500

    finally:
        try:
            if cur is not None:
                cur.close()
        except Exception:
            pass

@app.route("/push/unsubscribe", methods=["POST"])
@login_required
def push_unsubscribe():
    verify_csrf()

    conn = None
    cur = None

    try:
        user_id = g.utente["id"]
        data = request.get_json(silent=True) or {}
        endpoint = data.get("endpoint")

        if not endpoint:
            return jsonify({
                "ok": False,
                "error": "endpoint mancante"
            }), 400

        conn = get_db_connection()
        cur = get_cursor(conn)

        cur.execute(sql("""
            DELETE FROM push_subscriptions
            WHERE utente_id = ?
              AND endpoint = ?
        """), (
            user_id,
            endpoint
        ))

        conn.commit()

        security_log(
            "🔕 [push_unsubscribe] subscription rimossa",
            {
                "user_id": user_id,
                "endpoint": endpoint
            }
        )

        return jsonify({
            "ok": True
        })

    except Exception as e:
        try:
            if conn:
                conn.rollback()
        except Exception:
            pass

        log_exception_safe(
            "❌ [push_unsubscribe] errore",
            e,
            {
                "user_id": g.utente["id"] if getattr(g, "utente", None) else None
            },
            production=True
        )

        return jsonify({
            "ok": False,
            "error": "Errore durante la rimozione della subscription push."
        }), 500

    finally:
        try:
            if cur:
                cur.close()
        except Exception:
            pass

        try:
            if conn:
                conn.close()
        except Exception:
            pass

@app.route("/service-worker.js")
def service_worker():
    return app.send_static_file("service-worker.js")

# ==========================================================
# NOTIFICHE - ROTTE (AGGIUNTA)
# ==========================================================
@app.route('/notifiche')
@login_required
def notifiche():
    notifiche = get_notifiche_utente(g.utente['id'])
    return render_template('notifiche.html', notifiche=notifiche)


@app.route('/recensioni-ricevute', methods=["GET", "POST"])
@login_required
def mie_recensioni_ricevute():
    user_id = session.get('utente_id')
    if not user_id:
        flash("Devi accedere per visualizzare le recensioni.")
        return redirect(url_for("login"))

    # --- POST: gestione invio/modifica/eliminazione risposta alla recensione ---
    if request.method == "POST":
        azione = request.form.get("azione")
        if azione == "rispondi":
            id_recensione = request.form.get("id_recensione", type=int)
            testo = (request.form.get("testo_risposta") or "").strip()
            if id_recensione and testo:
                try:
                    aggiungi_o_modifica_risposta(id_recensione, user_id, testo)
                    flash("✅ Risposta inviata! Sarà visibile dopo approvazione.", "success")
                    invalidate_admin_counters()
                except Exception as e:
                    flash(f"❌ Errore durante il salvataggio della risposta: {e}", "danger")
            else:
                flash("Testo della risposta mancante o ID non valido.", "warning")

        elif azione == "modifica_risposta":
            id_risposta = request.form.get("id_risposta", type=int)
            testo = (request.form.get("testo_risposta") or "").strip()
            if id_risposta and testo:
                try:
                    aggiungi_o_modifica_risposta(id_risposta=id_risposta, testo=testo)
                    flash("✏️ Risposta modificata con successo (in attesa di approvazione).", "info")
                    invalidate_admin_counters()
                except Exception as e:
                    flash(f"Errore durante la modifica: {e}", "danger")
            else:
                flash("Testo della risposta mancante o ID non valido.", "warning")

        return redirect(url_for("mie_recensioni_ricevute"))

    # --- GET: mostra statistiche + elenco recensioni ricevute ---
    recensioni = get_recensioni_utente(user_id)
    media, totale = calcola_media_recensioni(user_id)

    return render_template(
        "mie_recensioni_ricevute.html",
        recensioni=recensioni,
        media=media,
        totale=totale
    )

# 🔹 Recensioni pubbliche visibili nel profilo utente
from flask import get_flashed_messages

@app.route("/recensioni_utente/<int:user_id>", methods=["GET", "POST"])
def recensioni_utente(user_id):
    """Mostra o gestisce le recensioni ricevute da un utente (profilo pubblico)."""
    if "utente_id" not in session:
        flash("Devi accedere per vedere i dettagli del profilo.", "warning")
        return redirect(url_for("login", next=request.path))

    flashed = get_flashed_messages(with_categories=True)
    session.modified = True

    conn = get_db_connection()

    c = get_cursor(conn)

    # ✅ verifica utente
    c.execute(sql("""
        SELECT * FROM utenti
        WHERE id = ?
          AND sospeso = 0
          AND (disattivato_admin IS NULL OR disattivato_admin = 0)
          AND attivo = 1
    """), (user_id,))
    user = c.fetchone()
    if not user:

        flash("Utente non trovato.", "error")
        return redirect(url_for("cerca"))

    # ✅ POST: nuova recensione
    if request.method == "POST":
        if "utente_id" not in session:
            flash("Devi accedere per lasciare una recensione.", "warning")

            return redirect(url_for("login"))

        # 🔒 Blocca recensioni se l’utente non ha caricato una foto profilo
        autore_id = session.get("utente_id")
        c.execute(sql("SELECT foto_profilo FROM utenti WHERE id = ?"), (autore_id,))
        fp_row = c.fetchone()

        if not fp_row or not fp_row["foto_profilo"]:
            flash("Per lasciare una recensione devi prima caricare una foto profilo.", "warning")

            return redirect(url_for("dashboard"))  # modifica se la tua pagina profilo ha un nome diverso


        id_autore = session.get("utente_id")
        voto = int(request.form.get("voto", 0))
        testo = request.form.get("testo", "").strip()

        esistente = get_recensione_autore_vs_destinatario(id_autore, user_id)
        if esistente:
            flash("⚠️ Hai già lasciato una recensione per questo utente. Puoi solo modificarla.", "warning")

            return redirect(url_for("recensioni_utente", user_id=user_id))

        try:
            stato = "approvato" if not testo else "in_attesa"

            conn.execute(sql(f"""
                INSERT INTO recensioni (id_autore, id_destinatario, voto, testo, stato, data)
                VALUES (?, ?, ?, ?, ?, {now_sql()})
            """), (id_autore, user_id, voto, testo, stato))
            conn.commit()

            # ✅ notifica automatica se solo stelline
            if stato == "approvato":
                # recupera username autore
                c.execute(sql("SELECT username FROM utenti WHERE id = ?"), (id_autore,))
                row = c.fetchone()
                username_autore = row["username"] if row and row["username"] else "utente"

                # ✅ salva notifica DB
                crea_notifica(
                    user_id,
                    f"@{username_autore} ti ha lasciato una valutazione ⭐ {voto}/5",
                    link=url_for("mie_recensioni_ricevute")
                )

                # ✅ invio realtime Socket.IO (badge notifiche)
                emit_update_notifications(user_id)

                flash("⭐ Recensione salvata!", "success")

            else:
                flash("✅ Recensione inviata! Sarà visibile dopo approvazione.", "success")
                invalidate_admin_counters()

                notifica_admin_evento(
                    titolo="Nuova recensione in attesa",
                    messaggio="Una nuova recensione è in attesa di approvazione.",
                    link=url_for("admin_recensioni", stato="in_attesa"),
                    push=True
                )

        except Exception as e:
            flash(f"❌ Errore durante il salvataggio della recensione: {e}", "error")

        finally:
            try:
                conn.close()
            except:
                pass


        return redirect(url_for("recensioni_utente", user_id=user_id))

    # ✅ GET: mostra recensioni
    recensioni = get_recensioni_utente(user_id)
    media, totale = calcola_media_recensioni(user_id)


    return render_template(
        "recensioni_utente.html",
        user=user,
        user_id=user_id,
        recensioni=recensioni,
        media=media,
        totale=totale,
        mia_recensione=None
    )


# ==========================================================
# 5️⃣ AUTENTICAZIONE
# ==========================================================
import json


def get_comune_info(comune_input):
    path = os.path.join(app.static_folder, "data", "comuni.json")

    with open(path, encoding="utf-8") as f:
        comuni = json.load(f)

    comune_input = comune_input.strip().lower()

    for c in comuni:
        if c["comune"].lower() == comune_input:
            return {
                "provincia": c.get("provincia"),
                "regione": c.get("regione")
            }

    return None

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        verify_csrf()

        nome = request.form['nome'].strip()
        cognome = request.form['cognome'].strip()
        citta = request.form['citta'].strip()
        email = request.form['email'].strip().lower()
        username = request.form['username'].strip().upper()
        password = request.form['password']
        conferma_password = request.form['conferma_password']
        accetto = request.form.get('accetto')
        codice_invito = request.form.get('codice_invito', '').strip()

        # ✅ Validazioni base
        if not accetto:
            flash("Devi accettare i termini e la privacy per continuare.")
            return redirect(url_for('register'))

        if password != conferma_password:
            flash("Le password non coincidono.")
            return redirect(url_for('register'))

        # 🔒 Codice beta obbligatorio
        CODICE_BETA = "LOCALCARE2026"   # ← puoi cambiarlo quando vuoi

        if codice_invito != CODICE_BETA:
            flash("Codice invito non valido.")
            return redirect(url_for('register'))

        # 🔎 Provincia dal JSON
        info = get_comune_info(citta)

        if not info:
            flash("Comune non valido. Selezionalo dall'elenco.")
            return redirect(url_for('register'))

        provincia = info["provincia"]
        regione = info["regione"]
        macro_area = provincia

        conn = get_db_connection()
        c = get_cursor(conn)

        # ✅ Controllo duplicati
        c.execute(sql("""
            SELECT * FROM utenti
            WHERE email = ?
               OR UPPER(username) = ?
        """), (email, username))

        existing_user = c.fetchone()
        if existing_user:
            if existing_user['email'] == email:
                flash("Questa email è già registrata.")
            else:
                flash("Questo ID utente è già in uso.")

            return redirect(url_for('register'))

        # 🔐 Sicurezza
        token = str(uuid.uuid4())
        hashed_pw = generate_password_hash(password)

        dek = get_random_bytes(32)
        dek_enc_b64, dek_nonce_b64 = encrypt_with_master(dek)
        salt_b64 = ""

        from nacl.public import PrivateKey
        x25519_priv = PrivateKey.generate()
        x25519_pub = x25519_priv.public_key

        cipher_priv = AES.new(dek, AES.MODE_GCM)
        priv_ct, priv_tag = cipher_priv.encrypt_and_digest(bytes(x25519_priv))

        x25519_priv_enc_b64 = gcm_pack(priv_ct, priv_tag)
        x25519_priv_nonce_b64 = base64.b64encode(cipher_priv.nonce).decode()
        x25519_pub_b64 = base64.b64encode(bytes(x25519_pub)).decode()

        # ✅ INSERT COMPLETO
        c.execute(sql("""
            INSERT INTO utenti (
                nome, cognome, email, username, password,
                citta, provincia, macro_area,
                attivo, token_verifica,
                key_salt, dek_enc, dek_nonce,
                dek_mk_enc, dek_mk_nonce,
                x25519_pub, x25519_priv_enc, x25519_priv_nonce,
                versione_consenso
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """), (
            nome, cognome, email, username, hashed_pw,
            citta, provincia, macro_area,
            0, token,
            salt_b64, dek_enc_b64, dek_nonce_b64,
            None, None,
            x25519_pub_b64, x25519_priv_enc_b64, x25519_priv_nonce_b64,
            "iubenda_2026_v1"
        ))

        conn.commit()


        conn.commit()

        invalidate_admin_counters()

        notifica_admin_evento(
            titolo="Nuovo utente registrato",
            messaggio=f"Nuovo utente registrato: @{username} — {nome} {cognome}",
            link=url_for("admin_utenti"),
            push=True
        )

        # 📧 Email conferma tramite Postmark
        link = build_external_url("conferma_email", token=token)

        email_inviata = _invia_email(
            destinazione=email,
            oggetto="Conferma account MyLocalCare",
            corpo=(
                f"Ciao {nome},\n\n"
                "per confermare il tuo account MyLocalCare apri questo link:\n\n"
                f"{link}\n\n"
                "Se non hai richiesto tu questa registrazione, ignora questa email.\n\n"
                "MyLocalCare"
            )
        )

        if email_inviata:
            flash("Registrazione completata! Controlla la tua email per confermare l'account.", "success")
        else:
            flash(
                "Registrazione completata, ma non siamo riusciti a inviare l'email di conferma. "
                "Contatta l'assistenza o riprova più tardi.",
                "warning"
            )

        return redirect(url_for('login'))

    return render_template('register.html')

@app.route("/termini")
def termini():
    return render_template("termini.html")


@app.route("/privacy")
def privacy():
    return render_template("privacy.html")

@app.route('/conferma/<token>')
def conferma_email(token):
    conn = get_db_connection()
    c = get_cursor(conn)

    try:
        # 1) Prima prova: conferma registrazione
        c.execute(sql("SELECT * FROM utenti WHERE token_verifica = ?"), (token,))
        utente = c.fetchone()

        if utente:
            c.execute(sql("UPDATE utenti SET attivo = 1, token_verifica = NULL WHERE id = ?"), (utente['id'],))
            conn.commit()

            crea_notifica(
                utente['id'],
                "📸 Completa il tuo profilo caricando una foto per essere visibile.",
                link=url_for('upload_foto')
            )

            flash("Email confermata! Ora puoi accedere.")
            return redirect(url_for('login'))

        # 2) Seconda prova: recupero password
        c.execute(sql("""
            SELECT token
            FROM password_reset_tokens
            WHERE token = ?
              AND usato = 0
        """), (token,))
        reset_token = c.fetchone()

        if reset_token:
            return redirect(url_for('reset_password', token=token))

        flash("Token non valido o già usato.")
        return redirect(url_for('login'))

    finally:
        try:
            c.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
# ==========================================================
# 🔐 LOGIN UTENTE + DECIFRATURA CHIAVI PERSONALI
# (VERSIONE PROFILING — NON MODIFICA LOGICA)
# ==========================================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        verify_csrf()

        email = request.form['email'].strip().lower()
        password = request.form['password']

        # ===============================
        # DB SELECT UTENTE
        # ===============================
        conn = get_db_connection()
        c = get_cursor(conn)

        c.execute(sql("SELECT * FROM utenti WHERE email = ?"), (email,))
        utente = c.fetchone()

        # 1️⃣ Utente inesistente
        if not utente:
            flash("Email o password non validi.", "error")
            return redirect(url_for('login'))

        # ---------------------------------------------------------
        # 🔐 BLOCCO LOCK
        # ---------------------------------------------------------
        from datetime import datetime, timezone, timedelta

        def parse_iso(dt):
            if not dt:
                return None
            try:
                return datetime.fromisoformat(dt)
            except Exception:
                return None

        lock_until = parse_iso(utente["lock_until"]) if "lock_until" in utente.keys() else None
        now = datetime.now(timezone.utc)

        if lock_until and lock_until > now:
            minuti = int((lock_until - now).total_seconds() // 60) + 1
            flash(f"Troppi tentativi falliti. Riprova tra {minuti} minuti.", "error")
            return redirect(url_for('login'))

        # Email non confermata
        if int(utente['attivo']) != 1:
            flash("Devi confermare l'email prima di accedere.", "warning")
            return redirect(url_for('login'))

        # ---------------------------------------------------------
        # PASSWORD CHECK
        # ---------------------------------------------------------
        if not check_password_hash(utente['password'], password):
            failed = (utente["failed_logins"] or 0) + 1
            lock_until = None

            if failed >= 5:
                lock_until = (datetime.now(timezone.utc) + timedelta(minutes=15)).isoformat()

            conn = get_db_connection()

            conn.execute(
                "UPDATE utenti SET failed_logins = ?, lock_until = ? WHERE id = ?",
                (failed, lock_until, utente["id"])
            )
            conn.commit()

            flash("Email o password non validi.", "error")
            return redirect(url_for('login'))

        # ---------------------------------------------------------
        # STATUS CHECK
        # ---------------------------------------------------------
        disattivato_admin = utente['disattivato_admin'] if 'disattivato_admin' in utente.keys() else 0
        if disattivato_admin == 1:
            flash("Account disattivato.", "error")
            return redirect(url_for('login'))

        sospeso = utente['sospeso'] if 'sospeso' in utente.keys() else 0
        if sospeso == 1:
            session.clear()
            session['utente_id'] = utente['id']
            session['sospeso'] = True
            return redirect(url_for('riattivazione_account'))

        # ---------------------------------------------------------
        # MACRO AREA
        # ---------------------------------------------------------
        citta = utente["citta"]
        provincia = get_provincia_from_comune(citta)

        if provincia:
            session["macro_area"] = provincia
        else:
            session["macro_area"] = "Italia"

        # ---------------------------------------------------------
        # DECRYPT MASTER (DEK)
        # ---------------------------------------------------------
        import base64

        try:
            dek = decrypt_with_master(utente['dek_enc'], utente['dek_nonce'])
            session['dek_b64'] = base64.b64encode(dek).decode()
        except Exception:
            flash("Errore chiave personale.", "error")
            return redirect(url_for("login"))

        # ---------------------------------------------------------
        # DECRYPT X25519
        # ---------------------------------------------------------
        try:
            x_priv_nonce = base64.b64decode(utente["x25519_priv_nonce"])
            x_priv_ct, x_priv_tag = gcm_unpack(utente["x25519_priv_enc"])

            cipher_xpriv = AES.new(dek, AES.MODE_GCM, nonce=x_priv_nonce)
            x_priv_bytes = cipher_xpriv.decrypt_and_verify(x_priv_ct, x_priv_tag)

            session["x25519_priv_b64"] = base64.b64encode(x_priv_bytes).decode()
            session["x25519_pub_b64"] = utente["x25519_pub"]

        except Exception:
            pass

        # ---------------------------------------------------------
        # RESET FAIL LOGIN
        # ---------------------------------------------------------
        conn = get_db_connection()

        conn.execute(
            "UPDATE utenti SET failed_logins = 0, lock_until = NULL WHERE id = ?",
            (utente["id"],)
        )
        conn.commit()

        # ---------------------------------------------------------
        # ADMIN SESSION
        # ---------------------------------------------------------
        import secrets

        if utente["ruolo"] == "admin":
            session_token = secrets.token_hex(32)
            expiry = (datetime.now(timezone.utc) + timedelta(hours=12)).isoformat()

            admin_security_version = int(
                utente["admin_security_version"]
                if "admin_security_version" in utente.keys() and utente["admin_security_version"] is not None
                else 0
            )

            conn = get_db_connection()

            conn.execute(sql("""
                UPDATE utenti
                SET admin_session_token = ?, admin_session_expiry = ?
                WHERE id = ?
            """), (session_token, expiry, utente["id"]))
            conn.commit()

            session["admin_session_token"] = session_token
            session["admin_security_version"] = admin_security_version

        browser_fingerprint = request.headers.get("User-Agent", "unknown")

        conn = get_db_connection()

        conn.execute(sql("""
            UPDATE utenti
            SET admin_browser_fingerprint = ?
            WHERE id = ?
        """), (browser_fingerprint, utente["id"]))
        conn.commit()

        session["admin_browser_fingerprint"] = browser_fingerprint

        # ---------------------------------------------------------
        # SESSION BASE
        # ---------------------------------------------------------
        session.permanent = True

        session['utente_id'] = utente['id']
        session['utente_username'] = utente['username']

        ensure_x25519_keys(utente['id'])

        from flask import get_flashed_messages
        get_flashed_messages()

        flash("Accesso effettuato con successo.", "success")

        next_page = request.args.get('next')
        if next_page:
            return redirect(next_page)

        return redirect(url_for('dashboard'))

    return render_template('login.html')

@app.route('/password_dimenticata', methods=['GET', 'POST'])
def password_dimenticata():
    # GET: mostra solo la pagina, senza inviare nulla
    if request.method == 'GET':
        return render_template('password_dimenticata.html')

    # POST: invio richiesta recupero accesso
    verify_csrf()

    email = request.form.get('email', '').strip().lower()

    if not email:
        flash("Inserisci un indirizzo email.", "error")
        return redirect(url_for('password_dimenticata'))

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(sql("SELECT * FROM utenti WHERE email = ?"), (email,))
        utente = cur.fetchone()

        # Non riveliamo se l'email esiste o no
        if not utente:
            flash(
                "Se l'indirizzo è registrato, riceverai un link per completare la procedura.",
                "info"
            )
            return redirect(url_for('login'))

        # Genera token semplice stile conferma account
        token = str(uuid.uuid4())

        reset_url = (
            app.config.get('APP_BASE_URL', 'https://www.mylocalcare.it').rstrip('/')
            + f"/conferma/{token}"
        )

        # Invalida eventuali token precedenti
        cur.execute(sql("""
            UPDATE password_reset_tokens
            SET usato = 1
            WHERE utente_id = ?
              AND usato = 0
        """), (utente['id'],))

        # Salva nuovo token
        cur.execute(sql(f"""
            INSERT INTO password_reset_tokens (utente_id, token, scadenza, usato)
            VALUES (?, ?, {epoch_now_sql()} + 3600, 0)
        """), (utente['id'], token))

        conn.commit()

        # Invia email tramite MailAPI Aruba
        email_inviata = _invia_email(
            destinazione=email,
            oggetto="Conferma account MyLocalCare",
            corpo=(
                f"Ciao {utente['nome']},\n\n"
                "per confermare il tuo account MyLocalCare apri questo link:\n\n"
                f"{reset_url}\n\n"
                "Se non hai richiesto tu questa registrazione, ignora questa email.\n\n"
                "MyLocalCare"
            )
        )

        if not email_inviata:
            flash("Errore nell'invio dell'email. Riprova più tardi.", "error")
            return redirect(url_for('password_dimenticata'))

        flash(
            "Se l'indirizzo è registrato, riceverai un link per completare la procedura.",
            "success"
        )
        return redirect(url_for('login'))

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass

        print("Errore recupero accesso:", repr(e), flush=True)
        flash("Si è verificato un errore. Riprova più tardi.", "error")
        return redirect(url_for('password_dimenticata'))

    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

@app.route('/account/<token>', methods=['GET', 'POST'])
@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    conn = get_db_connection()
    cur = get_cursor(conn)

    cur.execute(sql(f"""
        SELECT *
        FROM password_reset_tokens
        WHERE token = ?
          AND usato = 0
          AND scadenza >= {epoch_now_sql()}
    """), (token,))
    token_row = cur.fetchone()

    if not token_row:
        flash("Questo link è scaduto, non valido o già utilizzato.", "error")
        return redirect(url_for('password_dimenticata'))

    if request.method == 'POST':
        verify_csrf()

        password = request.form.get('password', '')
        conferma = request.form.get('conferma_password', '')

        if not password or not conferma:
            flash("Compila entrambi i campi password.", "error")
            return redirect(url_for('reset_password', token=token))

        if password != conferma:
            flash("Le password non coincidono.", "error")
            return redirect(url_for('reset_password', token=token))

        if len(password) < 8:
            flash("La password deve avere almeno 8 caratteri.", "error")
            return redirect(url_for('reset_password', token=token))

        # ✅ Recupera utente
        cur.execute(sql("SELECT * FROM utenti WHERE id = ?"), (token_row['utente_id'],))
        utente = cur.fetchone()

        if not utente:

            flash("Errore interno. Contatta il supporto.", "error")
            return redirect(url_for('login'))

        # ✅ Aggiorna SOLO la password (le chiavi restano invariate)
        pwd_hash = generate_password_hash(password)
        cur.execute(sql("UPDATE utenti SET password = ? WHERE email = ?"), (pwd_hash, email))

        # ✅ Marca token come usato
        cur.execute(sql("UPDATE password_reset_tokens SET usato = 1 WHERE token = ?"), (token,))
        conn.commit()


        # ✅ Login automatico
        session.clear()
        session['utente_id'] = utente['id']
        session['utente_username'] = utente['username']

                # ✅ Ripristina chiavi crittografiche in sessione come nel login
        try:
            import base64
            from Crypto.Cipher import AES
            from app import decrypt_with_master, gcm_unpack

            # 🔐 Decifra DEK con MASTER_SECRET
            dek = decrypt_with_master(utente['dek_enc'], utente['dek_nonce'])
            session['dek_b64'] = base64.b64encode(dek).decode()

            # 🔐 Decifra chiave privata X25519
            x_nonce = base64.b64decode(utente["x25519_priv_nonce"])
            x_ct, x_tag = gcm_unpack(utente["x25519_priv_enc"])

            cipher_x = AES.new(dek, AES.MODE_GCM, nonce=x_nonce)
            x_priv_bytes = cipher_x.decrypt_and_verify(x_ct, x_tag)

            session["x25519_priv_b64"] = base64.b64encode(x_priv_bytes).decode()
            session["x25519_pub_b64"] = utente["x25519_pub"]

            # ✅ Garantisce presenza chiavi (crea se mancanti)
            from app import ensure_x25519_keys
            ensure_x25519_keys(utente['id'])

        except Exception as e:
            print("Errore ripristino chiavi post-reset:", e)


        flash("Password aggiornata con successo!", "success")
        return redirect(url_for('dashboard'))

    return render_template('reset_password.html', token=token)

@app.route("/debug_chiavi_x25519")
@login_required
def debug_chiavi_x25519():
    if APP_RUNTIME_ROLE != "web" or os.getenv("APP_ENV", "production").lower() != "local":
        abort(404)

    return jsonify({
        "utente": g.utente["username"],
        "ha_dek_b64": bool(session.get("dek_b64")),
        "ha_pubblica_X25519": bool(session.get("x25519_pub_b64")),
        "ha_privata_X25519": bool(session.get("x25519_priv_b64"))
    })

@app.route('/logout')
def logout():
    # 🔥 Caso: logout dopo sospensione → messaggio speciale
    if session.get("sospensione_logout"):
        session.clear()
        flash(
            "Il tuo account è stato sospeso. Per riattivarlo effettua l’accesso e ti verrà mostrata la pagina di riattivazione.",
            "warning"
        )
        return redirect(url_for('login'))

    user_id = session.get("utente_id")

    # 🔵 Caso normale
    # Prima di cancellare la sessione, rimuoviamo anche le push subscription
    # collegate all'utente attualmente loggato.
    if user_id:
        conn = None
        cur = None

        try:
            conn = get_db_connection()
            cur = get_cursor(conn)

            # 🧹 Reset token e fingerprint admin nel DB
            cur.execute(sql("""
                UPDATE utenti
                SET admin_session_token = NULL,
                    admin_session_expiry = NULL,
                    admin_browser_fingerprint = NULL
                WHERE id = ?
            """), (user_id,))

            # 🔕 Logout = questo utente non deve più ricevere push
            # su subscription associate a questa sessione/account.
            cur.execute(sql("""
                DELETE FROM push_subscriptions
                WHERE utente_id = ?
            """), (user_id,))

            conn.commit()

            security_log(
                "🔕 [logout] push subscriptions rimosse",
                {
                    "user_id": user_id
                },
                production=True
            )

        except Exception as e:
            try:
                if conn:
                    conn.rollback()
            except Exception:
                pass

            log_exception_safe(
                "⚠️ [logout] errore pulizia sessione/push",
                e,
                {
                    "user_id": user_id
                },
                production=True
            )

        finally:
            try:
                if cur:
                    cur.close()
            except Exception:
                pass

            try:
                if conn:
                    conn.close()
            except Exception:
                pass

    session.pop("dek_b64", None)
    session.pop("id_priv_b64", None)
    session.pop("id_pub_b64", None)
    clear_admin_stepup()
    session.clear()

    flash('Sei uscito correttamente.', 'info')
    return redirect(url_for('login'))

# ==========================================================
# 6️⃣ ROTTE PUBBLICHE
# ==========================================================
@app.route('/', endpoint='home')
def landing():
    utente_id = session.get('utente_id')

    # 1️⃣ Se loggato → recupero macro_area dal DB
    if utente_id:
        conn = get_db_connection()
        cur = get_cursor(conn)
        cur.execute(sql("SELECT macro_area FROM utenti WHERE id = ?"), (utente_id,))
        row = cur.fetchone()


        if row and row["macro_area"]:
            session['macro_area'] = row["macro_area"]
            return redirect(url_for('home_v2'))

    # 2️⃣ Se non loggato ma ha già scelto macro_area
    if session.get('macro_area'):
        return redirect(url_for('home_v2'))

    # 3️⃣ Altrimenti mostra landing
    return render_template('landing.html')

@app.route('/home')
def home_v2():
    macro = session.get('macro_area') or "Milano"
    brand_name = f"{macro}Care"
    return render_template('home.html', brand_name=brand_name)

def get_provincia_from_comune(comune_input):
    path = os.path.join(app.static_folder, "data", "comuni.json")

    try:
        with open(path, encoding="utf-8") as f:
            comuni = json.load(f)
    except Exception:
        return None

    comune_norm = comune_input.strip().lower()

    for c in comuni:
        if c.get("comune", "").strip().lower() == comune_norm:
            # ⬅️ QUI
            return c.get("provincia_nome") or c.get("provincia")

    return None

@app.route('/set-macro-area', methods=['POST'])
def set_macro_area():
    comune = (request.form.get('macro_comune') or "").strip()

    if not comune:
        flash("Seleziona un comune dall’elenco per continuare.")
        return redirect(url_for('home'))  # ✅ NON landing

    provincia = get_provincia_from_comune(comune)
    if not provincia:
        flash("Comune non riconosciuto. Selezionalo dall’elenco.")
        return redirect(url_for('home'))  # ✅ NON landing

    # ✅ qui salvi sempre la PROVINCIA come macro_area
    session['macro_area'] = provincia

    # opzionale: se vuoi salvarla davvero
    regione = (request.form.get('macro_regione') or "").strip()
    if regione:
        session['macro_regione'] = regione

    return redirect(url_for('home_v2'))

from flask import request, render_template, session
import sqlite3, json, time

def get_filtri_categoria_da_db():
    conn = get_db_connection()
    cur = get_cursor(conn)

    cur.execute(sql("""
        SELECT categoria, filtro
        FROM filtri_categoria
        WHERE attivo = 1
        ORDER BY categoria ASC, ordine ASC, filtro ASC
    """))

    righe = cur.fetchall()
    conn.close()

    filtri = {}

    for r in righe:
        row = dict(r)
        categoria = row["categoria"]
        filtro = row["filtro"]

        if categoria not in filtri:
            filtri[categoria] = []

        filtri[categoria].append(filtro)

    return filtri

@app.route("/api/filtri-categoria")
@login_required
def api_filtri_categoria():
    return jsonify(get_filtri_categoria_da_db())

@app.route("/cerca")
def cerca():
    # 🔒 BLOCCO BETA — accesso solo utenti registrati
    if not session.get("utente_id"):
        flash(
            "MyLocalCare è attualmente in fase BetaTest privata. "
            "Registrati per ottenere l’accesso anticipato.",
            "warning"
        )
        return redirect(url_for("home"))

    raw_cat = request.args.get("categoria", "").strip()
    cat_slug = to_slug(raw_cat)

    categoria_label = raw_cat

    # lookup filtri: prova prima il valore originale, poi lo slug,
    # poi alcune equivalenze note usate in /cerca
    json_key_aliases = [
        raw_cat,
        cat_slug,
        raw_cat.lower(),
        cat_slug.lower()
    ]

    alias_map = {
        "operatori-benessere": "operatori benessere",
        "pet-sitter": "petsitter",
        "escursioni-sport": "escursioni & sport",
        "biglietti-spettacoli": "biglietti spettacoli",
        "libri-scuola": "libri scuola",
        "caffe-parole": "caffe & parole",
        "caffe-e-parole": "caffe & parole",

        # Nuove categorie
        "family-kids": "family & kids",
        "eventi-socialita": "eventi & socialita",
        "eventi-socialità": "eventi & socialità",
        "spazi-sale": "spazi & sale",
    }

    if cat_slug in alias_map:
        json_key_aliases.append(alias_map[cat_slug])

    if raw_cat.lower() in alias_map:
        json_key_aliases.append(alias_map[raw_cat.lower()])

    json_key = cat_slug

    zona = request.args.get("zona", "").strip()
    provincia_filtro = request.args.get("provincia", "").strip()
    filtri_attivi = request.args.getlist("filtri")

    # 🔹 NUOVO: tipo annuncio (offro / cerco)
    tipo_annuncio = request.args.get("tipo_annuncio", "").strip().lower()
    if tipo_annuncio not in ("offro", "cerco"):
        tipo_annuncio = ""

    # =========================================================
    # 🔒 PROVINCIA BASE (macro area)
    # =========================================================
    provincia_attiva = session.get("macro_area")

    if not provincia_attiva:
        utente_id = session.get("utente_id")
        if utente_id:
            conn_tmp = get_db_connection()
            cur_tmp = conn_tmp.cursor()
            cur_tmp.execute(sql("SELECT provincia FROM utenti WHERE id = ?"), (utente_id,))
            row = cur_tmp.fetchone()
            conn_tmp.close()

            if row and list(row.values())[0]:
                provincia_attiva = list(row.values())[0]
                session["macro_area"] = provincia_attiva

    provincia_query = provincia_filtro or provincia_attiva or "__INVALID__"

    # =========================================================
    # FILTRI CATEGORIA — letti da DB
    # =========================================================
    filtri_possibili = []

    conn_filtri = get_db_connection()
    cur_filtri = get_cursor(conn_filtri)

    for key in json_key_aliases:
        cur_filtri.execute(sql("""
            SELECT filtro
            FROM filtri_categoria
            WHERE categoria = ?
              AND attivo = 1
            ORDER BY ordine ASC, filtro ASC
        """), (key,))

        righe_filtri = cur_filtri.fetchall()

        if righe_filtri:
            filtri_possibili = [
                dict(r)["filtro"] for r in righe_filtri
            ]
            break

    conn_filtri.close()

    # =========================================================
    # DB
    # =========================================================
    conn = get_db_connection()

    c = get_cursor(conn)

    # =========================================================
    # SQL FLAGS (DEVONO STARE PRIMA DI ESSERE USATE)
    # =========================================================
    has_urgente_sql = f"""
    (
      CASE WHEN EXISTS (
        SELECT 1
        FROM attivazioni_servizi act
        JOIN servizi s ON s.id = act.servizio_id
        WHERE act.annuncio_id = a.id
          AND s.codice = 'annuncio_urgente'
          AND act.stato = 'attivo'
          AND act.data_inizio <= {now_sql()}
          AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
      ) THEN 1 ELSE 0 END
    ) AS has_urgente
    """

    affidabilita_top_sql = """
        CASE
          WHEN EXISTS (
            SELECT 1
            FROM attivazioni_servizi act
            JOIN servizi s ON s.id = act.servizio_id
            WHERE s.codice = 'badge_affidabilita'
              AND act.utente_id = a.utente_id
              AND act.stato = 'attivo'
          ) THEN 1
          WHEN
            COALESCE((
              SELECT AVG(r.voto)
              FROM recensioni r
              WHERE r.id_destinatario = a.utente_id
                AND r.stato = 'approvato'
            ), 0) >= 4.2
          AND
            COALESCE((
              SELECT COUNT(*)
              FROM recensioni r
              WHERE r.id_destinatario = a.utente_id
                AND r.stato = 'approvato'
            ), 0) >= 5
          THEN 1
          ELSE 0
        END AS affidabilita_top
    """

    # =========================================================
    # 🟡 VETRINA – ROTAZIONE CICLICA INVERSA (30s)
    # =========================================================
    bucket = int(time.time() // 30)

    query_vetrina = f"""
            SELECT DISTINCT
                a.*,
                u.id AS utente_id,
                u.username AS utente_username,
                u.nome AS nome_utente,
                u.cognome AS cognome_utente,
                u.foto_profilo,

            {has_urgente_sql},
            {affidabilita_top_sql},

            COALESCE(ROUND((
                SELECT AVG(r.voto)
                FROM recensioni r
                WHERE r.id_destinatario = a.utente_id
                  AND r.stato = 'approvato'
            ), 1), 0) AS media_recensioni,

            COALESCE((
                SELECT COUNT(*)
                FROM recensioni r
                WHERE r.id_destinatario = a.utente_id
                  AND r.stato = 'approvato'
            ), 0) AS numero_recensioni

        FROM annunci a
        JOIN utenti u ON a.utente_id = u.id
        WHERE
            a.stato = 'approvato'
            AND u.attivo = 1
            AND u.sospeso = 0
            AND (u.disattivato_admin IS NULL OR u.disattivato_admin = 0)
            AND a.provincia = ?
            AND EXISTS (
                SELECT 1
                FROM attivazioni_servizi act
                JOIN servizi s ON s.id = act.servizio_id
                WHERE act.annuncio_id = a.id
                  AND act.stato = 'attivo'
                  AND act.data_inizio <= {now_sql()}
                  AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
                  AND s.codice IN ('vetrina_annuncio', 'annuncio_urgente')
            )
    """

    params_vetrina = [provincia_query]

    if json_key:
        query_vetrina += " AND a.categoria = ?"
        params_vetrina.append(json_key)

    if tipo_annuncio:
        query_vetrina += " AND a.tipo_annuncio = ?"
        params_vetrina.append(tipo_annuncio)

    if zona:
        query_vetrina += " AND a.zona = ?"
        params_vetrina.append(zona)

    for f_att in filtri_attivi:
        query_vetrina += " AND a.filtri_categoria LIKE ?"
        params_vetrina.append(f"%{f_att}%")

    query_vetrina += " ORDER BY a.id ASC"

    c.execute(sql(query_vetrina), params_vetrina)
    rows = [dict(r) for r in c.fetchall()]

    annunci_vetrina = []
    n = len(rows)
    if n > 0:
        shift = bucket % n
        annunci_vetrina = (rows[-shift:] + rows[:-shift])[:6]

    # =========================================================
    # 🔥 BOOST SCORE – LISTA
    # =========================================================
    urgent_score_sql = f"""
    (
      CASE WHEN EXISTS (
        SELECT 1
        FROM attivazioni_servizi act
        JOIN servizi s ON s.id = act.servizio_id
        WHERE act.annuncio_id = a.id
          AND s.codice = 'annuncio_urgente'
          AND act.stato = 'attivo'
          AND act.data_inizio <= {now_sql()}
          AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
      ) THEN 200 ELSE 0 END
    ) AS urgent_score
    """
    boost_score_sql = f"""
    (
      CASE WHEN EXISTS (
        SELECT 1
        FROM attivazioni_servizi act
        JOIN servizi s ON s.id = act.servizio_id
        WHERE act.annuncio_id = a.id
          AND s.codice = 'boost_lista'
          AND act.stato = 'attivo'
          AND act.data_inizio <= {now_sql()}
          AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
      ) THEN 100 ELSE 0 END
    ) AS boost_score
    """
    has_evidenza_sql = f"""
    (
      CASE WHEN EXISTS (
        SELECT 1
        FROM attivazioni_servizi act
        JOIN servizi s ON s.id = act.servizio_id
        WHERE act.annuncio_id = a.id
          AND s.codice = 'badge_evidenza'
          AND act.stato = 'attivo'
          AND act.data_inizio <= {now_sql()}
          AND (act.data_fine IS NULL OR act.data_fine > {now_sql()})
      ) THEN 1 ELSE 0 END
    ) AS has_evidenza
    """

    query_annunci = f"""
        SELECT *
        FROM (
            SELECT
                a.*,
                u.id AS utente_id,
                u.username AS utente_username,
                u.nome AS nome_utente,
                u.cognome AS cognome_utente,
                u.email AS email_utente,
                u.foto_profilo,
                {urgent_score_sql},
                {boost_score_sql},
                {has_evidenza_sql},
                {has_urgente_sql},
                {affidabilita_top_sql},
                COALESCE(ROUND((
                    SELECT AVG(r.voto)
                    FROM recensioni r
                    WHERE r.id_destinatario = a.utente_id
                      AND r.stato = 'approvato'
                ), 1), 0) AS media_recensioni,
                COALESCE((
                    SELECT COUNT(*)
                    FROM recensioni r
                    WHERE r.id_destinatario = a.utente_id
                      AND r.stato = 'approvato'
                ), 0) AS numero_recensioni
            FROM annunci a
            JOIN utenti u ON a.utente_id = u.id
            WHERE a.stato = 'approvato'
              AND u.attivo = 1
              AND u.sospeso = 0
              AND (u.disattivato_admin IS NULL OR u.disattivato_admin = 0)
              AND u.foto_profilo IS NOT NULL
              AND u.foto_profilo != ''
              AND (u.ruolo IS NULL OR u.ruolo != 'admin')
              AND a.provincia = ?
    """

    params = [provincia_query]

    if json_key:
        query_annunci += " AND a.categoria = ?"
        params.append(json_key)

    if tipo_annuncio:
        query_annunci += " AND a.tipo_annuncio = ?"
        params.append(tipo_annuncio)

    if zona:
        query_annunci += " AND a.zona = ?"
        params.append(zona)

    for f_att in filtri_attivi:
        query_annunci += " AND a.filtri_categoria LIKE ?"
        params.append(f"%{f_att}%")

    # 🔒 CHIUSURA SUBQUERY + ORDER BY ESTERNO (PostgreSQL safe)
    query_annunci += """
        ) sub
        ORDER BY
          CASE
            WHEN urgent_score > 0 THEN 0
            WHEN boost_score > 0 THEN 1
            ELSE 2
          END ASC,

          CASE
            WHEN urgent_score > 0 OR boost_score > 0
            THEN (abs(id * 1103515245 + 12345) %% 97)
            ELSE NULL
          END ASC,

          data_pubblicazione DESC,
          affidabilita_top DESC
    """

    c.execute(sql(query_annunci), params)
    annunci = [dict(row) for row in c.fetchall()]

    categorie_con_foto_card = {"family-kids", "eventi-socialita", "spazi-sale"}

    def assegna_immagine_card(lista_annunci):
        for ann in lista_annunci:
            categoria_annuncio = ann.get("categoria")
            foto_card = (ann.get("foto_card") or "").strip()
            foto_profilo = (ann.get("foto_profilo") or "").strip()

            if categoria_annuncio in categorie_con_foto_card and foto_card:
                ann["immagine_card"] = foto_card
            else:
                ann["immagine_card"] = foto_profilo

    assegna_immagine_card(annunci)
    assegna_immagine_card(annunci_vetrina)

    return render_template(
        "cerca.html",
        categoria=json_key,
        categoria_label=categoria_label,
        zona=zona,
        filtri=filtri_attivi,
        tipo_annuncio=tipo_annuncio,
        annunci_vetrina=annunci_vetrina,
        annunci=annunci,
        filtri_possibili=filtri_possibili,
    )

@app.route("/notifica/<int:id>/apri")
@login_required
def apri_notifica(id):
    conn = get_db_connection()
    notifica = conn.execute(
        "SELECT link FROM notifiche WHERE id = ? AND id_utente = ?",
        (id, g.utente["id"])
    ).fetchone()
    if not notifica:
        flash("Notifica non trovata.", "error")
        return redirect(url_for("notifiche"))

    # Segna come letta
    conn.execute(sql(f"""
        UPDATE notifiche
        SET letta = 1,
            data_lettura = {now_sql()}
        WHERE id = ? AND id_utente = ?
    """), (id, g.utente["id"]))
    conn.commit()


    # 🔔 aggiorna il badge
    emit_update_notifications(g.utente["id"])

    # Reindirizza al link
    if notifica["link"]:
        return redirect(notifica["link"])
    else:
        return redirect(url_for("notifiche"))

@app.route("/elimina-risposta/<int:id>", methods=["POST"])
@login_required
def elimina_risposta_route(id):
    elimina_risposta(id, id_autore=g.utente["id"])
    flash("Risposta eliminata con successo ✅", "success")
    return redirect(request.referrer or url_for("mie_recensioni_ricevute"))

@app.route("/api/ai/aiuto-scrittura-annuncio", methods=["POST"])
@login_required
@foto_obbligatoria
def api_ai_aiuto_scrittura_annuncio():
    verify_csrf()

    data = request.get_json(silent=True) or {}

    titolo = (data.get("titolo") or "").strip()
    descrizione = (data.get("descrizione") or "").strip()
    lingua = (data.get("lingua") or "it").strip().lower()
    azione = (data.get("azione") or "improve").strip().lower()

    # Azioni supportate:
    # - improve: migliora nella lingua scelta dall'utente
    # - translate_it: durante la preview migliora comunque nella lingua scelta
    # - final_translate_it: traduzione finale in italiano prima di inserire nell'annuncio
    if azione not in ("improve", "translate_it", "final_translate_it"):
        azione = "improve"

    if lingua == "it":
        azione = "improve"

    if not descrizione:
        return jsonify({
            "ok": False,
            "error": "Descrizione mancante."
        }), 400

    if len(descrizione) > 1500:
        return jsonify({
            "ok": False,
            "error": "La descrizione supera il limite massimo di 1500 caratteri."
        }), 400

    if len(titolo) > 120:
        return jsonify({
            "ok": False,
            "error": "Il titolo supera il limite massimo di 120 caratteri."
        }), 400

    try:
        client = get_openai_client()

        istruzioni = """
Sei l'assistente di scrittura di MyLocalCare, una piattaforma italiana di annunci locali.

Devi aiutare l'utente a migliorare un annuncio.

Regole obbligatorie:
- Rispondi SOLO con JSON valido.
- Non aggiungere markdown.
- Non inventare qualifiche, certificazioni, esperienza, prezzi, disponibilità o dati personali non forniti.
- Non inserire dati sensibili.
- Mantieni un tono umano, chiaro, affidabile e semplice.
- Il titolo deve essere breve, naturale e adatto a un annuncio locale.
- La descrizione deve essere chiara, ordinata e credibile.
- Non usare linguaggio troppo commerciale o esagerato.
- Non usare emoji nel titolo o nella descrizione.

Gestione lingua:
- Rileva sempre la lingua effettiva del testo scritto dall'utente, indipendentemente dalla lingua scelta nell'interfaccia.
- Se il testo contiene parti significative in più lingue, imposta "testo_misto": true.
- Se il testo è sostanzialmente in una sola lingua, imposta "testo_misto": false.
- Il valore di "lingua_rilevata" deve essere un codice breve minuscolo, per esempio: "it", "en", "fr", "es", "ro", "ar", "uk", "ru", "de", "altro".

Gestione azione:
- Prima rileva sempre la lingua effettiva del testo originale.
- Se "lingua_scelta" è "it" e la lingua effettiva del testo originale NON è italiano, traduci e migliora titolo e descrizione in italiano, anche se l'azione ricevuta è "improve" o "translate_it".
- Se "lingua_scelta" è "it" e la lingua effettiva del testo originale è italiano, migliora titolo e descrizione in italiano.
- Se "lingua_scelta" è diversa da "it" e l'azione è "improve", migliora titolo e descrizione mantenendo la lingua effettiva del testo originale.
- Se "lingua_scelta" è diversa da "it" e l'azione è "translate_it", NON tradurre ancora in italiano: migliora titolo e descrizione mantenendo la lingua effettiva del testo originale.
- Se l'azione è "final_translate_it", traduci e migliora titolo e descrizione in italiano.
- Se l'azione è "final_translate_it" ma il testo è già completamente in italiano e "testo_misto" è false, puoi semplicemente migliorarlo in italiano senza traduzione artificiale.

Restituisci sempre esattamente queste chiavi JSON:
{
  "titolo": "...",
  "descrizione": "...",
  "lingua_rilevata": "...",
  "testo_misto": true
}
"""

        prompt_utente = {
            "lingua_scelta": lingua,
            "azione": azione,
            "titolo_originale": titolo,
            "descrizione_originale": descrizione
        }

        response = client.responses.create(
            model=os.getenv("OPENAI_TEXT_MODEL", "gpt-5.4-mini"),
            instructions=istruzioni,
            input=(
                "Restituisci una risposta in formato json valido.\n\n"
                "Dati dell'annuncio da migliorare:\n"
                + json.dumps(prompt_utente, ensure_ascii=False)
            ),
            text={
                "format": {
                    "type": "json_object"
                }
            },
            store=False
        )

        raw_text = (response.output_text or "").strip()

        try:
            result = json.loads(raw_text)
        except Exception:
            log_exception_safe(
                "❌ AI aiuto scrittura: risposta non JSON",
                extra={
                    "user_id": g.utente["id"] if g.utente else None,
                    "raw_preview": raw_text[:300]
                },
                production=True
            )

            return jsonify({
                "ok": False,
                "error": "Non sono riuscito a generare una proposta valida. Riprova."
            }), 502

        titolo_generato = (result.get("titolo") or "").strip()
        descrizione_generata = (result.get("descrizione") or "").strip()
        lingua_rilevata = (result.get("lingua_rilevata") or lingua or "altro").strip().lower()

        testo_misto_raw = result.get("testo_misto", False)
        if isinstance(testo_misto_raw, bool):
            testo_misto = testo_misto_raw
        elif isinstance(testo_misto_raw, str):
            testo_misto = testo_misto_raw.strip().lower() in ("true", "1", "yes", "si", "sì")
        else:
            testo_misto = False

        if not descrizione_generata:
            return jsonify({
                "ok": False,
                "error": "La proposta generata non contiene una descrizione valida."
            }), 502

        if not titolo_generato:
            titolo_generato = titolo or "Titolo suggerito"

        return jsonify({
            "ok": True,
            "titolo": titolo_generato[:120],
            "descrizione": descrizione_generata[:2500],
            "lingua_rilevata": lingua_rilevata[:20],
            "testo_misto": testo_misto
        })

    except Exception as e:
        log_exception_safe(
            "❌ Errore /api/ai/aiuto-scrittura-annuncio",
            e,
            {
                "user_id": g.utente["id"] if g.utente else None,
                "lingua": lingua,
                "azione": azione
            },
            production=True
        )

        return jsonify({
            "ok": False,
            "error": "Errore durante la generazione della proposta. Riprova più tardi."
        }), 500


@app.route('/api/operatori')
def api_operatori():
    categoria = request.args.get('categoria', '').strip()
    zona = request.args.get('zona')
    operatori = [dict(op) for op in get_operatori(categoria, zona)]
    return jsonify(operatori)

@app.route('/api/operatore/<int:id>')
def api_operatore(id):
    op = get_operatore_by_id(id)
    if op:
        return jsonify(dict(op)), 200
    return jsonify({"error": "Not found"}), 404

@app.route("/api/suggerimenti-comuni")
def suggerimenti_comuni():
    q = (request.args.get("q") or "").strip().lower()
    if len(q) < 2:
        return jsonify([])

    path = os.path.join(app.static_folder, "data", "comuni.json")

    try:
        with open(path, encoding="utf-8") as f:
            comuni = json.load(f)
    except Exception:
        return jsonify([])

    risultati = []

    # 1️⃣ MATCH CHE INIZIANO PER (Roma, Milano, ecc.)
    for c in comuni:
        nome = c.get("comune", "")
        if nome.lower().startswith(q):
            risultati.append({
                "comune": nome,
                "provincia": c.get("provincia"),
                "regione": c.get("regione")
            })

    # 2️⃣ MATCH CHE CONTENGONO (solo se servono)
    if len(risultati) < 10:
        for c in comuni:
            nome = c.get("comune", "")
            nome_lower = nome.lower()

            if q in nome_lower and not nome_lower.startswith(q):
                risultati.append({
                    "comune": nome,
                    "provincia": c.get("provincia"),
                    "regione": c.get("regione")
                })

            if len(risultati) >= 10:
                break

    return jsonify(risultati[:10])

@app.route("/api/servizi/<codice>/piani")
@login_required
def api_servizi_piani(codice):
    conn = get_db_connection()
    cur = get_cursor(conn)

    # servizio
    cur.execute(sql("""
        SELECT id, nome, descrizione
        FROM servizi
        WHERE codice = ? AND attivo = 1
    """), (codice,))
    servizio = cur.fetchone()

    if not servizio:
        return jsonify({"error": "Servizio non trovato"}), 404

    # piani
    cur.execute(sql("""
        SELECT
            id,
            durata_giorni,
            prezzo_cent,
            consigliato,
            evidenziato
        FROM servizi_piani
        WHERE servizio_id = ?
          AND attivo = 1
        ORDER BY ordine ASC, prezzo_cent ASC
    """), (servizio["id"],))

    piani = [
        {
            "id": r["id"],
            "durata": r["durata_giorni"],   # NULL → permanente
            "prezzo": r["prezzo_cent"],
            "consigliato": r["consigliato"],
            "evidenziato": r["evidenziato"],
            "permanente": r["durata_giorni"] is None
        }
        for r in cur.fetchall()
    ]



    return jsonify({
        "servizio": {
            "id": servizio["id"],
            "nome": servizio["nome"],
            "descrizione": servizio["descrizione"]
        },
        "piani": piani
    })

@app.route("/api/annunci/<int:annuncio_id>/servizi/<codice>")
@login_required
def api_annuncio_servizio_stato(annuncio_id, codice):
    conn = get_db_connection()
    cur = get_cursor(conn)

    # servizio
    cur.execute(sql("""
        SELECT id, ambito
        FROM servizi
        WHERE codice = ? AND attivo = 1
    """), (codice,))
    servizio = cur.fetchone()

    if not servizio:
        return jsonify({"error": "Servizio non trovato"}), 404

    # query dinamica in base all’ambito:
    # prende la MIGLIORE copertura attiva del servizio
    # priorità:
    # 1) permanente (data_fine NULL)
    # 2) scadenza più lontana
    # 3) attivazione più recente
    if servizio["ambito"] == "profilo":
        cur.execute(sql(f"""
            SELECT
                id,
                data_inizio,
                data_fine,
                stato,
                attivato_da
            FROM attivazioni_servizi
            WHERE servizio_id = ?
              AND utente_id = ?
              AND annuncio_id IS NULL
              AND stato = 'attivo'
              AND (data_fine IS NULL OR data_fine > {now_sql()})
            ORDER BY
                CASE WHEN data_fine IS NULL THEN 1 ELSE 0 END DESC,
                data_fine DESC,
                {order_datetime("data_inizio")} DESC
            LIMIT 1
        """), (servizio["id"], g.utente["id"]))
    else:
        cur.execute(sql(f"""
            SELECT
                id,
                data_inizio,
                data_fine,
                stato,
                attivato_da
            FROM attivazioni_servizi
            WHERE servizio_id = ?
              AND annuncio_id = ?
              AND utente_id = ?
              AND stato = 'attivo'
              AND (data_fine IS NULL OR data_fine > {now_sql()})
            ORDER BY
                CASE WHEN data_fine IS NULL THEN 1 ELSE 0 END DESC,
                data_fine DESC,
                {order_datetime("data_inizio")} DESC
            LIMIT 1
        """), (servizio["id"], annuncio_id, g.utente["id"]))

    att = cur.fetchone()

    if not att:
        return jsonify({
            "attivo": False,
            "stato": "non_attivo",
            "data_inizio": None,
            "data_fine": None,
            "permanente": False
        })

    def solo_data(v):
        if not v:
            return None
        return str(v)[:10]

    return jsonify({
        "attivo": True,
        "stato": "attivo",
        "data_inizio": solo_data(att["data_inizio"]),
        "data_fine": solo_data(att["data_fine"]),
        "permanente": not att["data_fine"],
        "attivato_da": att["attivato_da"]
    })

@app.route("/api/pacchetti/<codice>/piani")
@login_required
def api_pacchetti_piani(codice):
    conn = get_db_connection()
    cur = get_cursor(conn)

    # pacchetto
    cur.execute(sql("""
        SELECT id, nome, descrizione
        FROM pacchetti
        WHERE codice = ? AND attivo = 1
    """), (codice,))
    pacchetto = cur.fetchone()

    if not pacchetto:

        return jsonify({"error": "Pacchetto non trovato"}), 404

    # piani pacchetto
    cur.execute(sql("""
        SELECT
            id,
            durata_giorni,
            prezzo_cent,
            consigliato,
            evidenziato
        FROM pacchetti_piani
        WHERE pacchetto_id = ?
          AND attivo = 1
        ORDER BY ordine ASC, prezzo_cent ASC
    """), (pacchetto["id"],))

    piani = [
        {
            "id": r["id"],
            "durata": r["durata_giorni"],   # NULL → permanente
            "prezzo": r["prezzo_cent"],
            "consigliato": r["consigliato"],
            "evidenziato": r["evidenziato"],
            "permanente": r["durata_giorni"] is None
        }
        for r in cur.fetchall()
    ]



    return jsonify({
        "pacchetto": {
            "id": pacchetto["id"],
            "nome": pacchetto["nome"],
            "descrizione": pacchetto["descrizione"]
        },
        "piani": piani
    })

@app.route("/api/attiva", methods=["POST"])
@login_required
def api_attiva():
    data = request.get_json()

    annuncio_id = data.get("annuncio_id")
    piano_id = data.get("piano_id")
    tipo = data.get("tipo")  # "servizi" | "pacchetti"

    if not annuncio_id or not piano_id or tipo not in ("servizi", "pacchetti"):
        return jsonify({"error": "Dati non validi"}), 400

    conn = get_db_connection()

    cur = get_cursor(conn)

    try:
        # 1️⃣ recupero piano + prezzo
        if tipo == "servizi":
            cur.execute(sql("""
                SELECT p.id, p.prezzo_cent
                FROM servizi_piani p
                WHERE p.id = ? AND p.attivo = 1
            """), (piano_id,))
        else:
            cur.execute(sql("""
                SELECT p.id, p.prezzo_cent
                FROM pacchetti_piani p
                WHERE p.id = ? AND p.attivo = 1
            """), (piano_id,))

        piano = cur.fetchone()
        if not piano:
            return jsonify({"error": "Piano non valido"}), 404

        prezzo = int(piano["prezzo_cent"])
        if prezzo <= 0:
            return jsonify({"error": "Prezzo non valido"}), 400

        # 2️⃣ crea acquisto (pending)
        # 🔴 COSTRUZIONE DATI ACQUISTO (STANDARD UNICO)

        if tipo == "servizi":
            tipo_acquisto = "servizio"
            ref_id = piano_id   # per servizi: id del piano
        else:
            tipo_acquisto = "pacchetto"
            cur.execute(sql("""
                SELECT pacchetto_id
                FROM pacchetti_piani
                WHERE id = ?
            """), (piano_id,))
            row = cur.fetchone()
            if not row:
                return jsonify({"error": "Pacchetto non valido"}), 400
            ref_id = row["pacchetto_id"]

        # ✅ INSERT COMPLETO (UGUALE A crea-payment-intent)
        acquisto_id = insert_and_get_id(
            cur,
            f"""
            INSERT INTO acquisti
            (
                utente_id,
                tipo,
                ref_id,
                prezzo_id,
                metodo,
                importo_cent,
                stato,
                annuncio_id,
                created_at
            )
            VALUES (?, ?, ?, ?, 'stripe', ?, 'pending', ?, {now_sql()})
            """,
            (
                g.utente["id"],
                tipo_acquisto,
                ref_id,
                piano_id,
                prezzo,
                annuncio_id
            )
        )

        # 3️⃣ PaymentIntent Stripe
        intent = stripe.PaymentIntent.create(
            amount=prezzo,
            currency="eur",
            automatic_payment_methods={"enabled": True},
            metadata={
                "acquisto_id": acquisto_id,
                "piano_id": piano_id,
                "tipo": tipo,
                "annuncio_id": annuncio_id,
                "utente_id": g.utente["id"]
            }
        )

        # 4️⃣ salva riferimento Stripe
        cur.execute(sql("""
            UPDATE acquisti
            SET riferimento_esterno = ?
            WHERE id = ?
        """), (intent.id, acquisto_id))

        conn.commit()

        return jsonify({
            "ok": True,
            "client_secret": intent.client_secret
        })

    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500

    finally:
        try:
            conn.close()
        except:
            pass



@app.route("/api/crea-payment-intent", methods=["POST"])
@login_required
def crea_payment_intent():
    data = request.get_json() or {}

    annuncio_id = data.get("annuncio_id")
    piano_id = data.get("piano_id")
    tipo = data.get("tipo")  # "servizi" | "pacchetti"

    if not annuncio_id or not piano_id or tipo not in ("servizi", "pacchetti"):
        return jsonify({"error": "Dati non validi"}), 400

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        if tipo == "servizi":
            cur.execute(sql("""
                SELECT id, prezzo_cent
                FROM servizi_piani
                WHERE id = ? AND attivo = 1
            """), (piano_id,))

            piano = cur.fetchone()

            if not piano:
                return jsonify({"error": "Piano servizio non trovato"}), 404

            tipo_acquisto = "servizio"
            ref_id = piano["id"]

        else:
            cur.execute(sql("""
                SELECT id, pacchetto_id, prezzo_cent
                FROM pacchetti_piani
                WHERE id = ? AND attivo = 1
            """), (piano_id,))

            piano = cur.fetchone()

            if not piano:
                return jsonify({"error": "Piano pacchetto non trovato"}), 404

            tipo_acquisto = "pacchetto"
            ref_id = piano["pacchetto_id"]

        prezzo = int(piano["prezzo_cent"])

        if prezzo <= 0:
            return jsonify({"error": "Prezzo non valido"}), 400

        acquisto_id = insert_and_get_id(
            cur,
            f"""
            INSERT INTO acquisti
            (
                utente_id,
                tipo,
                ref_id,
                prezzo_id,
                metodo,
                importo_cent,
                stato,
                annuncio_id,
                created_at
            )
            VALUES (?, ?, ?, ?, 'stripe', ?, 'pending', ?, {now_sql()})
            """,
            (
                g.utente["id"],
                tipo_acquisto,
                ref_id,
                piano_id,
                prezzo,
                annuncio_id
            )
        )

        intent = stripe.PaymentIntent.create(
            amount=prezzo,
            currency="eur",
            automatic_payment_methods={"enabled": True},
            metadata={
                "acquisto_id": str(acquisto_id),
                "piano_id": str(piano_id),
                "tipo": str(tipo),
                "annuncio_id": str(annuncio_id),
                "utente_id": str(g.utente["id"])
            }
        )

        cur.execute(sql("""
            UPDATE acquisti
            SET riferimento_esterno = ?
            WHERE id = ?
        """), (intent.id, acquisto_id))

        conn.commit()

        security_log(
            "💳 [STRIPE] PaymentIntent creato con metadata",
            {
                "payment_intent": intent.id,
                "acquisto_id": acquisto_id,
                "tipo": tipo_acquisto,
                "ref_id": ref_id,
                "piano_id": piano_id,
                "annuncio_id": annuncio_id,
                "utente_id": g.utente["id"]
            },
            production=True
        )

        return jsonify({
            "client_secret": intent.client_secret
        })

    except Exception as e:
        conn.rollback()

        log_exception_safe(
            "❌ [STRIPE] errore creazione PaymentIntent",
            e,
            production=True
        )

        return jsonify({"error": "Errore creazione pagamento"}), 500

    finally:
        try:
            conn.close()
        except:
            pass

def gestisci_pagamento_confermato(payment_intent):
    # Stripe può passare sia dict sia StripeObject.
    # NON usare dict(payment_intent): su alcuni StripeObject genera KeyError(0).
    try:
        if isinstance(payment_intent, dict):
            riferimento_esterno = payment_intent.get("id")
            metadata_raw = payment_intent.get("metadata") or {}

        else:
            riferimento_esterno = getattr(payment_intent, "id", None)

            if hasattr(payment_intent, "get"):
                metadata_raw = payment_intent.get("metadata") or {}
            else:
                metadata_raw = getattr(payment_intent, "metadata", {}) or {}

        if hasattr(metadata_raw, "to_dict_recursive"):
            metadata = metadata_raw.to_dict_recursive()
        elif isinstance(metadata_raw, dict):
            metadata = dict(metadata_raw)
        else:
            metadata = {
                k: metadata_raw[k]
                for k in metadata_raw.keys()
            } if hasattr(metadata_raw, "keys") else {}

        acquisto_id = metadata.get("acquisto_id")

    except Exception as e:
        log_exception_safe(
            "❌ [STRIPE] impossibile leggere payment_intent/metadata",
            e,
            production=True
        )
        return

    security_log(
        "💳 [STRIPE] START gestisci_pagamento_confermato",
        {
            "payment_intent": riferimento_esterno,
            "acquisto_id": acquisto_id,
            "metadata_keys": list(metadata.keys()) if isinstance(metadata, dict) else []
        },
        production=True
    )

    if not acquisto_id:
        security_log(
            "⚠️ [STRIPE] metadata.acquisto_id mancante — tento recupero da riferimento_esterno",
            {
                "payment_intent": riferimento_esterno,
                "metadata_keys": list(metadata.keys()) if metadata else []
            },
            production=True
        )

        if not riferimento_esterno:
            security_log(
                "❌ [STRIPE] impossibile recuperare acquisto: payment_intent.id mancante",
                production=True
            )
            return

        conn_lookup = None
        cur_lookup = None

        try:
            conn_lookup = get_db_connection()
            cur_lookup = get_cursor(conn_lookup)

            cur_lookup.execute(sql("""
                SELECT id
                FROM acquisti
                WHERE riferimento_esterno = ?
                ORDER BY id DESC
                LIMIT 1
            """), (riferimento_esterno,))

            row_lookup = cur_lookup.fetchone()

            if not row_lookup:
                security_log(
                    "❌ [STRIPE] acquisto non trovato nemmeno tramite riferimento_esterno",
                    {
                        "payment_intent": riferimento_esterno
                    },
                    production=True
                )
                return

            acquisto_id = row_lookup["id"]

            security_log(
                "✅ [STRIPE] acquisto recuperato tramite riferimento_esterno",
                {
                    "payment_intent": riferimento_esterno,
                    "acquisto_id": acquisto_id
                },
                production=True
            )

        except Exception as e:
            log_exception_safe(
                "❌ [STRIPE] errore recupero acquisto tramite riferimento_esterno",
                e,
                {
                    "payment_intent": riferimento_esterno
                },
                production=True
            )
            return

        finally:
            try:
                if cur_lookup:
                    cur_lookup.close()
            except Exception:
                pass

            try:
                if conn_lookup:
                    conn_lookup.close()
            except Exception:
                pass

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        # SQLite: lock esplicito
        # PostgreSQL: la transazione è già gestita dalla connessione
        if not app.config.get("IS_POSTGRES"):
            cur.execute(sql("BEGIN IMMEDIATE"))

        # fonte di verità: l’acquisto
        cur.execute(sql("""
            SELECT id, utente_id, tipo, ref_id, prezzo_id, stato, annuncio_id
            FROM acquisti
            WHERE id = ?
        """), (int(acquisto_id),))

        acquisto = cur.fetchone()
        security_log(
            "💳 [STRIPE] acquisto letto",
            {
                "acquisto_id": acquisto["id"] if acquisto else None,
                "utente_id": acquisto["utente_id"] if acquisto else None,
                "tipo": acquisto["tipo"] if acquisto else None,
                "stato": acquisto["stato"] if acquisto else None,
                "annuncio_id": acquisto["annuncio_id"] if acquisto else None
            } if acquisto else None,
            production=True
        )

        if not acquisto:
            print("❌ Webhook Stripe: acquisto non trovato:", acquisto_id, flush=True)
            conn.rollback()
            return

        # idempotenza
        if acquisto["stato"] == "paid":
            conn.rollback()
            return

        utente_id = int(acquisto["utente_id"])
        tipo = acquisto["tipo"]                  # 'servizio' | 'pacchetto'
        ref_id = int(acquisto["ref_id"])         # servizio: id piano servizi_piani | pacchetto: pacchetto_id
        piano_id = acquisto["prezzo_id"]         # id servizi_piani o pacchetti_piani
        annuncio_id = acquisto["annuncio_id"]

        if annuncio_id is not None:
            annuncio_id = int(annuncio_id)

        # aggiorna acquisto
        cur.execute(sql("""
            UPDATE acquisti
            SET stato = 'paid',
                metodo = 'stripe',
                riferimento_esterno = ?
            WHERE id = ?
        """), (riferimento_esterno, int(acquisto_id)))

        security_log(
            "💳 [STRIPE] acquisto aggiornato a paid",
            {
                "acquisto_id": acquisto_id,
                "payment_intent": riferimento_esterno
            },
            production=True
        )

        # ===============================
        # SERVIZIO SINGOLO
        # ===============================
        if tipo == "servizio":
            piano_servizio_id = int(piano_id) if piano_id is not None else int(ref_id)

            cur.execute(sql("""
                SELECT servizio_id, durata_giorni, prezzo_cent
                FROM servizi_piani
                WHERE id = ?
            """), (piano_servizio_id,))
            piano = cur.fetchone()

            security_log(
                "💳 [STRIPE] piano servizio letto",
                {
                    "servizio_id": piano["servizio_id"] if piano else None,
                    "durata_giorni": piano["durata_giorni"] if piano else None,
                    "prezzo_cent": piano["prezzo_cent"] if piano else None
                } if piano else None,
                production=True
            )

            if not piano:
                raise Exception("Piano servizio non trovato")

            # storico acquisto servizio
            cur.execute(sql("""
                INSERT INTO acquisti_servizi
                (utente_id, servizio_id, metodo, importo, valuta, riferimento_esterno)
                VALUES (?, ?, 'stripe', ?, 'EUR', ?)
            """), (
                utente_id,
                int(piano["servizio_id"]),
                float(int(piano["prezzo_cent"]) / 100.0),
                riferimento_esterno
            ))

            security_log(
                "💳 [STRIPE] storico acquisto servizio inserito",
                {
                    "utente_id": utente_id,
                    "servizio_id": int(piano["servizio_id"]),
                    "payment_intent": riferimento_esterno
                },
                production=True
            )

            ok, msg, att_id = attiva_servizio(
                conn=conn,
                utente_id=utente_id,
                servizio_id=int(piano["servizio_id"]),
                annuncio_id=annuncio_id,
                durata_giorni=piano["durata_giorni"],
                acquisto_id=int(acquisto_id),
                attivato_da="stripe",
                note=f"Stripe PI {riferimento_esterno}"
            )

            security_log(
                "💳 [STRIPE] attivazione servizio completata",
                {
                    "ok": ok,
                    "attivazione_id": att_id,
                    "messaggio": msg
                },
                production=True
            )

            if not ok:
                raise Exception(msg)

            # se è annuncio_urgente, invia notifiche interne
            cur.execute(sql("""
                SELECT codice
                FROM servizi
                WHERE id = ?
            """), (int(piano["servizio_id"]),))
            servizio_row = cur.fetchone()

            if servizio_row and servizio_row["codice"] == "annuncio_urgente" and annuncio_id:
                try:
                    notifica_urgente(
                        annuncio_id=int(annuncio_id),
                        attivazione_id=int(att_id) if att_id else None,
                        eseguito_da="stripe",
                        conn=conn
                    )
                except Exception as e:
                    print(
                        f"⚠️ Errore notifica urgente (stripe servizio singolo): {e}",
                        flush=True
                    )

        # ===============================
        # PACCHETTO
        # ===============================
        elif tipo == "pacchetto":
            if piano_id is None:
                raise Exception("piano_id mancante su acquisti.prezzo_id (impossibile determinare durata pacchetto)")

            cur.execute(sql("""
                SELECT durata_giorni, prezzo_cent
                FROM pacchetti_piani
                WHERE id = ?
            """), (int(piano_id),))
            piano_p = cur.fetchone()

            if not piano_p:
                raise Exception("Piano pacchetto non trovato")

            durata_piano = piano_p["durata_giorni"]   # può essere NULL => permanente
            prezzo_tot_cent = int(piano_p["prezzo_cent"] or 0)

            cur.execute(sql("""
                SELECT servizio_id
                FROM pacchetti_servizi
                WHERE pacchetto_id = ?
            """), (ref_id,))
            servizi = cur.fetchall()

            if not servizi:
                raise Exception("Pacchetto senza servizi")

            quota = (prezzo_tot_cent / 100.0) / max(len(servizi), 1)

            for row in servizi:
                servizio_id = int(row["servizio_id"])

                # storico acquisto servizio (quota)
                cur.execute(sql("""
                    INSERT INTO acquisti_servizi
                    (utente_id, servizio_id, metodo, importo, valuta, riferimento_esterno)
                    VALUES (?, ?, 'stripe', ?, 'EUR', ?)
                """), (
                    utente_id,
                    servizio_id,
                    float(quota),
                    riferimento_esterno
                ))

                security_log(
                    "💳 [STRIPE] storico acquisto servizio pacchetto inserito",
                    {
                        "utente_id": utente_id,
                        "servizio_id": servizio_id,
                        "payment_intent": riferimento_esterno
                    },
                    production=True
                )

                ok, msg, att_id = attiva_servizio(
                    conn=conn,
                    utente_id=utente_id,
                    servizio_id=servizio_id,
                    annuncio_id=annuncio_id,
                    durata_giorni=durata_piano,
                    acquisto_id=int(acquisto_id),
                    attivato_da="stripe",
                    note=f"Stripe PI {riferimento_esterno} (pacchetto)"
                )

                security_log(
                    "💳 [STRIPE] attivazione servizio pacchetto completata",
                    {
                        "ok": ok,
                        "attivazione_id": att_id,
                        "messaggio": msg
                    },
                    production=True
                )

                if not ok and msg == "Servizio già attivo.":
                    continue

                if not ok:
                    raise Exception(msg)

                # se nel pacchetto c'è annuncio_urgente, invia notifiche interne
                cur.execute(sql("""
                    SELECT codice
                    FROM servizi
                    WHERE id = ?
                """), (servizio_id,))
                servizio_row = cur.fetchone()

                if servizio_row and servizio_row["codice"] == "annuncio_urgente" and annuncio_id:
                    try:
                        notifica_urgente(
                            annuncio_id=int(annuncio_id),
                            attivazione_id=int(att_id) if att_id else None,
                            eseguito_da="stripe",
                            conn=conn
                        )
                    except Exception as e:
                        print(
                            f"⚠️ Errore notifica urgente (stripe pacchetto): {e}",
                            flush=True
                        )

        else:
            raise Exception("Tipo acquisto non valido")

        conn.commit()

        security_log(
            "✅ [STRIPE] webhook completato — servizi attivati",
            {
                "acquisto_id": acquisto_id,
                "payment_intent": riferimento_esterno
            },
            production=True
        )

        # 📧 Email conferma pagamento aumento visibilità
        # Va inviata SOLO dopo il commit, così non rischiamo di confermare via mail
        # un pagamento non ancora registrato correttamente.
        try:
            cur.execute(sql("""
                SELECT email, nome, username
                FROM utenti
                WHERE id = ?
            """), (utente_id,))
            utente_mail = cur.fetchone()

            titolo_annuncio = None

            if annuncio_id:
                cur.execute(sql("""
                    SELECT titolo
                    FROM annunci
                    WHERE id = ?
                """), (annuncio_id,))
                annuncio_mail = cur.fetchone()

                if annuncio_mail:
                    titolo_annuncio = annuncio_mail["titolo"]

            importo_cent = 0
            durata_giorni = None
            tipo_label = "Aumento visibilità annuncio"

            if tipo == "servizio":
                cur.execute(sql("""
                    SELECT
                        sp.durata_giorni,
                        sp.prezzo_cent,
                        s.nome AS nome_servizio
                    FROM servizi_piani sp
                    JOIN servizi s ON s.id = sp.servizio_id
                    WHERE sp.id = ?
                """), (int(piano_id) if piano_id is not None else int(ref_id),))

                piano_mail = cur.fetchone()

                if piano_mail:
                    importo_cent = int(piano_mail["prezzo_cent"] or 0)
                    durata_giorni = piano_mail["durata_giorni"]
                    tipo_label = piano_mail["nome_servizio"] or "Servizio di aumento visibilità"

            elif tipo == "pacchetto":
                cur.execute(sql("""
                    SELECT
                        pp.durata_giorni,
                        pp.prezzo_cent,
                        p.nome AS nome_pacchetto
                    FROM pacchetti_piani pp
                    JOIN pacchetti p ON p.id = pp.pacchetto_id
                    WHERE pp.id = ?
                """), (int(piano_id),))

                piano_mail = cur.fetchone()

                if piano_mail:
                    importo_cent = int(piano_mail["prezzo_cent"] or 0)
                    durata_giorni = piano_mail["durata_giorni"]
                    tipo_label = piano_mail["nome_pacchetto"] or "Pacchetto aumento visibilità"

            if utente_mail and utente_mail["email"]:
                nome_destinatario = (
                    utente_mail["nome"]
                    or utente_mail["username"]
                    or "utente"
                )

                importo_testo = f"{importo_cent / 100:.2f} €".replace(".", ",")

                durata_testo = (
                    f"{durata_giorni} giorni"
                    if durata_giorni
                    else "durata prevista dal servizio acquistato"
                )

                corpo_email = (
                    f"Ciao {nome_destinatario},\n\n"
                    "ti confermiamo che il pagamento è stato ricevuto correttamente.\n\n"
                    f"Servizio acquistato: {tipo_label}\n"
                    f"Importo: {importo_testo}\n"
                    f"Durata: {durata_testo}\n"
                )

                if titolo_annuncio:
                    corpo_email += f"Annuncio: {titolo_annuncio}\n"

                corpo_email += (
                    "\nIl servizio è stato attivato sul tuo account MyLocalCare.\n\n"
                    "Grazie,\n"
                    "MyLocalCare"
                )

                _invia_email(
                    destinazione=utente_mail["email"],
                    oggetto="Conferma pagamento MyLocalCare",
                    corpo=corpo_email
                )

        except Exception as e:
            log_exception_safe(
                "⚠️ [STRIPE] pagamento confermato ma errore invio email conferma",
                e,
                {
                    "acquisto_id": acquisto_id,
                    "utente_id": utente_id,
                    "annuncio_id": annuncio_id,
                    "payment_intent": riferimento_esterno
                },
                production=True
            )

    except Exception as e:
        conn.rollback()
        security_log(
            "❌ [STRIPE] errore gestione pagamento confermato",
            {
                "error_type": type(e).__name__,
                "error": repr(e),
                "acquisto_id": acquisto_id,
                "payment_intent": riferimento_esterno
            },
            production=True
        )

        if os.getenv("APP_ENV", "production").lower() in ("local", "development"):
            traceback.print_exc()

    finally:
        try:
            conn.close()
        except Exception:
            pass

def attiva_servizio_by_id(conn, servizio_id, **kwargs):
    row = conn.execute(
        "SELECT codice FROM servizi WHERE id = ?",
        (servizio_id,)
    ).fetchone()

    if not row:
        print("Servizio non trovato:", servizio_id)
        return False

    return attiva_servizio(
        codice_servizio=row["codice"],
        **kwargs
    )


def attiva_servizio_da_piano(conn, piano_id, annuncio_id, utente_id, acquisto_id, metodo="admin", importo=0.0, valuta="EUR", riferimento_esterno=None):
    """
    Versione generica (non solo Stripe): attiva un singolo servizio da un piano
    e registra lo storico in acquisti_servizi secondo lo schema reale.
    """
    cur = get_cursor(conn)

    cur.execute(sql("""
        SELECT servizio_id, durata_giorni
        FROM servizi_piani
        WHERE id = ?
    """), (piano_id,))
    piano = cur.fetchone()

    if not piano:
        return

    servizio_id = int(piano["servizio_id"])

    # ✅ storico acquisto servizio (schema reale)
    cur.execute(sql("""
        INSERT INTO acquisti_servizi
        (utente_id, servizio_id, metodo, importo, valuta, riferimento_esterno)
        VALUES (?, ?, ?, ?, ?, ?)
    """), (int(utente_id), servizio_id, metodo, float(importo), (valuta or "EUR").upper(), riferimento_esterno))

    # ✅ attiva
    attiva_servizio(
        conn=conn,
        servizio_id=servizio_id,
        annuncio_id=annuncio_id,
        utente_id=utente_id,
        durata_giorni=piano["durata_giorni"],
        acquisto_id=acquisto_id,
        attivato_da=metodo
    )


def attiva_pacchetto_da_piano(conn, piano_id, annuncio_id, utente_id, acquisto_id, metodo="admin", importo_totale=0.0, valuta="EUR", riferimento_esterno=None):
    """
    Versione generica (non solo Stripe): attiva tutti i servizi del pacchetto
    e registra lo storico in acquisti_servizi secondo lo schema reale.
    """
    cur = get_cursor(conn)

    cur.execute(sql("""
        SELECT
            ps.servizio_id,
            COALESCE(ps.durata_override, sp.durata_giorni) AS durata_finale
        FROM pacchetti_servizi ps
        JOIN servizi_piani sp
          ON sp.servizio_id = ps.servizio_id
        WHERE ps.pacchetto_id = (
            SELECT pacchetto_id
            FROM pacchetti_piani
            WHERE id = ?
        )
    """), (piano_id,))

    servizi = cur.fetchall()
    if not servizi:
        return

    quota = float(importo_totale) / float(len(servizi)) if len(servizi) > 0 else 0.0

    for s in servizi:
        servizio_id = int(s["servizio_id"])

        # ✅ storico acquisto servizio (schema reale)
        cur.execute(sql("""
            INSERT INTO acquisti_servizi
            (utente_id, servizio_id, metodo, importo, valuta, riferimento_esterno)
            VALUES (?, ?, ?, ?, ?, ?)
        """), (int(utente_id), servizio_id, metodo, quota, (valuta or "EUR").upper(), riferimento_esterno))

        # ✅ attiva
        attiva_servizio(
            conn=conn,
            servizio_id=servizio_id,
            annuncio_id=annuncio_id,
            utente_id=utente_id,
            durata_giorni=s["durata_finale"],
            acquisto_id=acquisto_id,
            attivato_da=metodo
        )

# --- Modifica Profilo Utente ---
@app.route('/utente/modifica', methods=['GET', 'POST'])
@login_required
def modifica_profilo():
    conn = get_db_connection()
    c = get_cursor(conn)

    if request.method == 'POST':
        nome = request.form['nome'].strip()
        cognome = request.form['cognome'].strip()
        citta = request.form['citta'].strip()
        username = request.form['username'].strip()
        nuova_password = request.form.get('nuova_password', '').strip()
        conferma_password = request.form.get('conferma_password', '').strip()

        # 🔹 Controlla che lo username non sia già usato da altri
        c.execute(sql("SELECT id FROM utenti WHERE username = ? AND id != ?"), (username, g.utente['id']))
        altro = c.fetchone()
        if altro:
            flash("Questo username è già in uso. Scegline un altro.")

            return redirect(url_for('modifica_profilo'))

        # 🔹 Gestione cambio password (facoltativo)
        if nuova_password:
            if nuova_password != conferma_password:
                flash("Le password non coincidono.")

                return redirect(url_for('modifica_profilo'))
            hashed_pw = generate_password_hash(nuova_password)
            c.execute(
                "UPDATE utenti SET nome = ?, cognome = ?, citta = ?, username = ?, password = ? WHERE id = ?",
                (nome, cognome, citta, username, hashed_pw, g.utente['id'])
            )
        else:
            c.execute(
                "UPDATE utenti SET nome = ?, cognome = ?, citta = ?, username = ? WHERE id = ?",
                (nome, cognome, citta, username, g.utente['id'])
            )

        conn.commit()

        # 🔹 Aggiorna la sessione con i nuovi dati dell'utente
        session['utente_username'] = username
        session.modified = True


        flash("Profilo aggiornato con successo.")
        return redirect(url_for('dashboard'))

    # GET → mostra dati correnti
    cur = get_cursor(conn)
    cur.execute(sql("SELECT * FROM utenti WHERE id = ?"), (g.utente['id'],))
    utente = cur.fetchone()

    return render_template('modifica_profilo.html', utente=utente)

# ---------------------------
# IMPOSTAZIONI → PROFILO
# ---------------------------

@app.route("/impostazioni")
@login_required
def impostazioni():
    if not session.get("utente_id"):
        return redirect(url_for("login"))

    conn = get_db_connection()
    cur = get_cursor(conn)

    utente = cur.execute(
        sql("""
            SELECT email_notifiche
            FROM utenti
            WHERE id = ?
        """),
        (session["utente_id"],)
    ).fetchone()

    email_notifiche = 1

    if utente:
        email_notifiche = int(utente["email_notifiche"] or 0)

    return render_template(
        "impostazioni.html",
        email_notifiche=email_notifiche
    )

@app.route("/impostazioni/modifica-username", methods=["GET", "POST"])
@login_required
def modifica_username():
    if request.method == "POST":
        verify_csrf()

        nuovo = request.form.get("username", "").strip().upper()   # ✅ SALVA MAIUSCOLO

        if nuovo:
            conn = get_db_connection()
            cur = get_cursor(conn)

            # ✅ controllo duplicati case-insensitive
            cur.execute(sql("SELECT id FROM utenti WHERE UPPER(username)=?"), (nuovo,))
            esistente = cur.fetchone()

            if esistente and esistente[0] != session["utente_id"]:
                flash("Questo ID utente è già stato scelto.", "error")

                return redirect(url_for("modifica_username"))

            cur.execute(
                "UPDATE utenti SET username=? WHERE id=?",
                (nuovo, session["utente_id"])
            )
            conn.commit()


            # ✅ aggiorna sessione
            session['utente_username'] = nuovo
            session.modified = True

            flash("Username aggiornato", "success")
            return redirect(url_for("impostazioni"))

    return render_template("forms/modifica_username.html")

@app.route("/impostazioni/modifica-password", methods=["GET", "POST"])
@login_required
def modifica_password():
    # 🔐 Se l'utente è admin, il cambio password richiede uno step-up passkey recente.
    # Questo evita che una sessione admin già aperta possa cambiare password senza nuova verifica forte.
    if g.utente and "ruolo" in g.utente.keys() and g.utente["ruolo"] == "admin":
        if not admin_stepup_is_valid():
            next_url = url_for("modifica_password")
            flash(
                "Per modificare la password admin devi prima confermare la tua identità con passkey.",
                "warning"
            )
            return redirect(url_for("admin_unlock", next=next_url))

    if request.method == "POST":
        verify_csrf()

        pw_attuale = request.form.get("password_attuale", "")
        nuova_pw = request.form.get("nuova_password", "")
        conferma_pw = request.form.get("conferma_password", "")

        if not nuova_pw or nuova_pw != conferma_pw:
            flash("Le nuove password non coincidono.", "error")
            return redirect(url_for('modifica_password'))

        if len(nuova_pw) < 8:
            flash("La password deve avere almeno 8 caratteri.", "error")
            return redirect(url_for("modifica_password"))

        conn = get_db_connection()

        cur = get_cursor(conn)
        cur.execute(sql("SELECT * FROM utenti WHERE id = ?"), (session["utente_id"],))
        utente = cur.fetchone()

        if not utente:

            flash("Utente non trovato.", "error")
            return redirect(url_for("login"))

        # 🔹 Verifica password attuale
        if not check_password_hash(utente["password"], pw_attuale):

            flash("La password attuale non è corretta.", "error")
            return redirect(url_for("modifica_password"))

        # 🔹 Aggiorna SOLO l'hash della password
        hash_pw = generate_password_hash(nuova_pw)
        cur.execute(
            sql("UPDATE utenti SET password = ? WHERE id = ?"),
            (hash_pw, session["utente_id"])
        )
        conn.commit()

        # 🔐 Se è un admin, il cambio password è un evento critico:
        # invalida tutte le altre sessioni admin sensibili.
        if g.utente and "ruolo" in g.utente.keys() and g.utente["ruolo"] == "admin":
            bump_admin_security_version(
                user_id=session["utente_id"],
                reason="admin_password_changed"
            )

            # Dopo un cambio password, richiediamo un nuovo step-up alla prossima azione admin sensibile.
            clear_admin_stepup()

        flash("Password aggiornata con successo!", "success")
        return redirect(url_for("impostazioni"))

    return render_template("forms/modifica_password.html")


@app.route("/impostazioni/elimina-account")
@login_required
def elimina_account_step1():
    return render_template("impostazioni/elimina_account_step1.html")


@app.route("/impostazioni/sospendi-account", methods=["POST"])
@login_required
def sospendi_account():
    verify_csrf()

    conn = get_db_connection()

    cur = get_cursor(conn)

    # Aggiorna DB
    cur.execute(sql("UPDATE utenti SET sospeso=1 WHERE id=?"), (session["utente_id"],))
    conn.commit()

    # Recupera email e nome per invio notifica
    utente = cur.execute(
        sql("SELECT email, nome FROM utenti WHERE id=?"),
        (session["utente_id"],)
    ).fetchone()


    # Invia email
    invia_email_sospensione(utente["email"], utente["nome"])

    # Flag per messaggio logout
    session["sospensione_logout"] = True

    # Logout automatico
    return redirect(url_for("logout"))


@app.route("/impostazioni/elimina-account/confirm", methods=["GET", "POST"])
@login_required
def elimina_account_step2():
    if request.method == "POST":
        verify_csrf()

        user_id = session.get("utente_id")

        if not user_id:
            flash("Sessione non valida. Effettua nuovamente l'accesso.", "error")
            return redirect(url_for("login"))

        conn = get_db_connection()
        cur = get_cursor(conn)

        try:
            # =====================================================
            # 1) Rimuove notifiche dell'utente
            # =====================================================
            cur.execute(sql("""
                DELETE FROM notifiche
                WHERE id_utente = ?
            """), (user_id,))

            # =====================================================
            # 2) Rimuove eventuali subscription push dell'utente
            # =====================================================
            cur.execute(sql("""
                DELETE FROM push_subscriptions
                WHERE utente_id = ?
            """), (user_id,))

            # =====================================================
            # 3) Rimuove chat e messaggi dell'utente
            #    Così nessuno può più continuare a scrivergli
            # =====================================================
            cur.execute(sql("""
                DELETE FROM messaggi_chat
                WHERE mittente_id = ?
                   OR destinatario_id = ?
            """), (user_id, user_id))

            # Se esiste storico chiusure chat, rimuove anche quello
            cur.execute(sql("""
                DELETE FROM chat_chiusure
                WHERE admin_id = ?
                   OR user_id = ?
            """), (user_id, user_id))

            # =====================================================
            # 4) Scollega pagamenti/acquisti dagli annunci
            #    I pagamenti restano, ma non puntano più agli annunci eliminati
            # =====================================================
            cur.execute(sql("""
                UPDATE acquisti
                SET annuncio_id = NULL
                WHERE annuncio_id IN (
                    SELECT id
                    FROM annunci
                    WHERE utente_id = ?
                )
            """), (user_id,))

            # =====================================================
            # 5) Rimuove attivazioni servizi operative dell'utente
            #
            # Manteniamo gli acquisti/pagamenti nello storico,
            # ma rimuoviamo tutte le attivazioni ancora operative:
            # - servizi legati agli annunci dell'utente
            # - servizi legati direttamente al profilo utente
            # =====================================================

            # 5A) Attivazioni collegate agli annunci dell'utente
            cur.execute(sql("""
                DELETE FROM attivazioni_servizi
                WHERE annuncio_id IN (
                    SELECT id
                    FROM annunci
                    WHERE utente_id = ?
                )
            """), (user_id,))

            # 5B) Attivazioni collegate direttamente al profilo utente
            cur.execute(sql("""
                DELETE FROM attivazioni_servizi
                WHERE utente_id = ?
            """), (user_id,))

            # =====================================================
            # 6) Rimuove annunci dell'utente
            # =====================================================
            cur.execute(sql("""
                DELETE FROM annunci
                WHERE utente_id = ?
            """), (user_id,))

            # =====================================================
            # 7) Gestione recensioni
            #
            # - Le recensioni SCRITTE dall'utente verso altri restano.
            # - Le recensioni RICEVUTE dall'utente cancellato spariscono,
            #   perché il suo profilo non deve più esistere.
            # =====================================================

            # Prima rimuove eventuali risposte collegate alle recensioni ricevute dall'utente
            cur.execute(sql("""
                DELETE FROM risposte_recensioni
                WHERE id_recensione IN (
                    SELECT id
                    FROM recensioni
                    WHERE id_destinatario = ?
                )
            """), (user_id,))

            # Poi rimuove recensioni ricevute dall'utente cancellato
            cur.execute(sql("""
                DELETE FROM recensioni
                WHERE id_destinatario = ?
            """), (user_id,))

            # Eventuali risposte scritte dall'utente su recensioni di altri:
            # meglio rimuoverle perché l'account autore non esiste più come profilo attivo.
            cur.execute(sql("""
                DELETE FROM risposte_recensioni
                WHERE id_autore = ?
            """), (user_id,))

            # =====================================================
            # 8) Anonimizza l'utente invece di cancellare la riga
            #
            # Motivo:
            # - pagamenti/acquisti possono riferirsi ancora a utenti.id
            # - recensioni scritte ad altri possono riferirsi ancora a id_autore
            # - cancellare la riga romperebbe vincoli FK e storico
            # =====================================================
            anonimizzato = f"utente_eliminato_{user_id}"

            cur.execute(sql("""
                UPDATE utenti
                SET
                    nome = 'Utente',
                    cognome = 'eliminato',
                    email = ?,
                    username = ?,
                    password = '',
                    citta = NULL,
                    provincia = NULL,
                    macro_area = NULL,
                    telefono = NULL,
                    email_pubblica = NULL,
                    indirizzo_studio = NULL,
                    sito_web = NULL,
                    instagram = NULL,
                    facebook = NULL,
                    linkedin = NULL,
                    orari = NULL,
                    preferenze_contatto = NULL,
                    frase = NULL,
                    descrizione = NULL,
                    lingue = NULL,
                    foto_profilo = NULL,
                    copertina = NULL,
                    foto_galleria = NULL,
                    visibile_pubblicamente = 0,
                    visibile_in_chat = 0,
                    attivo = 0,
                    sospeso = 1,
                    disattivato_admin = 1,
                    token_verifica = NULL,
                    admin_session_token = NULL,
                    admin_session_expiry = NULL,
                    admin_browser_fingerprint = NULL,
                    x25519_pub = NULL,
                    x25519_priv_enc = NULL,
                    x25519_priv_nonce = NULL,
                    dek_enc = NULL,
                    dek_nonce = NULL,
                    dek_mk_enc = NULL,
                    dek_mk_nonce = NULL
                WHERE id = ?
            """), (
                f"deleted_user_{user_id}@mylocalcare.local",
                anonimizzato,
                user_id
            ))

            conn.commit()

        except Exception as e:
            conn.rollback()
            print("❌ Errore eliminazione account:", repr(e), flush=True)
            traceback.print_exc()
            flash("Errore durante l'eliminazione dell'account. Riprova o contatta l'assistenza.", "error")
            return redirect(url_for("elimina_account_step2"))

        session.clear()
        flash("Account eliminato definitivamente.", "success")
        return redirect(url_for("home"))

    return render_template("impostazioni/elimina_account_step2.html")

@app.route("/impostazioni/riattivazione-account")
@login_required
def riattivazione_account():
    if not session.get("sospeso"):
        return redirect(url_for("dashboard"))

    return render_template("impostazioni/riattiva_account.html")


@app.route("/impostazioni/riattiva-account", methods=["POST"])
@login_required
def riattiva_account():
    verify_csrf()

    user_id = session.get("utente_id")

    conn = get_db_connection()
    cur = get_cursor(conn)

    cur.execute(sql("UPDATE utenti SET sospeso=0 WHERE id=?"), (user_id,))
    conn.commit()

    # Per coerenza con il messaggio e per sicurezza, svuotiamo davvero la sessione
    session.clear()

    flash("Account riattivato! Per motivi di sicurezza effettua di nuovo il login.", "success")
    return redirect(url_for("login"))

# ----------------------------------------
# 🔒 CONTROLLO SOSPENSIONE AUTOMATICO
# ----------------------------------------
@app.before_request
def controllo_sospensione():

    rotte_escluse = [
        'login', 'logout', 'riattivazione_account', 'riattiva_account',
        'static'
    ]

    # Evita loop: non controllare per queste rotte
    if request.endpoint is None or request.endpoint in rotte_escluse:
        return

    # Se non è loggato: ignora
    if not session.get("utente_id"):
        return

    # Controlla stato nel DB
    conn = get_db_connection()

    c = get_cursor(conn)
    c.execute(sql("SELECT sospeso, disattivato_admin FROM utenti WHERE id=?"), (session["utente_id"],))
    stato = c.fetchone()


    # 🔒 Utente sospeso → attiva pagina riattivazione
    if stato and stato["sospeso"] == 1:
        session["sospeso"] = True
        return redirect(url_for("riattivazione_account"))

    # 🚫 Account disattivato dall’admin → blocco totale
    if stato and stato["disattivato_admin"] == 1:
        session.clear()
        flash("Il tuo account è stato disattivato dall’amministrazione.", "error")
        return redirect(url_for("login"))

# ---------------------------
# IMPOSTAZIONI → SICUREZZA
# ---------------------------

@app.route("/impostazioni/notifiche-email", methods=["GET", "POST"])
@login_required
def email_notifiche():
    if request.method == "POST":
        verify_csrf()

        attivo = 1 if request.form.get("email_notifiche") == "on" else 0
        conn = get_db_connection()
        cur = get_cursor(conn)
        cur.execute(sql("UPDATE utenti SET email_notifiche=? WHERE id=?"), (attivo, session["utente_id"]))
        conn.commit()
        flash("Preferenze aggiornate", "success")
        return redirect(url_for("impostazioni"))
    return render_template("forms/email_notifiche.html")

@app.route("/impostazioni/notifiche-email/toggle", methods=["POST"])
@login_required
def toggle_email_notifiche():
    verify_csrf()

    conn = get_db_connection()
    cur = get_cursor(conn)

    row = cur.execute(sql("""
        SELECT email_notifiche
        FROM utenti
        WHERE id = ?
    """), (session["utente_id"],)).fetchone()

    if not row:
        return jsonify({"ok": False, "error": "Utente non trovato"}), 404

    attuale = int(row["email_notifiche"] or 0)
    nuovo = 0 if attuale == 1 else 1

    cur.execute(sql("""
        UPDATE utenti
        SET email_notifiche = ?
        WHERE id = ?
    """), (nuovo, session["utente_id"]))

    conn.commit()

    return jsonify({
        "ok": True,
        "email_notifiche": nuovo
    })

# ---------------------------
# IMPOSTAZIONI → FOTO PROFILO
# ---------------------------

@app.route("/impostazioni/cambia-foto", methods=["GET", "POST"])
@login_required
def cambia_foto():
    if request.method == "POST":
        verify_csrf()

        file = request.files.get("foto")

        if not file or file.filename == "":
            flash("Carica una foto valida.", "error")
            return redirect(url_for("cambia_foto"))

        upload_dir = os.path.join("static", "uploads", "profile")
        os.makedirs(upload_dir, exist_ok=True)

        filename = f"{uuid.uuid4().hex}_{file.filename}"
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)

        # SALVA IL PERCORSO COERENTE
        percorso_db = f"uploads/profile/{filename}"

        conn = get_db_connection()
        cur = get_cursor(conn)
        cur.execute(sql("""
            UPDATE utenti SET foto_profilo=? WHERE id=?
        """), (percorso_db, session["utente_id"]))
        conn.commit()


        flash("Foto profilo aggiornata con successo!", "success")
        return redirect(url_for("impostazioni"))

    return render_template("forms/cambia_foto.html")


@app.route("/impostazioni/cambia-copertina", methods=["GET", "POST"])
@login_required
def cambia_copertina():
    if request.method == "POST":
        verify_csrf()

        file = request.files.get("copertina")

        if not file or file.filename == "":
            flash("Carica una copertina valida.", "error")
            return redirect(url_for("cambia_copertina"))

        upload_dir = os.path.join("static", "uploads", "profili", "copertine")
        os.makedirs(upload_dir, exist_ok=True)

        filename = f"{uuid.uuid4().hex}_{file.filename}"
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)

        percorso_db = f"uploads/profili/copertine/{filename}"

        conn = get_db_connection()
        cur = get_cursor(conn)
        cur.execute(sql("""
            UPDATE utenti SET copertina=? WHERE id=?
        """), (percorso_db, session["utente_id"]))
        conn.commit()


        flash("Copertina aggiornata con successo!", "success")
        return redirect(url_for("impostazioni"))

    return render_template("forms/cambia_copertina.html")

#NUOVO Annuncio
def contiene_contatti_nel_testo(testo):
    valore = str(testo or "")

    email_regex = r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}"
    telefono_regex = r"(?:\+?\d[\s().-]*){8,}"
    parole_contatto_regex = r"\b(?:whatsapp|wa\.me|telegram|tel\.|cellulare|telefono|chiamami|scrivimi al|contattami al)\b"

    return (
        re.search(email_regex, valore, re.IGNORECASE) is not None
        or re.search(telefono_regex, valore) is not None
        or re.search(parole_contatto_regex, valore, re.IGNORECASE) is not None
    )

@app.route("/nuovo-annuncio", methods=["GET", "POST"])
@login_required
@foto_obbligatoria
def nuovo_annuncio():
    if "utente_id" not in session:
        flash("Devi essere loggato per creare un annuncio.", "error")
        return redirect(url_for("login"))

    conn = get_db_connection()

    c = get_cursor(conn)

    # ✅ Verifica che l’utente sia attivo
    c.execute(
        "SELECT * FROM utenti WHERE id = ?",
        (session["utente_id"],)
    )
    utente = c.fetchone()

    if not utente or utente["attivo"] != 1:

        flash("Il tuo account deve essere approvato per pubblicare annunci.", "warning")
        return redirect(url_for("dashboard"))

    # ✅ Carica filtri categoria da DB
    filtri_per_categoria = get_filtri_categoria_da_db()

    # =========================================================
    # 📤 POST
    # =========================================================
    if request.method == "POST":

        # 🔹 CAMPI BASE
        categoria_raw = request.form.get("categoria", "")
        categoria = to_slug(categoria_raw)
        tipo_annuncio = request.form.get("tipo_annuncio", "").strip().lower()
        titolo = request.form.get("titolo", "").strip()
        descrizione = request.form.get("descrizione", "").strip()

        if not categoria:

            flash("Seleziona una categoria.", "warning")
            return redirect(url_for("nuovo_annuncio"))

        if not titolo:

            flash("Inserisci un titolo per l’annuncio.", "warning")
            return redirect(url_for("nuovo_annuncio"))

        if not descrizione:

            flash("Inserisci una descrizione dettagliata.", "warning")
            return redirect(url_for("nuovo_annuncio"))

        if contiene_contatti_nel_testo(descrizione):

            flash(
                "Non inserire telefono, email o altri contatti nella descrizione. "
                "Usa i campi Telefono ed Email nella sezione Dettagli e contatti.",
                "warning"
            )
            return redirect(url_for("nuovo_annuncio"))

        # 🔹 ZONA + PROVINCIA
        zona = request.form.get("zona", "").strip()
        provincia = request.form.get("provincia", "").strip() or None

        # 🔹 ALTRI CAMPI
        filtri = request.form.getlist("filtri_categoria")
        bio = request.form.get("bio_utente", "").strip()
        prezzo = request.form.get("prezzo", "").strip()
        telefono = request.form.get("telefono", "").strip()
        email = request.form.get("email", "").strip()
        username_modificato = request.form.get("username", utente["username"])

        categorie_con_foto_card = {
            "family-kids",
            "eventi-socialita",
            "spazi-sale"
        }

        foto_card_index_raw = request.form.get("foto_card_index", "").strip()
        # =====================================================
        # 🛡️ VALIDAZIONI
        # =====================================================

        if tipo_annuncio not in ("offro", "cerco"):

            flash("Devi selezionare se l’annuncio è 'Offro' oppure 'Cerco'.", "warning")
            return redirect(url_for("nuovo_annuncio"))

        if not zona:

            flash("Seleziona una zona o un comune dall’elenco.", "warning")
            return redirect(url_for("nuovo_annuncio"))

        # 🔒 1 annuncio per categoria per utente
        c.execute(sql("""
            SELECT id
            FROM annunci
            WHERE utente_id = ?
              AND categoria = ?
              AND stato IN ('in_attesa', 'approvato')
            LIMIT 1
        """), (utente["id"], categoria))

        esiste = c.fetchone()

        if esiste:

            flash(
                "Hai già un annuncio in questa categoria (in attesa o approvato). "
                "Per pubblicarne un altro, elimina o modifica quello esistente.",
                "warning"
            )
            return redirect(url_for("dashboard"))

        # =====================================================
        # 📸 UPLOAD MEDIA
        # =====================================================
        media_files = request.files.getlist("media")
        media_paths = []
        upload_dir = os.path.join("static", "uploads", "annunci")
        os.makedirs(upload_dir, exist_ok=True)

        for file in media_files:
            if file and file.filename:
                if not file.mimetype.startswith("image/"):
                    flash("Puoi caricare solo immagini, non video.", "warning")
                    return redirect(url_for("nuovo_annuncio"))

                filename = f"{uuid.uuid4().hex}_{file.filename}"
                file.save(os.path.join(upload_dir, filename))
                media_paths.append(f"uploads/annunci/{filename}")

        foto_card = None

        if categoria in categorie_con_foto_card and media_paths and foto_card_index_raw != "":
            try:
                foto_card_index = int(foto_card_index_raw)

                if 0 <= foto_card_index < len(media_paths):
                    foto_card = media_paths[foto_card_index]

            except ValueError:
                foto_card = None


        # =====================================================
        # 💾 INSERT DB
        # =====================================================
        c.execute(sql("""
            INSERT INTO annunci (
                utente_id,
                username,
                categoria,
                tipo_annuncio,
                titolo,
                descrizione,
                bio_utente,
                zona,
                provincia,
                filtri_categoria,
                media,
                foto_card,
                prezzo,
                telefono,
                email,
                stato
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'in_attesa')
        """), (
            utente["id"],
            username_modificato,
            categoria,
            tipo_annuncio,
            titolo,
            descrizione,
            bio,
            zona,
            provincia,
            ",".join(filtri),
            ",".join(media_paths),
            foto_card,
            prezzo,
            telefono,
            email
        ))

        conn.commit()


        # 👁️ Visibilità pubblica automatica
        conn = get_db_connection()
        conn.execute(
            "UPDATE utenti SET visibile_pubblicamente = 1 WHERE id = ?",
            (utente["id"],)
        )
        conn.commit()


        # 🔔 Aggiorna contatori admin
        invalidate_admin_counters()

        notifica_admin_evento(
            titolo="Nuovo annuncio in attesa",
            messaggio=f"Nuovo annuncio da approvare: {titolo}",
            link=url_for("admin_annunci", stato="in_attesa"),
            push=True
        )

        flash(
            "✅ Annuncio creato! Sarà pubblicato dopo approvazione dell’amministratore.",
            "success"
        )

        return redirect(url_for("dashboard"))

    # =========================================================
    # 📥 GET
    # =========================================================

    return render_template(
        "nuovo_annuncio.html",
        username=utente["username"],
        filtri_per_categoria=filtri_per_categoria
    )

@app.context_processor
def inject_servizi_utils():
    return dict(
        servizio_attivo_per_annuncio=servizio_attivo_per_annuncio,
        servizio_attivo_per_utente=servizio_attivo_per_utente
    )

@app.route("/annuncio/<int:id>")
def visualizza_annuncio_pubblico(id):
    conn = get_db_connection()

    c = get_cursor(conn)

    # 🔹 SQL condivisa: AFFIDABILITÀ TOP (identica a /cerca)
    affidabilita_top_sql = """
        CASE
          -- 🟢 OVERRIDE ADMIN
          WHEN EXISTS (
            SELECT 1
            FROM attivazioni_servizi act
            JOIN servizi s ON s.id = act.servizio_id
            WHERE s.codice = 'badge_affidabilita'
              AND act.utente_id = a.utente_id
              AND act.stato = 'attivo'
          ) THEN 1

          -- ⭐ REGOLA AUTOMATICA
          WHEN
            COALESCE((
              SELECT AVG(r.voto)
              FROM recensioni r
              WHERE r.id_destinatario = a.utente_id
                AND r.stato = 'approvato'
            ), 0) >= 4.2
          AND
            COALESCE((
              SELECT COUNT(*)
              FROM recensioni r
              WHERE r.id_destinatario = a.utente_id
                AND r.stato = 'approvato'
            ), 0) >= 5
          THEN 1

          ELSE 0
        END AS affidabilita_top
    """

    # 🔹 Query annuncio pubblico
    c.execute(sql(f"""
        SELECT
            a.*,
            a.tipo_annuncio,

            u.username,
            u.nome,
            u.cognome,
            a.email AS email,
            a.telefono AS telefono,
            NULL AS email_utente,
            NULL AS telefono_utente,
            u.foto_profilo,

            {affidabilita_top_sql},

            -- ⭐ MEDIA RECENSIONI
            COALESCE(ROUND((
                SELECT AVG(r.voto)
                FROM recensioni r
                WHERE r.id_destinatario = a.utente_id
                  AND r.stato = 'approvato'
            ), 1), 0) AS media_recensioni,

            -- 🔢 NUMERO RECENSIONI
            COALESCE((
                SELECT COUNT(*)
                FROM recensioni r
                WHERE r.id_destinatario = a.utente_id
                  AND r.stato = 'approvato'
            ), 0) AS numero_recensioni

        FROM annunci a
        JOIN utenti u ON a.utente_id = u.id
        WHERE a.id = ?
          AND u.attivo = 1
          AND u.sospeso = 0
          AND (u.disattivato_admin IS NULL OR u.disattivato_admin = 0)
          AND (u.ruolo IS NULL OR u.ruolo != 'admin')
    """), (id,))

    row = c.fetchone()


    # ❌ Annuncio non trovato
    if not row:
        return "Annuncio non trovato", 404

    annuncio = dict(row)
    annuncio["tipo_annuncio"] = (annuncio.get("tipo_annuncio") or "").lower()

    # 🔒 Annuncio non approvato → visibile solo al proprietario
    if annuncio["stato"] != "approvato":
        if not g.utente or g.utente["id"] != annuncio["utente_id"]:
            return "Annuncio non ancora pubblicato.", 403

    # ✅ SERVIZI — COME IN DASHBOARD
    contatti_attivi = servizio_attivo_per_utente(
        utente_id=annuncio["utente_id"],
        codice_servizio="contatti"
    )

    annuncio["contatti_visibili"] = bool(
        contatti_attivi and (
            (annuncio.get("email") or "").strip()
            or
            (annuncio.get("telefono") or "").strip()
        )
    )
    # 🔁 Gestione intelligente del tasto “Torna”
    ref = request.referrer or ""

    if "/modifica" in ref and session.get("last_annuncio_origin"):
        back_url = session["last_annuncio_origin"]
    else:
        if ref and "/modifica" not in ref:
            session["last_annuncio_origin"] = ref
        back_url = ref or session.get("last_annuncio_origin") or url_for("home")

    return render_template(
        "annuncio_pubblico.html",
        annuncio=annuncio,
        contatti_attivi=contatti_attivi,
        back_url=back_url
    )

    # --- Profilo pubblico dell’operatore ---
@app.route("/profilo_operatore/<int:id>")
def profilo_operatore(id):
    conn = get_db_connection()

    c = get_cursor(conn)

    c.execute(sql("SELECT * FROM operatori WHERE id = ?"), (id,))
    operatore = c.fetchone()


    if not operatore:
        return "Operatore non trovato", 404

    return render_template("profilo_operatore.html", operatore=operatore)

from services import servizio_attivo_per_utente

@app.route("/profilo/<int:id>")
def profilo_pubblico(id):
    """Visualizza il profilo pubblico completo di un utente (stile Facebook)."""
    conn = get_db_connection()

    c = get_cursor(conn)

    # 🔹 Carica dati utente
    c.execute(sql("""
        SELECT id, nome, cognome, email, username, citta, lingue, frase,
               telefono, email_pubblica, indirizzo_studio,
               sito_web, instagram, facebook, linkedin,
               orari, preferenze_contatto,
               visibile_pubblicamente, visibile_in_chat,
               media_recensioni, numero_recensioni,
               foto_profilo, copertina, foto_galleria,
               offro_1, offro_2, offro_3, offro_4, offro_5, offro_6, offro_7, offro_8, offro_9, offro_10, offro_11, offro_12, offro_13,
               cerco_1, cerco_2, cerco_3, cerco_4, cerco_5, cerco_6, cerco_7, cerco_8, cerco_9, cerco_10, cerco_11, cerco_12, cerco_13,
               esperienza_1, esperienza_2, esperienza_3,
               studio_1, studio_2, studio_3,
               certificazioni, descrizione
        FROM utenti
        WHERE id = ?
          AND sospeso = 0
          AND (disattivato_admin IS NULL OR disattivato_admin = 0)
          AND attivo = 1
          AND (ruolo IS NULL OR ruolo != 'admin')
    """), (id,))
    utente = c.fetchone()

    if not utente:

        flash("Profilo non trovato.", "error")
        return redirect(url_for("cerca"))

    utente = dict(utente)

    # 🧹 Normalizza None → ""
    for k, v in utente.items():
        if v is None:
            utente[k] = ""

    # 🔒 Controllo visibilità profilo
    if not bool(utente.get("visibile_pubblicamente", 0)):
        if not g.utente or g.utente["id"] != utente["id"]:

            flash("Questo profilo è privato.", "info")
            return redirect(url_for("cerca"))

    # =========================================================
    # 🔑 CONTROLLO SERVIZIO CONTATTI (PUNTO CHIAVE)
    # =========================================================
    servizio_contatti_attivo = servizio_attivo_per_utente(
        utente_id=utente["id"],
        codice_servizio="contatti"
    )

    # 🔹 Recensioni
    recensioni = get_recensioni_utente(id)
    media, totale = calcola_media_recensioni(id)

    # 🟡 Badge Fiducia Top automatico / manuale
    utente["affidabilita_top"] = 1 if (
        servizio_attivo_per_utente(
            utente_id=utente["id"],
            codice_servizio="badge_affidabilita"
        )
        or
        (
            float(media or 0) >= 4
            and int(totale or 0) >= 4
        )
    ) else 0

    # 🔹 Annunci
    c.execute(sql("""
        SELECT id, titolo, categoria, zona, prezzo, descrizione,
               media AS media_img, data_pubblicazione, filtri_categoria
        FROM annunci
        WHERE utente_id = ? AND stato = 'approvato'
        ORDER BY data_pubblicazione DESC
    """), (id,))
    annunci = [dict(r) for r in c.fetchall()]


    # =========================================================
    # 🔹 OFFRO / CERCO
    # =========================================================
    categorie = CATEGORIE_PREFERENZE

    offro_presenti = []
    cerco_presenti = []

    for i in range(1, len(categorie) + 1):
        try:
            utente[f"offro_{i}"] = int(utente.get(f"offro_{i}") or 0)
            utente[f"cerco_{i}"] = int(utente.get(f"cerco_{i}") or 0)
        except Exception:
            utente[f"offro_{i}"] = 0
            utente[f"cerco_{i}"] = 0

    for i, cat in enumerate(categorie, start=1):
        if utente[f"offro_{i}"] == 1:
            offro_presenti.append(cat)
        if utente[f"cerco_{i}"] == 1:
            cerco_presenti.append(cat)

    # =========================================================
    # ✅ RENDER
    # =========================================================
    return render_template(
        "dashboard.html",
        utente=utente,
        annunci=annunci,
        pubblico=True,
        media_recensioni=media,
        totale_recensioni=totale,

        # ✅ Recensioni da mostrare nella tab pubblica
        recensioni_ricevute=recensioni,
        recensioni=recensioni,
        recensioni_pubbliche=recensioni,

        offro_presenti=offro_presenti,
        cerco_presenti=cerco_presenti,
        servizio_contatti_attivo=servizio_contatti_attivo,
        page="profilo"
    )

@app.route("/utente/toggle_visibilita", methods=["POST"])
@login_required
def toggle_visibilita():
    conn = get_db_connection()
    c = get_cursor(conn)

    # Leggi stato attuale
    c.execute(sql("SELECT visibile_pubblicamente FROM utenti WHERE id = ?"), (g.utente["id"],))
    row = c.fetchone()
    if not row:

        flash("Utente non trovato.", "error")
        return redirect(url_for("dashboard"))

    stato_attuale = row["visibile_pubblicamente"]

    # Se vuole passare da visibile → invisibile, controlla che non abbia annunci attivi
    if stato_attuale == 1:
        c.execute(sql("""
            SELECT COUNT(*) FROM annunci
            WHERE utente_id = ? AND stato = 'approvato'
        """), (g.utente["id"],))
        annunci_attivi = fetchone_value(c.fetchone())

        if annunci_attivi > 0:

            flash("⚠️ Non puoi rendere invisibile il profilo mentre hai annunci pubblicati. "
                  "Archivia o elimina prima gli annunci approvati.", "warning")
            return redirect(request.referrer or url_for("dashboard"))

    # Alterna stato
    nuovo_stato = 0 if stato_attuale == 1 else 1
    c.execute(sql("UPDATE utenti SET visibile_pubblicamente = ? WHERE id = ?"), (nuovo_stato, g.utente["id"]))
    conn.commit()


    if nuovo_stato == 1:
        flash("✅ Il tuo profilo è ora visibile pubblicamente.", "success")
    else:
        flash("👁️‍🗨️ Il tuo profilo è ora nascosto dai risultati pubblici.", "info")

    return redirect(request.referrer or url_for("dashboard"))

@app.route("/ricerca-utenti")
def ricerca_utenti():
    conn = get_db_connection()
    c = get_cursor(conn)

    raw_nome = (request.args.get("username") or request.args.get("nome") or "").strip()
    zona = (request.args.get("zona") or "").strip()
    raw_cat = (request.args.get("categoria") or "").strip()

    # 👉 normalizzo categoria a slug
    cat_slug = to_slug(raw_cat)
    cat_index = CATEGORIA_TO_INDEX.get(cat_slug)

    print("DEBUG /ricerca-utenti →", {
        "raw_nome": raw_nome,
        "zona": zona,
        "raw_cat": raw_cat,
        "cat_slug": cat_slug,
        "cat_index": cat_index
    })

    query = """
        SELECT
            u.id, u.username, u.nome, u.cognome, u.citta, u.foto_profilo,
            u.visibile_pubblicamente,
            COALESCE(ROUND(AVG(r.voto), 1), 0) AS media_recensioni,
            COUNT(r.id) AS numero_recensioni
        FROM utenti u
        LEFT JOIN recensioni r
               ON r.id_destinatario = u.id AND r.stato = 'approvato'
        WHERE (u.visibile_pubblicamente = 1
               OR (u.visibile_in_chat = 1 AND u.id IN (
                   SELECT DISTINCT CASE
                       WHEN mittente_id = ? THEN destinatario_id
                       WHEN destinatario_id = ? THEN mittente_id
                   END
                   FROM messaggi_chat
                   WHERE mittente_id = ? OR destinatario_id = ?
               )))
          AND u.sospeso = 0
          AND (u.disattivato_admin IS NULL OR u.disattivato_admin = 0)
          AND u.attivo = 1
          AND u.foto_profilo IS NOT NULL
          AND u.foto_profilo != ''
          AND (u.ruolo IS NULL OR u.ruolo != 'admin')
    """

    params = [session.get("utente_id")] * 4

    # 🔍 filtro username
    if raw_nome:
        raw_nome_norm = raw_nome.lower()

        # Se l'utente inserisce una sola lettera, cerco SOLO username che iniziano con quella lettera.
        # Esempio: "b" trova "bubu", "barbara", "bruno", ma non "giangela".
        if len(raw_nome_norm) == 1:
            like = f"{raw_nome_norm}%"
        else:
            # Se inserisce più caratteri, cerco username che contengono quel testo.
            # Esempio: "bu" trova "bubu_topper_1000".
            like = f"%{raw_nome_norm}%"

        query += """
            AND LOWER(u.username) LIKE ?
        """
        params.append(like)

    # 📍 filtro zona
    if zona:
        query += " AND LOWER(u.citta) LIKE ?"
        params.append(f"%{zona.lower()}%")

    # ✅ filtro categoria CORRETTO (offro_X / cerco_X)
    if cat_index:
        query += f"""
            AND (
                u.offro_{cat_index} = 1
                OR u.cerco_{cat_index} = 1
            )
        """

    query += " GROUP BY u.id ORDER BY media_recensioni DESC"

    c.execute(sql(query), params)
    utenti = c.fetchall()


    logged_in = "utente_id" in session

    return render_template(
        "ricerca_utenti.html",
        utenti=utenti,
        logged_in=logged_in,
        nome=raw_nome,
        zona=zona,
        categoria=cat_slug,
        CATEGORY_MAP=CATEGORY_MAP
    )

# ==========================================================
# 6️⃣ CHAT TRA UTENTI
# ==========================================================
@app.route("/chat")
@login_required
def chat_threads_view():
    """Mostra tutte le chat dell’utente loggato"""
    threads = chat_threads(g.utente["id"])
    return render_template(
        "chat_threads.html",
        threads=threads,
        utente=g.utente,
        my_id=g.utente["id"]  # 👈 così il template sa qual è il mio id
    )

@app.route("/chat/threads_json")
@login_required
def chat_threads_json():
    """Restituisce i thread chat aggiornati in formato JSON per aggiornamenti live."""
    rows = chat_threads(g.utente["id"])   # 👈 come avevi prima
    threads = []

    for t in rows:
        d = dict(t)

        # 🔒 Se l'altro utente è l'admin → maschera i dati
        if is_admin(d.get("other_id")):
            d["other_nome"] = "MyLocalCare • Supporto"
            d["other_username"] = "support"
            d["other_foto"] = "img/support.png"

        threads.append(d)

    return jsonify(threads)

@app.route("/chat/<int:other_id>/json")
@login_required
def chat_conversazione_json(other_id):
    """Restituisce i messaggi della chat in formato JSON (per aggiornamenti live)."""
    user_id = g.utente["id"]
    after_id = request.args.get("after_id", type=int)

    # 🔹 Messaggi
    messaggi = chat_conversazione(user_id, other_id, after_id=after_id)
    chat_segna_letti(user_id, other_id)

    # 🔹 Recupero info "altro utente" per nome/avatar
    conn = get_db_connection()

    c = get_cursor(conn)
    c.execute(sql("""
        SELECT id, nome, cognome, username, foto_profilo
        FROM utenti
        WHERE id = ?
    """), (other_id,))
    altro = c.fetchone()


    # 🔒 Maschera admin
    if altro and is_admin(altro["id"]):
        other_display_name = "MyLocalCare • Supporto"
        other_avatar = "static/img/support.png"
    elif altro:
        other_display_name = f"{altro['nome']} {altro['cognome']}"
        other_avatar = altro["foto_profilo"]
    else:
        other_display_name = ""
        other_avatar = None

    def get_val(m, *keys):
        for k in keys:
            if k in m.keys():
                return m[k]
        return ""

    return jsonify({
        "ok": True,
        "me_id": user_id,
        "other_display_name": other_display_name,  # 👈 AGGIUNTO
        "other_avatar": other_avatar,              # 👈 AGGIUNTO
        "messages": [
            {
                "id": m["id"],
                "mittente_id": m["mittente_id"],
                "destinatario_id": m["destinatario_id"],
                "testo": m["testo"],
                "created_at": m["created_at"],
                "consegnato": m["consegnato"],
                "letto": m["letto"]
            }
            for m in messaggi
        ],
        "typing": typing_state.get((other_id, user_id), False)
    })

@app.route("/chat/<int:other_id>/older")
@login_required
def chat_load_older(other_id):

    user_id = g.utente["id"]
    before_id = request.args.get("before_id", type=int)

    if not before_id:
        return jsonify([])

    messaggi = chat_conversazione(
        user_id,
        other_id,
        limit=35,
        before_id=before_id
    )

    return jsonify([
        {
            "id": m["id"],
            "mittente_id": m["mittente_id"],
            "destinatario_id": m["destinatario_id"],
            "testo": m["testo"],
            "created_at": m["created_at"],
            "consegnato": m["consegnato"],
            "letto": m["letto"]
        }
        for m in messaggi
    ])

@app.route("/chat/unread_count")
def chat_unread_count():

    user = getattr(g, "utente", None)

    if not user:
        return jsonify({"count": 0})

    return jsonify({
        "count": chat_count_unread(user["id"])
    })

@app.route("/chat/<int:other_id>")
@login_required
@foto_obbligatoria
def chat_conversazione_view(other_id):
    """Mostra la pagina della chat tra l’utente loggato e un altro utente."""
    conn = get_db_connection()

    c = get_cursor(conn)

    # 🔹 Recupera l’altro utente (solo se non sospeso / non disattivato)
    c.execute(sql("""
        SELECT id, nome, cognome, username, foto_profilo
        FROM utenti
        WHERE id = ?
          AND sospeso = 0
          AND (disattivato_admin IS NULL OR disattivato_admin = 0)
          AND attivo = 1
    """), (other_id,))
    altro = c.fetchone()
    if not altro:

        return "Utente non disponibile", 404

    # 🔒 Maschera l'admin verso gli altri utenti
    if is_admin(altro["id"]):
        altro = dict(altro)
        altro["nome"] = "MyLocalCare"
        altro["cognome"] = "Supporto"
        altro["username"] = "support"
        altro["foto_profilo"] = "img/support.png"

    # 🔹 Recupera i messaggi esistenti
    messaggi = chat_conversazione(g.utente["id"], other_id)


    # 🔹 Segna come letti i messaggi ricevuti
    chat_segna_letti(g.utente["id"], other_id)
    socketio.emit(
        'update_unread_count',
        {'count': chat_count_unread(g.utente["id"])},
        room=f"user_{g.utente['id']}"
    )
    return render_template(
        "chat_conversazione.html",
        altro=altro,
        conversazione=messaggi,
        utente=g.utente,
        is_support=is_admin(other_id)
    )

typing_state = {}
pagina_attiva = {}

@app.route("/chat/<int:other_id>/chiudi", methods=["POST"])
@admin_required
def chiudi_chat(other_id):
    admin_id = g.utente["id"]

    conn = get_db_connection()

    conn.execute(sql(f"""
        INSERT INTO chat_chiusure (admin_id, user_id, closed_at)
        VALUES (?, ?, {now_sql()})
    """), (admin_id, other_id))

    conn.commit()


    flash("Chat chiusa correttamente.", "success")
    return redirect(url_for("utente_messaggi"))

def emit_incoming_call_later(caller_id, destinatario_id, room_name, room_url, delay=0.6):
    """
    Invia la chiamata con un piccolo ritardo per dare tempo
    alla socket del destinatario di riconnettersi/entrare nella room user_X.
    """
    socketio.sleep(delay)

    socketio.emit(
        "video_call_incoming",
        {
            "from": caller_id,
            "room_name": room_name,
            "room_url": room_url
        },
        room=f"user_{destinatario_id}"
    )

    socketio.emit(
        "video_busy",
        {
            "user_id": caller_id,
            "busy": True
        },
        room=f"user_{destinatario_id}"
    )

# ==========================================================
# 🎥 VIDEO CALLS — IMPOSTAZIONE GLOBALE ON/OFF
# ==========================================================
def is_video_calls_enabled():
    """
    Legge lo stato globale delle videochiamate.

    La tabella app_settings deve già esistere:
    - in produzione viene creata/aggiornata tramite init_db/migrazione;
    - nel runtime non eseguiamo CREATE TABLE.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        cur.execute(
            sql("SELECT valore FROM app_settings WHERE chiave = ?"),
            ("video_calls_enabled",)
        )
        row = cur.fetchone()

        # default: attive se la chiave non è ancora stata impostata
        if not row:
            return True

        return str(row["valore"]).strip() == "1"

    except Exception as e:
        log_exception_safe(
            "❌ Errore lettura app_settings.video_calls_enabled",
            e,
            production=True
        )

        # Fail-safe: se la tabella o la lettura hanno problemi,
        # meglio bloccare temporaneamente le videochiamate invece di aprirle senza controllo.
        return False

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass


def set_video_calls_enabled(enabled: bool):
    """
    Aggiorna lo stato globale delle videochiamate.

    La tabella app_settings deve già esistere:
    nel runtime non eseguiamo CREATE TABLE.
    """
    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        if app.config.get("IS_POSTGRES"):
            cur.execute("""
                INSERT INTO app_settings (chiave, valore)
                VALUES (%s, %s)
                ON CONFLICT (chiave)
                DO UPDATE SET valore = EXCLUDED.valore
            """, (
                "video_calls_enabled",
                "1" if enabled else "0"
            ))
        else:
            cur.execute("""
                INSERT INTO app_settings (chiave, valore)
                VALUES (?, ?)
                ON CONFLICT(chiave)
                DO UPDATE SET valore = excluded.valore
            """, (
                "video_calls_enabled",
                "1" if enabled else "0"
            ))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise

    finally:
        try:
            cur.close()
        except Exception:
            pass

        try:
            conn.close()
        except Exception:
            pass

@app.route("/admin/video-calls/toggle", methods=["POST"])
@admin_required
def admin_video_calls_toggle():
    enabled = request.form.get("enabled") == "1"

    set_video_calls_enabled(enabled)

    flash(
        "Videochiamate abilitate." if enabled else "Videochiamate disabilitate.",
        "success"
    )
    return redirect(url_for("admin_video_calls"))

@app.route("/video/start", methods=["POST"])
@login_required
def video_start():

    data = request.get_json()
    altro_id = data.get("altro_utente_id")

    if not altro_id:
        return jsonify({"error": "Utente non valido"}), 400

    if not is_video_calls_enabled():
        return jsonify({
            "error": "Videochiamate attualmente non disponibili."
        }), 503

    from datetime import datetime
    import time
    import requests

    conn = get_db_connection()
    cur = get_cursor(conn)

    try:
        # 🚫 BLOCCO SE CHIAMANTE GIÀ IN CHIAMATA ATTIVA
        cur.execute(sql(f"""
            SELECT id
            FROM video_call_log
            WHERE in_corso = 1
              AND (utente_1 = ? OR utente_2 = ?)
              AND last_ping IS NOT NULL
              AND last_ping >= {sql_now_minus_seconds(60)}
            LIMIT 1
        """), (g.utente["id"], g.utente["id"]))

        call_in_corso = cur.fetchone()

        if call_in_corso:
            return jsonify({
                "error": "Sei già in una videochiamata in corso."
            }), 409

        # 🚫 BLOCCO SE ANCHE IL DESTINATARIO È GIÀ IN CHIAMATA ATTIVA
        cur.execute(sql(f"""
            SELECT id
            FROM video_call_log
            WHERE in_corso = 1
              AND (utente_1 = ? OR utente_2 = ?)
              AND last_ping IS NOT NULL
              AND last_ping >= {sql_now_minus_seconds(60)}
            LIMIT 1
        """), (altro_id, altro_id))

        altro_in_chiamata = cur.fetchone()

        if altro_in_chiamata:
            return jsonify({
                "error": "L'utente è già in una videochiamata in corso."
            }), 409

        # 🔹 Recupero utenti
        me = cur.execute(
            "SELECT maggiorenne_verificato FROM utenti WHERE id = ?",
            (g.utente["id"],)
        ).fetchone()

        altro = cur.execute(
            "SELECT maggiorenne_verificato FROM utenti WHERE id = ?",
            (altro_id,)
        ).fetchone()

        if not altro:
            return jsonify({"error": "Utente non trovato"}), 404

        # 🔒 CONTROLLO BUDGET
        mese_corrente = datetime.now().strftime("%Y-%m")

        limite = cur.execute(sql("""
            SELECT bloccato
            FROM video_limiti_mensili
            WHERE mese = ?
        """), (mese_corrente,)).fetchone()

        if limite and limite["bloccato"] == 1:
            return jsonify({
                "error": "Il servizio video è temporaneamente sospeso per questo mese."
            }), 403

        # 🔞 VERIFICA MAGGIORENNE
        if me["maggiorenne_verificato"] != 1:
            return jsonify({"need_verifica": True}), 200

        # 🎥 CREAZIONE ROOM
        room_name = f"lc_{g.utente['id']}_{altro_id}_{int(time.time())}"

        headers = {
            "Authorization": f"Bearer {os.getenv('DAILY_API_KEY')}",
            "Content-Type": "application/json"
        }

        payload = {
            "name": room_name,
            "properties": {"max_participants": 2}
        }

        r = requests.post(
            "https://api.daily.co/v1/rooms",
            headers=headers,
            json=payload,
            timeout=5
        )

        if r.status_code != 200:
            return jsonify({"error": "Errore creazione room Daily"}), 500

        room_url = r.json()["url"]

        # 📝 LOG CHIAMATA
        cur.execute(sql(f"""
            INSERT INTO video_call_log (
                room_name,
                utente_1,
                utente_2,
                in_corso,
                last_ping
            )
            VALUES (?, ?, ?, 1, {now_sql()})
        """), (room_name, g.utente["id"], altro_id))

        conn.commit()

    finally:
        try:
            cur.close()
        except:
            pass
        try:
            conn.close()
        except:
            pass

    # 📞 NOTIFICA CHIAMATA CON PICCOLO RITARDO
    socketio.start_background_task(
        emit_incoming_call_later,
        g.utente["id"],
        altro_id,
        room_name,
        room_url,
        0.6
    )

    return jsonify({
        "room_name": room_name,
        "room_url": room_url
    })

@app.route("/video/verifica-maggiorenne", methods=["POST"])
@login_required
def verifica_maggiorenne():

    ip = request.remote_addr

    conn = get_db_connection()
    conn.execute(sql(f"""
        UPDATE utenti
        SET maggiorenne_verificato = 1,
            data_verifica_maggiorenne = {now_sql()},
            ip_verifica_maggiorenne = ?,
            versione_consenso = ?
        WHERE id = ?
    """), (ip, "v1.0_video", g.utente["id"]))

    conn.commit()


    return jsonify({"success": True})

@app.route("/video/check-maggiorenne")
@login_required
def video_check_maggiorenne():
    conn = get_db_connection()
    me = conn.execute(
        "SELECT maggiorenne_verificato FROM utenti WHERE id = ?",
        (g.utente["id"],)
    ).fetchone()

    return jsonify({
        "verified": me["maggiorenne_verificato"] == 1
    })

@app.route("/video/end", methods=["POST"])
def video_end():

    import json

    data = request.get_json(silent=True)

    if not data:
        try:
            data = json.loads(request.data.decode("utf-8") or "{}")
        except Exception:
            data = {}

    room_name = data.get("room_name")

    if not room_name:
        return jsonify({"error": "Room non valida"}), 400

    from datetime import datetime, timezone
    conn = get_db_connection()
    cur = get_cursor(conn)

    # 🔒 Prendi SOLO chiamata ancora attiva
    call = cur.execute(sql("""
        SELECT id, created_at
        FROM video_call_log
        WHERE room_name = ?
          AND in_corso = 1
        LIMIT 1
    """), (room_name,)).fetchone()

    if not call:
        return jsonify({"status": "already_closed"})

    # 🔥 created_at ora è datetime vero (TIMESTAMPTZ)
    start_time = call["created_at"]
    end_time = datetime.now(timezone.utc)

    durata_secondi = int((end_time - start_time).total_seconds())

    import math
    participant_minutes = math.ceil((durata_secondi / 60) * 2)

    FREE_MONTHLY = 10000
    COSTO_PER_MINUTO = 0.002

    mese = datetime.now().strftime("%Y-%m")

    used = cur.execute(sql("""
        SELECT minuti_totali
        FROM video_limiti_mensili
        WHERE mese = ?
    """), (mese,)).fetchone()

    used_minutes = used["minuti_totali"] if used else 0

    over_free = max(0, (used_minutes + participant_minutes) - FREE_MONTHLY)
    costo_cent = int(over_free * COSTO_PER_MINUTO * 100)

    # 🔴 CHIUSURA DEFINITIVA
    cur.execute(sql("""
        UPDATE video_call_log
        SET durata_secondi = ?,
            participant_minutes = ?,
            costo_stimato_cent = ?,
            in_corso = 0,
            ended_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """), (durata_secondi, participant_minutes, costo_cent, call["id"]))

    # 📊 UPDATE LIMITE
    cur.execute(sql("""
        INSERT INTO video_limiti_mensili (mese, minuti_totali, costo_totale_cent)
        VALUES (?, ?, ?)
        ON CONFLICT(mese) DO UPDATE SET
            minuti_totali = video_limiti_mensili.minuti_totali + EXCLUDED.minuti_totali,
            costo_totale_cent = video_limiti_mensili.costo_totale_cent + EXCLUDED.costo_totale_cent
    """), (mese, participant_minutes, costo_cent))

    # 🟢 Recupera utenti
    users = cur.execute(sql("""
        SELECT utente_1, utente_2
        FROM video_call_log
        WHERE id = ?
    """), (call["id"],)).fetchone()

    conn.commit()

    # 🟢 Notifica DOPO commit
    if users:
        socketio.emit("video_busy",
            {"user_id": users["utente_1"], "busy": False},
            room=f"user_{users['utente_2']}"
        )
        socketio.emit("video_busy",
            {"user_id": users["utente_2"], "busy": False},
            room=f"user_{users['utente_1']}"
        )

        # 🔴 Chiusura forzata per tutti nella room video
        socketio.emit(
            "video_call_ended",
            {"room": room_name},
            room=room_name
        )

    return jsonify({"status": "ok"})

# ==========================================================
# ❤️ PING CHIAMATA (mantiene viva la call)
# ==========================================================
@app.route("/video/ping", methods=["POST"])
def video_ping():

    room_name = request.json.get("room_name")

    if not room_name:
        return jsonify({"error": "Room non valida"}), 400

    conn = get_db_connection()

    conn.execute(sql(f"""
        UPDATE video_call_log
        SET last_ping = {now_sql()}
        WHERE room_name = ?
        AND in_corso = 1
    """), (room_name,))

    conn.commit()


    return jsonify({"ok": True})

# =====================================================
# 📞 CHIAMATA IN ARRIVO
# =====================================================

@socketio.on("video_call_left")
def handle_video_call_left(data):
    room_name = data.get("room")
    if not room_name:
        return

    # manda SOLO ai partecipanti della room
    socketio.emit(
        "video_call_left",
        {"room": room_name},
        room=room_name,
        include_self=False
    )


@socketio.on("join_video_room")
def handle_join_video_room(data):
    room_name = data.get("room")
    if not room_name:
        return
    join_room(room_name)

@socketio.on("video_busy")
def handle_video_busy(data):

    from flask import session

    user_id = session.get("utente_id")
    busy = bool(data.get("busy"))

    if not user_id:
        return

    socketio.emit(
        "video_busy",
        {
            "user_id": user_id,
            "busy": busy
        },
        include_self=False
    )

@socketio.on("video_call_rejected")
def handle_video_call_rejected(data):
    from flask import session

    room_name = data.get("room")
    from_user = data.get("from_user")

    if not room_name or not from_user:
        return

    # invia SOLO al chiamante
    socketio.emit(
        "video_call_rejected",
        {"room": room_name},
        room=f"user_{from_user}"
    )

    # 🔥 Notifica anche la room video (se qualcuno fosse già entrato)
    socketio.emit(
        "force_call_end",
        {"room": room_name},
        room=room_name
    )

# ==========================================================
# 🔴 EVENTI SOCKET.IO — CHAT IN TEMPO REALE
# ==========================================================


@socketio.on("check_video_status")
def check_video_status(data):
    user_id = data.get("user_id")

    conn = None
    try:
        conn = get_db_connection()

        conn.execute(sql(f"""
            UPDATE video_call_log
            SET in_corso = 0,
                ended_at = CURRENT_TIMESTAMP
            WHERE in_corso = 1
              AND last_ping IS NOT NULL
              AND last_ping < {sql_now_minus_seconds(60)}
        """))
        conn.commit()

        call = conn.execute(sql("""
            SELECT id FROM video_call_log
            WHERE in_corso = 1
            AND (utente_1 = ? OR utente_2 = ?)
            LIMIT 1
        """), (user_id, user_id)).fetchone()

        emit("video_status_result", {
            "busy": bool(call)
        })

    finally:
        if conn:
            try:
                conn.close()
            except:
                pass



def chat_count_unread(user_id):

    conn = get_db_connection()
    cur = get_cursor(conn)

    cur.execute(sql("""
        SELECT COUNT(*) AS count
        FROM messaggi_chat
        WHERE destinatario_id = ?
        AND letto = 0
    """), (user_id,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return 0

    return row["count"]

# ==========================================================
# 🔐 DB BOOTSTRAP — disattivato in runtime produzione
# ==========================================================
# Le tabelle/colonne di sicurezza admin devono essere create tramite migrazione
# o script dedicato, non automaticamente all'avvio dell'app.
#
# Tabelle/colonne richieste già verificate su Render:
# - utenti.admin_security_version
# - admin_passkeys
# - admin_recovery_codes
#
# Nota:
# le funzioni ensure_admin_* restano nel codice solo come riferimento/migrazione manuale,
# ma non vengono più chiamate automaticamente dal runtime web.

if app.config["IS_REALTIME_SERVER"]:
    register_socket_lifecycle_handlers(socketio, redis_client, chat_count_unread)

    register_chat_socket_handlers(
        socketio,
        app,
        get_db_connection=get_db_connection,
        get_cursor=get_cursor,
        sql=sql,
        chat_invia=chat_invia,
        chat_segna_letti=chat_segna_letti,
        emit_to_user_sids=emit_to_user_sids,
        chat_count_unread=chat_count_unread,
        set_open_chat=set_open_chat,
        get_open_chat=get_open_chat,
        clear_open_chat=clear_open_chat,
        invia_push=invia_push,
        recently_read_timers=recently_read_timers,
    )

    print("✅ Realtime handlers registrati (runtime=realtime)")
else:
    print("ℹ️ Realtime handlers NON registrati (runtime=web)")


@app.route("/chat-debug-page-open", methods=["POST"])
@login_required
def chat_debug_page_open():
    if os.getenv("APP_ENV", "production").lower() not in ("local", "development"):
        abort(404)

    try:
        data = request.get_json(silent=True) or {}

        privacy_debug("chat page open", {
            "user_id": session.get("utente_id"),
            "marker": data.get("marker"),
            "pathname": data.get("pathname")
        })

        return {"ok": True}, 200

    except Exception as e:
        privacy_debug("errore chat_debug_page_open", repr(e))
        return {"ok": False}, 500

@app.route("/chat-debug-socket-event", methods=["POST"])
@login_required
def chat_debug_socket_event():
    if os.getenv("APP_ENV", "production").lower() not in ("local", "development"):
        abort(404)

    try:
        data = request.get_json(silent=True) or {}

        privacy_debug("chat socket event", {
            "user_id": session.get("utente_id"),
            "event": data.get("event"),
            "pathname": data.get("pathname"),
            "page_id": data.get("page_id")
        })

        return {"ok": True}, 200

    except Exception as e:
        privacy_debug("errore chat_debug_socket_event", repr(e))
        return {"ok": False}, 500

@app.route("/webhook/stripe", methods=["POST"])
def webhook_stripe():
    try:
        security_log(
            "💳 [WEBHOOK] START /webhook/stripe",
            production=True
        )

        payload = request.data
        sig_header = request.headers.get("Stripe-Signature")
        endpoint_secret = os.environ.get("STRIPE_WEBHOOK_SECRET")

        security_log(
            "💳 [WEBHOOK] configurazione richiesta Stripe",
            {
                "sig_header_present": bool(sig_header),
                "endpoint_secret_present": bool(endpoint_secret)
            },
            production=True
        )

        if not endpoint_secret:
            print("❌ STRIPE_WEBHOOK_SECRET non trovato", flush=True)
            return "Webhook secret missing", 500

        try:
            event = stripe.Webhook.construct_event(
                payload=payload,
                sig_header=sig_header,
                secret=endpoint_secret
            )
            security_log(
                "💳 [WEBHOOK] evento Stripe ricevuto",
                {
                    "event_type": event["type"]
                },
                production=True
            )

        except ValueError as e:
            security_log(
                "❌ [WEBHOOK] payload Stripe non valido",
                {
                    "error_type": type(e).__name__,
                    "error": repr(e)
                },
                production=True
            )

            if os.getenv("APP_ENV", "production").lower() in ("local", "development"):
                traceback.print_exc()

            return "Invalid payload", 400

        except stripe.error.SignatureVerificationError as e:
            security_log(
                "❌ [WEBHOOK] firma Stripe non valida",
                {
                    "error_type": type(e).__name__,
                    "error": repr(e)
                },
                production=True
            )

            if os.getenv("APP_ENV", "production").lower() in ("local", "development"):
                traceback.print_exc()

            return "Invalid signature", 400

        if event["type"] == "payment_intent.succeeded":
            payment_intent = event["data"]["object"]
            security_log(
                "💳 [WEBHOOK] gestione pagamento confermato avviata",
                production=True
            )
            gestisci_pagamento_confermato(payment_intent)
            security_log(
                "✅ [WEBHOOK] gestione pagamento confermato completata",
                production=True
            )

        return "ok", 200

    except Exception as e:
        security_log(
            "❌ [WEBHOOK] eccezione non gestita",
            {
                "error_type": type(e).__name__,
                "error": repr(e)
            },
            production=True
        )

        if os.getenv("APP_ENV", "production").lower() in ("local", "development"):
            traceback.print_exc()

        return "Webhook error", 500

@app.route("/uploads/<path:filename>")
def uploaded_files(filename):
    return send_from_directory("/uploads", filename)

# =====================================================
# 🧹 AUTO CLEANUP VIDEO CALL FANTASMA
# =====================================================

from datetime import datetime, timedelta

def cleanup_video_calls():
    while True:
        conn = None
        try:
            with app.app_context():
                conn = get_db_connection()
                cur = get_cursor(conn)

                # 1️⃣ Trova call zombie
                zombies = cur.execute(sql(f"""
                    SELECT id, room_name, utente_1, utente_2
                    FROM video_call_log
                    WHERE in_corso = 1
                      AND last_ping IS NOT NULL
                      AND last_ping < {sql_now_minus_seconds(60)}
                """)).fetchall()

                if zombies:
                    # 2️⃣ Chiudi realmente le call
                    cur.execute(sql(f"""
                        UPDATE video_call_log
                        SET in_corso = 0,
                            ended_at = CURRENT_TIMESTAMP
                        WHERE in_corso = 1
                          AND last_ping IS NOT NULL
                          AND last_ping < {sql_now_minus_seconds(60)}
                    """))

                    conn.commit()

                    # 3️⃣ Notifica agli utenti che non sono più occupati
                    for z in zombies:
                        u1 = z["utente_1"]
                        u2 = z["utente_2"]

                        socketio.emit(
                            "video_busy",
                            {"user_id": u1, "busy": False},
                            room=f"user_{u2}"
                        )
                        socketio.emit(
                            "video_busy",
                            {"user_id": u2, "busy": False},
                            room=f"user_{u1}"
                        )

        except Exception as e:
            print("Errore cleanup video:", e)

        finally:
            # 🔥 QUESTO ERA IL PROBLEMA
            if conn:
                try:
                    conn.close()
                except:
                    pass

        # 🟢 Yield cooperativo per eventlet
        socketio.sleep(30)

# 🔥 Avvia cleanup UNA SOLA VOLTA all’avvio del worker
socketio.start_background_task(cleanup_video_calls)


# ==========================================================
# 7️⃣ AVVIO SERVER
# ==========================================================
if __name__ == "__main__":
    socketio.run(
        app,
        host="127.0.0.1",
        port=5050,
        debug=True,
        use_reloader=False
    )
